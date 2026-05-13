import openpyxl
import sys

GOV_PATH = r"C:\Users\scott\OneDrive - NorthMarq Capital, LLC\Team Briggs - Documents\Gv't Leased Research\Copy Government Master Document.xlsx"
DIA_PATH = r"C:\Users\scott\OneDrive - NorthMarq Capital, LLC\Team Briggs - Documents\Dialysis Research\Comps\Dialysis Comp Work MASTER.xlsx"

def list_sheets(path, label):
    wb = openpyxl.load_workbook(path, data_only=True, read_only=False)
    print(f"=== {label} ({len(wb.sheetnames)} sheets) ===")
    for s in wb.sheetnames:
        print(f"  - {s}")
    print()
    return wb

if __name__ == "__main__":
    target = sys.argv[1] if len(sys.argv) > 1 else "both"
    if target in ("gov", "both"):
        list_sheets(GOV_PATH, "GOV MASTER")
    if target in ("dia", "both"):
        list_sheets(DIA_PATH, "DIA MASTER")
