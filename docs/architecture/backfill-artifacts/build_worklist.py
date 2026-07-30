import json, datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

rows = json.load(open('/tmp/nomatch.json'))

# priority order for business (actionable) categories
PRIO = {
 "Client/Operator (dialysis)":1, "Cooperating Broker":2, "Buyer/Capital":3,
 "Title/Escrow":4, "Legal":5, "Lender/Bank":6, "Consultant/Env":7,
 "Government":8, "Other Business":9,
}
EXCLUDE = {"Personal (likely)", "Noise/Automated"}
ROLE_LOCALS = {"admin","info","team","notices","acquisitions","prorations","scheduling",
 "account","support","reply","alert","fcptdealteam","nationalnetlease","analystteam",
 "coteamfalcon","team-albert","isg","fanv-did-dl-cad"}

def suggest_name(email):
    local = email.split('@')[0]
    base = local.split('+')[0]
    if base.lower() in ROLE_LOCALS or base.isdigit():
        return "(shared/role mailbox - verify)"
    import re
    cleaned = re.sub(r'\d+$','', base)
    parts = re.split(r'[._\-]', cleaned)
    parts = [p for p in parts if p and not p.isdigit()]
    if not parts:
        return "(verify)"
    if len(parts)==1 and len(parts[0])>2:
        return parts[0].capitalize() + " (verify first name)"
    name = " ".join(p.capitalize() for p in parts if len(p)>=1)
    return name + " (verify)"

def org_from_domain(domain):
    d = domain.lower()
    m = {
     "firstam.com":"First American Title","fnf.com":"Fidelity National Financial",
     "ctt.com":"Chicago Title","stewart.com":"Stewart Title","ipx1031.com":"IPX1031",
     "cbre.com":"CBRE","stanjohnsonco.com":"Stan Johnson Co. (Northmarq)","jll.com":"JLL",
     "colliers.com":"Colliers","kidder.com":"Kidder Mathews","nmrk.com":"Newmark",
     "srsrealestatepartners.com":"SRS Real Estate","logiccre.com":"Logic CRE",
     "davita.com":"DaVita","arikkan.com":"Arikkan (DaVita RE)",
     "exchangeright.com":"ExchangeRight","fcpt.com":"Four Corners (FCPT)",
     "easterlyreit.com":"Easterly Government Properties","boydwatterson.com":"Boyd Watterson",
     "iracapital.com":"IRA Capital","stablewoodproperties.com":"Stablewood Properties",
     "presidiobay.com":"Presidio Bay","brickcapitalre.com":"Brick Capital",
     "jrwrealty.com":"JRW Realty","schusterdevco.com":"Schuster DevCo",
     "caspianrealty.com":"Caspian Realty","adler-realty.com":"Adler Realty",
     "adler-industrial.com":"Adler Industrial","valuenetlease.com":"ValueNet Lease",
     "foley.com":"Foley & Lardner","ltglegal.com":"LTG Legal","buchalter.com":"Buchalter",
     "hinshawlaw.com":"Hinshaw & Culbertson","kutakrock.com":"Kutak Rock",
     "briskinlaw.com":"Briskin Law","dewittllp.com":"DeWitt LLP","leo-law.com":"Leo Law",
     "msrlegal.com":"MSR Legal","wolinlawgroup.com":"Wolin Law","pattersonlawkc.com":"Patterson Law KC",
     "westmorelandparalegal.com":"Westmoreland Paralegal",
     "gsa.gov":"GSA (U.S.)","ssa.gov":"Social Security Admin","dea.gov":"DEA","ice.dhs.gov":"ICE / DHS",
     "rimkus.com":"Rimkus","aeiconsultants.com":"AEI Consultants","partneresi.com":"Partner ESI",
     "reisservice.com":"REIS Service",
    }
    return m.get(d, domain)

business = [r for r in rows if r["category"] not in EXCLUDE]
excluded = [r for r in rows if r["category"] in EXCLUDE]
business.sort(key=lambda r:(PRIO.get(r["category"],99), -r["touches"]))
excluded.sort(key=lambda r:(-r["touches"]))

wb = Workbook()
FONT="Arial"
hdr_fill = PatternFill("solid", fgColor="1F3864")
hdr_font = Font(name=FONT, bold=True, color="FFFFFF", size=11)
cat_fill = PatternFill("solid", fgColor="D9E1F2")
thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin,right=thin,top=thin,bottom=thin)
wrap = Alignment(vertical="center", wrap_text=True)
center = Alignment(horizontal="center", vertical="center")

def style_header(ws, headers, widths):
    for c,(h,w) in enumerate(zip(headers,widths),1):
        cell=ws.cell(1,c,h); cell.fill=hdr_fill; cell.font=hdr_font
        cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
        cell.border=border
        ws.column_dimensions[get_column_letter(c)].width=w
    ws.freeze_panes="A2"; ws.row_dimensions[1].height=28

# ---- Sheet 1: Add to Salesforce ----
ws = wb.active; ws.title="Add to Salesforce"
headers=["#","Priority Category","Suggested Name (VERIFY)","Email","Organization","Touches","Last Contact","In SF? (mark X)","Notes"]
widths=[5,26,30,34,30,9,13,13,26]
style_header(ws,headers,widths)
for i,r in enumerate(business,1):
    domain=r["email"].split('@')[1]
    row=[i, r["category"], suggest_name(r["email"]), r["email"], org_from_domain(domain),
         r["touches"], r["last_seen"], "", ""]
    for c,v in enumerate(row,1):
        cell=ws.cell(i+1,c,v); cell.font=Font(name=FONT,size=10); cell.border=border
        cell.alignment=center if c in (1,6,7,8) else wrap
    ws.cell(i+1,2).fill=cat_fill

# ---- Sheet 2: Personal / Excluded ----
ws2=wb.create_sheet("Personal & Noise (excluded)")
h2=["#","Category","Email","Touches","Last Contact","Note"]
w2=[5,20,34,9,13,40]
style_header(ws2,h2,w2)
for i,r in enumerate(excluded,1):
    note = "Automated/newsletter - ignore" if r["category"]=="Noise/Automated" else "Personal/family address - not a CRM contact"
    row=[i,r["category"],r["email"],r["touches"],r["last_seen"],note]
    for c,v in enumerate(row,1):
        cell=ws2.cell(i+1,c,v); cell.font=Font(name=FONT,size=10); cell.border=border
        cell.alignment=center if c in (1,4,5) else wrap

# ---- Sheet 3: Summary ----
ws3=wb.create_sheet("Summary", 0)
ws3.sheet_view.showGridLines=False
ws3["B2"]="Team Briggs - Salesforce Contact Gap Worklist"
ws3["B2"].font=Font(name=FONT,bold=True,size=16,color="1F3864")
ws3["B3"]=f"Correspondents active in 10+ yrs of email with NO matching Salesforce contact (by email). Generated {datetime.date(2026,7,30)}."
ws3["B3"].font=Font(name=FONT,size=10,italic=True,color="595959")
ws3["B5"]="How to use: work the 'Add to Salesforce' tab top-down (already ranked by relationship value). Add each to SF, then tell Claude to re-drain - newly-added contacts auto-link to their email history."
ws3["B5"].font=Font(name=FONT,size=10,color="404040"); ws3["B5"].alignment=Alignment(wrap_text=True,vertical="top")
ws3.merge_cells("B5:F6")

from collections import defaultdict
agg=defaultdict(lambda:[0,0])
for r in business:
    agg[r["category"]][0]+=1; agg[r["category"]][1]+=r["touches"]
rollup=sorted(agg.items(), key=lambda kv:PRIO.get(kv[0],99))
ws3["B8"]="Actionable (business) contacts by type"
ws3["B8"].font=Font(name=FONT,bold=True,size=12,color="1F3864")
th=["Category","Contacts","Total Touches"]
for c,h in enumerate(th,2):
    cell=ws3.cell(9,c,h); cell.fill=hdr_fill; cell.font=hdr_font; cell.border=border; cell.alignment=center
rr=10
for cat,(n,t) in rollup:
    ws3.cell(rr,2,cat).font=Font(name=FONT,size=10); ws3.cell(rr,2).border=border
    ws3.cell(rr,3,n).font=Font(name=FONT,size=10); ws3.cell(rr,3).border=border; ws3.cell(rr,3).alignment=center
    ws3.cell(rr,4,t).font=Font(name=FONT,size=10); ws3.cell(rr,4).border=border; ws3.cell(rr,4).alignment=center
    rr+=1
ws3.cell(rr,2,"TOTAL").font=Font(name=FONT,bold=True,size=10); ws3.cell(rr,2).border=border
ws3.cell(rr,3,f"=SUM(C10:C{rr-1})").font=Font(name=FONT,bold=True,size=10); ws3.cell(rr,3).border=border; ws3.cell(rr,3).alignment=center
ws3.cell(rr,4,f"=SUM(D10:D{rr-1})").font=Font(name=FONT,bold=True,size=10); ws3.cell(rr,4).border=border; ws3.cell(rr,4).alignment=center
ws3.cell(rr+2,2,f"Excluded (personal/family + automated): {len(excluded)} addresses - see tab 3.").font=Font(name=FONT,size=10,italic=True,color="595959")
for col,w in zip("ABCDEF",[3,30,12,15,15,15]): ws3.column_dimensions[col].width=w

wb.save('/tmp/TeamBriggs_Salesforce_Contact_Gaps.xlsx')
print("business rows:",len(business),"excluded:",len(excluded),"total:",len(rows))
