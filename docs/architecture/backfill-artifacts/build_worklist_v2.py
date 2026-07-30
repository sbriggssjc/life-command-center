import json, datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# v2 — keyed on party_kind from correspondent_kind (behavioral classifier),
# NOT the domain heuristic. Input nomatch_kind.json rows carry:
#   {email, touches, last_seen, party_kind, distinct_deal_subjects, category}
rows = json.load(open('/tmp/nomatch_kind.json'))

PRIO = {"Client/Operator (dialysis)":1,"Cooperating Broker":2,"Buyer/Capital":3,
 "Title/Escrow":4,"Legal":5,"Lender/Bank":6,"Consultant/Env":7,"Government":8,"Other Business":9}
ROLE_LOCALS = {"admin","info","team","notices","acquisitions","prorations","scheduling",
 "account","support","reply","alert","fcptdealteam","nationalnetlease","analystteam",
 "coteamfalcon","team-albert","isg","fanv-did-dl-cad"}
import re
def suggest_name(email):
    base=email.split('@')[0].split('+')[0]
    if base.lower() in ROLE_LOCALS or base.isdigit(): return "(shared/role mailbox - verify)"
    parts=[p for p in re.split(r'[._\-]', re.sub(r'\d+$','',base)) if p and not p.isdigit()]
    if not parts: return "(verify)"
    if len(parts)==1 and len(parts[0])>2: return parts[0].capitalize()+" (verify first name)"
    return " ".join(p.capitalize() for p in parts)+" (verify)"
ORG={"firstam.com":"First American Title","fnf.com":"Fidelity National Financial","ctt.com":"Chicago Title","stewart.com":"Stewart Title","ipx1031.com":"IPX1031","cbre.com":"CBRE","stanjohnsonco.com":"Stan Johnson Co. (Northmarq)","jll.com":"JLL","colliers.com":"Colliers","kidder.com":"Kidder Mathews","nmrk.com":"Newmark","srsrealestatepartners.com":"SRS Real Estate","logiccre.com":"Logic CRE","davita.com":"DaVita","arikkan.com":"Arikkan (DaVita RE)","exchangeright.com":"ExchangeRight","fcpt.com":"Four Corners (FCPT)","easterlyreit.com":"Easterly Government Properties","boydwatterson.com":"Boyd Watterson","iracapital.com":"IRA Capital","stablewoodproperties.com":"Stablewood Properties","presidiobay.com":"Presidio Bay","brickcapitalre.com":"Brick Capital","jrwrealty.com":"JRW Realty","schusterdevco.com":"Schuster DevCo","caspianrealty.com":"Caspian Realty","adler-realty.com":"Adler Realty","adler-industrial.com":"Adler Industrial","valuenetlease.com":"ValueNet Lease","foley.com":"Foley & Lardner","ltglegal.com":"LTG Legal","buchalter.com":"Buchalter","hinshawlaw.com":"Hinshaw & Culbertson","kutakrock.com":"Kutak Rock","briskinlaw.com":"Briskin Law","dewittllp.com":"DeWitt LLP","leo-law.com":"Leo Law","msrlegal.com":"MSR Legal","wolinlawgroup.com":"Wolin Law","pattersonlawkc.com":"Patterson Law KC","westmorelandparalegal.com":"Westmoreland Paralegal","gsa.gov":"GSA (U.S.)","ssa.gov":"Social Security Admin","dea.gov":"DEA","ice.dhs.gov":"ICE / DHS","rimkus.com":"Rimkus","aeiconsultants.com":"AEI Consultants","partneresi.com":"Partner ESI","reisservice.com":"REIS Service"}
def org(d): return ORG.get(d.lower(), d)

biz  = [r for r in rows if r["party_kind"]=="business"]
over = [r for r in rows if r["party_kind"]=="overlap"]
excl = [r for r in rows if r["party_kind"] in ("personal","noise","vendor")]
biz.sort(key=lambda r:(PRIO.get(r["category"],99), -r["touches"]))
over.sort(key=lambda r:-r["touches"])
excl.sort(key=lambda r:(-r["touches"]))
actionable = over + biz   # overlaps floated to top (surfaced for verify)

wb=Workbook(); FONT="Arial"
hf=PatternFill("solid",fgColor="1F3864"); hfont=Font(name=FONT,bold=True,color="FFFFFF",size=11)
over_fill=PatternFill("solid",fgColor="FCE4D6"); cat_fill=PatternFill("solid",fgColor="D9E1F2")
thin=Side(style="thin",color="BFBFBF"); bd=Border(left=thin,right=thin,top=thin,bottom=thin)
wrap=Alignment(vertical="center",wrap_text=True); ctr=Alignment(horizontal="center",vertical="center")
def hdr(ws,hs,ws_):
    for c,(h,w) in enumerate(zip(hs,ws_),1):
        x=ws.cell(1,c,h); x.fill=hf; x.font=hfont; x.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); x.border=bd
        ws.column_dimensions[get_column_letter(c)].width=w
    ws.freeze_panes="A2"; ws.row_dimensions[1].height=28

ws=wb.active; ws.title="Add to Salesforce"
H=["#","Kind","Priority Category","Suggested Name (VERIFY)","Email","Organization","Touches","Deal Threads","Last Contact","In SF?","Notes"]
W=[5,18,24,28,32,28,9,11,12,10,34]
hdr(ws,H,W)
for i,r in enumerate(actionable,1):
    dom=r["email"].split('@')[1]; ov=r["party_kind"]=="overlap"
    kind = "OVERLAP - verify" if ov else "Business"
    note = (f"Personal-domain address with {r.get('distinct_deal_subjects',0)} distinct deal threads - confirm if client/BD contact or Scott's personal real estate." if ov else "")
    row=[i, kind, r["category"], suggest_name(r["email"]), r["email"], org(dom), r["touches"], r.get("distinct_deal_subjects",0), r["last_seen"], "", note]
    for c,v in enumerate(row,1):
        x=ws.cell(i+1,c,v); x.font=Font(name=FONT,size=10,bold=ov and c==2); x.border=bd
        x.alignment=ctr if c in (1,2,7,8,9,10) else wrap
        if ov: x.fill=over_fill
        elif c==3: x.fill=cat_fill

ws2=wb.create_sheet("Personal & Noise (excluded)")
H2=["#","Kind","Email","Touches","Deal Threads","Last Contact","Why excluded"]; W2=[5,12,34,9,11,12,44]
hdr(ws2,H2,W2)
for i,r in enumerate(excl,1):
    why = ("Automated/newsletter/travel marketing - ignore" if r["party_kind"]=="noise" else "Facilities/service vendor (paint/pest/etc.) - not a BD contact" if r["party_kind"]=="vendor" else "No CRE-deal-thread signal - treated personal (verify if you disagree)")
    row=[i,r["party_kind"],r["email"],r["touches"],r.get("distinct_deal_subjects",0),r["last_seen"],why]
    for c,v in enumerate(row,1):
        x=ws2.cell(i+1,c,v); x.font=Font(name=FONT,size=10); x.border=bd; x.alignment=ctr if c in (1,2,4,5,6) else wrap

ws3=wb.create_sheet("Summary",0); ws3.sheet_view.showGridLines=False
ws3["B2"]="Team Briggs - Salesforce Contact Gap Worklist (v2, behavioral classifier)"
ws3["B2"].font=Font(name=FONT,bold=True,size=15,color="1F3864")
ws3["B3"]=f"No-match correspondents classified by BEHAVIOR (CRE deal-thread signal), not domain. Generated {datetime.date(2026,7,30)}."
ws3["B3"].font=Font(name=FONT,size=10,italic=True,color="595959")
ws3["B5"]=("party_kind: BUSINESS (non-personal domain) - add to SF.  OVERLAP (personal domain but active across >=3 distinct CRE deal "
 "threads) - VERIFY: real client/BD contact vs Scott's personal real estate.  PERSONAL / NOISE - excluded (see tab 3). "
 "Work 'Add to Salesforce' top-down; overlaps are floated to the top in peach. Add to SF, then re-drain to auto-link email history.")
ws3["B5"].font=Font(name=FONT,size=10,color="404040"); ws3["B5"].alignment=Alignment(wrap_text=True,vertical="top"); ws3.merge_cells("B5:G8")
from collections import Counter
c=Counter(r["party_kind"] for r in rows); t=Counter()
for r in rows: t[r["party_kind"]]+=r["touches"]
ws3["B10"]="By party_kind"; ws3["B10"].font=Font(name=FONT,bold=True,size=12,color="1F3864")
for j,h in enumerate(["party_kind","count","touches"],2):
    x=ws3.cell(11,j,h); x.fill=hf; x.font=hfont; x.border=bd; x.alignment=ctr
order=["business","overlap","personal","noise"]; rr=12
for k in order:
    ws3.cell(rr,2,k).font=Font(name=FONT,size=10,bold=k=="overlap"); ws3.cell(rr,2).border=bd
    ws3.cell(rr,3,c.get(k,0)).font=Font(name=FONT,size=10); ws3.cell(rr,3).border=bd; ws3.cell(rr,3).alignment=ctr
    ws3.cell(rr,4,t.get(k,0)).font=Font(name=FONT,size=10); ws3.cell(rr,4).border=bd; ws3.cell(rr,4).alignment=ctr
    if k=="overlap":
        for cc in (2,3,4): ws3.cell(rr,cc).fill=over_fill
    rr+=1
ws3.cell(rr,2,"TOTAL").font=Font(name=FONT,bold=True,size=10); ws3.cell(rr,2).border=bd
ws3.cell(rr,3,f"=SUM(C12:C{rr-1})").font=Font(name=FONT,bold=True,size=10); ws3.cell(rr,3).border=bd; ws3.cell(rr,3).alignment=ctr
ws3.cell(rr,4,f"=SUM(D12:D{rr-1})").font=Font(name=FONT,bold=True,size=10); ws3.cell(rr,4).border=bd; ws3.cell(rr,4).alignment=ctr
ws3.cell(rr+2,2,"Actionable = business + overlap = "+str(c.get('business',0)+c.get('overlap',0))+" contacts.").font=Font(name=FONT,size=10,italic=True,color="595959")
for col,w in zip("ABCDEFG",[3,26,12,14,14,10,10]): ws3.column_dimensions[col].width=w

wb.save('/tmp/TeamBriggs_Salesforce_Contact_Gaps.xlsx')
print("actionable:",len(actionable),"(business",len(biz),"+ overlap",len(over),") excluded:",len(excl))
