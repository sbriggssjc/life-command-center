#!/usr/bin/env python3
# Valley MOB master sheet — canonical Northmarq BOV/OM structure, Northridge layout,
# role-based consistent styling: navy column-headers + totals, bold-navy section headers,
# pale-blue totals, peach input cells. Formula-driven. Year k -> col (3+k); Years 1..10 = cols 4..13.
import openpyxl, datetime
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter as CL

NAVY="FF003DA5"; PALE="FFE0E8F4"; PEACH="FFFFF2CC"; TEXT="FF191919"; MUTED="FF666666"
WHITE="FFFFFFFF"; GOLD="FFFCEFC8"; TITLEF="Calibri Light"; BODY="Calibri"
def F(sz=10,b=False,color=TEXT,name=BODY,it=False): return Font(name=name,size=sz,bold=b,color=color,italic=it)
def fill(c): return PatternFill("solid",fgColor=c)
thin=Side(style="thin",color="FFBFBFBF"); med=Side(style="medium",color=NAVY)
CEN=Alignment("center",vertical="center",wrap_text=True)
L=Alignment("left",vertical="center",wrap_text=True); R=Alignment("right",vertical="center")
LT=Alignment("left",vertical="top",wrap_text=True)
MONEY='"$"#,##0'; M2='"$"#,##0.00'; PCT='0.00%'; PCT1='0.0%'; NUM='#,##0'; MULT='0.00"x"'; YRS='0.00" yrs"'
BB=Border(left=thin,right=thin,top=thin,bottom=thin)

wb=openpyxl.Workbook()

def titlebar(ws,l1,l2,lastcol):
    ws.merge_cells(start_row=1,start_column=2,end_row=1,end_column=lastcol)
    ws.merge_cells(start_row=2,start_column=2,end_row=2,end_column=lastcol)
    a=ws.cell(1,2,l1); a.font=F(15,True,WHITE,TITLEF); a.fill=fill(NAVY); a.alignment=Alignment("left",vertical="center",indent=1)
    b=ws.cell(2,2,l2); b.font=F(10,False,WHITE,BODY); b.fill=fill(NAVY); b.alignment=Alignment("left",vertical="center",indent=1)
    ws.row_dimensions[1].height=26; ws.row_dimensions[2].height=16
    ws.column_dimensions['A'].width=2

def sec(ws,row,text,lastcol):
    c=ws.cell(row,2,text.upper()); c.font=F(11,True,NAVY,TITLEF)
    for col in range(2,lastcol+1): ws.cell(row,col).border=Border(bottom=med)
    ws.row_dimensions[row].height=18

def chead(ws,row,vals,c0=2):
    for i,v in enumerate(vals):
        c=ws.cell(row,c0+i,v); c.font=F(9,True,WHITE); c.fill=fill(NAVY); c.alignment=CEN; c.border=BB
    ws.row_dimensions[row].height=22

def totalrow(ws,row,c0,c1):
    for col in range(c0,c1+1):
        cc=ws.cell(row,col); cc.fill=fill(PALE)
        cc.font=Font(name=BODY,size=cc.font.sz or 10,bold=True,color=NAVY)

# ===================== TERMS =====================
ws=wb.active; ws.title="Terms"; ws.sheet_properties.tabColor="003DA5"
LAST=6
titlebar(ws,"207 FOB JAMES DRIVE — VALLEY, ALABAMA","Four-Tenant Medical Office Building   ·   8,900 SF   ·   100% Occupied",LAST)
ws.column_dimensions['B'].width=30
for col in ['C','D','E','F']: ws.column_dimensions[col].width=23
def kv(row,label,value,fmt=None,vbold=False,mergeval=None):
    lc=ws.cell(row,2,label); lc.font=F(10,False,MUTED); lc.alignment=L
    vc=ws.cell(row,3,value); vc.font=F(10,vbold,TEXT); vc.alignment=L if isinstance(value,str) else R
    if fmt: vc.number_format=fmt
    if mergeval: ws.merge_cells(start_row=row,start_column=3,end_row=row,end_column=mergeval); vc.alignment=L

r=4
sec(ws,r,"Real Estate",LAST); r+=1
for lbl,v,fmt,mv in [("Ownership Interest:","Fee Simple (Leased Fee — sold subject to the 4 leases)",None,6),
("Parcel ID / APN:","12-17-06-13-1-001-014.001  ⚠ confirm",None,6),("Address:","207 Fob James Drive",None,6),
("Building Unit Addresses:","205 / 207 / 209 Fob James Drive (one parcel)",None,6),
("City:","Valley",None,None),("County:","Chambers",None,None),("State:","Alabama",None,None),
("Zip Code:","36854",None,None),("Land (Acres):",2.0,None,None)]:
    kv(r,lbl,v,fmt,mergeval=mv); r+=1
acres_row=r-1
kv(r,"Land (SF):","=C%d*43560"%acres_row,NUM); r+=1
kv(r,"Built:","1994  ⚠ confirm",mergeval=6); r+=1
kv(r,"Buildings:","One (1)"); r+=1
RBA=r; kv(r,"Rentable Building Area (SF):",8900,NUM); r+=1
kv(r,"Floors:","One (1)"); r+=1
kv(r,"Construction:","Brick / Metal",mergeval=6); r+=1
kv(r,"Use:","Multi-Tenant Medical Office (MOB)",mergeval=6); r+=1
kv(r,"Parking Spaces:","⚠ confirm"); r+=1
kv(r,"Parking Ratio (per 1,000 SF):","⚠ confirm"); r+=1
kv(r,"Occupancy:",1.0,PCT1); r+=2

sec(ws,r,"Lease Abstract",LAST); r+=1
chead(ws,r,["Lease Abstract","Tenant 1","Tenant 2","Tenant 3","Tenant 4"]); r+=1
D=datetime.datetime
ABS=[("Tenant (Legal):",["Ivan Lewis Slavich, D.O.","Jeffrey Shelley","CareSouth HHA Holdings of Valley, LLC","A Step Above After School LLC"],None),
("DBA / Trade Name:",["Dr. Ivan L. Slavich III — Internal Medicine","Shelley Orthodontics","Enhabit Home Health","A Step Above After School Care"],None),
("Suite / Address:",["207 & 209 Fob James Dr","207 Fob James Dr","207 Fob James Dr","205 Fob James Dr"],None),
("Guaranty:",["None stated","None stated","None stated","Harlee Crowder (personal)"],None),
("Use:",["Internal Medicine","Orthodontics","Home Health","After-School Care"],None),
("Commencement:",[D(2024,6,1),D(2024,7,1),D(2023,5,1),D(2025,6,1)],"mm/dd/yyyy"),
("Expiration:",[D(2029,5,31),D(2029,6,30),D(2028,4,30),D(2030,12,31)],"mm/dd/yyyy"),
("Current Term:",["Sixty (60) Months","Sixty (60) Months","Sixty (60) Months","Sixty-Seven (67) Months"],None),
("Renewal Options:",["1 × 3-Year","Confirm original","1 × 3-Year","None stated"],None),
("Lease Structure:",["Modified Gross","Modified Gross","Modified Gross","Modified Gross"],None),
("Taxes / Ins / CAM:",["LL; tenant reimburses share","Util incl. in rent","Pro-rata per LL","Tenant pays own electric"],None),
("Tenant Maintenance:",["Minor ≤ $250","Minor ≤ $250","Minor ≤ $75","Minor ≤ $250"],None),
("Landlord Maintenance:",["Structural, HVAC, roof, parking","Structural / major","External bldg & structural","Structural / major"],None),
("Early Termination:",["After mo. 25, 3-mo notice","After mo. 25, 3-mo notice","If LL default uncured","After 2028; 60-day + $2,000"],None),
("Escalation:",["≈ +2.0% / yr","+$50/mo per yr (≈ +2.6%)","≈ +2.0% / yr","Stepped (see Rent)"],None)]
for lbl,vals,fmt in ABS:
    lc=ws.cell(r,2,lbl); lc.font=F(9,False,MUTED); lc.alignment=LT; lc.border=Border(bottom=thin)
    for i,v in enumerate(vals):
        c=ws.cell(r,3+i,v); c.font=F(9,False,TEXT); c.alignment=LT; c.border=Border(bottom=thin)
        if fmt: c.number_format=fmt; c.alignment=CEN
    ws.row_dimensions[r].height=24; r+=1
SFrow=r
lc=ws.cell(r,2,"SF Leased:"); lc.font=F(9,True,MUTED); lc.alignment=L
for i,sf in enumerate([3150,1450,2800,1500]):
    c=ws.cell(r,3+i,sf); c.number_format=NUM; c.font=F(9,True); c.alignment=CEN
r+=1
lc=ws.cell(r,2,"% of Building:"); lc.font=F(9,False,MUTED); lc.alignment=L
for i in range(4):
    col=CL(3+i); c=ws.cell(r,3+i,"=%s%d/$C$%d"%(col,SFrow,RBA)); c.number_format=PCT1; c.font=F(9); c.alignment=CEN
r+=1
AR=r; lc=ws.cell(r,2,"Current Annual Rent:"); lc.font=F(9,True,MUTED); lc.alignment=L
for i,ref in enumerate(["='Rent Roll'!H5","='Rent Roll'!H6","='Rent Roll'!H7","='Rent Roll'!H8"]):
    c=ws.cell(r,3+i,ref); c.number_format=MONEY; c.font=F(9,True); c.alignment=CEN
r+=1
lc=ws.cell(r,2,"Rent $/SF:"); lc.font=F(9,False,MUTED); lc.alignment=L
for i in range(4):
    col=CL(3+i); c=ws.cell(r,3+i,"=%s%d/%s%d"%(col,AR,col,SFrow)); c.number_format=M2; c.font=F(9); c.alignment=CEN
r+=2

sec(ws,r,"Executive Summary",LAST); r+=1
def es(label,value,fmt=None,bold=False,inp=False,merge=False):
    global r
    lc=ws.cell(r,2,label); lc.font=F(10,False,MUTED); lc.alignment=L
    vc=ws.cell(r,3,value); vc.font=F(10,bold,TEXT); vc.alignment=L if isinstance(value,str) else R
    if fmt: vc.number_format=fmt
    if inp: vc.fill=fill(PEACH); vc.alignment=R
    if merge: ws.merge_cells(start_row=r,start_column=3,end_row=r,end_column=6); vc.alignment=L
    r+=1
es("Tenants:","Slavich 35% · Enhabit 31% · Daycare 17% · Shelley 16% (by SF)",merge=True)
es("Owner / Lessor:","Kitchens Family Revocable Trust",merge=True)
es("Interest:","=C5",merge=True)
es("Location:","Valley, Chambers County, Alabama",merge=True)
es("Rentable Building Area (SF):","=C%d"%RBA,NUM)
es("Occupancy:",1.0,PCT1)
es("Use:","Multi-Tenant Medical Office",merge=True)
es("Average Lease Term Remaining:","='Rent Roll'!I9",YRS)
es("Expense Structure:","Modified Gross",merge=True)
es("Rental Increases:","Yes — annual",merge=True)
NOIe=r; es("In-Place NOI:","='Pro Forma'!D26",MONEY,bold=True)
es("NOI / SF:","=C%d/C%d"%(NOIe,RBA),M2)
STABe=r; es("Stabilized NOI (Yr 2):","='Pro Forma'!E26",MONEY)
r+=1
ws.cell(r,2,"Pricing — Proposal & Trade Range").font=F(10,True,NAVY); r+=1
ASKrow=r; es("Ask Price:",1387000,MONEY,bold=True,inp=True)
es("In-Place Cap:","=C%d/C%d"%(NOIe,ASKrow),PCT)
es("Stabilized Cap:","=C%d/C%d"%(STABe,ASKrow),PCT)
es("Ask PPSF:","=C%d/C%d"%(ASKrow,RBA),MONEY)
r+=1
chead(ws,r,["Trade Scenario","Cap Rate","Price","$ / SF"]); r+=1
for nm,cap in [("Trade — Mid",0.0800),("Trade — Conservative",0.0850)]:
    ws.cell(r,2,nm).font=F(10); ws.cell(r,2).alignment=L
    cc=ws.cell(r,3,cap); cc.number_format=PCT; cc.font=F(10); cc.alignment=CEN; cc.fill=fill(PEACH)
    pc=ws.cell(r,4,"=$C$%d/C%d"%(NOIe,r)); pc.number_format=MONEY; pc.font=F(10,True); pc.alignment=R
    sc=ws.cell(r,5,"=D%d/$C$%d"%(r,RBA)); sc.number_format=MONEY; sc.font=F(10); sc.alignment=R
    r+=1
ws.cell(r,2,"Trade prices = in-place NOI ÷ cap. Ask reflects BOV positioning (peach = editable input).").font=F(8,False,MUTED,it=True)
ws.merge_cells(start_row=r,start_column=2,end_row=r,end_column=6); r+=2

sec(ws,r,"Diligence Flags (internal)",LAST); r+=1
for mark,txt in [("✓","Slavich expiration = May 31, 2029 (60-mo term; lease ¶ typo states 2028)"),
("✓","Enhabit expiration = Apr 30, 2028 (lease ¶ typo states 'April 20')"),
("✓","Shelley 5-yr extension confirmed executed (Thomas Kitchens)"),
("✓","205 Fob James correct per Daycare lease (units 205/207/209, one parcel)"),
("✓","Listing executed by Thomas Kitchens, Trustee"),
("⚠","Parcel #, year built, parking, tax map ID — confirm county records"),
("⚠","General R&M $2,454 & HVAC $350 — seller flagged; verify normalized"),
("⚠","Utility reimbursement structure complex — confirm pro-rata settlement"),
("⚠","Bank statements / deposit records not yet received")]:
    m=ws.cell(r,2,mark); m.font=F(10,True,"FF2E7D32" if mark=="✓" else "FFB26A00"); m.alignment=CEN
    t=ws.cell(r,3,txt); t.font=F(9); t.alignment=L; ws.merge_cells(start_row=r,start_column=3,end_row=r,end_column=6); r+=1
ws.freeze_panes="A3"

# ===================== RENT =====================
rs=wb.create_sheet("Rent"); rs.sheet_properties.tabColor="003DA5"
for col,w in [('B',16),('C',22),('D',13),('E',13),('F',12),('G',20)]: rs.column_dimensions[col].width=w
titlebar(rs,"RENT — CONTRACTED RENT SCHEDULES","Contracted lease period   ·   Gold = renewal-period assumption (+3%)",7)
def rentblk(start,title,rows,sf,note=None):
    rr=start; sec(rs,rr,title,7); rr+=1
    chead(rs,rr,["Lease Year","Period","Monthly Rent","Annual Rent","Rate/SF/Mo","Status"]); rr+=1
    f=rr
    for yr,per,mo,ann,status,renew in rows:
        rs.cell(rr,2,yr).font=F(9); rs.cell(rr,2).alignment=L
        rs.cell(rr,3,per).font=F(9); rs.cell(rr,3).alignment=L
        cm=rs.cell(rr,4,mo); cm.number_format=MONEY; cm.font=F(9); cm.alignment=R
        ca=rs.cell(rr,5,ann); ca.number_format=MONEY; ca.font=F(9); ca.alignment=R
        ps=rs.cell(rr,6,"=D%d/%d"%(rr,sf)); ps.number_format=M2; ps.font=F(9); ps.alignment=CEN
        st=rs.cell(rr,7,status); st.font=F(9,False,"FFB26A00" if renew else TEXT); st.alignment=L
        fc=GOLD if renew else (PALE if (rr-f)%2 else WHITE)
        for col in range(2,8): rs.cell(rr,col).fill=fill(fc); rs.cell(rr,col).border=Border(left=thin,right=thin,bottom=thin)
        rr+=1
    if note: rs.cell(rr,2,note).font=F(8,False,MUTED,it=True); rs.merge_cells(start_row=rr,start_column=2,end_row=rr,end_column=7); rr+=1
    return rr+1
slav=[("Year 1","Jun 2024 – May 2025",3300,39600,"Contracted",0),("Year 2","Jun 2025 – May 2026",3366,40392,"Contracted",0),("Year 3","Jun 2026 – May 2027",3433,41196,"Contracted",0),("Year 4","Jun 2027 – May 2028",3500,42000,"Contracted",0),("Year 5","Jun 2028 – May 2029",3570,42840,"Contracted",0),("Year 6 (Renewal)","Jun 2029 – May 2030",3677,44124,"Renewal @ +3%",1),("Year 7","Jun 2030 – May 2031",3787,45444,"Renewal @ +3%",1)]
n=rentblk(4,"Tenant 1 — Ivan Lewis Slavich, D.O.  ·  3,150 SF",slav,3150,"Slavich pays $630/mo toward garbage, power & sewer (modified gross).")
shel=[("Year 1","Jul 2024 – Jun 2025",1900,22800,"Contracted",0),("Year 2","Jul 2025 – Jun 2026",1950,23400,"Contracted",0),("Year 3","Jul 2026 – Jun 2027",2000,24000,"Contracted",0),("Year 4","Jul 2027 – Jun 2028",2050,24600,"Contracted",0),("Year 5","Jul 2028 – Jun 2029",2100,25200,"Contracted",0),("Year 6 (Renewal)","Jul 2029 – Jun 2030",2163,25956,"Renewal @ +3%",1),("Year 7","Jul 2030 – Jun 2031",2228,26736,"Renewal @ +3%",1)]
n=rentblk(n,"Tenant 2 — Jeffrey Shelley · Shelley Orthodontics  ·  1,450 SF",shel,1450,"$150/mo of payment is utility reimbursement (gross rent shown).")
enh=[("Year 1","May 2023 – Apr 2024",3630,43560,"Contracted",0),("Year 2","May 2024 – Apr 2025",3700,44400,"Contracted",0),("Year 3","May 2025 – Apr 2026",3775,45300,"Contracted",0),("Year 4","May 2026 – Apr 2027",3850,46200,"Contracted",0),("Year 5","May 2027 – Apr 2028",3930,47160,"Contracted",0),("Year 6 (Renewal)","May 2028 – Apr 2029",4048,48576,"Renewal @ +3%",1),("Year 7","May 2029 – Apr 2030",4169,50028,"Renewal @ +3%",1)]
n=rentblk(n,"Tenant 3 — CareSouth HHA / Enhabit Home Health  ·  2,800 SF",enh,2800)
day=[("2025 (Ramp)","Jun–Dec 2025 (7 mo)",1000,7000,"Contracted",0),("2026","Jan–Dec 2026",1500,18000,"Contracted",0),("2027","Jan–Dec 2027",1500,18000,"Contracted",0),("2028","Jan–Dec 2028",1500,18000,"Contracted",0),("2029","Jan–Dec 2029",1650,19800,"Contracted",0),("2030","Jan–Dec 2030",1750,21000,"Contracted",0),("2031 (Renewal)","Jan–Dec 2031",1803,21636,"Renewal @ +3%",1)]
n=rentblk(n,"Tenant 4 — A Step Above After School LLC  ·  1,500 SF",day,1500,"Day Care pays own electricity directly; $2,000 advance deposit.")
rs.freeze_panes="A3"

# ===================== RENT ROLL =====================
rr=wb.create_sheet("Rent Roll"); rr.sheet_properties.tabColor="003DA5"
for col,w in [('B',34),('C',10),('D',11),('E',12),('F',12),('G',13),('H',13),('I',12)]: rr.column_dimensions[col].width=w
titlebar(rr,"RENT ROLL — IN-PLACE (FY 2026)","Contracted rents per executed leases   ·   100% occupied",9)
chead(rr,4,["Tenant","SF","% of Total","Lease Start","Lease End","Term Rem. (yrs)","Annual Rent","Rent $/SF"])
data=[("Ivan Lewis Slavich, D.O.",3150,D(2024,6,1),D(2029,5,31),40861),
("Shelley Orthodontics (J. Shelley)",1450,D(2024,7,1),D(2029,6,30),23700),
("CareSouth / Enhabit Home Health",2800,D(2023,5,1),D(2028,4,30),45900),
("A Step Above After School LLC",1500,D(2025,6,1),D(2030,12,31),18000)]
row=5; f=row
for tn,sf,st,en,ann in data:
    fc=PALE if (row-f)%2 else WHITE
    rr.cell(row,2,tn).font=F(9); rr.cell(row,2).alignment=L
    c=rr.cell(row,3,sf); c.number_format=NUM; c.font=F(9); c.alignment=R
    c=rr.cell(row,4,"=C%d/$C$9"%row); c.number_format=PCT1; c.font=F(9); c.alignment=CEN
    c=rr.cell(row,5,st); c.number_format="mm/dd/yyyy"; c.font=F(9); c.alignment=CEN
    c=rr.cell(row,6,en); c.number_format="mm/dd/yyyy"; c.font=F(9); c.alignment=CEN
    c=rr.cell(row,7,"=(F%d-NOW())/365"%row); c.number_format='0.00'; c.font=F(9); c.alignment=CEN
    c=rr.cell(row,8,ann); c.number_format=MONEY; c.font=F(9); c.alignment=R
    c=rr.cell(row,9,"=H%d/C%d"%(row,row)); c.number_format=M2; c.font=F(9); c.alignment=R
    for col in range(2,10): rr.cell(row,col).fill=fill(fc); rr.cell(row,col).border=Border(left=thin,right=thin,bottom=thin)
    row+=1
rr.cell(9,2,"TOTAL / WEIGHTED").alignment=L
rr.cell(9,3,"=SUM(C5:C8)").number_format=NUM
rr.cell(9,4,"=C9/C9").number_format=PCT1
rr.cell(9,7,"=SUMPRODUCT(C5:C8,G5:G8)/C9").number_format='0.00'
rr.cell(9,8,"=SUM(H5:H8)").number_format=MONEY
rr.cell(9,9,"=G9").number_format='0.00'
totalrow(rr,9,2,9)
for col in range(3,10): rr.cell(9,col).alignment=CEN
rr.cell(9,2).font=F(10,True,NAVY)
row=11; sec(rr,row,"In-Place NOI",9); row+=1
def rrk(label,formula,fmt=MONEY,bold=False,total=False):
    global row
    rr.cell(row,2,label).font=F(10,bold,MUTED); rr.cell(row,2).alignment=L
    c=rr.cell(row,3,formula); c.number_format=fmt; c.font=F(10,bold,NAVY if total else TEXT); c.alignment=R
    if total: totalrow(rr,row,2,3); rr.cell(row,2).font=F(10,True,NAVY)
    row+=1
rrk("Gross Scheduled Rent (FY 2026):","=H9")
rrk("Vacancy & Credit Loss (3%):","=-C12*0.03")
EGI=row; rrk("Effective Gross Income:","=C12+C13",bold=True)
rrk("Total Operating Expenses:","=-'Pro Forma'!D24")
rrk("NET OPERATING INCOME:","=C%d+C%d"%(EGI,row-1),bold=True,total=True)
rr.freeze_panes="A5"

# ===================== PRO FORMA =====================
pf=wb.create_sheet("Pro Forma"); pf.sheet_properties.tabColor="003DA5"
pf.column_dimensions['B'].width=32
for i in range(3,15): pf.column_dimensions[CL(i)].width=11
titlebar(pf,"PRO FORMA — TEN-YEAR HOLD","207 Fob James Drive, Valley, AL   ·   Modified-Gross MOB",14)
chead(pf,6,["Investment Year","Year 0","Year 1","Year 2","Year 3","Year 4","Year 5","Year 6","Year 7","Year 8","Year 9","Year 10","10-Yr Total"]); pf.row_dimensions[6].height=20
pf.cell(7,2,"Twelve-Months Ending").font=F(8,False,MUTED); pf.cell(7,2).alignment=L
pf.cell(7,3,"=C100").number_format="mmm-yy"; pf.cell(7,3).font=F(8); pf.cell(7,3).alignment=CEN  # patched later -> acq date
pf.cell(7,4,"=EDATE(C7,12)-1").number_format="mmm-yy"; pf.cell(7,4).font=F(8); pf.cell(7,4).alignment=CEN
for i in range(5,14):
    pf.cell(7,i,"=EDATE(%s7,12)"%CL(i-1)).number_format="mmm-yy"; pf.cell(7,i).font=F(8); pf.cell(7,i).alignment=CEN
YC=set(range(4,14))  # Years 1..10 -> cols D..M
def pfrow(rw,label,vals,fmt=MONEY,bold=False,total=False,pct=False):
    lc=pf.cell(rw,2,label); lc.font=F(9,bold,NAVY if total else TEXT); lc.alignment=L
    for i,v in enumerate(vals):
        c=pf.cell(rw,3+i,v); c.font=F(9,bold,NAVY if total else TEXT); c.alignment=R
        if v in (None,"—"): c.value="—"; c.alignment=CEN
        else: c.number_format=(PCT if pct else fmt)
    if total: totalrow(pf,rw,2,14)
revs={"  Ivan Slavich, D.O.":[40861,41665,42490,43589,44124,45444,46812,48216,49668,51156],
"  Shelley Orthodontics":[23700,24300,24900,25578,25956,26736,27540,28368,29220,30084],
"  CareSouth / Enhabit":[45900,46840,48104,48576,50028,51528,53076,54672,56172,57852],
"  A Step Above (Day Care)":[18000,18000,18000,19800,21000,21636,22284,22956,23640,24348]}
row=8
pf.cell(row,2,"REVENUE:").font=F(10,True,NAVY); row+=1
rev0=row
for k,v in revs.items():
    pfrow(row,k,["—"]+v+["=SUM(D%d:M%d)"%(row,row)]); row+=1
rev1=row-1
gsr=row; pfrow(row,"Scheduled Base Rent:",["—"]+["=SUM(%s%d:%s%d)"%(CL(3+i),rev0,CL(3+i),rev1) for i in range(1,11)]+["=SUM(D%d:M%d)"%(row,row)],bold=True,total=True); row+=1
vac=row; pfrow(row,"Vacancy & Credit Loss (3%)",["—"]+["=-%s%d*0.03"%(CL(3+i),gsr) for i in range(1,11)]+["=SUM(D%d:M%d)"%(row,row)]); row+=1
grev=row; pfrow(row,"Gross Revenue:",["—"]+["=%s%d+%s%d"%(CL(3+i),gsr,CL(3+i),vac) for i in range(1,11)]+["=SUM(D%d:M%d)"%(row,row)],bold=True,total=True); row+=2
pf.cell(row,2,"EXPENSES:").font=F(10,True,NAVY); row+=1
exp0=row
for k,v in {"  Property Taxes (2.0%)":[7359,7506,7656,7809,7966,8125,8287,8453,8622,8795],
"  Insurance (2.0%)":[3475,3544,3615,3688,3761,3837,3913,3992,4072,4153],
"  City License & Tax":[2400]*10,"  Lawn Maintenance":[3400]*10,
"  HVAC (2.0%)":[350,357,364,371,379,386,394,402,410,418],
"  General R&M (2.5%)":[2454,2515,2578,2643,2709,2776,2846,2917,2990,3065]}.items():
    pfrow(row,k,["—"]+v+["=SUM(D%d:M%d)"%(row,row)]); row+=1
exp1=row-1
toe=row; pfrow(row,"Total Operating Expenses:",["—"]+["=SUM(%s%d:%s%d)"%(CL(3+i),exp0,CL(3+i),exp1) for i in range(1,11)]+["=SUM(D%d:M%d)"%(row,row)],bold=True,total=True); row+=1
capr=row; pfrow(row,"Capital Reserves",["—"]+[0]*10+["=SUM(D%d:M%d)"%(row,row)]); row+=1
noi=row; pfrow(row,"NET OPERATING INCOME:",["—"]+["=%s%d-%s%d-%s%d"%(CL(3+i),grev,CL(3+i),toe,CL(3+i),capr) for i in range(1,11)]+["=SUM(D%d:M%d)"%(row,row)],bold=True,total=True); row+=1
pfrow(row,"RENTAL INCREASES:",["—","—"]+["=(%s%d-%s%d)/%s%d"%(CL(3+i),noi,CL(2+i),noi,CL(2+i),noi) for i in range(2,11)]+["—"],pct=True); row+=2

# ---- returns + assumptions ----
CIOstart=row
ACQ0=CIOstart+30
RA_date=ACQ0+1; RA_cap=ACQ0+2; RA_price=ACQ0+3; RA_tax=ACQ0+4; RA_cpi=ACQ0+5
FIN0=ACQ0+6
RF_ltv=FIN0+1; RF_eq=FIN0+2; RF_loan=FIN0+3; RF_rate=FIN0+4; RF_term=FIN0+5; RF_amort=FIN0+6; RF_mc=FIN0+7
DIS0=FIN0+8
RD_excap=DIS0+1; RD_sale=DIS0+2; RD_fees=DIS0+3; RD_payoff=DIS0+4; RD_net=DIS0+5
PRICE="$C$%d"%RA_price; LEVEQ="$C$%d"%RF_eq; LOANAMT="$C$%d"%RF_loan
def subh(text):
    global row
    pf.cell(row,2,text).font=F(10,True,NAVY); pf.cell(row,2).alignment=L; row+=1
def grid(label,valfn,fmt=MONEY,bold=False,total=False,pct=False):
    global row
    pf.cell(row,2,label).font=F(9,bold,NAVY if (bold or total) else TEXT); pf.cell(row,2).alignment=L
    for c in range(3,14):
        v=valfn(c); cell=pf.cell(row,c,v); cell.font=F(9,bold,NAVY if total else TEXT); cell.alignment=R
        if v in (None,"—"): cell.value="—"; cell.alignment=CEN
        else: cell.number_format=(PCT if pct else fmt)
    if total: totalrow(pf,row,2,13)
    out=row; row+=1; return out
def line(label,cval,fmt=MONEY,bold=False,inp=False):
    global row
    pf.cell(row,2,label).font=F(9,bold,NAVY if bold else TEXT); pf.cell(row,2).alignment=L
    cell=pf.cell(row,3,cval); cell.number_format=fmt; cell.font=F(9,bold,TEXT); cell.alignment=R
    if inp: cell.fill=fill(PEACH)
    out=row; row+=1; return out

subh("Cash Investment Outcomes")
eqU=grid("Equity Requirement", lambda c: "=-%s"%PRICE if c==3 else "—")
dispU=grid("Disposition Proceeds", lambda c: "=$C$%d-$C$%d"%(RD_sale,RD_fees) if c==13 else "—")
ncfU=grid("Net Cash Flows", lambda c: ("=C%d"%eqU if c==3 else ("=%s%d+%s%d"%(CL(c),noi,CL(c),dispU) if c==13 else "=%s%d"%(CL(c),noi))), bold=True)
grid("Cumulative Cash Flow", lambda c: ("—" if c==3 else ("=D%d"%ncfU if c==4 else "=%s%d+%s%d"%(CL(c-1),row,CL(c),ncfU))))
capU=grid("Capitalization Rate", lambda c: ("=%s%d/%s"%(CL(c),noi,PRICE) if c in YC else "—"), fmt=PCT, pct=True)
grid("Cumulative Return (Pre-Sale)", lambda c: ("—" if c==3 else ("=D%d"%capU if c==4 else "=%s%d+%s%d"%(CL(c-1),row,CL(c),capU))), fmt=PCT, pct=True)
line("Average Capitalization Rate","=AVERAGE(D%d:M%d)"%(capU,capU),PCT,bold=True)
line("Equity Realization Multiple","=SUM(D%d:M%d)/-C%d"%(ncfU,ncfU,ncfU),MULT,bold=True)
line("IRR (Unleveraged)","=IFERROR(XIRR(C%d:M%d,$C$7:$M$7),\"n/a\")"%(ncfU,ncfU),PCT,bold=True)
row+=1
subh("Cash Flow After Debt Service")
prin=grid("Principal", lambda c: ("=SUMIFS(Amort!$F:$F,Amort!$C:$C,\"<=\"&%s$7,Amort!$C:$C,\">\"&%s$7)"%(CL(c),CL(c-1)) if c in YC else "—"))
intr=grid("Interest", lambda c: ("=SUMIFS(Amort!$G:$G,Amort!$C:$C,\"<=\"&%s$7,Amort!$C:$C,\">\"&%s$7)"%(CL(c),CL(c-1)) if c in YC else "—"))
tds=grid("Total Debt Service", lambda c: ("=%s%d+%s%d"%(CL(c),prin,CL(c),intr) if c in YC else "—"), bold=True)
cfads=grid("CASH FLOW AFTER DEBT SERVICE:", lambda c: ("=%s%d-%s%d"%(CL(c),noi,CL(c),tds) if c in YC else "—"), bold=True, total=True)
eqb=grid("Cumulative Equity Build-Up", lambda c: ("—" if c==3 else ("=D%d"%prin if c==4 else "=%s%d+%s%d"%(CL(c-1),row,CL(c),prin))))
grid("Leverage (Loan ÷ Value)", lambda c: ("=(%s-%s%d)/%s"%(LOANAMT,CL(c),eqb,PRICE) if c in YC else "—"), fmt=PCT1, pct=True)
row+=1
subh("Leveraged Investment Outcomes")
leqL=grid("Equity Requirement", lambda c: "=-%s"%LEVEQ if c==3 else "—")
ldispL=grid("Disposition Proceeds", lambda c: "=$C$%d"%RD_net if c==13 else "—")
lncf=grid("Net Cash Flow", lambda c: ("=C%d"%leqL if c==3 else ("=%s%d+%s%d"%(CL(c),cfads,CL(c),ldispL) if c==13 else "=%s%d"%(CL(c),cfads))), bold=True)
grid("Cumulative Cash Flow", lambda c: ("—" if c==3 else ("=D%d"%lncf if c==4 else "=%s%d+%s%d"%(CL(c-1),row,CL(c),lncf))))
coc=grid("Cash on Cash (Pre-Sale)", lambda c: ("=%s%d/%s"%(CL(c),cfads,LEVEQ) if c in YC else "—"), fmt=PCT, pct=True)
grid("Cumulative Return (Pre-Sale)", lambda c: ("—" if c==3 else ("=D%d"%coc if c==4 else "=%s%d+%s%d"%(CL(c-1),row,CL(c),coc))), fmt=PCT, pct=True)
line("Average Cash on Cash (Pre-Sale)","=AVERAGE(D%d:M%d)"%(coc,coc),PCT,bold=True)
line("Equity Realization Multiple","=SUM(D%d:M%d)/-C%d"%(lncf,lncf,lncf),MULT,bold=True)
line("IRR (Leveraged)","=IFERROR(XIRR(C%d:M%d,$C$7:$M$7),\"n/a\")"%(lncf,lncf),PCT,bold=True)
row+=1
assert row==ACQ0, "ACQ row mismatch %d vs %d"%(row,ACQ0)
subh("Acquisition Assumptions")
line("Acquisition Date",datetime.datetime(2026,1,1),"mmm-yyyy",inp=True)
line("Going-In Cap Rate","=D%d/C%d"%(noi,RA_price),PCT)
line("Acquisition Price","=Terms!C%d"%ASKrow,MONEY,bold=True)
line("Anticipated Tax Increases",0.02,PCT1,inp=True)
line("Anticipated CPI / Cost Increases",0.02,PCT1,inp=True)
subh("Financing Assumptions")
line("Loan-to-Value",0.65,PCT1,inp=True)
line("Equity Requirement","=C%d-C%d"%(RA_price,RF_loan),MONEY)
line("Initial Loan Amount","=C%d*C%d"%(RA_price,RF_ltv),MONEY)
line("Interest Rate",0.065,PCT1,inp=True)
line("Term (Years)",10,NUM,inp=True)
line("Amortization (Years)",25,NUM,inp=True)
line("Mortgage Constant","=IF(C%d=0,0,PMT(C%d/12,C%d*12,-1)*12)"%(RF_amort,RF_rate,RF_amort),PCT)
subh("Disposition Assumptions")
line("Exit Capitalization Rate",0.08,PCT1,inp=True)
line("Sales Price (Yr 10 NOI ÷ Exit Cap)","=M%d/C%d"%(noi,RD_excap),MONEY)
line("Fees & Expenses (5%)","=C%d*0.05"%RD_sale,MONEY)
line("Loan Payoff (Yr 10)","=Amort!H133",MONEY)
line("Net Proceeds","=C%d-C%d-C%d"%(RD_sale,RD_fees,RD_payoff),MONEY,bold=True)
pf.cell(7,3).value="=C%d"%RA_date  # year-0 ending = acquisition date
pf.freeze_panes="C7"
pf.page_setup.orientation="landscape"; pf.page_setup.fitToWidth=1; pf.page_setup.fitToHeight=0
pf.sheet_properties.pageSetUpPr=openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)

# ===================== AMORT =====================
am=wb.create_sheet("Amort"); am.sheet_properties.tabColor="003DA5"
am.column_dimensions['B'].width=10
for col in ['C','D','E','F','G','H']: am.column_dimensions[col].width=15
titlebar(am,"AMORTIZATION SCHEDULE","65% LTV · 6.50% · 25-yr amortization · 10-yr hold",8)
sec(am,4,"Financing Assumptions",8)
for i,(lbl,v,fmt,inp) in enumerate([("Acquisition Price","='Pro Forma'!C%d"%RA_price,MONEY,False),
("Loan-to-Value","='Pro Forma'!C%d"%RF_ltv,PCT1,False),("Initial Loan Amount","=C6*C7",MONEY,False),
("Interest Rate","='Pro Forma'!C%d"%RF_rate,PCT1,False),("Amortization (months)",300,NUM,True),("Hold (months)",120,NUM,True)]):
    am.cell(6+i,2,lbl).font=F(9,False,MUTED); am.cell(6+i,2).alignment=L
    c=am.cell(6+i,3,v); c.number_format=fmt; c.font=F(9,True); c.alignment=R
    if inp: c.fill=fill(PEACH)
LOAN="$C$8"; RATE="$C$9"; AMO="$C$10"
chead(am,13,["Pmt #","Date","Beginning Balance","Payment","Principal","Interest","Ending Balance"])
s=14
am.cell(s,2,1).font=F(9); am.cell(s,2).alignment=CEN
am.cell(s,3,"='Pro Forma'!C%d"%RA_date).number_format="mmm-yyyy"; am.cell(s,3).font=F(9); am.cell(s,3).alignment=CEN
am.cell(s,4,"=%s"%LOAN).number_format=MONEY; am.cell(s,4).font=F(9)
am.cell(s,5,"=PMT(%s/12,%s,-%s)"%(RATE,AMO,LOAN)).number_format=MONEY; am.cell(s,5).font=F(9)
am.cell(s,7,"=D%d*(%s/12)"%(s,RATE)).number_format=MONEY; am.cell(s,7).font=F(9)
am.cell(s,6,"=E%d-G%d"%(s,s)).number_format=MONEY; am.cell(s,6).font=F(9)
am.cell(s,8,"=D%d-F%d"%(s,s)).number_format=MONEY; am.cell(s,8).font=F(9)
for k in range(1,120):
    rw=s+k
    am.cell(rw,2,"=B%d+1"%(rw-1)).alignment=CEN; am.cell(rw,2).font=F(8)
    am.cell(rw,3,"=EDATE(C%d,1)"%(rw-1)).number_format="mmm-yyyy"; am.cell(rw,3).font=F(8); am.cell(rw,3).alignment=CEN
    am.cell(rw,4,"=H%d"%(rw-1)).number_format=MONEY; am.cell(rw,4).font=F(8)
    am.cell(rw,5,"=PMT(%s/12,%s,-%s)"%(RATE,AMO,LOAN)).number_format=MONEY; am.cell(rw,5).font=F(8)
    am.cell(rw,7,"=D%d*(%s/12)"%(rw,RATE)).number_format=MONEY; am.cell(rw,7).font=F(8)
    am.cell(rw,6,"=E%d-G%d"%(rw,rw)).number_format=MONEY; am.cell(rw,6).font=F(8)
    am.cell(rw,8,"=D%d-F%d"%(rw,rw)).number_format=MONEY; am.cell(rw,8).font=F(8)
am.freeze_panes="A14"
wb.save("/tmp/valley_out.xlsx")
print("BUILD OK. noi=%d ncfU=%d cfads=%d lncf=%d ACQ0=%d price=%d"%(noi,ncfU,cfads,lncf,ACQ0,RA_price))
  