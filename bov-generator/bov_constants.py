"""Shared constants and helpers for BOV Master Sheet v2."""
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles.differential import DifferentialStyle

# ── Colors ────────────────────────────────────────────────────────────────────
NAVY  = "003DA5"
PALE  = "E0E8F4"
GOLD  = "FFF2CC"   # renewal / option periods
WHITE = "FFFFFF"
TEXT  = "191919"
MUTED = "6A748C"
TOTBG = "D6E4F5"
YELL  = "FFFACD"
INPC  = "003DA5"

# ── Fills ─────────────────────────────────────────────────────────────────────
F_NAVY  = PatternFill("solid", fgColor=NAVY)
F_PALE  = PatternFill("solid", fgColor=PALE)
F_GOLD  = PatternFill("solid", fgColor=GOLD)
F_WHITE = PatternFill("solid", fgColor=WHITE)
F_TOT   = PatternFill("solid", fgColor=TOTBG)
F_YELL  = PatternFill("solid", fgColor=YELL)

# ── Fonts — 10pt minimum everywhere ──────────────────────────────────────────
def ft(nm="Calibri", sz=10, b=False, i=False, c=TEXT, ul=None):
    return Font(name=nm, size=sz, bold=b, italic=i, color=c, underline=ul)

FT_COVER  = ft("Calibri", 22, b=True, c=NAVY)
FT_TITLE  = ft("Calibri", 14, b=True, c=NAVY)
FT_SHDR   = ft("Calibri", 10, b=True, c=WHITE)          # section bar on NAVY
FT_CHDR   = ft("Calibri Light", 10, b=True, c=WHITE)    # col header on NAVY
FT_CHDR_F = ft("Calibri Light", 10, b=True, i=True, c=WHITE)  # formula col hdr
FT_LABEL  = ft("Calibri", 10, b=True, c=TEXT)
FT_DATA   = ft("Calibri", 10, c=TEXT)
FT_INPUT  = ft("Calibri", 10, c=INPC)     # blue text = user enters
FT_FORM   = ft("Calibri", 10, i=True, c=MUTED)  # italic muted = calculated
FT_TOTAL  = ft("Calibri", 10, b=True, c=TEXT)
FT_NOTE   = ft("Calibri", 10, i=True, c=MUTED)  # footnotes / notes (10pt min)
FT_LINK   = ft("Calibri", 10, c="0563C1", ul="single")  # hyperlink
FT_CALLOUT= ft("Calibri", 16, b=True, c=WHITE)  # callout box values
FT_BRAND  = ft("Calibri", 11, c="555555")

# ── Borders ───────────────────────────────────────────────────────────────────
_t = Side(style='thin',   color="CCCCCC")
_m = Side(style='medium', color=NAVY)
BD_THIN  = Border(bottom=_t)
BD_ALL   = Border(left=_t, right=_t, top=_t, bottom=_t)
BD_NM    = Border(top=_m, bottom=_m)   # NM-branded top/bottom
BD_BOT_M = Border(bottom=_m)

# ── Alignments ────────────────────────────────────────────────────────────────
AL_L  = Alignment(horizontal='left',   vertical='center', wrap_text=True)
AL_C  = Alignment(horizontal='center', vertical='center')
AL_R  = Alignment(horizontal='right',  vertical='center')
AL_TL = Alignment(horizontal='left',   vertical='top',    wrap_text=True)
AL_I1 = Alignment(horizontal='left',   vertical='center', indent=1, wrap_text=True)

# ── Number formats ────────────────────────────────────────────────────────────
D0  = '$#,##0'
D2  = '$#,##0.00'
P2  = '0.00%'
P1  = '0.0%'
MX  = '0.00x'
DT  = 'MM/DD/YYYY'
N0  = '#,##0'
N1  = '#,##0.0'
PCT = '0.0%'

# ── Core helpers ──────────────────────────────────────────────────────────────
def sec(ws, row, label, col_start=1, ncols=14):
    """Navy fill section header, merged across ncols."""
    c = ws.cell(row=row, column=col_start, value=label)
    c.font = FT_SHDR; c.fill = F_NAVY; c.alignment = AL_L
    for j in range(col_start + 1, col_start + ncols):
        ws.cell(row=row, column=j).fill = F_NAVY
    ws.merge_cells(start_row=row, start_column=col_start,
                   end_row=row, end_column=col_start + ncols - 1)
    ws.row_dimensions[row].height = 16

def inp(ws, row, col, val=None, fmt=None):
    """Yellow-fill blue-text input cell. CF clears yellow when filled."""
    c = ws.cell(row=row, column=col, value=val)
    c.font = FT_INPUT; c.fill = F_YELL; c.alignment = AL_L
    if fmt: c.number_format = fmt
    return c

def frm(ws, row, col, formula, fmt=None, align=AL_R):
    """Italic muted formula cell."""
    c = ws.cell(row=row, column=col, value=formula)
    c.font = FT_FORM; c.alignment = align
    if fmt: c.number_format = fmt
    return c

def lbl(ws, row, col, val, indent=0, bold=True):
    c = ws.cell(row=row, column=col, value=val)
    c.font = FT_LABEL if bold else FT_DATA
    c.alignment = Alignment(horizontal='left', vertical='center',
                             wrap_text=True, indent=indent)
    return c

def dat(ws, row, col, val, fmt=None, align=AL_R):
    c = ws.cell(row=row, column=col, value=val)
    c.font = FT_DATA; c.alignment = align
    if fmt: c.number_format = fmt
    return c

def merge(ws, r1, c1, r2, c2):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)

def h(ws, row, ht): ws.row_dimensions[row].height = ht
def w(ws, col, wd): ws.column_dimensions[get_column_letter(col)].width = wd

def add_cf_clear(ws, cell_range):
    """Conditional format: yellow input cells turn white when filled."""
    first = cell_range.split(':')[0]
    ws.conditional_formatting.add(cell_range,
        FormulaRule(
            formula=[f'NOT(ISBLANK({first}))'],
            fill=PatternFill(fill_type="solid", fgColor="FFFFFF")
        ))

def hyperlink_cell(ws, row, col, sheet_name, display=None):
    """Internal tab hyperlink using HYPERLINK formula."""
    label = display or sheet_name
    c = ws.cell(row=row, column=col,
                value=f'=HYPERLINK("#\'{sheet_name}\'!A1","{label}")')
    c.font = FT_LINK; c.alignment = AL_L
    return c

def callout_boxes(ws, row, cf_row, label_prefix=""):
    """
    Draw three Navy callout boxes (IRR | ERM | CoCR) spanning cols C–N (3–14).
    row      = first row of boxes (label row); row+1 = value row
    cf_row   = row number containing the cash flow series C:M
    """
    cf_r = cf_row
    boxes = [
        (3,  6,  f"{label_prefix}IRR",
         f'=IFERROR(IF(C{cf_r}="","",IRR(C{cf_r}:M{cf_r})),"")', P2),
        (7,  10, "EQUITY REALIZATION MULTIPLE",
         f'=IFERROR(IF(OR(C{cf_r}="",C{cf_r}=0),"",SUM(D{cf_r}:M{cf_r})/(-C{cf_r})),"")', MX),
        (11, 14, "AVG CASH-ON-CASH RETURN",
         f'=IFERROR(IF(OR(C{cf_r}="",C{cf_r}=0),"",AVERAGE(D{cf_r}:L{cf_r})/(-C{cf_r})),"")', P2),
    ]
    h(ws, row,   16)
    h(ws, row+1, 30)
    for c1, c2, label, formula, fmt in boxes:
        # label row
        for col in range(c1, c2 + 1):
            ws.cell(row=row, column=col).fill = F_NAVY
        lc = ws.cell(row=row, column=c1, value=label)
        lc.font = FT_SHDR; lc.fill = F_NAVY; lc.alignment = AL_C
        merge(ws, row, c1, row, c2)
        # value row
        for col in range(c1, c2 + 1):
            ws.cell(row=row+1, column=col).fill = F_NAVY
        vc = ws.cell(row=row+1, column=c1, value=formula)
        vc.font = FT_CALLOUT; vc.fill = F_NAVY; vc.alignment = AL_C
        vc.number_format = fmt
        merge(ws, row+1, c1, row+1, c2)
    # label col B
    ws.cell(row=row,   column=2).fill = F_NAVY
    ws.cell(row=row+1, column=2).fill = F_NAVY
