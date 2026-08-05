"""
comps_width_postpass.py — re-apply the shared column-width contract to a comps
workbook AFTER the LibreOffice recalc, as the LAST mutation before conformance.

Why this exists (Prompt 48): `comps_generator._autofit_no_wrap` correctly writes
ONE shared width per header across On Market / Sold *before* save, but the export
path then runs LibreOffice (`recalc_runner.recalc_and_validate` →
`calculateAll()` + `store()`), and LibreOffice **re-optimizes column widths when
it stores the file** — even for columns openpyxl wrote with customWidth. A shared
column that is populated on one sheet but blank on the other desyncs (e.g. PATIENTS
is blank on On Market → LibreOffice leaves the header floor, but has counts on Sold
→ LibreOffice fits to content), so the post-recalc conformance gate sees mismatched
widths and 500s. openpyxl gets the last word here without dropping LibreOffice's
cached formula values.

The constraint: re-opening the recalc'd workbook with openpyxl and re-saving would
STRIP LibreOffice's cached formula results (openpyxl writes formulas but no cached
`<v>` values), so the delivered file would show blanks until the user recalcs and
the validator's recalc-error check (data_only) would read `None` everywhere. So we
do NOT round-trip the workbook through openpyxl. Instead we:

  1. MEASURE the recalc'd content with openpyxl `data_only=True` (formula columns now
     hold cached values, so their rendered width is real, not the pre-recalc floor),
     using the SAME width contract the renderer/validator share
     (`disp_len` / `min_content_width` / `target_column_width`).
  2. Surgically rewrite ONLY the `<cols>` column-width definitions inside the
     recalc'd `.xlsx` (string-replace the `<cols>…</cols>` block in each sheet's XML
     in the zip), leaving every cell value and cached result byte-for-byte intact.

Both sheets get identical widths for every shared header, every column still fits,
and the cached values survive — so the conformance validator (run next) sees the
shared-contract widths AND real cached values.
"""

import re
import shutil
import tempfile
import zipfile
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from validate_comps_output import (
    HEADER_ROW,
    disp_len,
    min_content_width,
    target_column_width,
    _n,
)

_NS_MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"


def _avg_row(ws):
    """Row of the AVG/TOTALS bar (col A == 'AVG'), or None (mirror of the
    renderer/validator helpers)."""
    for r in range(HEADER_ROW + 1, ws.max_row + 1):
        if str(ws.cell(r, 1).value).strip().upper() == "AVG":
            return r
    return None


def compute_shared_widths(path, sheet_names):
    """Measure the recalc'd workbook and return {sheet_name: {col_idx: width}} using
    the shared width contract. A header shared across sheets gets ONE width (the max
    content width anywhere), so On Market and Sold line up. Cached formula values are
    read via data_only=True, so computed columns measure their real rendered width."""
    wb = load_workbook(path, data_only=True)
    try:
        present = [s for s in sheet_names if s in wb.sheetnames]
        shared = {}                 # normalized header -> longest content anywhere
        per_sheet = {}              # sheet -> {col_idx: normalized_header}
        for name in present:
            ws = wb[name]
            avg = _avg_row(ws)
            measure_last = avg or ws.max_row     # header..AVG bar (matches renderer)
            cols = {}
            for c in range(1, ws.max_column + 1):
                hdr = ws.cell(HEADER_ROW, c).value
                if hdr in (None, ""):
                    continue
                key = _n(hdr)
                longest = 0
                for r in range(HEADER_ROW, measure_last + 1):
                    longest = max(longest, disp_len(ws.cell(r, c)))
                # Floor for computed columns (still a lower bound even post-recalc).
                longest = max(longest, min_content_width(hdr))
                cols[c] = key
                shared[key] = max(shared.get(key, 0), longest)
            per_sheet[name] = cols
    finally:
        wb.close()

    out = {}
    for name, cols in per_sheet.items():
        out[name] = {c: target_column_width(shared[key], key) for c, key in cols.items()}
    return out


def _sheet_xml_targets(zf):
    """Map worksheet name -> the zip path of its sheet XML (via workbook.xml + rels)."""
    wb_xml = zf.read("xl/workbook.xml").decode("utf-8")
    # name -> r:id
    name_rid = {}
    for m in re.finditer(r"<sheet\b([^>]*?)/?>", wb_xml):
        attrs = dict(re.findall(r'([\w:]+)="([^"]*)"', m.group(1)))
        nm = attrs.get("name")
        rid = attrs.get("r:id") or attrs.get("id")
        if nm and rid:
            name_rid[nm] = rid
    # r:id -> target
    rels = zf.read("xl/_rels/workbook.xml.rels").decode("utf-8")
    rid_target = {}
    for m in re.finditer(r"<Relationship\b([^>]*?)/?>", rels):
        attrs = dict(re.findall(r'([\w:]+)="([^"]*)"', m.group(1)))
        rid = attrs.get("Id")
        tgt = attrs.get("Target")
        if rid and tgt:
            rid_target[rid] = tgt
    out = {}
    for nm, rid in name_rid.items():
        tgt = rid_target.get(rid)
        if not tgt:
            continue
        # Normalize to a zip path (targets are like "worksheets/sheetN.xml" or
        # "/xl/worksheets/sheetN.xml").
        p = tgt.lstrip("/")
        if p.startswith("xl/"):
            zip_path = p
        else:
            zip_path = "xl/" + p
        out[nm] = zip_path
    return out


def _fmt_width(w):
    """Format a width so openpyxl reads it back exactly (raw passthrough)."""
    return ("%g" % round(float(w), 4))


def _build_cols_block(existing_block, widths_by_idx):
    """Return a fresh `<cols>…</cols>` string: preserve each existing `<col>`'s
    attributes (notably `style`) per column index, override width/customWidth for the
    indices we sized, drop bestFit, and emit one `<col>` per index."""
    idx_attr = {}
    if existing_block:
        for m in re.finditer(r"<col\b([^>]*?)/?>", existing_block):
            attrs = dict(re.findall(r'(\w+)="([^"]*)"', m.group(1)))
            try:
                mn = int(attrs.get("min", "0"))
                mx = int(attrs.get("max", attrs.get("min", "0")))
            except ValueError:
                continue
            for i in range(mn, mx + 1):
                idx_attr[i] = dict(attrs)

    for i, w in widths_by_idx.items():
        a = idx_attr.get(i, {})
        a["width"] = _fmt_width(w)
        a["customWidth"] = "1"
        a.pop("bestFit", None)
        idx_attr[i] = a

    if not idx_attr:
        return "<cols></cols>"

    parts = []
    for i in sorted(idx_attr):
        a = idx_attr[i]
        a["min"] = str(i)
        a["max"] = str(i)
        order = ["min", "max", "width", "style", "customWidth", "hidden",
                 "outlineLevel", "collapsed"]
        keys = order + [k for k in a if k not in order]
        attr_str = " ".join(f'{k}="{a[k]}"' for k in keys if k in a)
        parts.append(f"<col {attr_str}/>")
    return "<cols>" + "".join(parts) + "</cols>"


def _rewrite_sheet_cols(xml_text, widths_by_idx):
    """Replace (or insert) the `<cols>` block in one sheet's XML string. All other
    bytes — including every cell value and cached `<v>` — are left untouched."""
    new_cols = _build_cols_block(_extract_cols(xml_text), widths_by_idx)
    if re.search(r"<cols\b[^>]*>.*?</cols>", xml_text, re.S):
        return re.sub(r"<cols\b[^>]*>.*?</cols>", lambda _m: new_cols, xml_text, count=1, flags=re.S)
    if re.search(r"<cols\b[^>]*/>", xml_text):
        return re.sub(r"<cols\b[^>]*/>", lambda _m: new_cols, xml_text, count=1)
    # No <cols> present — insert immediately before <sheetData>.
    return re.sub(r"(<sheetData\b)", lambda m: new_cols + m.group(1), xml_text, count=1)


def _extract_cols(xml_text):
    m = re.search(r"<cols\b[^>]*>.*?</cols>", xml_text, re.S)
    return m.group(0) if m else ""


def reapply_shared_widths(path, sheet_names=("On Market", "Sold")):
    """Re-apply the shared width contract to `path` AFTER recalc, in place, without
    dropping cached formula values. Returns {sheet: {col_idx: width}} applied.

    Rewrites ONLY the `<cols>` column-width XML in each requested sheet. Sheets not
    present in the workbook are skipped. Safe to call on the single-sheet lease/BOV
    path (pass its sheet name) — it just applies auto-fit widths to that one sheet.
    """
    path = str(Path(path))
    widths = compute_shared_widths(path, sheet_names)
    if not widths:
        return {}

    with zipfile.ZipFile(path, "r") as zf:
        targets = _sheet_xml_targets(zf)
        names = zf.namelist()
        infos = {i.filename: i for i in zf.infolist()}
        data = {name: zf.read(name) for name in names}

    for sheet_name, cols in widths.items():
        zip_path = targets.get(sheet_name)
        if not zip_path or zip_path not in data:
            continue
        xml_text = data[zip_path].decode("utf-8")
        data[zip_path] = _rewrite_sheet_cols(xml_text, cols).encode("utf-8")

    # Rewrite the zip preserving order + compression; only the touched sheets differ.
    fd, tmp_name = tempfile.mkstemp(suffix=".xlsx", dir=str(Path(path).parent))
    import os as _os
    _os.close(fd)
    try:
        with zipfile.ZipFile(tmp_name, "w") as zf:
            for name in names:
                info = infos[name]
                zf.writestr(info, data[name], compress_type=info.compress_type)
        shutil.move(tmp_name, path)
    finally:
        if Path(tmp_name).exists():
            Path(tmp_name).unlink()

    return widths
