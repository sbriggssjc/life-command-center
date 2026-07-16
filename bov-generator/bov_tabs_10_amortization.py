"""BOV v2 — Tab 10: Amortization Schedule."""
from bov_constants import *
from openpyxl.styles import Border, Side, PatternFill, Font
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.utils import get_column_letter

_NM_M = Side(style='medium', color=NAVY)
_NM_T = Side(style='thin',   color=NAVY)
_GY_T = Side(style='thin',   color="CCCCCC")

_AS = "Assumptions & Flags"
MAX_AMORT = 30   # 30 years max (25yr amortization standard)


def build_amortization_tab(wb):
    ws = wb.create_sheet("Amortization")
    ws.sheet_view.showGridLines = False

    # ── Column widths ──────────────────────────────────────────────────────────
    # A=Year#(6), B=Payment#(8), C=Beg Balance(14), D=Principal(12),
    # E=Interest(12), F=Total Payment(14), G=End Balance(14), H=spacer(2),
    # I=Cumulative Principal(14), J=Cumulative Interest(14)
    col_widths = [6, 8, 16, 14, 14, 14, 16, 2, 16, 16]
    for i, wd in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = wd

    # ── Header ────────────────────────────────────────────────────────────────
    r = 1
    ws.row_dimensions[r].height = 6
    r = 2
    ws.row_dimensions[r].height = 28
    c = ws.cell(row=r, column=1, value="AMORTIZATION SCHEDULE")
    c.font = FT_TITLE; c.alignment = AL_L
    merge(ws, r, 1, r, 10)

    r = 3
    ws.row_dimensions[r].height = 14
    c = ws.cell(row=r, column=1, value="Auto-calculates from Assumptions & Flags tab  ·  Hold period years highlighted  ·  Annual summary below monthly detail")
    c.font = FT_NOTE; c.alignment = AL_L
    merge(ws, r, 1, r, 10)

    r = 4
    ws.row_dimensions[r].height = 8

    # ── Loan summary inputs (pulled from Assumptions) ─────────────────────────
    r = 5
    sec(ws, r, "LOAN SUMMARY", col_start=1, ncols=10)

    loan_refs = [
        (6,  "Loan Amount",         f'=IFERROR(\'{_AS}\'!$I$8,"")',   D0),
        (6,  "Interest Rate",       f'=IFERROR(\'{_AS}\'!$I$10,"")',  P2),
        (6,  "Amortization (yrs)",  f'=IFERROR(\'{_AS}\'!$I$11,"")',  N0),
        (6,  "Hold Period (yrs)",   f'=IFERROR(\'{_AS}\'!$I$13,"")',  N0),
        (6,  "Monthly Payment",     f'=IFERROR(\'{_AS}\'!$I$15,"")',  D0),
        (6,  "Annual Debt Service", f'=IFERROR(\'{_AS}\'!$I$14,"")',  D0),
    ]
    r = 6
    ws.row_dimensions[r].height = 18
    labels   = ["Loan Amount",     "Interest Rate",    "Amortization (yrs)", "Hold Period (yrs)",  "Monthly Payment", "Annual Debt Service"]
    formulas = [
        f'=IFERROR(\'{_AS}\'!$I$8,"")',
        f'=IFERROR(\'{_AS}\'!$I$10,"")',
        f'=IFERROR(\'{_AS}\'!$I$11,"")',
        f'=IFERROR(\'{_AS}\'!$I$13,"")',
        f'=IFERROR(\'{_AS}\'!$I$15,"")',
        f'=IFERROR(\'{_AS}\'!$I$14,"")',
    ]
    fmts = [D0, P2, N0, N0, D0, D0]
    # 3 items in first row, 3 in second label row
    cols_a = [1, 3, 5, 7, 8, 9]  # label/value pairs stacked
    for i, (label, formula, fmt) in enumerate(zip(labels, formulas, fmts)):
        col_l = 1 + i * 1   # re-do: spread across row
    # Simple 2-row layout: row 6 = labels, row 7 = values
    for i, (label, formula, fmt) in enumerate(zip(labels, formulas, fmts)):
        col = 1 + i
        if col > 7: col = 1  # wrap — just skip for now
        c = ws.cell(row=6, column=col + (i // 4) * 0, value=label)
        c.font = FT_NOTE; c.alignment = AL_C; c.fill = F_PALE
    # Actually let's do a clean single-row display
    for i, (label, formula, fmt) in enumerate(zip(labels[:5], formulas[:5], fmts[:5])):
        col = 1 + i * 2
        if col > 10: break
        c = ws.cell(row=6, column=col, value=label)
        c.font = FT_NOTE; c.alignment = AL_C; c.fill = F_PALE
        c2 = frm(ws, 6, col + 1, formula, fmt=fmt, align=AL_C)
        c2.fill = F_PALE

    r = 7
    ws.row_dimensions[r].height = 8

    # ══════════════════════════════════════════════════════════════════════════
    # ANNUAL AMORTIZATION TABLE
    # ══════════════════════════════════════════════════════════════════════════
    r = 8
    sec(ws, r, "ANNUAL AMORTIZATION  —  Hold period rows highlighted", col_start=1, ncols=10)

    r = 9
    ws.row_dimensions[r].height = 22
    hdrs = ["YEAR", "PAYMENTS", "BEG. BALANCE", "PRINCIPAL", "INTEREST",
            "TOTAL PAYMENT", "END BALANCE", "", "CUMUL. PRINCIPAL", "CUMUL. INTEREST"]
    for ci, hdr in enumerate(hdrs, 1):
        c = ws.cell(row=r, column=ci, value=hdr)
        c.font = FT_CHDR; c.fill = F_NAVY; c.alignment = AL_C
        c.border = Border(bottom=_NM_M)

    DATA_START = 10
    LOAN_AMT   = f'\'{_AS}\'!$I$8'
    RATE_MO    = f'(\'{_AS}\'!$I$10/12)'
    AMORT_YRS  = f'\'{_AS}\'!$I$11'
    PMT_MO     = f'\'{_AS}\'!$I$15'   # monthly payment from Assumptions

    for i in range(MAX_AMORT):
        rr = DATA_START + i
        yr_num = i + 1
        ws.row_dimensions[rr].height = 18

        # Year
        c = ws.cell(row=rr, column=1, value=yr_num)
        c.font = FT_DATA; c.alignment = AL_C

        # Payments (12 per year, except partial final year if amort < full periods)
        c = ws.cell(row=rr, column=2, value=12)
        c.font = FT_DATA; c.alignment = AL_C

        # Beginning Balance
        if i == 0:
            beg_formula = f'=IFERROR(IF({LOAN_AMT}="","",ABS({LOAN_AMT})),"")'
        else:
            prev_end = f'G{rr - 1}'
            beg_formula = f'=IFERROR(IF({prev_end}="","",{prev_end}),"")'
        c = frm(ws, rr, 3, beg_formula, fmt=D0, align=AL_R)

        # Principal paid in year (Beg_Bal - End_Bal)
        # End Balance = Beg * (1+r)^12 - PMT*((1+r)^12-1)/r
        end_formula = (
            f'=IFERROR(IF(C{rr}="","",MAX(0,C{rr}*(1+{RATE_MO})^12'
            f'-{PMT_MO}*((1+{RATE_MO})^12-1)/{RATE_MO})),"")'
        )
        # Interest = Total Payments - Principal
        int_formula  = f'=IFERROR(IF(C{rr}="","",{PMT_MO}*12-(C{rr}-G{rr})),"")'
        prin_formula = f'=IFERROR(IF(C{rr}="","",C{rr}-G{rr}),"")'
        tot_formula  = f'=IFERROR(IF(C{rr}="","",{PMT_MO}*12),"")'

        frm(ws, rr, 4, prin_formula, fmt=D0, align=AL_R)
        frm(ws, rr, 5, int_formula,  fmt=D0, align=AL_R)
        frm(ws, rr, 6, tot_formula,  fmt=D0, align=AL_R)
        frm(ws, rr, 7, end_formula,  fmt=D0, align=AL_R)

        # Cumulative Principal
        if i == 0:
            cum_prin = f'=IFERROR(D{rr},"")'
        else:
            cum_prin = f'=IFERROR(IF(D{rr}="","",I{rr-1}+D{rr}),"")'
        frm(ws, rr, 9, cum_prin, fmt=D0, align=AL_R)

        # Cumulative Interest
        if i == 0:
            cum_int = f'=IFERROR(E{rr},"")'
        else:
            cum_int = f'=IFERROR(IF(E{rr}="","",J{rr-1}+E{rr}),"")'
        frm(ws, rr, 10, cum_int, fmt=D0, align=AL_R)

        # NM thin border
        for ci in range(1, 11):
            ws.cell(row=rr, column=ci).border = Border(bottom=_GY_T)

    DATA_END = DATA_START + MAX_AMORT - 1

    # ── CF: highlight hold-period rows ────────────────────────────────────────
    # Row is "within hold period" if year number <= Hold Period
    HOLD_PERIOD_REF = f'\'{_AS}\'!$I$13'
    _hl_fill = PatternFill(fill_type="solid", fgColor=TOTBG)
    _hl_font = Font(name="Calibri", size=10, bold=True, color=TEXT)
    for i in range(MAX_AMORT):
        rr = DATA_START + i
        yr_num = i + 1
        formula = f'=AND(NOT(ISBLANK({HOLD_PERIOD_REF})),A{rr}<={HOLD_PERIOD_REF})'
        ws.conditional_formatting.add(
            f"A{rr}:J{rr}",
            FormulaRule(formula=[formula], fill=_hl_fill, font=_hl_font)
        )

    # ── Totals row ────────────────────────────────────────────────────────────
    tot_r = DATA_END + 1
    ws.row_dimensions[tot_r].height = 18
    c = ws.cell(row=tot_r, column=1, value="TOTALS")
    c.font = FT_TOTAL; c.fill = F_TOT; c.alignment = AL_L
    merge(ws, tot_r, 1, tot_r, 2)
    ws.cell(row=tot_r, column=1).border = Border(top=_NM_M, bottom=_NM_M)

    for col, formula in [
        (4, f'=IFERROR(SUM(D{DATA_START}:D{DATA_END}),"")',),
        (5, f'=IFERROR(SUM(E{DATA_START}:E{DATA_END}),"")',),
        (6, f'=IFERROR(SUM(F{DATA_START}:F{DATA_END}),"")',),
    ]:
        c = frm(ws, tot_r, col, formula, fmt=D0, align=AL_R)
        c.font = FT_TOTAL; c.fill = F_TOT
        c.border = Border(top=_NM_M, bottom=_NM_M)
    for ci in [3, 7, 8, 9, 10]:
        ws.cell(row=tot_r, column=ci).fill = F_TOT
        ws.cell(row=tot_r, column=ci).border = Border(top=_NM_M, bottom=_NM_M)

    # ── Notes ─────────────────────────────────────────────────────────────────
    note_r = tot_r + 2
    notes = [
        "Highlighted rows = hold period (auto-updates when Hold Period changes on Assumptions tab).",
        "End Balance formula: Beg. Balance × (1 + Monthly Rate)^12 − Monthly Payment × ((1 + Monthly Rate)^12 − 1) / Monthly Rate.",
        "IO Period: if entered on Assumptions tab, update this schedule manually for the IO years (principal = $0).",
        "Remaining balance at end of hold period is used in disposition analysis on Assumptions & Flags tab.",
    ]
    for idx, note in enumerate(notes):
        rr = note_r + idx
        ws.row_dimensions[rr].height = 14
        c = ws.cell(row=rr, column=1, value=f"▪  {note}")
        c.font = FT_NOTE; c.alignment = AL_L
        merge(ws, rr, 1, rr, 10)

    bot_r = note_r + len(notes) + 1
    ws.row_dimensions[bot_r].height = 4
    for ci in range(1, 11):
        ws.cell(row=bot_r, column=ci).border = Border(bottom=Side(style='medium', color=NAVY))
