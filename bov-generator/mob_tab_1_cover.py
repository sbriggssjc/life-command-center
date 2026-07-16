"""Cover tab for BOV Master MOB Sheet — Multi-Tenant MOB / Commercial."""
from bov_constants import *

ALL_TABS = [
    ("Cover",               "Property identification and workbook guide"),
    ("Executive Summary",   "Investment snapshot · occupancy · pricing strategy · engagement proposal"),
    ("Real Estate",         "Leg 1 — Physical property, location, and market fundamentals"),
    ("Rent Roll",           "Leg 2 — Consolidated single-page tenant and rent summary"),
    ("Lease Abstract",      "Leg 2 — Executed lease terms and obligations by tenant"),
    ("Rent Schedule",       "Leg 2 — Year-by-year rent schedule by tenant"),
    ("Credit",              "Leg 3 — Tenant / guarantor credit, operations, and unit economics"),
    ("Pro Forma",           "10-year investment model: Revenue → NOI → Returns"),
    ("Assumptions & Flags", "All key inputs, per-tenant data, discrepancy flags"),
    ("Sensitivity Analysis","Returns across different cap rate / pricing / occupancy scenarios"),
    ("Amortization",        "Loan amortization schedule — auto-updates with assumptions"),
]


def build_mob_cover(wb):
    ws = wb.create_sheet("Cover")
    ws.sheet_view.showGridLines = False

    # Columns: A=margin(2), B=label/content(28), C=content/link(44), D=margin(2)
    for col, wd in [(1, 2), (2, 28), (3, 44), (4, 2)]:
        w(ws, col, wd)

    # ── Top Navy bar ──────────────────────────────────────────────────────────
    h(ws, 1, 10)
    for col in range(1, 5):
        ws.cell(row=1, column=col).fill = F_NAVY
    merge(ws, 1, 1, 1, 4)

    # ── Brand row ─────────────────────────────────────────────────────────────
    h(ws, 2, 24)
    c = ws.cell(row=2, column=2, value="NorthMarq  ·  Investment Sales")
    c.font = ft("Calibri", 11, b=True, c=NAVY); c.alignment = AL_L
    c2 = ws.cell(row=2, column=3, value="Team Briggs")
    c2.font = ft("Calibri", 11, b=True, c=MUTED); c2.alignment = AL_R

    # Pale accent band
    h(ws, 3, 8); h(ws, 4, 8); h(ws, 5, 8)
    for r in range(3, 6):
        for col in range(1, 5):
            ws.cell(row=r, column=col).fill = F_PALE
    merge(ws, 3, 1, 5, 4)

    # ── Main title ────────────────────────────────────────────────────────────
    h(ws, 6, 52)
    c = ws.cell(row=6, column=2, value="Broker Opinion of Value")
    c.font = FT_COVER; c.alignment = Alignment(horizontal='left', vertical='center')
    merge(ws, 6, 2, 6, 3)

    h(ws, 7, 22)
    c = ws.cell(row=7, column=2,
                value="Multi-Tenant MOB / Commercial  ·  Investment Analysis Workbook")
    c.font = FT_BRAND; c.alignment = AL_L
    merge(ws, 7, 2, 7, 3)
    h(ws, 8, 14)

    # ── Property fields ───────────────────────────────────────────────────────
    fields = [
        (9,  "PROPERTY NAME",         "[Property Name  —  City, State]"),
        (11, "PROPERTY ADDRESS",       None),
        (13, "CLIENT / SELLER",        None),
        (15, "ASSET TYPE",             "Multi-Tenant MOB / Commercial"),
        (17, "ANALYSIS DATE",          None),
        (19, "PREPARED BY",            "Scott Briggs  ·  SVP, Commercial Investment Sales  ·  NorthMarq"),
        (21, "ANALYST",                "Sarah Martin"),
    ]
    for row, label, value in fields:
        c = ws.cell(row=row, column=2, value=label)
        c.font = ft("Calibri", 10, b=True, c=MUTED); c.alignment = AL_L
        h(ws, row, 14)
        h(ws, row + 1, 22)
        if value:
            v = ws.cell(row=row + 1, column=2, value=value)
            v.font = ft("Calibri", 11, b=True, c=TEXT); v.alignment = AL_L
        else:
            v = inp(ws, row + 1, 2, fmt=DT if "DATE" in label else None)
            v.alignment = AL_L
        merge(ws, row + 1, 2, row + 1, 3)
    h(ws, 23, 14)

    # ── Three legs of the stool + tab guide ──────────────────────────────────
    sec(ws, 24, "THREE LEGS OF THE STOOL  —  WORKBOOK STRUCTURE", col_start=2, ncols=2)

    legs = [
        ("LEG 1  —  REAL ESTATE FUNDAMENTALS",
         "Underlying value of the physical property; highest & best use; "
         "contractual vs. prevailing market rents; long-term value generator.",
         [("Real Estate", None)]),
        ("LEG 2  —  ECONOMICS OF THE LEASE",
         "Contractual terms, obligations, and protections imposed on or granted to "
         "the Landlord across all tenants. Better terms → greater value.",
         [("Rent Roll", None), ("Lease Abstract", None), ("Rent Schedule", None)]),
        ("LEG 3  —  TENANT & GUARANTOR CREDIT",
         "Corporate credit quality, unit-level performance, and importance of each "
         "location to the overall tenant enterprise.",
         [("Credit", None)]),
    ]

    r = 25
    for leg_label, leg_desc, tabs in legs:
        h(ws, r, 16)
        c = ws.cell(row=r, column=2, value=leg_label)
        c.font = ft("Calibri", 10, b=True, c=NAVY); c.fill = F_PALE; c.alignment = AL_L
        ws.cell(row=r, column=3).fill = F_PALE
        merge(ws, r, 2, r, 3)
        r += 1

        h(ws, r, 36)
        c = ws.cell(row=r, column=2, value=leg_desc)
        c.font = FT_NOTE; c.fill = F_WHITE; c.alignment = AL_TL
        merge(ws, r, 2, r, 3)
        r += 1

        for sheet_name, display in tabs:
            h(ws, r, 18)
            ws.cell(row=r, column=2, value="→").font = ft("Calibri", 10, c=MUTED)
            ws.cell(row=r, column=2).alignment = AL_C
            hyperlink_cell(ws, r, 3, sheet_name)
            r += 1
        h(ws, r, 8); r += 1

    # ── Synthesis tabs ────────────────────────────────────────────────────────
    h(ws, r, 16)
    c = ws.cell(row=r, column=2, value="SYNTHESIS  —  ANALYSIS & MODEL")
    c.font = ft("Calibri", 10, b=True, c=NAVY); c.fill = F_PALE; c.alignment = AL_L
    ws.cell(row=r, column=3).fill = F_PALE
    merge(ws, r, 2, r, 3)
    r += 1

    for sheet_name, desc in [
        ("Executive Summary",    "Investment snapshot + occupancy + pricing strategy + engagement proposal"),
        ("Pro Forma",            "10-year investment model"),
        ("Assumptions & Flags",  "All key inputs, per-tenant data, and discrepancy flags"),
        ("Sensitivity Analysis", "Returns across different cap rate / pricing / occupancy scenarios"),
        ("Amortization",         "Loan amortization schedule"),
    ]:
        h(ws, r, 18)
        ws.cell(row=r, column=2, value="→").font = ft("Calibri", 10, c=MUTED)
        ws.cell(row=r, column=2).alignment = AL_C
        hyperlink_cell(ws, r, 3, sheet_name)
        r += 1

    h(ws, r, 10); r += 1

    # ── Workbook tab guide ────────────────────────────────────────────────────
    sec(ws, r, "WORKBOOK TAB GUIDE", col_start=2, ncols=2); r += 1
    h(ws, r, 18)
    c = ws.cell(row=r, column=2, value="TAB")
    c.font = FT_CHDR; c.fill = F_NAVY; c.alignment = AL_C
    c2 = ws.cell(row=r, column=3, value="CONTENTS")
    c2.font = FT_CHDR; c2.fill = F_NAVY; c2.alignment = AL_L
    r += 1

    for i, (tab_name, tab_desc) in enumerate(ALL_TABS):
        h(ws, r, 26)
        bg = F_PALE if i % 2 == 0 else F_WHITE
        c = ws.cell(row=r, column=2, value=tab_name)
        c.font = FT_LABEL; c.fill = bg; c.alignment = AL_L
        d = ws.cell(row=r, column=3, value=tab_desc)
        d.font = FT_DATA; d.fill = bg; d.alignment = AL_TL
        r += 1

    h(ws, r, 10); r += 1

    # ── Color legend ──────────────────────────────────────────────────────────
    sec(ws, r, "COLOR CONVENTIONS", col_start=2, ncols=2); r += 1
    legend = [
        (PALE,  NAVY,  "NM Blue fill",  "Contracted / executed lease periods"),
        (GOLD,  TEXT,  "Gold fill",     "Renewal / option / projection periods"),
        (YELL,  TEXT,  "Yellow fill",   "Input cells — clears automatically when filled"),
        (WHITE, INPC,  "Blue text",     "User-entry fields"),
        (WHITE, MUTED, "Muted italic",  "Calculated / formula cells — do not edit"),
    ]
    for bg, fg, name, desc in legend:
        h(ws, r, 18)
        s = ws.cell(row=r, column=2, value=f"  {name}")
        s.fill = PatternFill("solid", fgColor=bg)
        s.font = ft("Calibri", 10, b=True, c=fg); s.alignment = AL_L
        from openpyxl.styles import Border, Side
        s.border = Border(
            left=Side(style='thin', color="CCCCCC"),
            right=Side(style='thin', color="CCCCCC"),
            top=Side(style='thin', color="CCCCCC"),
            bottom=Side(style='thin', color="CCCCCC"),
        )
        d = ws.cell(row=r, column=3, value=desc)
        d.font = FT_DATA; d.alignment = AL_L
        r += 1

    # Bottom Navy bar
    h(ws, r, 10)
    for col in range(1, 5):
        ws.cell(row=r, column=col).fill = F_NAVY
    merge(ws, r, 1, r, 4)

    # CF clear on input cells (property fields)
    for row in [10, 12, 14, 18]:
        add_cf_clear(ws, f"B{row}:B{row}")
