"""
comps_generator.py — Briggs CRE Comps population engine (the shared comps tool).

The comps analog of the BOV generator: ONE deterministic engine that fills the
Briggs comps template's INPUT columns from structured comp rows and leaves the
formula-protected columns (RENT/SF, all $/SF, all CAP, TERM, BPS, PRICE ADJ, DOM,
EFF. RENT/SF, #) untouched — so the output is identical no matter which access
point prepared the rows (per Comps_Column_Mapping.md's critical rule).

Design (mirrors the BOV pattern):
  • The LLM at each door maps the raw CoStar/Salesforce export → structured rows
    using the shared column mapping + normalization (Comps_Column_Mapping.md).
  • THIS engine writes those rows into the template's input cells, header-driven
    and formula-safe, then LibreOffice recalcs so the analytics compute.

Header-driven: columns are located by their row-5 header text (normalized), not
by hardcoded letters — robust to template column shifts. A column is treated as
FORMULA (never written) when the template's first data row already holds a
formula there. Rows start at DATA_START_ROW.

Input contract (all keys optional; omit what you don't have — never guess).
Keys are the Briggs template column names; the aliases in _ALIASES let a caller
pass the query_comps field names straight through (st→state, init_price→
initial_price, yr_built→built, rba_sf→rba, lease_exp→exp, annual_noi/annual_rent→
rent, list_date→on_market, sale_price→sold_price, sale_date→date …). Only columns
the template actually has are written; anything else lands in `unknown_keys`.
  Sales:  { "comp_type":"sales",
            "on_market":[ {address,city,state,rba,tenant,lease_type,exp,annual_noi,
                           initial_price,cur_price,on_market,bumps,options,built,
                           notes}, ... ],
            "sold":[ {…on_market fields…, last_price, sale_price, sale_date}, ... ] }
  Lease:  { "comp_type":"lease",
            "comps":[ {property_type,source,address,city,state,suite_space,
                       sf_leased,annual_rent,lease_type,lease_comm,exp,
                       execution_date,ti_sf,free_rent_mos,rent_bumps,built,
                       renovated,notes}, ... ] }

buyer / seller / financing are OPT-IN only — omit them unless the user explicitly
asks for buyer/seller/financing in the comps (not part of the default column set,
and there is no template column for them today, so they are otherwise left out).

Dates accept 'YYYY-MM-DD' or 'MM/DD/YYYY' (written as true Excel dates). Numbers
accept plain values or strings with $ , %. Text is written verbatim (already
normalized upstream).
"""

import os
import re
from datetime import datetime, date
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

DATA_START_ROW = 6

# --- Estimated-value rendering -------------------------------------------------
# When a comp row carries an "is estimated" flag, the corresponding value cell is
# marked so an estimated cap can never be mistaken for a verified one (the guardrail
# both the rent-imputation and gov-NOI-modeling passes require). We keep the cell
# NUMERIC (so RENT/SF, PPSF, and every cap still compute) and only (a) append a
# literal " (est.)" suffix to its number format and (b) apply a light amber fill.
_EST_FILL = PatternFill("solid", fgColor="FFF3CD")
# (flag key on the row) -> (template header token whose value cell gets marked).
# NOTE: nothing is flagged. Team Briggs policy is reliable-or-exclude for BOTH gov NOI and
# dialysis rent — an unreliable estimate is filtered out upstream by the comps reliability gate,
# not color-coded, and the standard reliability disclaimer covers the accuracy of what is shown.
_ESTIMATE_FLAGS = ()
# Recognized non-column metadata keys — present to drive rendering/provenance, not
# to be written into a template column, so they must NOT count as "unknown".
_META_KEYS = {
    "rent_is_imputed", "rent_source", "rent_psf_basis", "actual_annual_rent",
    "noi_is_modeled", "noi_modeled_source",
}


def _is_truthy(v) -> bool:
    """A jsonb/bool/string flag is 'on' for True, 't', 'true', 'yes', '1'."""
    if isinstance(v, bool):
        return v
    return str(v).strip().lower() in ("t", "true", "yes", "1")


def _mark_estimated(cell):
    """Flag a written numeric value cell as estimated: append ' (est.)' to its number
    format (keeps it numeric so formulas still read it) + light fill. Idempotent."""
    if not isinstance(cell.value, (int, float)):
        return  # only mark a real numeric value; never a blank or text cell
    fmt = cell.number_format or "General"
    if "(est.)" not in fmt:
        base = "0" if fmt in ("General", "@") else fmt
        cell.number_format = base + '" (est.)"'
    cell.fill = _EST_FILL
TEMPLATE_DIR = Path(os.environ.get("COMPS_TEMPLATE_DIR", Path(__file__).parent / "templates"))

SALES_TEMPLATE = "Comps Blank Template - Briggs.xlsx"
LEASE_TEMPLATE = "Lease Comps Template - Briggs.xlsx"
# Dialysis-specific sales template: identical to SALES_TEMPLATE but with CHAIRS and
# PATIENTS input columns inserted immediately after RBA (SF) on both On Market and Sold
# sheets (formula-protected columns shift accordingly). Selected when the caller flags the
# request dialysis (payload vertical == 'dialysis' or dialysis == true). Chairs/patients are
# the most-recent counts, per the dialysis comp standard. Header-driven, so no other change.
DIALYSIS_SALES_TEMPLATE = "Comps Blank Template - Briggs - Dialysis.xlsx"
# Government-specific sales template: Agency-first column order with the government
# nuances (GOV LEVEL, USE, GOV SF LEASED vs total RBA, GOV OCCP %, GROSS RENT, NOI +
# NOI/SF, EXPIR/TERMIN., TERM REM/FIRM REM, GUARANTOR, BUMPS/DROP, ASK history). Rent/SF
# is gross rent ÷ leased SF; PPSF and every cap are on TOTAL RBA / whole-building NOI.
# Selected when the caller flags the request government (vertical == 'government' or
# government == true). Header-driven, so the shared engine needs no other change.
GOV_SALES_TEMPLATE = "Comps Blank Template - Briggs - Government.xlsx"


class CompsError(Exception):
    def __init__(self, message, status=422):
        super().__init__(message)
        self.status = status


def _norm(h) -> str:
    """Normalize a header or key to a canonical token: lowercase, alnum→_, trimmed.
    'RBA (SF)'→'rba_sf', 'SUITE / SPACE'→'suite_space', 'TI ($/SF)'→'ti_sf',
    'LEASE COMM.'→'lease_comm', 'CITY'→'city'."""
    s = re.sub(r"[^a-z0-9]+", "_", str(h or "").lower()).strip("_")
    return _ALIASES.get(s, s)


# Row-key aliases → canonical Briggs header token. Lets a caller pass the comps-engine /
# query_comps field names straight through to the master-order template columns
# (TENANT, LAND, BUILT, RBA, CHAIRS, PATIENTS, RENT, EXP, EXPENSES, RENEWAL OPTIONS,
#  SOLD PRICE, DATE, INITIAL/LAST PRICE, ON MARKET, DOM ...).
_ALIASES = {
    "chair_count": "chairs", "chair_ct": "chairs",
    "patient_count": "patients", "patient_ct": "patients",
    "year_built": "built", "yr_built": "built",
    "st": "state", "init_price": "initial_price",
    "building_sf": "rba", "rba_sf": "rba", "building_size": "rba",
    "sale_price": "sold_price",
    "sale_date": "date",
    "lease_expiration": "exp", "lease_exp": "exp",
    "lease_type": "expenses", "expense_structure": "expenses",
    "options": "renewal_options", "renewal_option": "renewal_options", "renewal_option_text": "renewal_options",
    "annual_rent": "rent", "annual_noi": "rent",
    "land_acres": "land", "land_area": "land",
    "list_date": "on_market", "on_market_date": "on_market", "listing_date": "on_market",
    "cur_price": "last_price",   # on-market current ask = LAST PRICE column
    "sold_sf": "sold_sf", "price_per_sf": "sold_sf",
}


def _to_number(v):
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return v
    s = str(v).strip().replace("$", "").replace(",", "").replace("%", "")
    try:
        return float(s)
    except ValueError:
        return None


def _to_date(v):
    if v is None or v == "":
        return None
    if isinstance(v, (datetime, date)):
        return v
    s = str(v).strip()
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%Y/%m/%d"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None  # unparseable → leave blank (never write junk into a date cell)


# Which normalized headers are DATE columns (write as true Excel dates) and which
# are NUMERIC (write as floats). Everything else is text. Formula columns are
# detected from the template and never written regardless of this map.
# NOTE: aliases are applied by _norm() BEFORE these sets are consulted, so list the
# canonical (post-alias) tokens here — e.g. 'exp', 'date', 'on_market', 'rent', 'built'.
_DATE_KEYS = {"exp", "date", "on_market", "lease_exp", "list_date", "sale_date", "lease_comm", "execution_date",
              # government sales tokens
              "expir", "termin", "sold_date"}
_NUM_KEYS = {
    "land", "built", "rba", "chairs", "patients", "rent",
    "sold_price", "initial_price", "last_price",
    "annual_noi", "init_price", "cur_price", "sale_price", "rba_sf",
    "sf_leased", "annual_rent", "ti_sf", "free_rent_mos", "yr_built", "renovated",
    # government sales tokens (GOV SF LEASED, GOV OCCP %, GROSS RENT, NOI, ASK history)
    "gov_sf_leased", "gov_occp", "gross_rent", "noi", "initial_ask", "last_ask",
}


def _is_government(payload: dict) -> bool:
    """True when the caller flags a government comp request → use the government sales
    template (Agency-first, GOV LEVEL/USE/GOV SF LEASED/NOI columns). Accepts
    vertical=='government', government==true, or asset_type/property_type mentioning
    'government'/'gov'. Checked BEFORE dialysis so a mislabeled row can't fall through."""
    if not isinstance(payload, dict):
        return False
    if payload.get("government") is True:
        return True
    for k in ("vertical", "asset_type", "property_type"):
        v = str(payload.get(k, "")).lower()
        if v in ("government", "gov") or "government" in v:
            return True
    return False


def _is_dialysis(payload: dict) -> bool:
    """True when the caller flags a dialysis comp request → use the dialysis sales
    template (CHAIRS/PATIENTS columns after RBA). Accepts vertical=='dialysis',
    dialysis==true, or asset_type/property_type containing 'dialysis'."""
    if not isinstance(payload, dict):
        return False
    if payload.get("dialysis") is True:
        return True
    for k in ("vertical", "asset_type", "property_type"):
        if "dialysis" in str(payload.get(k, "")).lower():
            return True
    return False


def _header_map(ws, header_row=5):
    """Return {normalized_header: (col_idx, is_formula)}. is_formula is True when
    the first data row already holds a formula in that column (→ never overwrite)."""
    out = {}
    for c in range(1, ws.max_column + 1):
        h = ws.cell(header_row, c).value
        if h in (None, ""):
            continue
        cell = ws.cell(DATA_START_ROW, c)
        is_formula = (cell.data_type == "f") or (isinstance(cell.value, str) and cell.value.startswith("="))
        out[_norm(h)] = (c, is_formula)
    return out


def _write_rows(ws, rows):
    """Write structured rows into the sheet's INPUT columns (header-matched,
    formula-safe). Returns (written_count, skipped_formula_keys, unknown_keys)."""
    hmap = _header_map(ws)
    skipped_formula, unknown = set(), set()
    for i, row in enumerate(rows or []):
        r = DATA_START_ROW + i
        for key, val in (row or {}).items():
            k = _norm(key)
            if k in _META_KEYS:
                continue                 # rendering/provenance metadata — not a column
            if k not in hmap:
                unknown.add(k)
                continue
            col, is_formula = hmap[k]
            if is_formula:
                skipped_formula.add(k)   # protect calculated columns — never write
                continue
            if val is None or val == "":
                continue
            if k in _DATE_KEYS:
                val = _to_date(val)
            elif k in _NUM_KEYS:
                val = _to_number(val)
            if val is None:
                continue
            ws.cell(r, col).value = val
        # After the row's values are written, mark any estimated value cells so an
        # imputed rent / modeled NOI reads as estimated (formula-safe: stays numeric).
        for flag_key, target_token in _ESTIMATE_FLAGS:
            if _is_truthy(_row_get(row, flag_key)) and target_token in hmap:
                col, is_formula = hmap[target_token]
                if not is_formula:
                    _mark_estimated(ws.cell(r, col))
    return len(rows or []), sorted(skipped_formula), sorted(unknown)


def _row_get(row, canon):
    """Value from a row dict whose key normalizes (post-alias) to `canon`."""
    for k, v in (row or {}).items():
        if _norm(k) == canon:
            return v
    return None


def _sort_rows(rows, sheet):
    """Workflow sort: Sold by DATE desc; On Market / Available by cap asc; Lease by
    execution date desc. Missing keys sort last so blanks never lead the table."""
    rows = list(rows or [])
    if sheet == "Sold":
        rows.sort(key=lambda r: (_to_date(_row_get(r, "date")) is None,
                                 _to_date(_row_get(r, "date")) or datetime.min), reverse=True)
    elif sheet in ("On Market", "Available"):
        def cap(r):
            rent = _to_number(_row_get(r, "rent")); price = _to_number(_row_get(r, "last_price"))
            return rent / price if (rent and price) else None
        rows.sort(key=lambda r: (cap(r) is None, cap(r) or 0))
    elif sheet == "Lease Comps":
        rows.sort(key=lambda r: (_to_date(_row_get(r, "execution_date")) is None,
                                 _to_date(_row_get(r, "execution_date")) or datetime.min), reverse=True)
    return rows


def _trim_to_totals(ws, n):
    """Delete the unused blank rows between the last written comp and the template's
    AVG/TOTALS bar so the bar sits directly beneath the data, and rewrite the bar's
    AVERAGE/COUNT ranges to the trimmed row count (Workflow step 5)."""
    tot = None
    for r in range(DATA_START_ROW, ws.max_row + 1):
        if str(ws.cell(r, 1).value).strip().upper() == "AVG":
            tot = r
            break
    if tot is None:
        return
    old_last = tot - 1                       # last pre-filled data row (e.g. 105)
    capacity = old_last - DATA_START_ROW + 1
    if n <= 0 or n >= capacity:
        return
    del_start = DATA_START_ROW + n
    ws.delete_rows(del_start, old_last - del_start + 1)
    tot = DATA_START_ROW + n
    new_last = DATA_START_ROW + n - 1        # = 5 + n
    for c in range(1, ws.max_column + 1):
        v = ws.cell(tot, c).value
        if isinstance(v, str) and v.startswith("="):
            ws.cell(tot, c).value = v.replace(str(old_last), str(new_last))


def populate_comps(payload: dict, out_path: str, template_dir: Path = None) -> dict:
    """Fill the Briggs comps template from a structured payload. Returns a summary
    { comp_type, sheets:{name:count}, skipped_formula_keys, unknown_keys, out_path }.
    Does NOT recalc — the caller runs LibreOffice recalc (same as the BOV flow)."""
    tdir = Path(template_dir or TEMPLATE_DIR)
    comp_type = str(payload.get("comp_type", "")).lower()
    if comp_type == "sales":
        if _is_government(payload):
            tpl = tdir / GOV_SALES_TEMPLATE
        elif _is_dialysis(payload):
            tpl = tdir / DIALYSIS_SALES_TEMPLATE
        else:
            tpl = tdir / SALES_TEMPLATE
    elif comp_type == "lease":
        tpl = tdir / LEASE_TEMPLATE
    else:
        raise CompsError("comp_type must be 'sales' or 'lease'")
    if not tpl.exists():
        raise CompsError(f"template not found: {tpl}", status=500)

    wb = load_workbook(tpl)
    summary = {"comp_type": comp_type, "sheets": {}, "skipped_formula_keys": [], "unknown_keys": []}
    skipped_all, unknown_all = set(), set()

    if comp_type == "sales":
        for sheet, key in (("On Market", "on_market"), ("Sold", "sold")):
            if sheet in wb.sheetnames and payload.get(key):
                rows = _sort_rows(payload[key], sheet)
                n, sk, un = _write_rows(wb[sheet], rows)
                _trim_to_totals(wb[sheet], n)
                summary["sheets"][sheet] = n
                skipped_all.update(sk); unknown_all.update(un)
    else:  # lease
        rows = _sort_rows(payload.get("comps") or payload.get("lease_comps") or [], "Lease Comps")
        if "Lease Comps" in wb.sheetnames and rows:
            n, sk, un = _write_rows(wb["Lease Comps"], rows)
            _trim_to_totals(wb["Lease Comps"], n)
            summary["sheets"]["Lease Comps"] = n
            skipped_all.update(sk); unknown_all.update(un)

    if not summary["sheets"]:
        raise CompsError("no comp rows supplied (sales: on_market/sold; lease: comps)")

    wb.save(out_path)
    wb.close()
    summary["skipped_formula_keys"] = sorted(skipped_all)
    summary["unknown_keys"] = sorted(unknown_all)
    summary["out_path"] = out_path
    return summary
