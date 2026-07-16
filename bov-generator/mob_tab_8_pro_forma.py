"""BOV MOB — Tab 8: Pro Forma (10-year investment model, Multi-Tenant).

Assumptions & Flags cell references used here:
  Left side (col C):
    C9  = Building SF (GLA)
    C14 = Close Date
    C18/27/36/45/54 = T1-T5 Tenant Names
    C21/30/39/48/57 = T1-T5 Year 1 Annual Rent
    C22/31/40/49/58 = T1-T5 Rent Escalation %
    C24/33/42/51/60 = T1-T5 Reimbursements
    C63 = Vacancy / Credit Loss %
    C69 = RE Taxes  C70=Insurance  C71=CAM  C72=Mgmt Fee%
    C73=G&A  C74=R&M  C75=Utilities  C76=Cap Reserves
    C85 = Year 1 NOI (for Valuation Matrix)
    C94 = Purchase Price
  Right side (col I):
    I9  = Down Payment / Equity
    I14 = Annual Debt Service
    I20 = Exit Cap Rate
    I25 = Net Reversion (Unleveraged)
    I26 = Remaining Loan Balance
    I27 = Net Reversion After Debt

Row map (key callout references):
  Row 41 = UNLEVERAGED CASH FLOW series (C=Yr0, D:M=Yr1-10)
  Row 43 = Unleveraged callout LABELS  → 'Assumptions & Flags' refs C44, G44, K44
  Row 44 = Unleveraged callout VALUES
  Row 56 = LEVERAGED CASH FLOW series
  Row 58 = Leveraged callout LABELS    → 'Assumptions & Flags' refs C59, G59, K59
  Row 59 = Leveraged callout VALUES
"""
from bov_constants import *
from openpyxl.styles import Border, Side, Font, PatternFill
from openpyxl.utils import get_column_letter

_NM_M = Side(style='medium', color=NAVY)
_NM_T = Side(style='thin',   color=NAVY)
_GY_T = Side(style='thin',   color="CCCCCC")
_AS   = "Assumptions & Flags"

LABEL_COL = 1
NCOLS     = 14
YR0_COL   = 3
YR_COLS   = list(range(4, 14))   # cols 4-13 = Year 1-10

# Tenant assumption row references (Year 1 rent, escalation)
T_RENT_ROWS = [21, 30, 39, 48, 57]   # C21, C30, C39, C48, C57
T_ESC_ROWS  = [22, 31, 40, 49, 58]   # C22, C31, C40, C49, C58
T_REIMB_ROWS= [24, 33, 42, 51, 60]   # C24, C33, C42, C51, C60
T_NAME_ROWS = [18, 27, 36, 45, 54]   # C18, C27, C36, C45, C54


def build_mob_pro_forma_tab(wb):
    ws = wb.create_sheet("Pro Forma")
    ws.sheet_view.showGridLines = False

    # ── Column widths ──────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 1.5
    ws.column_dimensions["C"].width = 12
    for col in range(4, 14):
        ws.column_dimensions[get_column_letter(col)].width = 11
    ws.column_dimensions["N"].width = 22

    # ── Header ────────────────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 6
    ws.row_dimensions[2].height = 28
    c = ws.cell(row=2, column=1, value="PRO FORMA  —  10-YEAR INVESTMENT MODEL  (MULTI-TENANT MOB)")
    c.font = FT_TITLE; c.alignment = AL_L
    merge(ws, 2, 1, 2, NCOLS)

    ws.row_dimensions[3].height = 14
    c = ws.cell(row=3, column=1,
                value="Multi-Tenant MOB / Commercial  ·  All inputs on Assumptions & Flags tab  ·  Col C = Year 0; Cols D–M = Operating Years 1–10")
    c.font = FT_NOTE; c.alignment = AL_L
    merge(ws, 3, 1, 3, NCOLS)

    ws.row_dimensions[4].height = 8

    # ── Column headers ────────────────────────────────────────────────────────
    r = 5
    ws.row_dimensions[r].height = 18
    for ci in range(1, NCOLS + 1): ws.cell(row=r, column=ci).fill = F_NAVY
    c = ws.cell(row=r, column=YR0_COL, value="Year 0 / Purchase")
    c.font = FT_CHDR; c.fill = F_NAVY; c.alignment = AL_C
    for i, col in enumerate(YR_COLS):
        c = ws.cell(row=r, column=col, value=f"Year {i+1}")
        c.font = FT_CHDR; c.fill = F_NAVY; c.alignment = AL_C

    # ── Year-end dates ────────────────────────────────────────────────────────
    r = 6
    ws.row_dimensions[r].height = 14
    c = ws.cell(row=r, column=1, value="Year-End Date")
    c.font = FT_NOTE; c.fill = F_PALE; c.alignment = AL_L
    ws.cell(row=r, column=YR0_COL).fill = F_PALE
    for i, col in enumerate(YR_COLS):
        n = i + 1
        c = frm(ws, r, col,
                f'=IFERROR(DATE(YEAR(\'{_AS}\'!$C$14)+{n},MONTH(\'{_AS}\'!$C$14),DAY(\'{_AS}\'!$C$14)-1),"")',
                fmt=DT, align=AL_C)
        c.fill = F_PALE; c.font = FT_NOTE
    ws.cell(row=r, column=NCOLS).fill = F_PALE

    # ══════════════════════════════════════════════════════════════════════════
    # SECTION 1: REVENUE
    # ══════════════════════════════════════════════════════════════════════════
    r = 7
    sec(ws, r, "REVENUE", col_start=1, ncols=NCOLS)

    # T1-T5 Base Rent rows (8-12)
    TENANT_RENT_ROWS = []
    for t_idx in range(5):
        r = 8 + t_idx
        ws.row_dimensions[r].height = 18
        rent_ref  = T_RENT_ROWS[t_idx]
        esc_ref   = T_ESC_ROWS[t_idx]
        name_ref  = T_NAME_ROWS[t_idx]
        TENANT_RENT_ROWS.append(r)

        # Label: pulls tenant name from Assumptions
        c = ws.cell(row=r, column=1,
                    value=f'=IFERROR(IF(\'{_AS}\'!$C${name_ref}="","Tenant {t_idx+1}",\'{_AS}\'!$C${name_ref}&" — Base Rent"),"")')
        c.font = FT_DATA; c.alignment = AL_L

        # Year 1 = from Assumptions
        frm(ws, r, 4,
            f'=IFERROR(IF(\'{_AS}\'!$C${rent_ref}="","",\'{_AS}\'!$C${rent_ref}),"")',
            fmt=D0, align=AL_R)
        # Years 2-10: escalate
        for i, col in enumerate(YR_COLS[1:], 2):
            prev = get_column_letter(col - 1)
            frm(ws, r, col,
                f'=IFERROR(IF({prev}{r}="","",{prev}{r}*(1+\'{_AS}\'!$C${esc_ref})),"")',
                fmt=D0, align=AL_R)

    # Thin separator
    ws.row_dimensions[13].height = 4
    for ci in range(1, NCOLS + 1):
        ws.cell(row=13, column=ci).border = Border(bottom=_GY_T)

    # Gross Potential Rent (row 14)
    r = 14
    ws.row_dimensions[r].height = 18
    lbl(ws, r, 1, "Gross Potential Rent")
    ws.cell(row=r, column=1).fill = F_PALE
    GPR_ROW = r
    for col in YR_COLS:
        cl = get_column_letter(col)
        c = frm(ws, r, col,
                f'=IFERROR(IFERROR({cl}8,0)+IFERROR({cl}9,0)+IFERROR({cl}10,0)+IFERROR({cl}11,0)+IFERROR({cl}12,0),"")',
                fmt=D0, align=AL_R)
        c.fill = F_PALE; c.font = FT_LABEL
    ws.cell(row=r, column=NCOLS).fill = F_PALE

    # Vacancy / Credit Loss (row 15)
    r = 15
    ws.row_dimensions[r].height = 18
    lbl(ws, r, 1, "Vacancy / Credit Loss", bold=False)
    VAC_ROW = r
    for col in YR_COLS:
        cl = get_column_letter(col)
        frm(ws, r, col,
            f'=IFERROR(IF(OR({cl}{GPR_ROW}="",\'{_AS}\'!$C$63=""),"",-{cl}{GPR_ROW}*\'{_AS}\'!$C$63),"")',
            fmt=D0, align=AL_R)
    ws.cell(row=r, column=NCOLS, value="Ref: Assumptions C63").font = FT_NOTE

    # Effective Base Rent (row 16)
    r = 16
    ws.row_dimensions[r].height = 18
    lbl(ws, r, 1, "Effective Base Rent")
    EBR_ROW = r
    for col in YR_COLS:
        cl = get_column_letter(col)
        c = frm(ws, r, col,
                f'=IFERROR(IF({cl}{GPR_ROW}="","",{cl}{GPR_ROW}+IFERROR({cl}{VAC_ROW},0)),"")',
                fmt=D0, align=AL_R)
        c.font = FT_LABEL

    # Total Tenant Reimbursements (row 17)
    r = 17
    ws.row_dimensions[r].height = 18
    lbl(ws, r, 1, "Total Tenant Reimbursements", bold=False)
    REIMB_ROW = r
    reimb_parts = "+".join([f'IFERROR(\'{_AS}\'!$C${rr},0)' for rr in T_REIMB_ROWS])
    for col in YR_COLS:
        frm(ws, r, col,
            f'=IFERROR({reimb_parts},"")',
            fmt=D0, align=AL_R)
    ws.cell(row=r, column=NCOLS, value="Sum T1-T5 from Assumptions").font = FT_NOTE

    # Other Income (row 18, input)
    r = 18
    ws.row_dimensions[r].height = 18
    lbl(ws, r, 1, "Other Income", bold=False)
    OTHER_ROW = r
    for col in YR_COLS:
        inp(ws, r, col, fmt=D0)
        ws.cell(row=r, column=col).alignment = AL_R
    add_cf_clear(ws, f"D{OTHER_ROW}:M{OTHER_ROW}")

    # EGI (row 20)
    ws.row_dimensions[19].height = 4
    r = 20
    ws.row_dimensions[r].height = 20
    lbl(ws, r, 1, "EFFECTIVE GROSS INCOME (EGI)")
    ws.cell(row=r, column=1).fill = F_TOT
    EGI_ROW = r
    for col in YR_COLS:
        cl = get_column_letter(col)
        c = frm(ws, r, col,
                f'=IFERROR(IFERROR({cl}{EBR_ROW},0)+IFERROR({cl}{REIMB_ROW},0)+IFERROR({cl}{OTHER_ROW},0),"")',
                fmt=D0, align=AL_R)
        c.font = FT_TOTAL; c.fill = F_TOT
        c.border = Border(top=_GY_T, bottom=_NM_T)

    # ══════════════════════════════════════════════════════════════════════════
    # SECTION 2: EXPENSES
    # ══════════════════════════════════════════════════════════════════════════
    ws.row_dimensions[21].height = 6
    sec(ws, 22, "OPERATING EXPENSES", col_start=1, ncols=NCOLS)

    # Expense rows 23-30 — all flat from Assumptions except Mgmt Fee (% of EGI)
    EXPENSE_REFS = [
        (23, "Real Estate Taxes",          f'\'{_AS}\'!$C$69', D0, False),
        (24, "Insurance",                  f'\'{_AS}\'!$C$70', D0, False),
        (25, "CAM / Janitorial",           f'\'{_AS}\'!$C$71', D0, False),
        (26, "Management Fee",             None,                D0, False),  # % of EGI
        (27, "G&A / Admin",                f'\'{_AS}\'!$C$73', D0, False),
        (28, "Repairs & Maintenance",      f'\'{_AS}\'!$C$74', D0, False),
        (29, "Utilities",                  f'\'{_AS}\'!$C$75', D0, False),
        (30, "Capital / Replacement Reserves", f'\'{_AS}\'!$C$76', D0, False),
    ]
    MGMT_ROW = 26
    EXP_ROWS = [23, 24, 25, 26, 27, 28, 29, 30]

    for row_n, label, ref, fmt_, bold in EXPENSE_REFS:
        ws.row_dimensions[row_n].height = 18
        lbl(ws, row_n, 1, label, bold=False)
        for col in YR_COLS:
            cl = get_column_letter(col)
            if row_n == MGMT_ROW:
                # Mgmt fee = % of EGI
                formula = f'=IFERROR(IF(OR({cl}{EGI_ROW}="",\'{_AS}\'!$C$72=""),"",' \
                          f'{cl}{EGI_ROW}*\'{_AS}\'!$C$72),"")'
            else:
                formula = f'=IFERROR(IF({ref}="","",{ref}),"")'
            frm(ws, row_n, col, formula, fmt=D0, align=AL_R)
    ws.cell(row=26, column=NCOLS, value="% of EGI — from Assumptions C72").font = FT_NOTE

    # Total Expenses (row 32)
    ws.row_dimensions[31].height = 4
    r = 32
    ws.row_dimensions[r].height = 18
    lbl(ws, r, 1, "TOTAL OPERATING EXPENSES")
    ws.cell(row=r, column=1).fill = F_TOT
    EXP_ROW = r
    for col in YR_COLS:
        cl = get_column_letter(col)
        c = frm(ws, r, col,
                f'=IFERROR(SUM({cl}23:{cl}30),"")',
                fmt=D0, align=AL_R)
        c.font = FT_TOTAL; c.fill = F_TOT
        c.border = Border(top=_GY_T, bottom=_NM_T)

    # ══════════════════════════════════════════════════════════════════════════
    # SECTION 3: NOI
    # ══════════════════════════════════════════════════════════════════════════
    ws.row_dimensions[33].height = 6
    r = 34
    ws.row_dimensions[r].height = 22
    NOI_ROW = r
    c = ws.cell(row=r, column=1, value="NET OPERATING INCOME (NOI)")
    c.font = Font(name="Calibri", size=11, bold=True, color=NAVY)
    c.fill = F_PALE
    for col in range(1, NCOLS + 1): ws.cell(row=r, column=col).fill = F_PALE
    for col in YR_COLS:
        cl = get_column_letter(col)
        c = ws.cell(row=r, column=col,
                    value=f'=IFERROR({cl}{EGI_ROW}-{cl}{EXP_ROW},"")')
        c.number_format = D0
        c.font = Font(name="Calibri", size=11, bold=True, color=NAVY)
        c.alignment = AL_R; c.fill = F_PALE
        c.border = Border(top=_NM_M, bottom=_NM_M)
    ws.cell(row=r, column=NCOLS).fill = F_PALE

    # ══════════════════════════════════════════════════════════════════════════
    # SECTION 4: UNLEVERAGED ANALYSIS
    # ══════════════════════════════════════════════════════════════════════════
    ws.row_dimensions[35].height = 6
    sec(ws, 36, "UNLEVERAGED INVESTMENT ANALYSIS", col_start=1, ncols=NCOLS)

    # Purchase Price Year 0 (row 37)
    r = 37
    ws.row_dimensions[r].height = 18
    lbl(ws, r, 1, "Purchase Price (Year 0, negative outflow)")
    UL_CF_ROW = r
    frm(ws, r, YR0_COL,
        f'=IFERROR(IF(\'{_AS}\'!$C$94="","",-\'{_AS}\'!$C$94),"")',
        fmt=D0, align=AL_R)
    ws.cell(row=r, column=NCOLS, value="Col C = Year 0 outflow").font = FT_NOTE

    # Annual NOI (row 38)
    r = 38
    ws.row_dimensions[r].height = 18
    lbl(ws, r, 1, "Annual NOI Cash Flows (Years 1-10)")
    UL_NOI_ROW = r
    for col in YR_COLS:
        cl = get_column_letter(col)
        frm(ws, r, col, f'=IFERROR({cl}{NOI_ROW},"")', fmt=D0, align=AL_R)

    # Net Reversion (row 39)
    r = 39
    ws.row_dimensions[r].height = 18
    lbl(ws, r, 1, "Net Reversion (Year 10)")
    NET_REV_UL_ROW = r
    frm(ws, r, 13,   # col M = Year 10
        f'=IFERROR(\'{_AS}\'!$I$25,"")',
        fmt=D0, align=AL_R)
    ws.cell(row=r, column=NCOLS, value="Gross − 5% bkg − 1% costs").font = FT_NOTE

    # Thin separator
    ws.row_dimensions[40].height = 4
    for ci in range(1, NCOLS + 1):
        ws.cell(row=40, column=ci).border = Border(bottom=_GY_T)

    # UNLEVERAGED CASH FLOW series — row 41
    r = 41
    ws.row_dimensions[r].height = 18
    lbl(ws, r, 1, "UNLEVERAGED CASH FLOW (for IRR)")
    ws.cell(row=r, column=1).fill = F_TOT
    UL_IRR_ROW = r
    c = frm(ws, r, YR0_COL,
            f'=IFERROR(C{UL_CF_ROW},"")', fmt=D0, align=AL_R)
    c.fill = F_TOT; c.font = FT_TOTAL
    for col in YR_COLS[:-1]:   # Y1-Y9: NOI only
        cl = get_column_letter(col)
        c = frm(ws, r, col, f'=IFERROR({cl}{UL_NOI_ROW},"")', fmt=D0, align=AL_R)
        c.fill = F_TOT; c.font = FT_TOTAL
    c = frm(ws, r, 13,   # Y10: NOI + Net Reversion
            f'=IFERROR(IF(M{UL_NOI_ROW}="","",M{UL_NOI_ROW}+M{NET_REV_UL_ROW}),"")',
            fmt=D0, align=AL_R)
    c.fill = F_TOT; c.font = FT_TOTAL
    for col in [YR0_COL] + YR_COLS:
        ws.cell(row=r, column=col).border = Border(top=_GY_T, bottom=_NM_T)

    # Unleveraged Callout Boxes — labels at row 43, values at row 44
    ws.row_dimensions[42].height = 6
    callout_boxes(ws, row=43, cf_row=UL_IRR_ROW, label_prefix="UNLEVERAGED ")
    # callout_boxes sets rows 43 (labels) and 44 (values)

    # ══════════════════════════════════════════════════════════════════════════
    # SECTION 5: DEBT SERVICE & LEVERAGED ANALYSIS
    # ══════════════════════════════════════════════════════════════════════════
    ws.row_dimensions[45].height = 6
    sec(ws, 46, "CASH FLOW AFTER DEBT SERVICE  —  LEVERAGED ANALYSIS", col_start=1, ncols=NCOLS)

    # NOI (row 47)
    r = 47
    ws.row_dimensions[r].height = 18
    lbl(ws, r, 1, "Net Operating Income")
    DS_NOI_ROW = r
    for col in YR_COLS:
        cl = get_column_letter(col)
        frm(ws, r, col, f'=IFERROR({cl}{NOI_ROW},"")', fmt=D0, align=AL_R)

    # Debt Service (row 48)
    r = 48
    ws.row_dimensions[r].height = 18
    lbl(ws, r, 1, "Annual Debt Service", bold=False)
    DS_ROW = r
    for col in YR_COLS:
        frm(ws, r, col,
            f'=IFERROR(IF(\'{_AS}\'!$I$14="","",\'{_AS}\'!$I$14),"")',
            fmt=D0, align=AL_R)
    ws.cell(row=r, column=NCOLS, value="From Assumptions — I14").font = FT_NOTE

    # CFADS (row 50)
    ws.row_dimensions[49].height = 4
    r = 50
    ws.row_dimensions[r].height = 18
    lbl(ws, r, 1, "CASH FLOW AFTER DEBT SERVICE")
    ws.cell(row=r, column=1).fill = F_TOT
    CFADS_ROW = r
    for col in YR_COLS:
        cl = get_column_letter(col)
        c = frm(ws, r, col,
                f'=IFERROR(IF({cl}{DS_NOI_ROW}="","",{cl}{DS_NOI_ROW}-{cl}{DS_ROW}),"")',
                fmt=D0, align=AL_R)
        c.font = FT_TOTAL; c.fill = F_TOT
        c.border = Border(top=_GY_T, bottom=_NM_T)

    # DCR (row 51)
    r = 51
    ws.row_dimensions[r].height = 18
    lbl(ws, r, 1, "Debt Coverage Ratio (DCR)")
    for col in YR_COLS:
        cl = get_column_letter(col)
        frm(ws, r, col,
            f'=IFERROR(IF(OR({cl}{DS_ROW}="",{cl}{DS_ROW}=0),"",{cl}{DS_NOI_ROW}/{cl}{DS_ROW}),"")',
            fmt=MX, align=AL_R)

    ws.row_dimensions[52].height = 6

    # Equity Investment Year 0 (row 53)
    r = 53
    ws.row_dimensions[r].height = 18
    lbl(ws, r, 1, "Equity Investment (Year 0, negative)")
    EQ_CF_ROW = r
    frm(ws, r, YR0_COL,
        f'=IFERROR(IF(\'{_AS}\'!$I$9="","",-\'{_AS}\'!$I$9),"")',
        fmt=D0, align=AL_R)
    ws.cell(row=r, column=NCOLS, value="Down payment from I9").font = FT_NOTE

    # Net Reversion After Debt (row 54)
    r = 54
    ws.row_dimensions[r].height = 18
    lbl(ws, r, 1, "Net Reversion After Debt Payoff")
    NET_REV_LEV_ROW = r
    frm(ws, r, 13,
        f'=IFERROR(\'{_AS}\'!$I$27,"")',
        fmt=D0, align=AL_R)
    ws.cell(row=r, column=NCOLS, value="Net Rev − loan balance").font = FT_NOTE

    # Thin separator
    ws.row_dimensions[55].height = 4
    for ci in range(1, NCOLS + 1):
        ws.cell(row=55, column=ci).border = Border(bottom=_GY_T)

    # LEVERAGED CASH FLOW series — row 56
    r = 56
    ws.row_dimensions[r].height = 18
    lbl(ws, r, 1, "LEVERAGED CASH FLOW (for IRR)")
    ws.cell(row=r, column=1).fill = F_TOT
    LEV_IRR_ROW = r
    c = frm(ws, r, YR0_COL,
            f'=IFERROR(C{EQ_CF_ROW},"")', fmt=D0, align=AL_R)
    c.fill = F_TOT; c.font = FT_TOTAL
    for col in YR_COLS[:-1]:    # Y1-Y9 CFADS
        cl = get_column_letter(col)
        c = frm(ws, r, col, f'=IFERROR({cl}{CFADS_ROW},"")', fmt=D0, align=AL_R)
        c.fill = F_TOT; c.font = FT_TOTAL
    c = frm(ws, r, 13,   # Y10 CFADS + Net Reversion after debt
            f'=IFERROR(IF(M{CFADS_ROW}="","",M{CFADS_ROW}+M{NET_REV_LEV_ROW}),"")',
            fmt=D0, align=AL_R)
    c.fill = F_TOT; c.font = FT_TOTAL
    for col in [YR0_COL] + YR_COLS:
        ws.cell(row=r, column=col).border = Border(top=_GY_T, bottom=_NM_T)

    # Leveraged Callout Boxes — labels at row 58, values at row 59
    ws.row_dimensions[57].height = 6
    callout_boxes(ws, row=58, cf_row=LEV_IRR_ROW, label_prefix="LEVERAGED ")
    # callout_boxes sets rows 58 (labels) and 59 (values)

    # ══════════════════════════════════════════════════════════════════════════
    # SECTION 6: DISPOSITION ANALYSIS
    # ══════════════════════════════════════════════════════════════════════════
    ws.row_dimensions[60].height = 6
    sec(ws, 61, "DISPOSITION ANALYSIS  —  Year 10 Exit (detail reference)", col_start=1, ncols=NCOLS)

    disp_items = [
        (62, "Exit NOI (Year 10)",             f'=IFERROR(M{NOI_ROW},"")',                                    D0,  False),
        (63, "Exit Cap Rate",                   f'=IFERROR(\'{_AS}\'!$I$20,"")',                               P2,  False),
        (64, "GROSS SALE PROCEEDS",             f'=IFERROR(IF(OR(C62="",C63="",C63=0),"",C62/C63),"")',       D0,  True),
        (65, "Less: Brokerage Fee (5%)",        f'=IFERROR(IF(C64="","",-C64*0.05),"")',                       D0,  False),
        (66, "Less: Transaction Costs (1%)",    f'=IFERROR(IF(C64="","",-C64*0.01),"")',                       D0,  False),
        (67, "NET REVERSION (Unleveraged)",     f'=IFERROR(IF(C64="","",C64+C65+C66),"")',                    D0,  True),
        (68, "Less: Remaining Loan Balance",    f'=IFERROR(\'{_AS}\'!$I$26,"")',                               D0,  False),
        (69, "NET REVERSION AFTER DEBT PAYOFF", f'=IFERROR(IF(C67="","",C67-C68),"")',                        D0,  True),
    ]
    for row_n, label, formula, fmt_, is_total in disp_items:
        ws.row_dimensions[row_n].height = 18
        lbl(ws, row_n, 1, label, bold=is_total)
        if is_total: ws.cell(row=row_n, column=1).fill = F_TOT
        c = frm(ws, row_n, 3, formula, fmt=fmt_, align=AL_R)
        if is_total:
            c.font = FT_TOTAL; c.fill = F_TOT
            c.border = Border(top=_GY_T, bottom=_NM_M)

    ws.row_dimensions[70].height = 14
    c = ws.cell(row=70, column=1,
                value="Brokerage: 5%  ·  Transaction Costs: 1%  ·  Gross Proceeds = Exit NOI ÷ Exit Cap  ·  All inputs on Assumptions & Flags tab")
    c.font = FT_NOTE; c.alignment = AL_L
    merge(ws, 70, 1, 70, NCOLS)

    # ══════════════════════════════════════════════════════════════════════════
    # SECTION 7: PRICE AT CAP RATE  —  Dynamic, centered on Asking Cap
    # ══════════════════════════════════════════════════════════════════════════
    ws.row_dimensions[71].height = 6
    sec(ws, 72, "PRICE AT CAP RATE  —  Year 1 NOI  ·  Center = Asking Cap (Assumptions C95)  ·  ±50 bps in 10 bp steps", col_start=1, ncols=NCOLS)

    # 11 cap rate columns: asking cap ±50 bps in 10 bp steps (cols 3-13)
    _VM_OFFSETS  = [-0.005, -0.004, -0.003, -0.002, -0.001, 0.000, 0.001, 0.002, 0.003, 0.004, 0.005]
    _VM_ASK_COL  = 8   # asking cap is center column (col 8 = col 3 + offset_index 5)
    _VM_GOLD_HDR = PatternFill("solid", fgColor="FFC000")   # gold for asking-cap header cell
    _VM_GOLD_VAL = PatternFill("solid", fgColor="FFF2CC")   # light gold for asking-cap value cells

    r = 73
    ws.row_dimensions[r].height = 18
    c = ws.cell(row=r, column=1, value="GOING-IN CAP RATE  →")
    c.font = FT_LABEL; c.alignment = AL_L
    for i, offset in enumerate(_VM_OFFSETS):
        col = 3 + i
        is_ask = (col == _VM_ASK_COL)
        if offset == 0.0:
            formula = f'=IFERROR(\'{_AS}\'!$C$95,"")'
        elif offset < 0:
            formula = f'=IFERROR(\'{_AS}\'!$C$95{offset:.3f},"")'
        else:
            formula = f'=IFERROR(\'{_AS}\'!$C$95+{offset:.3f},"")'
        c = frm(ws, r, col, formula, fmt=P2, align=AL_C)
        c.font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
        c.fill = _VM_GOLD_HDR if is_ask else F_NAVY

    r = 74
    ws.row_dimensions[r].height = 18
    c = ws.cell(row=r, column=1, value="PRICE AT CAP RATE  (NOI ÷ Cap)")
    c.font = FT_DATA; c.alignment = AL_L
    for i in range(11):
        col = 3 + i
        cl  = get_column_letter(col)
        is_ask = (col == _VM_ASK_COL)
        c = frm(ws, r, col,
            f'=IFERROR(IF(OR(D{NOI_ROW}="",{cl}73=0),"",D{NOI_ROW}/{cl}73),"")',
            fmt=D0, align=AL_R)
        c.font = FT_TOTAL
        if is_ask:
            c.fill = _VM_GOLD_VAL

    r = 75
    ws.row_dimensions[r].height = 18
    c = ws.cell(row=r, column=1, value="Price per SF")
    c.font = FT_DATA; c.alignment = AL_L
    for i in range(11):
        col = 3 + i
        cl  = get_column_letter(col)
        is_ask = (col == _VM_ASK_COL)
        c = frm(ws, r, col,
            f'=IFERROR(IF(OR({cl}74="",\'{_AS}\'!$C$9="",\'{_AS}\'!$C$9=0),"",{cl}74/\'{_AS}\'!$C$9),"")',
            fmt=D0, align=AL_R)
        if is_ask:
            c.fill = _VM_GOLD_VAL

    # NM bottom accent
    ws.row_dimensions[77].height = 4
    for ci in range(1, NCOLS + 1):
        ws.cell(row=77, column=ci).border = Border(bottom=Side(style='medium', color=NAVY))
