"""
BOV Master Sheet — Multi-Tenant MOB — Main assembly script.
Accepts an output_path argument for Railway service use.
"""
import openpyxl

from mob_tab_1_cover          import build_mob_cover
from mob_tab_2_execsum        import build_mob_exec_summary
from mob_tab_3_real_estate    import build_mob_real_estate_tab
from mob_tab_4_rent_roll      import build_rent_roll_tab
from mob_tab_5_lease_abstract import build_mob_lease_abstract_tab
from mob_tab_6_rent_schedule  import build_mob_rent_schedule_tab
from mob_tab_7_credit         import build_mob_credit_tab
from mob_tab_8_pro_forma      import build_mob_pro_forma_tab
from mob_tab_9_assumptions    import build_mob_assumptions_tab
from mob_tab_10_sensitivity   import build_mob_sensitivity_tab
from mob_tab_11_amortization  import build_mob_amortization_tab

TAB_COLORS = {
    "Cover":               "003DA5",
    "Executive Summary":   "003DA5",
    "Real Estate":         "1F7A8C",
    "Rent Roll":           "1F7A8C",
    "Lease Abstract":      "1F7A8C",
    "Rent Schedule":       "1F7A8C",
    "Credit":              "1F7A8C",
    "Pro Forma":           "5C6BC0",
    "Assumptions & Flags": "5C6BC0",
    "Sensitivity Analysis":"5C6BC0",
    "Amortization":        "5C6BC0",
}

FREEZE_PANES = {
    "Rent Roll":           "B8",
    "Rent Schedule":       "B16",
    "Pro Forma":           "B7",
    "Amortization":        "A10",
    "Assumptions & Flags": "B5",
}


def build_mob(output_path: str) -> None:
    """Build the MOB workbook and write it to output_path."""
    wb = openpyxl.Workbook()
    default_sheet = wb.active

    build_mob_cover(wb)
    build_mob_exec_summary(wb)
    build_mob_real_estate_tab(wb)
    build_rent_roll_tab(wb)
    build_mob_lease_abstract_tab(wb)
    build_mob_rent_schedule_tab(wb)
    build_mob_credit_tab(wb)
    build_mob_pro_forma_tab(wb)
    build_mob_assumptions_tab(wb)
    build_mob_sensitivity_tab(wb)
    build_mob_amortization_tab(wb)

    if default_sheet.title == "Sheet" and len(wb.sheetnames) > 1:
        del wb[default_sheet.title]

    for name, color in TAB_COLORS.items():
        if name in wb.sheetnames:
            wb[name].sheet_properties.tabColor = color

    for name, cell in FREEZE_PANES.items():
        if name in wb.sheetnames:
            wb[name].freeze_panes = cell

    wb.save(output_path)


if __name__ == "__main__":
    import sys
    path = sys.argv[1] if len(sys.argv) > 1 else "/home/claude/BOV_Master_MOB_Briggs.xlsx"
    print(f"Building MOB → {path}")
    build_mob(path)
    print("Done.")
