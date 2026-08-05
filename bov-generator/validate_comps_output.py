"""
validate_comps_output.py — conformance validator for a produced Briggs comps workbook.

The comps renderer (`comps_generator.py::populate_comps`) is the single source of
truth for HOW a comps workbook is built; this module is the independent CHECK that
a produced `.xlsx` actually conforms to the Briggs comps standard before it is
delivered. It exists because nothing previously verified the output, so divergent
workbooks (hand-rolled layouts, wrong sheets, blank CHAIRS, untrimmed 100-row
grids, a literal typed into a formula column) could ship unnoticed (Prompt 37).

What it asserts (per vertical — sales/dialysis/government auto-detected from the
sheet headers, or forced via `vertical=`):

  1. **Sheet set** is EXACTLY {Cover, Index, On Market, Sold} (the sales shape) —
     exact names, no invented "summary"/"methodology" tabs, none missing.
  2. **Header row (row 5)** on On Market / Sold matches the canonical column list
     for that vertical (dialysis includes CHAIRS/PATIENTS; government is the
     Agency-first set).
  3. **Formula-protected columns** (RENT/SF, SOLD/SF, SOLD CAP, TERM, INITIAL/LAST
     CAP, BID-ASK, PRICE CHG, DOM, plus the gov NOI/SF, TERM REM, FIRM REM, PPSF,
     ASK PSF, % ASK ACHIEVED …) still hold a FORMULA in every data row — never a
     literal value.
  4. **The AVG/TOTALS bar sits directly beneath the last data row** (trim applied —
     no blank filler rows between the data and the bar) and its AVERAGE/COUNT
     ranges end on that last data row (so the averages cover exactly the comps
     written, no more).
  5. **Recalc reports 0 errors** — no Excel error literal (#VALUE!, #DIV/0!, #REF!,
     #NAME?, #NULL!, #NUM!, #N/A) in any cached cell value (the export path recalcs
     via LibreOffice before validating, so the cached values are authoritative).

Returns a `ValidationResult` (pass/fail + the specific violations). The export path
(`main.py::generate_comps` and the local `populate_comps` fallback) calls
`validate_comps_file(path)` and treats a non-conforming workbook as an ERROR — a
divergent file is never delivered.

CLI:  python -m validate_comps_output <workbook.xlsx> [--vertical sales|dialysis|government]
"""

import argparse
import re
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import List, Optional

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

HEADER_ROW = 5
DATA_START_ROW = 6

# ── Shared column-width contract (the ONE source of truth) ────────────────────
# Both the renderer (`comps_generator._autofit_no_wrap`) and this validator size /
# check column widths through THESE helpers so a correctly-built workbook always
# passes: same padding, same min/max clamp, same computed-column floor, same
# rendered-length measurement (Prompt 46 — reconcile auto-fit ↔ validator).
AUTOFIT_PAD = 2          # small breathing room past the longest content
AUTOFIT_MIN = 4          # never collapse a column narrower than this
AUTOFIT_MAX = 50         # sane ceiling so one long cell can't blow the layout out
WIDTH_TOL = 0.51         # float slack when comparing widths (recalc/round trip)


def _tok(key) -> str:
    """Normalize any header spelling (UPPER 'RENT/SF' or lower 'rent_sf') to a
    single lowercase-underscore token so the width floor is identical no matter
    which side (renderer uses _norm, validator uses _n) supplies the key."""
    return re.sub(r"[^a-z0-9]+", "_", str(key or "").lower()).strip("_")


def min_content_width(key: str) -> int:
    """Floor width for COMPUTED (formula/date) columns. Their values are written
    post-autofit by the LibreOffice recalc, so a pre-recalc measurement only sees
    the header (formulas contribute 0), sizing TERM/DOM/caps/$-SF/dates too narrow
    and clipping the value. Give each a floor wide enough for its formatted result.
    Consulted identically by the renderer and the validator so both agree."""
    k = _tok(key)
    if "term" in k or k.endswith("_rem") or k in ("t_rem", "f_rem"):   # TERM / T-Rem / F-Rem
        return 7
    if k == "dom":                                        # DOM (days on market)
        return 6
    if k in ("date", "com", "exp", "termn", "termin", "expir", "on_market",
             "sold_date", "lease_comm", "lease_exp", "commence", "expire"):  # mm/dd/yyyy dates
        return 11
    if "price" in k or "ask" in k:                        # SOLD/INITIAL/LAST PRICE / ASK ($)
        return 11
    if "cap" in k or k.endswith("_sf") or "psf" in k or "rent" in k:  # %/cap and $-per-SF
        return 7
    return 0


def target_column_width(longest_content, key: str) -> float:
    """The ONE width formula: floor the measured content for computed columns, add
    padding, clamp. Renderer sets this; validator recomputes the same to check."""
    longest = max(int(longest_content or 0), min_content_width(key))
    return float(max(AUTOFIT_MIN, min(AUTOFIT_MAX, longest + AUTOFIT_PAD)))


# Sheet set for a SALES comps workbook (all three sales verticals share it).
SALES_SHEETS = ["Cover", "Index", "On Market", "Sold"]

EXCEL_ERRORS = ("#VALUE!", "#DIV/0!", "#REF!", "#NAME?", "#NULL!", "#NUM!", "#N/A")


def _n(h) -> str:
    """Normalize a header for comparison: upper, collapse whitespace, strip."""
    return re.sub(r"\s+", " ", str(h or "").strip().upper())


# ── Canonical column specs (the standard) ─────────────────────────────────────
# One entry per vertical. `headers` is the exact ordered row-5 column list for a
# sheet; `formula` is the subset of those columns that MUST hold a formula in
# every data row. `count_col` is the header the AVG-bar COUNTA() anchors on (its
# range end == the last data row); `presence_col` is an INPUT column that a real
# comp always populates (used to prove the grid was trimmed — no blank data rows).
#
# These lists are the standard, hardcoded on purpose: a validator that re-read the
# template it is checking could never catch a drifted template. They mirror
# bov-generator/templates/*.xlsx (the single source of truth) as of Prompt 37.

_SALES_ONMARKET = [
    "#", "TENANT", "ADDRESS", "CITY", "STATE", "LAND", "BUILT", "RBA",
    "RENT", "RENT/SF", "EXP", "TERM", "EXPENSES", "BUMPS", "OPTIONS",
    "INITIAL PRICE", "INITIAL CAP", "LAST PRICE", "LAST CAP", "PRICE CHG",
    "ON MARKET", "DOM", "STATUS",
]
_SALES_SOLD = [
    "#", "TENANT", "ADDRESS", "CITY", "STATE", "LAND", "BUILT", "RBA",
    "RENT", "RENT/SF", "EXP", "TERM", "EXPENSES", "BUMPS", "OPTIONS",
    "SOLD PRICE", "SOLD/SF", "SOLD CAP", "DATE", "INITIAL PRICE", "INITIAL CAP",
    "LAST PRICE", "LAST CAP", "BID-ASK SPREAD", "PRICE CHG", "ON MARKET", "DOM",
]
_SALES_OM_FORMULA = ["#", "RENT/SF", "TERM", "INITIAL CAP", "LAST CAP", "PRICE CHG", "DOM"]
_SALES_SOLD_FORMULA = ["#", "RENT/SF", "TERM", "SOLD/SF", "SOLD CAP", "INITIAL CAP",
                       "LAST CAP", "BID-ASK SPREAD", "PRICE CHG", "DOM"]


def _insert_after(seq, anchor, items):
    out = list(seq)
    i = out.index(anchor)
    out[i + 1:i + 1] = items
    return out


# Dialysis = the generic sales set with CHAIRS + PATIENTS inserted after RBA.
_DIA_ONMARKET = _insert_after(_SALES_ONMARKET, "RBA", ["CHAIRS", "PATIENTS"])
_DIA_SOLD = _insert_after(_SALES_SOLD, "RBA", ["CHAIRS", "PATIENTS"])

# Government = the Agency-first set.
_GOV_ONMARKET = [
    "#", "AGENCY", "ADDRESS", "CITY", "STATE", "GOV LEVEL", "LAND", "BUILT", "USE",
    "GOV SF LEASED", "RBA", "GOV OCCP %", "GROSS RENT", "RENT/SF", "NOI", "NOI/SF",
    "EXPIR", "TERMIN.", "TERM REM", "FIRM REM", "EXPENSES", "GUARANTOR",
    "BUMPS/DROP", "OPTIONS", "INITIAL ASK", "INITIAL CAP", "LAST ASK", "LAST CAP",
    "ASK PSF", "PRICE CHG", "ON MARKET", "DOM", "STATUS",
]
_GOV_SOLD = [
    "#", "AGENCY", "ADDRESS", "CITY", "STATE", "GOV LEVEL", "LAND", "BUILT", "USE",
    "GOV SF LEASED", "RBA", "GOV OCCP %", "GROSS RENT", "RENT/SF", "NOI", "NOI/SF",
    "EXPIR", "TERMIN.", "TERM REM", "FIRM REM", "EXPENSES", "GUARANTOR",
    "BUMPS/DROP", "OPTIONS", "SOLD PRICE", "PPSF", "SOLD CAP", "SOLD DATE",
    "INITIAL ASK", "INITIAL CAP", "LAST ASK", "LAST CAP", "% ASK ACHIEVED",
    "BID-ASK SPREAD", "PRICE CHG", "ON MARKET", "DOM",
]
_GOV_OM_FORMULA = ["#", "RENT/SF", "NOI/SF", "TERM REM", "FIRM REM", "INITIAL CAP",
                   "LAST CAP", "ASK PSF", "PRICE CHG", "DOM"]
_GOV_SOLD_FORMULA = ["#", "RENT/SF", "NOI/SF", "TERM REM", "FIRM REM", "PPSF", "SOLD CAP",
                     "INITIAL CAP", "LAST CAP", "% ASK ACHIEVED", "BID-ASK SPREAD",
                     "PRICE CHG", "DOM"]

SPECS = {
    "sales": {
        "sheets": SALES_SHEETS,
        "On Market": {"headers": _SALES_ONMARKET, "formula": _SALES_OM_FORMULA,
                      "count_col": "TENANT", "presence_col": "ADDRESS"},
        "Sold": {"headers": _SALES_SOLD, "formula": _SALES_SOLD_FORMULA,
                 "count_col": "TENANT", "presence_col": "ADDRESS"},
    },
    "dialysis": {
        "sheets": SALES_SHEETS,
        "On Market": {"headers": _DIA_ONMARKET, "formula": _SALES_OM_FORMULA,
                      "count_col": "TENANT", "presence_col": "ADDRESS"},
        "Sold": {"headers": _DIA_SOLD, "formula": _SALES_SOLD_FORMULA,
                 "count_col": "TENANT", "presence_col": "ADDRESS"},
    },
    "government": {
        "sheets": SALES_SHEETS,
        "On Market": {"headers": _GOV_ONMARKET, "formula": _GOV_OM_FORMULA,
                      "count_col": "AGENCY", "presence_col": "ADDRESS"},
        "Sold": {"headers": _GOV_SOLD, "formula": _GOV_SOLD_FORMULA,
                 "count_col": "AGENCY", "presence_col": "ADDRESS"},
    },
}


@dataclass
class ValidationResult:
    ok: bool
    vertical: Optional[str]
    violations: List[str] = field(default_factory=list)
    checks: List[str] = field(default_factory=list)   # names of checks that PASSED

    def as_dict(self) -> dict:
        return {"ok": self.ok, "vertical": self.vertical,
                "violations": self.violations, "passed_checks": self.checks}

    def __str__(self) -> str:
        head = f"{'PASS' if self.ok else 'FAIL'} (vertical={self.vertical})"
        if self.ok:
            return head + f" — {len(self.checks)} checks passed"
        return head + "\n" + "\n".join(f"  ✗ {v}" for v in self.violations)


def _detect_vertical(wb) -> Optional[str]:
    """Infer the vertical from the On Market / Sold header row: AGENCY in col B →
    government; CHAIRS & PATIENTS present → dialysis; else generic sales."""
    ws = wb["On Market"] if "On Market" in wb.sheetnames else (
        wb["Sold"] if "Sold" in wb.sheetnames else None)
    if ws is None:
        return None
    headers = {_n(ws.cell(HEADER_ROW, c).value) for c in range(1, ws.max_column + 1)}
    if _n(ws.cell(HEADER_ROW, 2).value) == "AGENCY":
        return "government"
    if "CHAIRS" in headers and "PATIENTS" in headers:
        return "dialysis"
    return "sales"


def _header_index(ws, header) -> Optional[int]:
    tgt = _n(header)
    for c in range(1, ws.max_column + 1):
        if _n(ws.cell(HEADER_ROW, c).value) == tgt:
            return c
    return None


def _find_avg_row(ws) -> Optional[int]:
    for r in range(DATA_START_ROW, ws.max_row + 1):
        if _n(ws.cell(r, 1).value) == "AVG":
            return r
    return None


def _is_formula(cell) -> bool:
    return cell.data_type == "f" or (isinstance(cell.value, str) and cell.value.startswith("="))


def _range_ends(formula: str) -> List[int]:
    """Return the end-row numbers of every A1 range (X6:X105 → 105) in a formula."""
    return [int(m) for m in re.findall(r"[A-Z]+\d+:[A-Z]+(\d+)", str(formula or ""))]


def _disp_len(cell) -> int:
    """Rendered character width of a cell (mirror of comps_generator._display_len):
    dates at display width, formulas contribute nothing, numbers comma-formatted."""
    from datetime import datetime, date
    v = cell.value
    if v is None or v == "":
        return 0
    if isinstance(v, str) and v.startswith("="):
        return 0
    if isinstance(v, (datetime, date)):
        return 10
    if isinstance(v, (int, float)):
        return len(("{:,.0f}".format(v)) if float(v).is_integer() else ("{:,.2f}".format(v)))
    return len(str(v))


# Public alias — the renderer imports this so both sides measure rendered width
# the exact same way (dates at display width, formulas → 0, numbers comma-formatted).
disp_len = _disp_len


def validate_comps_workbook(wb, vertical: Optional[str] = None) -> ValidationResult:
    """Validate an already-open workbook. `vertical` forces the spec; otherwise it
    is auto-detected from the headers."""
    vertical = vertical or _detect_vertical(wb)
    res = ValidationResult(ok=True, vertical=vertical)

    def fail(msg):
        res.ok = False
        res.violations.append(msg)

    if vertical not in SPECS:
        fail(f"unrecognized comps vertical (could not match sheets/headers to a known standard): {vertical!r}")
        return res
    spec = SPECS[vertical]

    # 1 — exact sheet set
    if set(wb.sheetnames) != set(spec["sheets"]):
        extra = sorted(set(wb.sheetnames) - set(spec["sheets"]))
        missing = sorted(set(spec["sheets"]) - set(wb.sheetnames))
        parts = []
        if missing:
            parts.append(f"missing {missing}")
        if extra:
            parts.append(f"unexpected {extra}")
        fail(f"sheet set != {spec['sheets']}: " + "; ".join(parts))
    else:
        res.checks.append("sheet_set")

    sheet_widths = {}   # sheet -> {normalized header: column width} for the shared-width check
    for sheet in ("On Market", "Sold"):
        if sheet not in wb.sheetnames:
            continue  # already reported by the sheet-set check
        ws = wb[sheet]
        s = spec[sheet]

        # 2 — canonical header row (row 5)
        actual = []
        for c in range(1, ws.max_column + 1):
            v = ws.cell(HEADER_ROW, c).value
            if v not in (None, ""):
                actual.append(_n(v))
        expected = [_n(h) for h in s["headers"]]
        if actual != expected:
            fail(f"[{sheet}] header row 5 does not match the canonical {vertical} columns. "
                 f"expected {expected} got {actual}")
        else:
            res.checks.append(f"{sheet}:headers")

        # locate the AVG bar & last data row (needed by checks 3 + 4)
        avg_row = _find_avg_row(ws)
        if avg_row is None:
            fail(f"[{sheet}] no AVG/TOTALS bar found (col A == 'AVG')")
            continue
        last_data_row = avg_row - 1
        if last_data_row < DATA_START_ROW:
            fail(f"[{sheet}] AVG bar at row {avg_row} leaves no data rows")
            continue

        # 3 — formula-protected columns hold a formula in EVERY data row
        for hdr in s["formula"]:
            col = _header_index(ws, hdr)
            if col is None:
                fail(f"[{sheet}] formula-protected column {hdr!r} not found in header row")
                continue
            bad = [r for r in range(DATA_START_ROW, last_data_row + 1)
                   if not _is_formula(ws.cell(r, col))]
            if bad:
                fail(f"[{sheet}] column {hdr!r} holds a literal (not a formula) at row(s) "
                     f"{bad[:8]}{'…' if len(bad) > 8 else ''}")
            else:
                res.checks.append(f"{sheet}:formula:{hdr}")

        # 4 — trim applied: no blank data rows above the bar, AVG ranges end on
        #     the last data row. A sheet with ZERO comps (e.g. a sales set with
        #     only Sold rows leaves On Market unused) is the pristine template
        #     grid — legitimately untrimmed, so the trim/range checks don't apply.
        pcol = _header_index(ws, s["presence_col"])
        populated = 0
        if pcol is not None:
            populated = sum(1 for r in range(DATA_START_ROW, last_data_row + 1)
                            if ws.cell(r, pcol).value not in (None, ""))
        if pcol is not None and populated == 0:
            res.checks.append(f"{sheet}:empty_unused")
            continue  # nothing written → skip trim + AVG-range enforcement

        if pcol is not None:
            blanks = [r for r in range(DATA_START_ROW, last_data_row + 1)
                      if ws.cell(r, pcol).value in (None, "")]
            if blanks:
                fail(f"[{sheet}] grid not trimmed to the AVG bar — blank {s['presence_col']} "
                     f"at data row(s) {blanks[:8]}{'…' if len(blanks) > 8 else ''} above the bar (row {avg_row})")
            else:
                res.checks.append(f"{sheet}:trimmed")

        bad_ranges = []
        for c in range(1, ws.max_column + 1):
            v = ws.cell(avg_row, c).value
            if isinstance(v, str) and v.startswith("="):
                for end in _range_ends(v):
                    if end != last_data_row:
                        bad_ranges.append((_n(ws.cell(HEADER_ROW, c).value) or f"col{c}", end))
        if bad_ranges:
            fail(f"[{sheet}] AVG bar AVERAGE/COUNT range(s) do not end on the last data row "
                 f"({last_data_row}): {bad_ranges[:8]}{'…' if len(bad_ranges) > 8 else ''}")
        else:
            res.checks.append(f"{sheet}:avg_ranges")

        # 6 — no wrapped cells anywhere in the header/data/AVG span (Prompt 43).
        wrapped = []
        for r in range(HEADER_ROW, avg_row + 1):
            for c in range(1, ws.max_column + 1):
                al = ws.cell(r, c).alignment
                if al is not None and al.wrap_text:
                    wrapped.append(ws.cell(r, c).coordinate)
        if wrapped:
            fail(f"[{sheet}] wrapped cell(s) present (wrap_text must be off): "
                 f"{wrapped[:8]}{'…' if len(wrapped) > 8 else ''}")
        else:
            res.checks.append(f"{sheet}:no_wrap")

        # 7 — every column is wide enough for its longest content (auto-fit),
        #     and record widths for the cross-sheet shared-width check.
        widths, too_narrow = {}, []
        for c in range(1, ws.max_column + 1):
            hdr = ws.cell(HEADER_ROW, c).value
            if hdr in (None, ""):
                continue
            longest = 0
            for r in range(HEADER_ROW, avg_row + 1):
                longest = max(longest, _disp_len(ws.cell(r, c)))
            # Apply the SAME computed-column floor the renderer sizes to (formula
            # cells measure 0 pre-recalc), so a formula-protected column the
            # renderer can only floor-size is checked consistently on both sides.
            longest = max(longest, min_content_width(hdr))
            w = ws.column_dimensions[get_column_letter(c)].width
            widths[_n(hdr)] = w
            # width must cover the content (capped at the renderer's 50-char ceiling)
            if w is None or w + WIDTH_TOL < min(longest, AUTOFIT_MAX):
                too_narrow.append((_n(hdr), round(w or 0, 1), longest))
        if too_narrow:
            fail(f"[{sheet}] column(s) narrower than their content (not auto-fit): "
                 f"{too_narrow[:8]}{'…' if len(too_narrow) > 8 else ''}")
        else:
            res.checks.append(f"{sheet}:widths_fit")
        sheet_widths[sheet] = widths

    # 8 — shared columns share ONE width across On Market and Sold (tabs line up).
    if "On Market" in sheet_widths and "Sold" in sheet_widths:
        om, sold = sheet_widths["On Market"], sheet_widths["Sold"]
        mism = [(h, round(om[h] or 0, 1), round(sold[h] or 0, 1))
                for h in (set(om) & set(sold))
                if abs((om[h] or 0) - (sold[h] or 0)) > WIDTH_TOL]
        if mism:
            fail(f"shared column widths differ between On Market and Sold: "
                 f"{sorted(mism)[:8]}{'…' if len(mism) > 8 else ''}")
        else:
            res.checks.append("shared_widths_match")

    return res


def validate_comps_file(path: str, vertical: Optional[str] = None,
                        check_recalc_errors: bool = True) -> ValidationResult:
    """Open a produced comps `.xlsx` and validate it. Structural checks read the
    formula workbook; the recalc-error check (5) reads the cached data_only values
    (populated by the LibreOffice recalc the export path runs before validating)."""
    p = Path(path)
    if not p.exists():
        return ValidationResult(ok=False, vertical=None, violations=[f"file not found: {path}"])

    wb = load_workbook(p, data_only=False)
    try:
        res = validate_comps_workbook(wb, vertical=vertical)
    finally:
        wb.close()

    if check_recalc_errors:
        wb_vals = load_workbook(p, data_only=True)
        try:
            errs = []
            for sn in wb_vals.sheetnames:
                ws = wb_vals[sn]
                if not hasattr(ws, "iter_rows"):
                    continue
                for row in ws.iter_rows():
                    for cell in row:
                        if isinstance(cell.value, str):
                            for e in EXCEL_ERRORS:
                                if e in cell.value:
                                    errs.append(f"{sn}!{cell.coordinate}={e}")
                                    break
        finally:
            wb_vals.close()
        if errs:
            res.ok = False
            res.violations.append(
                f"recalc reports {len(errs)} formula error(s): "
                f"{errs[:8]}{'…' if len(errs) > 8 else ''}")
        else:
            res.checks.append("recalc_zero_errors")

    return res


class CompsConformanceError(Exception):
    """Raised by the export path when a produced workbook fails conformance."""
    def __init__(self, result: ValidationResult):
        self.result = result
        super().__init__("comps workbook failed conformance:\n" + "\n".join(result.violations))


def assert_conforms(path: str, vertical: Optional[str] = None,
                    check_recalc_errors: bool = True) -> ValidationResult:
    """Validate and raise CompsConformanceError on any violation. Returns the
    (passing) result on success."""
    res = validate_comps_file(path, vertical=vertical, check_recalc_errors=check_recalc_errors)
    if not res.ok:
        raise CompsConformanceError(res)
    return res


def _main(argv=None) -> int:
    ap = argparse.ArgumentParser(description="Validate a produced Briggs comps workbook for conformance.")
    ap.add_argument("workbook", help="path to the produced comps .xlsx")
    ap.add_argument("--vertical", choices=["sales", "dialysis", "government"],
                    help="force the vertical spec (default: auto-detect from headers)")
    ap.add_argument("--no-recalc-check", action="store_true",
                    help="skip the cached-value Excel-error scan (check 5)")
    args = ap.parse_args(argv)

    res = validate_comps_file(args.workbook, vertical=args.vertical,
                              check_recalc_errors=not args.no_recalc_check)
    print(res)
    return 0 if res.ok else 1


if __name__ == "__main__":
    sys.exit(_main())
