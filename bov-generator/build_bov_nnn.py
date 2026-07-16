"""
BOV Master Sheet — Single-Tenant NNN — Main assembly script.
Accepts an output_path argument for Railway service use.
"""
import openpyxl

from bov_tabs_1_cover_execsum import build_cover, build_exec_summary
from bov_tabs_3_real_estate   import build_real_estate_tab
from bov_tabs_4_lease_abstract import build_lease_abstract_tab
from bov_tabs_5_rent_schedule  import build_rent_schedule_tab
from bov_tabs_6_credit         import build_credit_tab
from bov_tabs_7_pro_forma      import build_pro_forma_tab
from bov_tabs_8_assumptions    import build_assumptions_tab
from bov_tabs_9_sensitivity    import build_sensitivity_tab
from bov_tabs_10_amortization  import build_amortization_tab

TAB_COLORS = {
    "Cover":               "003DA5",
    "Executive Summary":   "003DA5",
    "Real Estate":         "1F7A8C",
    "Lease Abstract":      "1F7A8C",
    "Rent Schedule":       "1F7A8C",
    "Credit":              "1F7A8C",
    "Pro Forma":           "5C6BC0",
    "Assumptions & Flags": "5C6BC0",
    "Sensitivity Analysis":"5C6BC0",
    "Amortization":        "5C6BC0",
}

FREEZE_PANES = {
    "Rent Schedule":       "B16",
    "Pro Forma":           "B7",
    "Amortization":        "A10",
    "Assumptions & Flags": "B5",
}


def build_nnn(output_path: str) -> None:
    """Build the NNN workbook and write it to output_path."""
    wb = openpyxl.Workbook()
    default_sheet = wb.active

    build_cover(wb)
    build_exec_summary(wb)
    build_real_estate_tab(wb)
    build_lease_abstract_tab(wb)
    build_rent_schedule_tab(wb)
    build_credit_tab(wb)
    build_pro_forma_tab(wb)
    build_assumptions_tab(wb)
    build_sensitivity_tab(wb)
    build_amortization_tab(wb)

    if default_sheet.title == "Sheet" and len(wb.sheetnames) > 1:
        del wb[default_sheet.title]

    for name, color in TAB_COLORS.items():
        if name in wb.sheetnames:
            wb[name].sheet_properties.tabColor = color

    for name, cell in FREEZE_PANES.items():
        if name in wb.sheetnames:
            wb[name].freeze_panes = cell

    wb.save(output_path)


if __name__ == "__main__":
    import sys
    path = sys.argv[1] if len(sys.argv) > 1 else "/home/claude/BOV_Master_NNN_Briggs.xlsx"
    print(f"Building NNN → {path}")
    build_nnn(path)
    print("Done.")
