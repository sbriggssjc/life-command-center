"""BOV v2 (NNN) — Tab 9: Sensitivity Analysis.

6-metric × 11-going-in-cap sensitivity matrix.
Helper cash flow rows (rows 23-33 unlev, 35-45 lev) enable LibreOffice-safe IRR().
Each helper row holds real annual cash flows; IRR() references a cell range,
never an array literal — required for LibreOffice recalculation compatibility.

NNN Assumptions cell refs:
  C33 = Year 1 NOI
  C36 = Purchase Price
  C37 = Going-In Cap Rate
  C25 = Annual Rent Escalation
  I20 = Exit Cap Rate
  I13 = Hold Period (years)
  I10 = Interest Rate
  I11 = Amortization (years)
  I9  = Equity / Down Payment
"""
from bov_constants import *
from openpyxl.styles import Border, Side, Font, PatternFill
from openpyxl.utils import get_column_letter

_NM_M = Side(style='medium', color=NAVY)
_NM_T = Side(style='thin',   color="CCCCCC")
_GY_T = Side(style='thin',   color="CCCCCC")

_AS = "Assumptions & Flags"

# NNN Assumptions references
_NOI   = f"'{_AS}'!$C$33"
_GCAP  = f"'{_AS}'!$C$37"
_ESC   = f"'{_AS}'!$C$25"
_XCAP  = f"'{_AS}'!$I$20"
_HOLD  = f"'{_AS}'!$I$13"
_RATE  = f"'{_AS}'!$I$10"
_AMORT = f"'{_AS}'!$I$11"

# 11 going-in cap rate columns: center = asking cap (C37), ±50bps in 10bp steps
CAP_OFFSETS = [-0.005, -0.004, -0.003, -0.002, -0.001, 0.000, 0.001, 0.002, 0.003, 0.004, 0.005]
ASK_IDX     = 5           # index of 0.000 offset
ASK_COL     = 3 + ASK_IDX  # = column 8 (Excel col H)
DATA_COLS   = list(range(3, 14))   # columns 3–13

# Row assignments (display section)
CAP_HDR_ROW = 7
M1_ROW      = 9    # Avg Going-In Yield (hold-weighted)
M2_ROW      = 10   # Unleveraged Equity Realization Multiple
M3_ROW      = 11   # Unleveraged IRR
M4_ROW      = 13   # Avg Cash-on-Cash (Leveraged)
M5_ROW      = 14   # Leveraged Equity Realization Multiple
M6_ROW      = 15   # Leveraged IRR

# Helper block row ranges (11 rows each: Y0 through Y10)
UL_HELP_START  = 23
UL_HELP_END    = 33
LEV_HELP_START = 35
LEV_HELP_END   = 45

# Gold fills for asking-cap column
_GOLD_HDR = PatternFill("solid", fgColor="FFC000")
_GOLD_VAL = PatternFill("solid", fgColor="FFF2CC")


def _cap_formula(offset):
    """Return the Excel formula for the cap rate header cell."""
    if offset == 0.0:
        return f'=IFERROR({_GCAP},"")'
    elif offset > 0:
        return f'=IFERROR({_GCAP}+{round(offset, 4)},"")'
    else:
        return f'=IFERROR({_GCAP}-{round(abs(offset), 4)},"")'


def build_sensitivity_tab(wb):
    ws = wb.create_sheet("Sensitivity Analysis")
    ws.sheet_view.showGridLines = False

    # ── Column widths ──────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 3
    for col in DATA_COLS:
        ws.column_dimensions[get_column_letter(col)].width = 10

    # ── Header ────────────────────────────────────────────────────────────────
    r = 1;  ws.row_dimensions[r].height = 6
    r = 2;  ws.row_dimensions[r].height = 28
    c = ws.cell(row=r, column=1, value="SENSITIVITY ANALYSIS")
    c.font = FT_TITLE; c.alignment = AL_L
    merge(ws, r, 1, r, 13)

    r = 3;  ws.row_dimensions[r].height = 14
    c = ws.cell(row=r, column=1,
                value="Six return metrics × 11 going-in cap rate scenarios  ·  Base case from Assumptions & Flags  ·  Asking cap highlighted gold")
    c.font = FT_NOTE; c.alignment = AL_L
    merge(ws, r, 1, r, 13)

    r = 4;  ws.row_dimensions[r].height = 8

    # ── Base case inputs strip ─────────────────────────────────────────────────
    r = 5;  ws.row_dimensions[r].height = 16
    sec(ws, r, "BASE CASE INPUTS  —  From Assumptions & Flags", col_start=1, ncols=12)

    r = 6;  ws.row_dimensions[r].height = 16
    strip_items = [
        ("Year 1 NOI",     f'=IFERROR({_NOI},"")',                     D0),
        ("Purchase Price", f'=IFERROR(\'{_AS}\'!$C$36,"")',             D0),
        ("Going-In Cap",   f'=IFERROR({_GCAP},"")',                    P2),
        ("Escalation",     f'=IFERROR({_ESC},"")',                     P2),
        ("Exit Cap",       f'=IFERROR({_XCAP},"")',                    P2),
        ("Hold (Years)",   f'=IFERROR({_HOLD},"")',                    "0"),
    ]
    for i, (label, formula, fmt) in enumerate(strip_items):
        col = 3 + i * 2
        c = ws.cell(row=r, column=col, value=label)
        c.font = FT_NOTE; c.alignment = AL_C; c.fill = F_PALE
        frm(ws, r, col + 1, formula, fmt=fmt, align=AL_C).fill = F_PALE

    # ══════════════════════════════════════════════════════════════════════════
    # CAP RATE HEADER ROW (row 7)
    # ══════════════════════════════════════════════════════════════════════════
    r = CAP_HDR_ROW   # = 7
    ws.row_dimensions[r].height = 20

    c = ws.cell(row=r, column=1, value="METRIC  ↓  |  Going-In Cap Rate  →")
    c.font = FT_LABEL; c.fill = F_NAVY; c.alignment = AL_L

    for idx, offset in enumerate(CAP_OFFSETS):
        col = 3 + idx
        frm(ws, r, col, _cap_formula(offset), fmt=P2, align=AL_C)
        cell = ws.cell(row=r, column=col)
        cell.font = FT_CHDR
        cell.fill = _GOLD_HDR if col == ASK_COL else F_NAVY

    # ══════════════════════════════════════════════════════════════════════════
    # SECTION A: UNLEVERAGED RETURNS
    # ══════════════════════════════════════════════════════════════════════════
    r = 8;  ws.row_dimensions[r].height = 16
    sec(ws, r, "UNLEVERAGED RETURNS", col_start=1, ncols=12)

    # M1: Avg Going-In Yield (cap rate growth-weighted over hold period)
    r = M1_ROW;  ws.row_dimensions[r].height = 18
    c = ws.cell(row=r, column=1, value="Avg Going-In Yield (hold-weighted)")
    c.font = FT_LABEL; c.alignment = AL_L
    for col in DATA_COLS:
        cl  = get_column_letter(col)
        cap = f"${cl}${CAP_HDR_ROW}"
        formula = (
            f'=IFERROR(IF(OR({_NOI}="",{cap}=0),"",IF({_ESC}=0,{cap},'
            f'{cap}*(POWER(1+{_ESC},{_HOLD})-1)/({_ESC}*{_HOLD}))),"")'
        )
        c = frm(ws, r, col, formula, fmt=P2, align=AL_C)
        c.border = Border(bottom=_GY_T)
        if col == ASK_COL:
            c.fill = _GOLD_VAL

    # M2: Unleveraged Equity Realization Multiple
    r = M2_ROW;  ws.row_dimensions[r].height = 18
    c = ws.cell(row=r, column=1, value="Unleveraged Equity Realization Multiple")
    c.font = FT_LABEL; c.alignment = AL_L
    for col in DATA_COLS:
        cl  = get_column_letter(col)
        cap = f"${cl}${CAP_HDR_ROW}"
        formula = (
            f'=IFERROR(IF(OR({_NOI}="",{cap}=0),"",'
            f'SUM(${cl}${UL_HELP_START + 1}:${cl}${UL_HELP_END})/({_NOI}/{cap})),"")'
        )
        c = frm(ws, r, col, formula, fmt=MX, align=AL_C)
        c.border = Border(bottom=_GY_T)
        if col == ASK_COL:
            c.fill = _GOLD_VAL

    # M3: Unleveraged IRR (references helper block A cell range)
    r = M3_ROW;  ws.row_dimensions[r].height = 18
    c = ws.cell(row=r, column=1, value="Unleveraged IRR")
    c.font = Font(name="Calibri", bold=True, size=10, color=NAVY)
    c.alignment = AL_L; c.fill = F_PALE
    for col in DATA_COLS:
        cl  = get_column_letter(col)
        cap = f"${cl}${CAP_HDR_ROW}"
        formula = (
            f'=IFERROR(IF(OR({_NOI}="",{cap}=0),"",'
            f'IRR(${cl}${UL_HELP_START}:${cl}${UL_HELP_END})),"")'
        )
        c = frm(ws, r, col, formula, fmt=P2, align=AL_C)
        c.font = Font(name="Calibri", bold=True, size=10, color=NAVY)
        c.fill = _GOLD_VAL if col == ASK_COL else F_PALE
        c.border = Border(bottom=_NM_T)

    # ══════════════════════════════════════════════════════════════════════════
    # SECTION B: LEVERAGED RETURNS
    # ══════════════════════════════════════════════════════════════════════════
    r = 12;  ws.row_dimensions[r].height = 16
    sec(ws, r, "LEVERAGED RETURNS  —  65% LTV  ·  Rate and Amortization from Assumptions", col_start=1, ncols=12)

    # M4: Avg Cash-on-Cash Return (Leveraged)
    r = M4_ROW;  ws.row_dimensions[r].height = 18
    c = ws.cell(row=r, column=1, value="Avg Cash-on-Cash Return (Leveraged)")
    c.font = FT_LABEL; c.alignment = AL_L
    for col in DATA_COLS:
        cl  = get_column_letter(col)
        cap = f"${cl}${CAP_HDR_ROW}"
        loan    = f"(0.65*{_NOI}/{cap})"
        ds_ann  = f"(-PMT({_RATE}/12,{_AMORT}*12,{loan})*12)"
        avg_noi = f"IF({_ESC}=0,{_NOI},{_NOI}*(POWER(1+{_ESC},{_HOLD})-1)/({_ESC}*{_HOLD}))"
        equity  = f"(0.35*{_NOI}/{cap})"
        formula = (
            f'=IFERROR(IF(OR({_NOI}="",{cap}=0),"",'
            f'({avg_noi}-{ds_ann})/{equity}),"")'
        )
        c = frm(ws, r, col, formula, fmt=P2, align=AL_C)
        c.border = Border(bottom=_GY_T)
        if col == ASK_COL:
            c.fill = _GOLD_VAL

    # M5: Leveraged Equity Realization Multiple
    r = M5_ROW;  ws.row_dimensions[r].height = 18
    c = ws.cell(row=r, column=1, value="Leveraged Equity Realization Multiple")
    c.font = FT_LABEL; c.alignment = AL_L
    for col in DATA_COLS:
        cl  = get_column_letter(col)
        cap = f"${cl}${CAP_HDR_ROW}"
        equity = f"(0.35*{_NOI}/{cap})"
        formula = (
            f'=IFERROR(IF(OR({_NOI}="",{cap}=0),"",'
            f'SUM(${cl}${LEV_HELP_START + 1}:${cl}${LEV_HELP_END})/{equity}),"")'
        )
        c = frm(ws, r, col, formula, fmt=MX, align=AL_C)
        c.border = Border(bottom=_GY_T)
        if col == ASK_COL:
            c.fill = _GOLD_VAL

    # M6: Leveraged IRR (references helper block B cell range)
    r = M6_ROW;  ws.row_dimensions[r].height = 18
    c = ws.cell(row=r, column=1, value="Leveraged IRR")
    c.font = Font(name="Calibri", bold=True, size=10, color=NAVY)
    c.alignment = AL_L; c.fill = F_PALE
    for col in DATA_COLS:
        cl  = get_column_letter(col)
        cap = f"${cl}${CAP_HDR_ROW}"
        formula = (
            f'=IFERROR(IF(OR({_NOI}="",{cap}=0),"",'
            f'IRR(${cl}${LEV_HELP_START}:${cl}${LEV_HELP_END})),"")'
        )
        c = frm(ws, r, col, formula, fmt=P2, align=AL_C)
        c.font = Font(name="Calibri", bold=True, size=10, color=NAVY)
        c.fill = _GOLD_VAL if col == ASK_COL else F_PALE
        c.border = Border(bottom=_NM_T)

    # ── Display section bottom accent ──────────────────────────────────────────
    r = 16;  ws.row_dimensions[r].height = 4
    for ci in range(1, 14):
        ws.cell(row=r, column=ci).border = Border(bottom=_NM_M)

    # ── Notes above helper blocks ──────────────────────────────────────────────
    r = 17;  ws.row_dimensions[r].height = 8
    r = 18;  ws.row_dimensions[r].height = 12
    c = ws.cell(row=r, column=1,
                value="Helper rows 23-33 (unleveraged) and 35-45 (leveraged): annual cash flows per cap scenario — enables LibreOffice-safe IRR(cell range)")
    c.font = FT_NOTE; c.alignment = AL_L; c.fill = F_PALE
    merge(ws, r, 1, r, 13)

    r = 19;  ws.row_dimensions[r].height = 12
    c = ws.cell(row=r, column=1,
                value="Assumptions: 65% LTV  ·  Rate from I10  ·  Amortization from I11  ·  Hold from I13  ·  Exit Cap from I20  ·  Disposition: 6% gross proceeds")
    c.font = FT_NOTE; c.alignment = AL_L
    merge(ws, r, 1, r, 13)

    r = 20;  ws.row_dimensions[r].height = 14
    ws.cell(row=r, column=1, value="YEAR").font = FT_NOTE
    for col in DATA_COLS:
        cl = get_column_letter(col)
        c = ws.cell(row=r, column=col, value=f"={cl}{CAP_HDR_ROW}")
        c.number_format = P2; c.font = FT_NOTE; c.alignment = AL_C

    r = 21;  ws.row_dimensions[r].height = 4
    r = 22;  ws.row_dimensions[r].height = 10
    ws.cell(row=r, column=1, value="UNLEVERAGED CASH FLOWS").font = FT_NOTE

    # ══════════════════════════════════════════════════════════════════════════
    # HELPER BLOCK A: UNLEVERAGED CASH FLOWS (ROWS 23-33, Y0-Y10)
    # ══════════════════════════════════════════════════════════════════════════
    for yr in range(11):
        rr = UL_HELP_START + yr
        ws.row_dimensions[rr].height = 12
        ws.cell(row=rr, column=1, value=f"Y{yr}").font = FT_NOTE

        for col in DATA_COLS:
            cl  = get_column_letter(col)
            cap = f"${cl}${CAP_HDR_ROW}"

            if yr == 0:
                # Year 0: negative purchase price outflow
                formula = (
                    f'=IFERROR(IF(OR({_NOI}="",{cap}=0),"",-{_NOI}/{cap}),"")'
                )
            elif yr < 10:
                # Years 1-9: escalated NOI
                pow_n = yr - 1
                noi_yr = _NOI if pow_n == 0 else f'{_NOI}*(1+{_ESC})^{pow_n}'
                formula = (
                    f'=IFERROR(IF(${cl}${UL_HELP_START}="","",{noi_yr}),"")'
                )
            else:
                # Year 10: Y10 NOI + net reversion (6% disposition costs)
                formula = (
                    f'=IFERROR(IF(${cl}${UL_HELP_START}="","",'
                    f'{_NOI}*(1+{_ESC})^9+({_NOI}*(1+{_ESC})^9/{_XCAP})*0.94),"")'
                )
            c = ws.cell(row=rr, column=col, value=formula)
            c.number_format = D0; c.font = FT_NOTE; c.alignment = AL_R

    # ── Separator before leveraged block ──────────────────────────────────────
    r = 34;  ws.row_dimensions[r].height = 10
    ws.cell(row=r, column=1, value="LEVERAGED CASH FLOWS").font = FT_NOTE

    # ══════════════════════════════════════════════════════════════════════════
    # HELPER BLOCK B: LEVERAGED CASH FLOWS (ROWS 35-45, Y0-Y10)
    # ══════════════════════════════════════════════════════════════════════════
    for yr in range(11):
        rr = LEV_HELP_START + yr
        ws.row_dimensions[rr].height = 12
        ws.cell(row=rr, column=1, value=f"Y{yr}").font = FT_NOTE

        for col in DATA_COLS:
            cl  = get_column_letter(col)
            cap    = f"${cl}${CAP_HDR_ROW}"
            loan   = f"(0.65*{_NOI}/{cap})"
            ds_ann = f"(-PMT({_RATE}/12,{_AMORT}*12,{loan})*12)"

            if yr == 0:
                # Year 0: negative equity (35% down payment)
                formula = (
                    f'=IFERROR(IF(OR({_NOI}="",{cap}=0),"",-0.35*{_NOI}/{cap}),"")'
                )
            elif yr < 10:
                # Years 1-9: NOI - annual debt service
                pow_n  = yr - 1
                noi_yr = _NOI if pow_n == 0 else f'{_NOI}*(1+{_ESC})^{pow_n}'
                formula = (
                    f'=IFERROR(IF(${cl}${LEV_HELP_START}="","",{noi_yr}-{ds_ann}),"")'
                )
            else:
                # Year 10: Y10 NOI - DS + net reversion after debt payoff
                noi_10   = f'{_NOI}*(1+{_ESC})^9'
                net_rev  = f'({noi_10}/{_XCAP})*0.94'
                rem_loan = f'(-PV({_RATE}/12,({_AMORT}-{_HOLD})*12,PMT({_RATE}/12,{_AMORT}*12,{loan})))'
                formula  = (
                    f'=IFERROR(IF(${cl}${LEV_HELP_START}="","",'
                    f'{noi_10}-{ds_ann}+{net_rev}-{rem_loan}),"")'
                )
            c = ws.cell(row=rr, column=col, value=formula)
            c.number_format = D0; c.font = FT_NOTE; c.alignment = AL_R

    # ── Final bottom accent ────────────────────────────────────────────────────
    bot_r = LEV_HELP_END + 2
    ws.row_dimensions[bot_r].height = 4
    for ci in range(1, 14):
        ws.cell(row=bot_r, column=ci).border = Border(bottom=_NM_M)
