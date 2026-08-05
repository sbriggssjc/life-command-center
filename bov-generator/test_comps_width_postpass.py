"""
test_comps_width_postpass.py — the post-recalc shared-width re-apply (Prompt 48)
resyncs a LibreOffice-desynced shared column WITHOUT dropping cached formula values.

LibreOffice cannot be driven in the test env, so the post-recalc state is
synthesized by zip surgery: a shared column (PATIENTS) is given DIFFERENT widths on
On Market vs Sold (mimicking LibreOffice's content-fit re-optimization on store),
and cached `<v>` values are injected into formula cells (mimicking the results
recalc computed). The test then asserts the pre-pass workbook FAILS conformance on
the shared-width mismatch, and that after `reapply_shared_widths` it PASSES while the
injected cached values survive byte-for-byte.
"""

import re
import sys
import zipfile
from pathlib import Path

import pytest
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

_HERE = Path(__file__).parent
if str(_HERE) not in sys.path:
    sys.path.insert(0, str(_HERE))

from comps_generator import populate_comps
from validate_comps_output import validate_comps_file
from comps_width_postpass import reapply_shared_widths


def _dialysis_payload():
    def sold(i):
        return {"tenant": f"DaVita {i}", "address": f"{i} Old Camp Rd", "city": "The Villages",
                "state": "FL", "land": 1.2, "built": 2015, "rba": 8000 + i, "chairs": 18,
                "patients": 120 + i, "rent": 250000, "exp": "2035-01-01", "expenses": "NNN",
                "bumps": "2%", "renewal_options": "2x5", "sold_price": 4000000 + i * 1000,
                "date": "2024-06-01", "initial_price": 4200000, "last_price": 4100000,
                "on_market": "2023-01-01"}

    def om(i):
        # NO patient count — listings never carry patient counts (the desync trigger)
        return {"tenant": f"Fresenius {i}", "address": f"{i} Market St", "city": "Ocala",
                "state": "FL", "land": 1.0, "built": 2010, "rba": 7000 + i, "chairs": 16,
                "rent": 230000, "exp": "2032-01-01", "expenses": "NNN", "bumps": "2%",
                "renewal_options": "2x5", "initial_price": 3800000, "last_price": 3700000,
                "on_market": "2024-02-01"}

    return {"comp_type": "sales", "dialysis": True,
            "on_market": [om(i) for i in range(1, 6)],
            "sold": [sold(i) for i in range(1, 7)]}


def _sheet_xml_path(zf, name):
    wb = zf.read("xl/workbook.xml").decode()
    rid = dict(re.findall(r'<sheet name="([^"]+)"[^>]*r:id="([^"]+)"', wb))[name]
    rels = zf.read("xl/_rels/workbook.xml.rels").decode()
    # map Id -> Target
    idmap = {}
    for m in re.finditer(r"<Relationship\b([^>]*)/?>", rels):
        a = dict(re.findall(r'([\w:]+)="([^"]*)"', m.group(1)))
        idmap[a["Id"]] = a["Target"]
    p = idmap[rid].lstrip("/")
    return p if p.startswith("xl/") else "xl/" + p


def _find_header_col(xml, header):
    for m in re.finditer(r'<c r="([A-Z]+)5"[^>]*>(.*?)</c>', xml, re.S):
        if re.sub(r"<[^>]+>", "", m.group(2)).strip().upper() == header:
            return m.group(1)
    return None


def _synthesize_postrecalc(src, dst):
    """Take a freshly populated (pre-recalc) workbook and produce a synthetic
    post-recalc file: PATIENTS desynced (10 on On Market, 13 on Sold) + cached
    formula values injected. Returns (patients_col_letter,)."""
    with zipfile.ZipFile(src) as zf:
        names = zf.namelist()
        infos = {i.filename: i for i in zf.infolist()}
        data = {n: zf.read(n) for n in names}
        om_path = _sheet_xml_path(zf, "On Market")
        sold_path = _sheet_xml_path(zf, "Sold")

    def inject_cached(xml):
        xml = xml.replace('<f>IFERROR(ROW()-5,"")</f><v />',
                          '<f>IFERROR(ROW()-5,"")</f><v>1</v>')
        return re.sub(r"(<f>[^<]*</f>)<v />", r"\1<v>42.5</v>", xml)

    def set_width(xml, col_letter, width):
        idx = column_index_from_string(col_letter)
        pat = re.compile(r'<col\b[^>]*min="%d"[^>]*max="%d"[^>]*/>' % (idx, idx))
        return pat.sub(lambda m: re.sub(r'width="[^"]*"', f'width="{width}"', m.group(0)),
                       xml, count=1)

    om = inject_cached(data[om_path].decode())
    sold = inject_cached(data[sold_path].decode())
    pcol = _find_header_col(om, "PATIENTS")
    om = set_width(om, pcol, 10)
    sold = set_width(sold, pcol, 13)
    data[om_path] = om.encode()
    data[sold_path] = sold.encode()

    with zipfile.ZipFile(dst, "w") as zf:
        for n in names:
            zf.writestr(infos[n], data[n], compress_type=infos[n].compress_type)
    return pcol


def _count_cached(path, sheet="On Market"):
    wb = load_workbook(path, data_only=True)
    try:
        ws = wb[sheet]
        a6 = ws["A6"].value
        n = sum(1 for row in ws.iter_rows() for c in row if c.value == 42.5)
    finally:
        wb.close()
    return a6, n


def test_postpass_resyncs_shared_width_and_preserves_cached_values(tmp_path):
    pre = tmp_path / "pre.xlsx"
    populate_comps(_dialysis_payload(), str(pre))

    post = tmp_path / "postrecalc.xlsx"
    _synthesize_postrecalc(str(pre), str(post))

    # 1 — the synthetic post-recalc desync FAILS conformance exactly like production.
    res = validate_comps_file(str(post), check_recalc_errors=True)
    assert not res.ok
    assert any("shared column widths differ" in v and "PATIENTS" in v for v in res.violations), \
        res.violations

    a6_before, n_before = _count_cached(str(post))
    assert a6_before == 1 and n_before > 0

    # 2 — the post-pass re-applies the shared width (last mutation before validate).
    applied = reapply_shared_widths(str(post))
    om_widths, sold_widths = applied["On Market"], applied["Sold"]
    # PATIENTS shares ONE width across both sheets now.
    pcol_idx = column_index_from_string(_find_header_col(
        zipfile.ZipFile(str(post)).read("xl/worksheets/sheet3.xml").decode(), "PATIENTS"))
    assert om_widths[pcol_idx] == sold_widths[pcol_idx]

    # 3 — conformance now PASSES.
    res2 = validate_comps_file(str(post), check_recalc_errors=True)
    assert res2.ok, res2.violations

    # 4 — the cached formula values recalc computed SURVIVED the width rewrite.
    a6_after, n_after = _count_cached(str(post))
    assert a6_after == 1
    assert n_after == n_before

    # 5 — the workbook still opens and still carries FORMULAS (not stripped).
    wbf = load_workbook(str(post), data_only=False)
    try:
        assert str(wbf["On Market"]["A6"].value).startswith("=")
        assert set(wbf.sheetnames) == {"Cover", "Index", "On Market", "Sold"}
    finally:
        wbf.close()


def test_postpass_idempotent(tmp_path):
    pre = tmp_path / "pre.xlsx"
    populate_comps(_dialysis_payload(), str(pre))
    post = tmp_path / "postrecalc.xlsx"
    _synthesize_postrecalc(str(pre), str(post))
    first = reapply_shared_widths(str(post))
    second = reapply_shared_widths(str(post))
    assert first == second
    assert validate_comps_file(str(post), check_recalc_errors=True).ok


if __name__ == "__main__":
    sys.exit(pytest.main([__file__, "-q"]))
