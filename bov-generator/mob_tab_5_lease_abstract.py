"""BOV MOB — Tab 5: Lease Abstract (Leg 2).
Layout: Short-form summary table at TOP (5 tenant columns), then 5 per-tenant
long-form abstract sections stacked below.
"""
from bov_constants import *
from openpyxl.styles import Border, Side
from openpyxl.utils import get_column_letter

_NM_M = Side(style='medium', color=NAVY)
_NM_T = Side(style='thin',   color="CCCCCC")

# Tenant Assumptions cell map (Name row, SF row from 'Assumptions & Flags')
TENANT_MAP = [
    {"name_cell": "C18", "sf_cell": "C20", "label": "TENANT 1"},
    {"name_cell": "C27", "sf_cell": "C29", "label": "TENANT 2"},
    {"name_cell": "C36", "sf_cell": "C38", "label": "TENANT 3"},
    {"name_cell": "C45", "sf_cell": "C47", "label": "TENANT 4"},
    {"name_cell": "C54", "sf_cell": "C56", "label": "TENANT 5"},
]

# Short-form provisions (one value per tenant column)
SF_PROVISIONS = [
    "Tenant (Lease)",
    "Guarantor",
    "Lease Type",
    "Lease Commencement",
    "Lease Expiration",
    "Remaining Lease Term",
    "Suite / Unit",
    "Leased SF",
    "Year 1 Base Rent",
    "Rent / SF",
    "Rent Escalations",
    "Renewal Options",
    "Renewal Rent Method",
    "Expense Structure",
    "LL Responsibilities",
    "Early Termination",
    "Assignment / Subletting",
    "Key Lease Strengths",
    "Key Lease Risks",
]

# Long-form rows: (article_section, clause_provision) — same for each tenant
LONG_ROWS = [
    ("Lease Identification",    "Lease Type / Form"),
    ("Lease Identification",    "Execution Date"),
    ("Lease Identification",    "Effective / Commencement Date"),
    ("Lease Identification",    "Lease Expiration Date"),
    ("Lease Identification",    "Landlord of Record"),
    ("Lease Identification",    "Tenant of Record"),
    ("Lease Identification",    "Guarantor (if any)"),
    (None, None),
    ("Premises",                "Demised Premises / Suite"),
    ("Premises",                "Leased SF (per lease)"),
    ("Premises",                "Leased SF (per survey)"),
    ("Premises",                "Permitted Use"),
    ("Premises",                "Prohibited Uses"),
    (None, None),
    ("Rent",                    "Base Rent — Year 1"),
    ("Rent",                    "Annual Rent Escalations"),
    ("Rent",                    "Rent Commencement Date"),
    ("Rent",                    "Rent Abatement / Free Rent"),
    (None, None),
    ("Expense Structure",       "Lease Structure (NNN / Gross / MG)"),
    ("Expense Structure",       "Real Estate Taxes — Responsibility"),
    ("Expense Structure",       "Insurance — Responsibility"),
    ("Expense Structure",       "CAM / Maintenance — Responsibility"),
    ("Expense Structure",       "Expense Cap (if any)"),
    ("Expense Structure",       "Landlord Obligations"),
    (None, None),
    ("Options & Renewals",      "Number of Renewal Options"),
    ("Options & Renewals",      "Option Term Length"),
    ("Options & Renewals",      "Renewal Rent — Method"),
    ("Options & Renewals",      "Renewal Notice Requirement"),
    ("Options & Renewals",      "Option to Purchase"),
    ("Options & Renewals",      "Right of First Refusal (ROFR)"),
    (None, None),
    ("Assignment & Subletting", "Assignment Rights"),
    ("Assignment & Subletting", "Subletting Rights"),
    ("Assignment & Subletting", "Change of Control Provisions"),
    (None, None),
    ("Termination & Default",   "Early Termination Right"),
    ("Termination & Default",   "Termination Fee / Penalty"),
    ("Termination & Default",   "Default / Cure Periods"),
    ("Termination & Default",   "Co-Tenancy / Go-Dark Provision"),
    (None, None),
    ("Other Provisions",        "TI Allowance / Landlord Work"),
    ("Other Provisions",        "Signage Rights"),
    ("Other Provisions",        "Parking Allocation"),
    ("Other Provisions",        "Subordination / SNDA / Estoppel"),
    ("Other Provisions",        "Holdover Provisions"),
]


def build_mob_lease_abstract_tab(wb):
    ws = wb.create_sheet("Lease Abstract")
    ws.sheet_view.showGridLines = False

    # ── Column widths ─────────────────────────────────────────────────────────
    # A(1)=margin  B(2)=article/provision label  C-G(3-7)=5 tenant cols (short-form)
    # Long-form uses B-H: B=article, C=clause, D=page(8), E=lease section(16),
    #                      F=source(16), G:H=operative language(merged 18+18)
    col_widths = [2, 26, 18, 8, 16, 16, 18, 18]
    for i, wd in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = wd

    # ── Tab header ────────────────────────────────────────────────────────────
    r = 1;  ws.row_dimensions[r].height = 6
    r = 2;  ws.row_dimensions[r].height = 28
    c = ws.cell(row=r, column=2, value="LEASE ABSTRACT  —  Multi-Tenant")
    c.font = FT_TITLE; c.alignment = AL_L
    merge(ws, r, 2, r, 8)

    r = 3;  ws.row_dimensions[r].height = 14
    c = ws.cell(row=r, column=2,
                value="Leg 2 of 3  ·  Executed lease terms by tenant  ·  Short-form 5-tenant summary at top  ·  Full per-tenant abstracts below")
    c.font = FT_NOTE; c.alignment = AL_L
    merge(ws, r, 2, r, 8)

    r = 4;  ws.row_dimensions[r].height = 8

    # ══════════════════════════════════════════════════════════════════════════
    # SHORT-FORM SUMMARY TABLE (TOP)
    # ══════════════════════════════════════════════════════════════════════════
    SF_HDR = 5
    ws.row_dimensions[SF_HDR].height = 16
    sec(ws, SF_HDR, "SHORT-FORM SUMMARY  —  Key Lease Terms Across All Tenants", col_start=2, ncols=6)

    # Column header row: PROVISION + 5 tenant name columns
    SF_COL_HDR = 6
    ws.row_dimensions[SF_COL_HDR].height = 24

    # PROVISION label
    c = ws.cell(row=SF_COL_HDR, column=2, value="PROVISION")
    c.font = FT_CHDR; c.fill = F_NAVY; c.alignment = AL_C
    c.border = Border(bottom=_NM_M)

    # Tenant name columns — formula to pull from Assumptions & Flags
    AS = "'Assumptions & Flags'"
    for ti, t in enumerate(TENANT_MAP):
        col = 3 + ti
        frm(ws, SF_COL_HDR, col,
            f'=IFERROR(IF({AS}!{t["name_cell"]}="","{t["label"]}",{AS}!{t["name_cell"]}),"{t["label"]}")')
        ws.cell(row=SF_COL_HDR, column=col).font  = FT_CHDR
        ws.cell(row=SF_COL_HDR, column=col).fill  = F_NAVY
        ws.cell(row=SF_COL_HDR, column=col).alignment = AL_C
        ws.cell(row=SF_COL_HDR, column=col).border = Border(bottom=_NM_M)

    # Data rows
    SF_DATA_START = SF_COL_HDR + 1
    for idx, prov in enumerate(SF_PROVISIONS):
        rr = SF_DATA_START + idx
        ws.row_dimensions[rr].height = 32
        c = ws.cell(row=rr, column=2, value=prov)
        c.font = FT_LABEL; c.alignment = AL_TL
        ws.cell(row=rr, column=2).border = Border(bottom=_NM_T)
        # 5 input cells
        for col in range(3, 8):
            inp(ws, rr, col)
            ws.cell(row=rr, column=col).alignment = AL_TL
            ws.cell(row=rr, column=col).border = Border(bottom=_NM_T)

    SF_DATA_END = SF_DATA_START + len(SF_PROVISIONS) - 1
    # CF clear
    for col in range(3, 8):
        cl = get_column_letter(col)
        add_cf_clear(ws, f"{cl}{SF_DATA_START}:{cl}{SF_DATA_END}")

    # SF bottom accent
    sf_bot = SF_DATA_END + 1
    ws.row_dimensions[sf_bot].height = 4
    for ci in range(2, 8):
        ws.cell(row=sf_bot, column=ci).border = Border(top=_NM_M)

    sep_r = sf_bot + 1
    ws.row_dimensions[sep_r].height = 14

    # ══════════════════════════════════════════════════════════════════════════
    # PER-TENANT LONG-FORM SECTIONS (STACKED)
    # ══════════════════════════════════════════════════════════════════════════
    LF_HDR = sep_r + 1
    ws.row_dimensions[LF_HDR].height = 16
    sec(ws, LF_HDR, "LONG-FORM ABSTRACTS  —  Full article-by-article reference per tenant", col_start=2, ncols=6)

    # Long-form column header (re-used at top of each tenant section)
    # B=ITEM/ARTICLE  C:G=DETAIL (merged)
    LF_COL_HDR = LF_HDR + 1
    ws.row_dimensions[LF_COL_HDR].height = 20
    lf_hdrs = ["ARTICLE / SECTION", "CLAUSE / PROVISION", "PAGE", "DOCUMENT SOURCE", "OPERATIVE LANGUAGE"]
    lf_hdr_cols = [2, 3, 4, 5, 6]
    # Use cols 2=article, 3=clause, 4=page, 5=source, 6=operative (cols C-G, with B as article)
    # Actually for long-form in MOB: each section is a single-tenant standalone
    # B=article(26), C=clause(18), D=page(8), E=lease section(16),
    # F=source(16), G:H=operative language(merged 18+18)
    lf_hdr_labels = ["ARTICLE / SECTION", "CLAUSE / PROVISION", "PAGE", "LEASE SECTION", "SOURCE", "OPERATIVE LANGUAGE"]
    for ci, hdr in enumerate(lf_hdr_labels, 2):
        c = ws.cell(row=LF_COL_HDR, column=ci, value=hdr)
        c.font = FT_CHDR; c.fill = F_NAVY; c.alignment = AL_C
        c.border = Border(bottom=_NM_M)
    merge(ws, LF_COL_HDR, 7, LF_COL_HDR, 8)

    current_row = LF_COL_HDR + 1

    for ti, tenant in enumerate(TENANT_MAP):
        # ── Tenant section header ──────────────────────────────────────────
        ws.row_dimensions[current_row].height = 20
        tenant_hdr_formula = (
            f'=IFERROR(IF({AS}!{tenant["name_cell"]}="","'
            f'{tenant["label"]} — Full Lease Abstract",'
            f'"LEASE ABSTRACT — "&{AS}!{tenant["name_cell"]}),"{tenant["label"]} — Full Lease Abstract")'
        )
        frm(ws, current_row, 2, tenant_hdr_formula)
        ws.cell(row=current_row, column=2).font  = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
        ws.cell(row=current_row, column=2).fill  = PatternFill("solid", fgColor="1F7A8C")  # Teal
        ws.cell(row=current_row, column=2).alignment = AL_L
        merge(ws, current_row, 2, current_row, 8)
        current_row += 1

        # ── Long-form rows for this tenant ────────────────────────────────
        sec_data_start = current_row
        row_num    = 0
        data_rnum  = 0
        current_art = None

        for entry in LONG_ROWS:
            rr  = sec_data_start + row_num
            art, clause = entry
            if art is None:
                ws.row_dimensions[rr].height = 6
                row_num += 1
                continue
            ws.row_dimensions[rr].height = 44
            data_rnum += 1

            # B: article / section
            if art != current_art:
                c2 = ws.cell(row=rr, column=2, value=art)
                c2.font = FT_LABEL; c2.alignment = AL_TL
                current_art = art
            else:
                ws.cell(row=rr, column=2).alignment = AL_TL

            # C: clause / provision
            c3 = ws.cell(row=rr, column=3, value=clause)
            c3.font = FT_DATA; c3.alignment = AL_TL

            # D: page ref (input, centered)
            inp(ws, rr, 4); ws.cell(row=rr, column=4).alignment = AL_C

            # E: lease section ref (input, e.g. "Section 1.1(i)(a)-(c)")
            inp(ws, rr, 5); ws.cell(row=rr, column=5).alignment = AL_C

            # F: document source (input)
            inp(ws, rr, 6); ws.cell(row=rr, column=6).alignment = AL_TL

            # G:H merged: operative language (input, wide)
            inp(ws, rr, 7); ws.cell(row=rr, column=7).alignment = AL_TL
            merge(ws, rr, 7, rr, 8)

            for ci in range(2, 9):
                ws.cell(row=rr, column=ci).border = Border(bottom=_NM_T)

            row_num += 1

        sec_data_end = sec_data_start + row_num - 1

        # CF clear on input columns within this tenant's section
        for col in [4, 5, 6, 7]:
            cl = get_column_letter(col)
            add_cf_clear(ws, f"{cl}{sec_data_start}:{cl}{sec_data_end}")

        current_row = sec_data_end + 1

        # Tenant section bottom accent
        ws.row_dimensions[current_row].height = 4
        for ci in range(2, 9):
            ws.cell(row=current_row, column=ci).border = Border(top=_NM_M)
        current_row += 1

        # Gap between tenant sections (except after last)
        if ti < len(TENANT_MAP) - 1:
            ws.row_dimensions[current_row].height = 10
            current_row += 1

    # ── Notes row ─────────────────────────────────────────────────────────────
    note_r = current_row + 1
    ws.row_dimensions[note_r].height = 14
    c = ws.cell(row=note_r, column=2,
                value="Page: lease page number  ·  Lease Section: cite the specific section reference (e.g. 'Section 1.1(i)(a)')  ·  Source: 'Executed Lease', 'Amendment No. 1', etc.  ·  Operative Language: quote or paraphrase the actual lease provision")
    c.font = FT_NOTE; c.alignment = AL_L
    merge(ws, note_r, 2, note_r, 8)

    # NM bottom accent
    bot_r = note_r + 2
    ws.row_dimensions[bot_r].height = 4
    for ci in range(2, 9):
        ws.cell(row=bot_r, column=ci).border = Border(bottom=_NM_M)
