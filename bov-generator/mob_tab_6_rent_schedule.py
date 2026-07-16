"""BOV MOB — Tab 6: Rent Schedule (Leg 2 continued).
5 stacked per-tenant rent schedules with 15 rows each.
Current-lease-year row auto-highlights in light blue.
"""
from bov_constants import *
from openpyxl.styles import Border, Side, PatternFill, Font
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter

CURR_FILL = PatternFill("solid", fgColor="D6E4F5")   # NM light blue
_NM_M     = Side(style='medium', color=NAVY)
_NM_T     = Side(style='thin',   color=NAVY)

MAX_ROWS_PER_TENANT = 15   # 15 lease period rows per tenant

# Tenant cell map in 'Assumptions & Flags'
TENANT_MAP = [
    {"label": "TENANT 1", "name": "C18", "suite": "C19", "sf": "C20",
     "rent": "C21", "esc": "C22"},
    {"label": "TENANT 2", "name": "C27", "suite": "C28", "sf": "C29",
     "rent": "C30", "esc": "C31"},
    {"label": "TENANT 3", "name": "C36", "suite": "C37", "sf": "C38",
     "rent": "C39", "esc": "C40"},
    {"label": "TENANT 4", "name": "C45", "suite": "C46", "sf": "C47",
     "rent": "C48", "esc": "C49"},
    {"label": "TENANT 5", "name": "C54", "suite": "C55", "sf": "C56",
     "rent": "C57", "esc": "C58"},
]

AS = "'Assumptions & Flags'"


def _tenant_section(ws, current_row, tenant_idx, tenant):
    """Write one per-tenant rent schedule section. Returns next available row."""

    # ── Teal tenant header ────────────────────────────────────────────────────
    ws.row_dimensions[current_row].height = 20
    hdr_formula = (
        f'=IFERROR(IF({AS}!{tenant["name"]}="",'
        f'"{tenant["label"]} — Rent Schedule",'
        f'"RENT SCHEDULE — "&{AS}!{tenant["name"]}),"{tenant["label"]} — Rent Schedule")'
    )
    frm(ws, current_row, 2, hdr_formula)
    ws.cell(row=current_row, column=2).font      = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    ws.cell(row=current_row, column=2).fill      = PatternFill("solid", fgColor="1F7A8C")
    ws.cell(row=current_row, column=2).alignment = AL_L
    merge(ws, current_row, 2, current_row, 10)
    current_row += 1

    # ── Tenant info block ─────────────────────────────────────────────────────
    info_rows = [
        ("Suite / Unit",     f"={AS}!{tenant['suite']}"),
        ("Tenant Name",      f"={AS}!{tenant['name']}"),
        ("Leased SF",        f"={AS}!{tenant['sf']}"),
        ("Year 1 Base Rent", f"={AS}!{tenant['rent']}"),
        ("Escalation %",     f"={AS}!{tenant['esc']}"),
    ]
    for label, formula in info_rows:
        ws.row_dimensions[current_row].height = 16
        lbl(ws, current_row, 2, label)
        merge(ws, current_row, 2, current_row, 4)
        frm(ws, current_row, 5, formula)
        ws.cell(row=current_row, column=5).alignment = AL_L
        merge(ws, current_row, 5, current_row, 7)
        current_row += 1

    ws.row_dimensions[current_row].height = 6
    current_row += 1

    # ── Column headers ────────────────────────────────────────────────────────
    ws.row_dimensions[current_row].height = 22
    hdrs = ["YR", "START DATE", "END DATE", "LEASE PERIOD",
            "ANNUAL RENT", "MONTHLY RENT", "RENT / SF", "ESCALATION", "NOTES"]
    for ci, hdr in enumerate(hdrs, 2):
        c = ws.cell(row=current_row, column=ci, value=hdr)
        c.font = FT_CHDR; c.fill = F_NAVY; c.alignment = AL_C
        c.border = Border(bottom=_NM_M)
    current_row += 1

    # ── Data rows ─────────────────────────────────────────────────────────────
    DATA_START = current_row
    # Tenant SF cell reference (absolute) for Rent/SF denominator
    sf_abs = f"{AS}!${tenant['sf'][0]}${tenant['sf'][1:]}"

    for i in range(MAX_ROWS_PER_TENANT):
        rr = DATA_START + i
        ws.row_dimensions[rr].height = 18

        # YR#
        yr_c = ws.cell(row=rr, column=2, value=i + 1)
        yr_c.font = FT_DATA; yr_c.alignment = AL_C

        # Start Date
        inp(ws, rr, 3, fmt=DT); ws.cell(row=rr, column=3).alignment = AL_C

        # End Date
        inp(ws, rr, 4, fmt=DT); ws.cell(row=rr, column=4).alignment = AL_C

        # Lease Period
        inp(ws, rr, 5); ws.cell(row=rr, column=5).alignment = AL_L

        # Annual Rent
        inp(ws, rr, 6, fmt=D0); ws.cell(row=rr, column=6).alignment = AL_R

        # Monthly Rent = Annual / 12
        frm(ws, rr, 7,
            f'=IFERROR(IF(F{rr}="","",F{rr}/12),"")',
            fmt=D0, align=AL_R)

        # Rent / SF = Annual / Tenant SF
        frm(ws, rr, 8,
            f'=IFERROR(IF(OR(F{rr}="",{sf_abs}="",{sf_abs}=0),"",F{rr}/{sf_abs}),"")',
            fmt=D2, align=AL_R)

        # Escalation %
        inp(ws, rr, 9, fmt=P2); ws.cell(row=rr, column=9).alignment = AL_C

        # Notes
        inp(ws, rr, 10); ws.cell(row=rr, column=10).alignment = AL_TL

        for ci in range(2, 11):
            ws.cell(row=rr, column=ci).border = Border(bottom=_NM_T)

    DATA_END = DATA_START + MAX_ROWS_PER_TENANT - 1

    # CF: clear yellow when cells are filled
    for col in [3, 4, 5, 6, 9, 10]:
        cl = get_column_letter(col)
        add_cf_clear(ws, f"{cl}{DATA_START}:{cl}{DATA_END}")

    # CF: highlight current lease year row
    _hl_fill = PatternFill(fill_type="solid", fgColor=TOTBG)
    _hl_font = Font(name="Calibri", size=10, bold=True, color=TEXT)
    for i in range(MAX_ROWS_PER_TENANT):
        rr = DATA_START + i
        formula = f'AND(NOT(ISBLANK($C{rr})),NOT(ISBLANK($D{rr})),$C{rr}<=TODAY(),$D{rr}>=TODAY())'
        ws.conditional_formatting.add(
            f"B{rr}:J{rr}",
            FormulaRule(formula=[formula], fill=_hl_fill, font=_hl_font)
        )

    current_row = DATA_END + 1

    # Medium bottom border
    ws.row_dimensions[current_row].height = 4
    for ci in range(2, 11):
        ws.cell(row=current_row, column=ci).border = Border(top=_NM_M)
    current_row += 1

    # Totals row
    ws.row_dimensions[current_row].height = 18
    c = ws.cell(row=current_row, column=2, value="TOTALS / AVERAGES")
    c.font = FT_TOTAL; c.fill = F_TOT; c.alignment = AL_L
    merge(ws, current_row, 2, current_row, 5)
    ws.cell(row=current_row, column=2).border = Border(top=_NM_M, bottom=_NM_M)

    frm(ws, current_row, 6,
        f'=IFERROR(SUMIF(F{DATA_START}:F{DATA_END},"<>",F{DATA_START}:F{DATA_END}),"")',
        fmt=D0, align=AL_R)
    ws.cell(row=current_row, column=6).font   = FT_TOTAL
    ws.cell(row=current_row, column=6).fill   = F_TOT
    ws.cell(row=current_row, column=6).border = Border(top=_NM_M, bottom=_NM_M)
    for ci in [7, 8, 9, 10]:
        ws.cell(row=current_row, column=ci).fill  = F_TOT
        ws.cell(row=current_row, column=ci).border = Border(top=_NM_M, bottom=_NM_M)
    current_row += 1

    return current_row


def build_mob_rent_schedule_tab(wb):
    ws = wb.create_sheet("Rent Schedule")
    ws.sheet_view.showGridLines = False

    # freeze at B16 (after first tenant's header + info block + col header)
    ws.freeze_panes = "B16"

    # ── Column widths ─────────────────────────────────────────────────────────
    col_widths = [2, 6, 13, 13, 22, 14, 13, 10, 12, 34]
    for i, wd in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = wd

    # ── Tab header ────────────────────────────────────────────────────────────
    r = 1;  ws.row_dimensions[r].height = 6
    r = 2;  ws.row_dimensions[r].height = 28
    c = ws.cell(row=r, column=2, value="RENT SCHEDULE  —  Multi-Tenant")
    c.font = FT_TITLE; c.alignment = AL_L
    merge(ws, r, 2, r, 10)

    r = 3;  ws.row_dimensions[r].height = 14
    c = ws.cell(row=r, column=2,
                value="Leg 2 of 3  ·  Contractual rent by lease period per tenant  ·  Current year auto-highlighted  ·  See Rent Roll tab for consolidated single-page view")
    c.font = FT_NOTE; c.alignment = AL_L
    merge(ws, r, 2, r, 10)

    r = 4;  ws.row_dimensions[r].height = 8

    # ── Property header ───────────────────────────────────────────────────────
    r = 5
    sec(ws, r, "PROPERTY IDENTIFICATION", col_start=2, ncols=9)

    prop_rows = [
        ("Property Address", 6),
        ("Building SF (GLA)", 7),
    ]
    for label, rr in prop_rows:
        ws.row_dimensions[rr].height = 16
        lbl(ws, rr, 2, label)
        merge(ws, rr, 2, rr, 4)
        inp(ws, rr, 5)
        merge(ws, rr, 5, rr, 7)
        ws.cell(row=rr, column=5).alignment = AL_L
        add_cf_clear(ws, f"E{rr}:E{rr}")

    ws.row_dimensions[8].height = 8

    # ── Per-tenant sections ───────────────────────────────────────────────────
    current_row = 9
    for ti, tenant in enumerate(TENANT_MAP):
        current_row = _tenant_section(ws, current_row, ti, tenant)
        # Gap between tenants (except after last)
        if ti < len(TENANT_MAP) - 1:
            ws.row_dimensions[current_row].height = 12
            current_row += 1

    # ── Notes & legend ────────────────────────────────────────────────────────
    note_r = current_row + 1
    notes = [
        "LEASE PERIOD column:  Initial Term · Extension Term · Option 1, Option 2, etc.",
        "For rent tied to FMV, CPI, or other variable methods, leave Annual Rent blank and describe in Notes column.",
        "Current lease year row auto-highlights in blue based on today's date (Start Date ≤ Today ≤ End Date).",
        "See the Rent Roll tab for a consolidated single-page view of all tenant rents.",
    ]
    for idx, note in enumerate(notes):
        rr = note_r + idx
        ws.row_dimensions[rr].height = 14
        c = ws.cell(row=rr, column=2, value=f"▪  {note}")
        c.font = FT_NOTE; c.alignment = AL_L
        merge(ws, rr, 2, rr, 10)

    bot_r = note_r + len(notes) + 1
    ws.row_dimensions[bot_r].height = 4
    for ci in range(2, 11):
        ws.cell(row=bot_r, column=ci).border = Border(bottom=_NM_M)
