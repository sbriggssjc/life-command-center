"""BOV MOB — Tab 9: Assumptions & Flags (Multi-Tenant).

Left column cell map (col A=label, col C=value):
  Property:
    C9  = Building SF (GLA)
    C14 = Estimated Close Date

  Tenant 1 (rows 17-24):
    C18 = T1 Tenant Name
    C19 = T1 Suite / Unit
    C20 = T1 SF (Leased)
    C21 = T1 Annual Rent (Year 1)
    C22 = T1 Rent Escalation %
    C23 = T1 Lease Type (NNN / MG / Gross)
    C24 = T1 Reimbursements ($/yr)

  Tenant 2 (rows 26-33):
    C27 = T2 Tenant Name
    C28 = T2 Suite / Unit
    C29 = T2 SF (Leased)
    C30 = T2 Annual Rent (Year 1)
    C31 = T2 Rent Escalation %
    C32 = T2 Lease Type
    C33 = T2 Reimbursements ($/yr)

  Tenant 3 (rows 35-42):
    C36 = T3 Tenant Name
    C37 = T3 Suite / Unit
    C38 = T3 SF (Leased)
    C39 = T3 Annual Rent (Year 1)
    C40 = T3 Rent Escalation %
    C41 = T3 Lease Type
    C42 = T3 Reimbursements ($/yr)

  Tenant 4 (rows 44-51):
    C45 = T4 Tenant Name
    C46 = T4 Suite / Unit
    C47 = T4 SF (Leased)
    C48 = T4 Annual Rent (Year 1)
    C49 = T4 Rent Escalation %
    C50 = T4 Lease Type
    C51 = T4 Reimbursements ($/yr)

  Tenant 5 (rows 53-60):
    C54 = T5 Tenant Name
    C55 = T5 Suite / Unit
    C56 = T5 SF (Leased)
    C57 = T5 Annual Rent (Year 1)
    C58 = T5 Rent Escalation %
    C59 = T5 Lease Type
    C60 = T5 Reimbursements ($/yr)

  Portfolio summary (rows 62-67):
    C63 = Vacancy / Credit Loss %
    C64 = Portfolio Avg Escalation % (for Sensitivity)
    C65 = Occupied SF (formula: sum T1-T5 SF)
    C66 = Occupancy Rate (formula)

  Expenses (rows 68-77):
    C69 = RE Taxes ($/yr)
    C70 = Insurance ($/yr)
    C71 = CAM / Janitorial ($/yr)
    C72 = Management Fee (%)
    C73 = G&A / Admin ($/yr)
    C74 = Repairs & Maintenance ($/yr)
    C75 = Utilities ($/yr)
    C76 = Capital / Replacement Reserves ($/yr)

  NOI waterfall (rows 78-86):
    C79 = Gross Potential Rent (formula: sum T1-T5 rents)
    C80 = Less: Vacancy / Credit Loss (formula)
    C81 = Effective Base Rent (formula)
    C82 = Total Reimbursements (formula: sum T reimbursements)
    C83 = Effective Gross Income (formula)
    C84 = Total Operating Expenses (formula)
    C85 = NET OPERATING INCOME — Year 1  ← KEY REFERENCE

  Pricing (rows 87-97):
    C88 = Asking Cap Rate
    C89 = Trade Range Low Cap Rate
    C90 = Trade Range High Cap Rate
    C91 = Implied Asking Price (formula: C85 / C88)
    C94 = Purchase Price  ← KEY REFERENCE
    C95 = Going-In Cap Rate (formula: C85 / C94)
    C96 = Price Per SF (formula: C94 / C9)

Right column cell map (col G=label, col I=value):
  I5-I17: Debt / Leverage (same as NNN)
    I6  = Purchase Price (link to C94)
    I7  = LTV %
    I8  = Loan Amount (formula)
    I9  = Down Payment / Equity (formula)
    I10 = Interest Rate
    I11 = Amortization (years)
    I12 = IO Period (years)
    I13 = Hold Period (years)
    I14 = Annual Debt Service (formula)
    I15 = Monthly Debt Service (formula)
    I16 = DCR Year 1 (formula: C85 / I14)
    I17 = DSCR at 1.20x NOI Required

  I19-I27: Disposition
    I20 = Exit Cap Rate
    I21 = Exit NOI Year 10 (formula: C85*(1+C64)^I13)
    I22 = GROSS SALE PROCEEDS (formula)
    I23 = Less: Brokerage (5%)
    I24 = Less: Transaction Costs (1%)
    I25 = NET REVERSION (Unleveraged)
    I26 = Remaining Loan Balance
    I27 = NET REVERSION AFTER DEBT

  I29-I35: Return Summary (links to Pro Forma)
    I30 = UL IRR  → 'Pro Forma'!C44
    I31 = UL ERM  → 'Pro Forma'!G44
    I32 = UL CoCR → 'Pro Forma'!K44
    I33 = LEV IRR → 'Pro Forma'!C59
    I34 = LEV ERM → 'Pro Forma'!G59
    I35 = LEV CoCR→ 'Pro Forma'!K59
"""
from bov_constants import *
from openpyxl.styles import Border, Side
from openpyxl.utils import get_column_letter

_NM_M = Side(style='medium', color=NAVY)
_GY_T = Side(style='thin',   color="CCCCCC")

# Tenant data rows: (name_row, suite_row, sf_row, rent_row, esc_row, type_row, reimb_row)
TENANT_ROWS = [
    (18, 19, 20, 21, 22, 23, 24),   # T1
    (27, 28, 29, 30, 31, 32, 33),   # T2
    (36, 37, 38, 39, 40, 41, 42),   # T3
    (45, 46, 47, 48, 49, 50, 51),   # T4
    (54, 55, 56, 57, 58, 59, 60),   # T5
]


def build_mob_assumptions_tab(wb):
    ws = wb.create_sheet("Assumptions & Flags")
    ws.sheet_view.showGridLines = False

    # ── Column widths ──────────────────────────────────────────────────────────
    # A=label(30), B=spacer(1.5), C=value(18), D=spacer(1.5), E=notes(34)
    # F=spacer(1.5), G=label(30), H=spacer(1.5), I=value(18), J=spacer(1.5), K=notes(34)
    col_widths = [30, 1.5, 18, 1.5, 34, 1.5, 30, 1.5, 18, 1.5, 34]
    for i, wd in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = wd

    # ── Header ────────────────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 6
    ws.row_dimensions[2].height = 28
    c = ws.cell(row=2, column=1, value="ASSUMPTIONS & FLAGS  —  MULTI-TENANT MOB / COMMERCIAL")
    c.font = FT_TITLE; c.alignment = AL_L
    merge(ws, 2, 1, 2, 11)

    ws.row_dimensions[3].height = 14
    c = ws.cell(row=3, column=1,
                value="All key inputs for the Pro Forma  ·  Blue text = broker enters  ·  Yellow = fill before delivering  ·  5 tenant slots; leave blank if suite is vacant")
    c.font = FT_NOTE; c.alignment = AL_L
    merge(ws, 3, 1, 3, 11)

    ws.row_dimensions[4].height = 8

    # ── Helpers ───────────────────────────────────────────────────────────────
    def row_inp(rr, label, val=None, fmt=None, note=None):
        ws.row_dimensions[rr].height = 18
        lbl(ws, rr, 1, label)
        c = inp(ws, rr, 3, val, fmt)
        c.alignment = AL_R
        if note:
            ws.cell(row=rr, column=5, value=note).font = FT_NOTE
            ws.cell(row=rr, column=5).alignment = AL_L
        add_cf_clear(ws, f"C{rr}:C{rr}")
        ws.cell(row=rr, column=1).border = Border(bottom=_GY_T)
        ws.cell(row=rr, column=3).border = Border(bottom=_GY_T)

    def row_frm(rr, label, formula, fmt=None, note=None, bold=False):
        ws.row_dimensions[rr].height = 18
        lbl(ws, rr, 1, label, bold=bold)
        c = frm(ws, rr, 3, formula, fmt=fmt, align=AL_R)
        if bold:
            c.font = FT_TOTAL
            ws.cell(row=rr, column=3).fill = F_TOT
            ws.cell(row=rr, column=1).fill = F_TOT
        if note:
            ws.cell(row=rr, column=5, value=note).font = FT_NOTE
            ws.cell(row=rr, column=5).alignment = AL_L
        ws.cell(row=rr, column=1).border = Border(bottom=_GY_T)
        ws.cell(row=rr, column=3).border = Border(bottom=_GY_T)

    def row_inp_r(rr, label, val=None, fmt=None, note=None):
        ws.row_dimensions[rr].height = max(ws.row_dimensions[rr].height if ws.row_dimensions[rr].height else 18, 18)
        lbl(ws, rr, 7, label)
        c = inp(ws, rr, 9, val, fmt)
        c.alignment = AL_R
        if note:
            ws.cell(row=rr, column=11, value=note).font = FT_NOTE
            ws.cell(row=rr, column=11).alignment = AL_L
        add_cf_clear(ws, f"I{rr}:I{rr}")
        ws.cell(row=rr, column=7).border = Border(bottom=_GY_T)
        ws.cell(row=rr, column=9).border = Border(bottom=_GY_T)

    def row_frm_r(rr, label, formula, fmt=None, note=None, bold=False):
        ws.row_dimensions[rr].height = max(ws.row_dimensions[rr].height if ws.row_dimensions[rr].height else 18, 18)
        lbl(ws, rr, 7, label, bold=bold)
        c = frm(ws, rr, 9, formula, fmt=fmt, align=AL_R)
        if bold:
            c.font = FT_TOTAL
            ws.cell(row=rr, column=9).fill = F_TOT
            ws.cell(row=rr, column=7).fill = F_TOT
        if note:
            ws.cell(row=rr, column=11, value=note).font = FT_NOTE
            ws.cell(row=rr, column=11).alignment = AL_L
        ws.cell(row=rr, column=7).border = Border(bottom=_GY_T)
        ws.cell(row=rr, column=9).border = Border(bottom=_GY_T)

    # ══════════════════════════════════════════════════════════════════════════
    # LEFT COLUMN — PROPERTY, TENANTS, NOI WATERFALL, PRICING
    # ══════════════════════════════════════════════════════════════════════════

    # SECTION: Property Identification (rows 5-14)
    sec(ws, 5, "PROPERTY IDENTIFICATION", col_start=1, ncols=5)
    row_inp(6,  "Property Address / Name",    note="Used on Cover and Executive Summary")
    row_inp(7,  "Property Type",              note='e.g., "Multi-Tenant MOB" or "Strip Center"')
    row_inp(8,  "Market / Submarket",         note="City, State — MSA name")
    row_inp(9,  "Total Building SF (GLA)",    fmt=N0, note="Gross leasable area — confirm with survey")
    row_inp(10, "Land Area (Acres)",           fmt="0.00", note="Per survey or tax record")
    row_inp(11, "Year Built",                  note="Four-digit year")
    row_inp(12, "Zoning",                      note="Confirm permitted use")
    row_inp(13, "Property Owner / Seller",     note="Legal entity and/or trade name")
    row_inp(14, "Estimated Close Date",        fmt=DT, note="Used for year-end date row on Pro Forma")

    # Spacer
    ws.row_dimensions[15].height = 6
    ws.row_dimensions[16].height = 16
    c = ws.cell(row=16, column=1,
                value="TENANT RENT ROLL  —  Year 1 inputs  ·  Leave row blank if suite is vacant  ·  Lease Type: NNN | MG | Gross")
    c.font = FT_SHDR; c.fill = F_NAVY; c.alignment = AL_L
    for ci in [2, 3, 4, 5]: ws.cell(row=16, column=ci).fill = F_NAVY
    merge(ws, 16, 1, 16, 5)
    ws.row_dimensions[16].height = 16

    # Per-tenant sections
    tenant_labels = ["TENANT 1", "TENANT 2", "TENANT 3", "TENANT 4", "TENANT 5"]
    attr_labels = [
        ("Tenant / Operator Name",    None,  None),
        ("Suite / Unit",              None,  None),
        ("Leased SF",                 N0,    "Exclude common area"),
        ("Annual Rent — Year 1 ($)",  D0,    "Per executed lease / rent roll"),
        ("Rent Escalation (%)",        P2,    "Fixed bump or avg CPI — label in Notes"),
        ("Lease Type",                None,  "NNN | Modified Gross | Gross"),
        ("Tenant Reimbursements ($/yr)", D0, "NNN=full pass; MG/Gross=partial or $0"),
    ]

    for t_idx, (t_rows) in enumerate(TENANT_ROWS):
        name_r, suite_r, sf_r, rent_r, esc_r, type_r, reimb_r = t_rows
        hdr_r = name_r - 1

        ws.row_dimensions[hdr_r].height = 16
        c = ws.cell(row=hdr_r, column=1, value=tenant_labels[t_idx])
        c.font = ft("Calibri", 10, b=True, c=NAVY); c.fill = F_PALE; c.alignment = AL_L
        for ci in [2, 3, 4, 5]: ws.cell(row=hdr_r, column=ci).fill = F_PALE
        merge(ws, hdr_r, 1, hdr_r, 5)

        rows_in_order = [name_r, suite_r, sf_r, rent_r, esc_r, type_r, reimb_r]
        for rr, (label, fmt, note) in zip(rows_in_order, attr_labels):
            row_inp(rr, label, fmt=fmt, note=note)

        # Spacer after each tenant except last
        if t_idx < 4:
            spacer_r = reimb_r + 1
            ws.row_dimensions[spacer_r].height = 6

    # SECTION: Portfolio Summary (rows 62-67)
    ws.row_dimensions[61].height = 6
    sec(ws, 62, "PORTFOLIO SUMMARY  —  Formulas; update if data changed", col_start=1, ncols=5)
    row_inp(63, "Vacancy / Credit Loss (%)",     fmt=P2, note="5-10% typical; 0% if fully occupied")
    row_inp(64, "Portfolio Avg Escalation (%)",  fmt=P2, note="Blended rate for Sensitivity analysis")
    row_frm(65, "Occupied SF",
            '=IFERROR(IFERROR(C20,0)+IFERROR(C29,0)+IFERROR(C38,0)+IFERROR(C47,0)+IFERROR(C56,0),"")',
            fmt=N0, note="Sum of leased SF — T1 through T5")
    row_frm(66, "Occupancy Rate",
            '=IFERROR(IF(OR(C9="",C9=0),"",C65/C9),"")',
            fmt=P2, note="Occupied SF / Total Building SF")

    # SECTION: Operating Expenses (rows 68-77)
    ws.row_dimensions[67].height = 6
    sec(ws, 68, "OPERATING EXPENSES  —  Year 1 inputs; flat for Pro Forma (adjust as needed)", col_start=1, ncols=5)
    row_inp(69, "Real Estate Taxes ($/yr)",        fmt=D0, note="LL-responsible; NNN = pass-through")
    row_inp(70, "Insurance ($/yr)",                fmt=D0, note="LL-responsible")
    row_inp(71, "CAM / Janitorial ($/yr)",         fmt=D0, note="Common area maintenance")
    row_inp(72, "Management Fee (%)",              fmt=P2, note="% of Effective Gross Income; 3-5% typical")
    row_inp(73, "G&A / Admin ($/yr)",              fmt=D0, note="General & administrative")
    row_inp(74, "Repairs & Maintenance ($/yr)",    fmt=D0, note="Routine R&M")
    row_inp(75, "Utilities ($/yr)",                fmt=D0, note="Common area / LL-responsible")
    row_inp(76, "Capital / Replacement Reserves ($/yr)", fmt=D0, note="$0.15–$0.25/SF typical")

    # SECTION: NOI Waterfall (rows 78-86)
    ws.row_dimensions[77].height = 6
    sec(ws, 78, "YEAR 1 NOI WATERFALL  —  All formulas; resolve to Pro Forma Year 1", col_start=1, ncols=5)

    row_frm(79, "Gross Potential Rent",
            '=IFERROR(IFERROR(C21,0)+IFERROR(C30,0)+IFERROR(C39,0)+IFERROR(C48,0)+IFERROR(C57,0),"")',
            fmt=D0, note="Sum of T1-T5 Year 1 rents")
    row_frm(80, "Less: Vacancy / Credit Loss",
            '=IFERROR(IF(OR(C79="",C63=""),"",-C79*C63),"")',
            fmt=D0, note="Negative value")
    row_frm(81, "Effective Base Rent",
            '=IFERROR(IF(C79="","",C79+IFERROR(C80,0)),"")',
            fmt=D0)
    row_frm(82, "Total Tenant Reimbursements",
            '=IFERROR(IFERROR(C24,0)+IFERROR(C33,0)+IFERROR(C42,0)+IFERROR(C51,0)+IFERROR(C60,0),"")',
            fmt=D0, note="Sum of T1-T5 reimbursements")
    row_frm(83, "Effective Gross Income (EGI)",
            '=IFERROR(IF(C81="","",C81+IFERROR(C82,0)),"")',
            fmt=D0)
    row_frm(84, "Total Operating Expenses",
            '=IFERROR(IFERROR(C69,0)+IFERROR(C70,0)+IFERROR(C71,0)'
            '+IFERROR(C83,0)*IFERROR(C72,0)'
            '+IFERROR(C73,0)+IFERROR(C74,0)+IFERROR(C75,0)+IFERROR(C76,0),"")',
            fmt=D0, note="Incl. Mgmt Fee % of EGI")

    # C85 = NET OPERATING INCOME — KEY
    ws.row_dimensions[85].height = 22
    lbl(ws, 85, 1, "NET OPERATING INCOME  —  Year 1", bold=True)
    ws.cell(row=85, column=1).fill = F_PALE
    for ci in [2, 3, 4, 5]: ws.cell(row=85, column=ci).fill = F_PALE
    c = frm(ws, 85, 3,
            '=IFERROR(IF(C83="","",C83-IFERROR(C84,0)),"")',
            fmt=D0, align=AL_R)
    c.font = ft("Calibri", 11, b=True, c=NAVY)
    ws.cell(row=85, column=3).fill = F_PALE
    from openpyxl.styles import Border as Bdr, Side as Sd
    _nm = Sd(style='medium', color=NAVY)
    ws.cell(row=85, column=3).border = Bdr(top=_nm, bottom=_nm)

    # SECTION: Pricing (rows 87-97)
    ws.row_dimensions[86].height = 6
    sec(ws, 87, "PRICING STRATEGY  ·  FUNGIBLE INPUTS", col_start=1, ncols=5)

    row_inp(88, "Recommended Asking Cap Rate",  fmt=P2, note="Drives Recommended Asking Price")
    row_inp(89, "Trade Range — Low Cap Rate",   fmt=P2, note="Low = higher price")
    row_inp(90, "Trade Range — High Cap Rate",  fmt=P2, note="High = lower price")
    row_frm(91, "Implied Asking Price (NOI / Cap)",
            '=IFERROR(IF(OR(C85="",C88="",C88=0),"",C85/C88),"")',
            fmt=D0, note="Auto-calculates from NOI / Cap Rate")

    ws.row_dimensions[92].height = 6
    sec(ws, 93, "ACQUISITION PRICING", col_start=1, ncols=5)

    # C94 = Purchase Price — KEY
    row_inp(94, "Purchase Price ($)",   fmt=D0, note="Fungible — change to model different scenarios")
    row_frm(95, "Going-In Cap Rate",
            '=IFERROR(IF(OR(C85="",C94="",C94=0),"",C85/C94),"")',
            fmt=P2, note="Y1 NOI / Purchase Price")
    row_frm(96, "Price Per SF",
            '=IFERROR(IF(OR(C94="",C9="",C9=0),"",C94/C9),"")',
            fmt=D2, note="Uses Total Building SF (C9)")

    # Discrepancy Flags (rows 97-109)
    ws.row_dimensions[97].height = 6
    sec(ws, 98, "DISCREPANCY FLAGS  —  Note any source conflicts or assumptions below", col_start=1, ncols=5)
    ws.row_dimensions[99].height = 18
    ws.cell(row=99, column=1, value="Flag / Assumption").font = FT_CHDR
    ws.cell(row=99, column=1).fill = F_NAVY
    ws.cell(row=99, column=3, value="Description / Resolution").font = FT_CHDR
    ws.cell(row=99, column=3).fill = F_NAVY
    merge(ws, 99, 3, 99, 5)
    for i in range(10):
        rr = 100 + i
        ws.row_dimensions[rr].height = 18
        inp(ws, rr, 1); ws.cell(row=rr, column=1).alignment = AL_TL
        inp(ws, rr, 3); ws.cell(row=rr, column=3).alignment = AL_TL
        merge(ws, rr, 3, rr, 5)
        ws.cell(row=rr, column=1).border = Border(bottom=_GY_T)
        ws.cell(row=rr, column=3).border = Border(bottom=_GY_T)

    # ══════════════════════════════════════════════════════════════════════════
    # RIGHT COLUMN — DEBT / LEVERAGE, DISPOSITION, RETURN SUMMARY
    # ══════════════════════════════════════════════════════════════════════════

    # SECTION: Debt / Leverage (rows 5-17)
    c = ws.cell(row=5, column=7, value="DEBT / LEVERAGE ASSUMPTIONS")
    c.font = FT_SHDR; c.fill = F_NAVY; c.alignment = AL_L
    for ci in [7, 8, 9, 10, 11]: ws.cell(row=5, column=ci).fill = F_NAVY
    merge(ws, 5, 7, 5, 11)
    ws.row_dimensions[5].height = 16

    row_inp_r(6,  "Purchase Price ($)",      fmt=D0,  note="Link to C94 or enter here; fungible")
    row_inp_r(7,  "LTV (%)",                 fmt=P2,  note="Default 65%")
    row_frm_r(8,  "Loan Amount",
              '=IFERROR(IF(OR(I6="",I7=""),"",I6*I7),"")',
              fmt=D0, note="Loan = Price × LTV")
    row_frm_r(9,  "Down Payment / Equity",
              '=IFERROR(IF(OR(I6="",I7=""),"",I6*(1-I7)),"")',
              fmt=D0, note="Equity = Price × (1 − LTV)")
    row_inp_r(10, "Interest Rate",           fmt=P2,  note="Default 6.5%")
    row_inp_r(11, "Amortization (years)",    fmt=N0,  note="Default 25 years")
    row_inp_r(12, "IO Period (years)",       fmt=N0,  note="Enter 0 if fully amortizing")
    row_inp_r(13, "Hold Period (years)",     fmt=N0,  note="Default 10")
    row_frm_r(14, "Annual Debt Service",
              '=IFERROR(IF(OR(I8="",I10="",I11=""),"",ABS(I8)*(I10/12)*(1+I10/12)^(I11*12)/((1+I10/12)^(I11*12)-1)*12),"")',
              fmt=D0, note="P&I — fully amortizing")
    row_frm_r(15, "Monthly Debt Service",
              '=IFERROR(IF(I14="","",I14/12),"")', fmt=D0)
    row_frm_r(16, "DCR — Year 1",
              '=IFERROR(IF(OR(C85="",I14="",I14=0),"",C85/I14),"")',
              fmt=MX, note="Must exceed 1.20x; refs NOI at C85")
    row_frm_r(17, "DSCR at 1.20x NOI Required",
              '=IFERROR(IF(I14="","",I14*1.20),"")', fmt=D0)

    ws.row_dimensions[18].height = 6

    # SECTION: Disposition (rows 19-27)
    c = ws.cell(row=19, column=7, value="DISPOSITION / EXIT ASSUMPTIONS")
    c.font = FT_SHDR; c.fill = F_NAVY; c.alignment = AL_L
    for ci in [7, 8, 9, 10, 11]: ws.cell(row=19, column=ci).fill = F_NAVY
    merge(ws, 19, 7, 19, 11)
    ws.row_dimensions[19].height = 16

    row_inp_r(20, "Exit Cap Rate",           fmt=P2, note="Applied to projected Year-10 NOI")
    row_frm_r(21, "Exit NOI (Year 10 Projected)",
              '=IFERROR(IF(OR(C85="",C64=""),"",C85*(1+C64)^I13),"")',
              fmt=D0, note="Y1 NOI × (1 + Portfolio Avg Esc)^Hold")
    row_frm_r(22, "GROSS SALE PROCEEDS",
              '=IFERROR(IF(OR(I21="",I20="",I20=0),"",I21/I20),"")',
              fmt=D0, bold=True)
    row_frm_r(23, "Less: Brokerage (5%)",
              '=IFERROR(IF(I22="","",-I22*0.05),"")', fmt=D0)
    row_frm_r(24, "Less: Transaction Costs (1%)",
              '=IFERROR(IF(I22="","",-I22*0.01),"")', fmt=D0)
    row_frm_r(25, "NET REVERSION (Unleveraged)",
              '=IFERROR(IF(I22="","",I22+I23+I24),"")',
              fmt=D0, bold=True)
    row_frm_r(26, "Remaining Loan Balance at Exit",
              '=IFERROR(IF(OR(I8="",I10="",I11="",I13=""),"",ABS(I8)*(1+I10/12)^(I13*12)-(ABS(I8)*(I10/12)*(1+I10/12)^(I11*12)/((1+I10/12)^(I11*12)-1))*(((1+I10/12)^(I13*12)-1)/(I10/12))),"")',
              fmt=D0, note="Outstanding balance at hold period end")
    row_frm_r(27, "NET REVERSION AFTER DEBT",
              '=IFERROR(IF(OR(I25="",I26=""),"",I25-I26),"")',
              fmt=D0, bold=True)

    ws.row_dimensions[28].height = 6

    # SECTION: Return Summary (rows 29-35)
    c = ws.cell(row=29, column=7, value="RETURN SUMMARY REFERENCE")
    c.font = FT_SHDR; c.fill = F_NAVY; c.alignment = AL_L
    for ci in [7, 8, 9, 10, 11]: ws.cell(row=29, column=ci).fill = F_NAVY
    merge(ws, 29, 7, 29, 11)
    ws.row_dimensions[29].height = 16

    def row_ref_r(rr, label, formula, fmt=None):
        ws.row_dimensions[rr].height = 18
        lbl(ws, rr, 7, label, bold=False)
        frm(ws, rr, 9, formula, fmt=fmt, align=AL_R)
        ws.cell(row=rr, column=7).border = Border(bottom=_GY_T)
        ws.cell(row=rr, column=9).border = Border(bottom=_GY_T)

    # Pro Forma MOB callout rows: UL at 43-44, LEV at 58-59
    row_ref_r(30, "Unleveraged IRR",   '=IFERROR(\'Pro Forma\'!C44,"N/A")', fmt=P2)
    row_ref_r(31, "Unleveraged ERM",   '=IFERROR(\'Pro Forma\'!G44,"N/A")', fmt=MX)
    row_ref_r(32, "Unleveraged CoCR",  '=IFERROR(\'Pro Forma\'!K44,"N/A")', fmt=P2)
    row_ref_r(33, "Leveraged IRR",     '=IFERROR(\'Pro Forma\'!C59,"N/A")', fmt=P2)
    row_ref_r(34, "Leveraged ERM",     '=IFERROR(\'Pro Forma\'!G59,"N/A")', fmt=MX)
    row_ref_r(35, "Leveraged CoCR",    '=IFERROR(\'Pro Forma\'!K59,"N/A")', fmt=P2)

    # NM bottom accent
    ws.row_dimensions[112].height = 4
    for ci in range(1, 12):
        ws.cell(row=112, column=ci).border = Border(bottom=Side(style='medium', color=NAVY))
