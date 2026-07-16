"""Executive Summary tab for BOV Master MOB Sheet — Multi-Tenant MOB / Commercial."""
from bov_constants import *

_A   = "'Assumptions & Flags'"
_NOI = f"{_A}!$C$85"
_SF  = f"{_A}!$C$9"
_GPR = f"{_A}!$C$79"
_OCC = f"{_A}!$C$66"

# Tenant assumption row indices (Assumptions & Flags sheet):
# Tenant Name rows:   C18, C27, C36, C45, C54
# Suite rows:         C19, C28, C37, C46, C55
# SF rows:            C20, C29, C38, C47, C56
# Annual Rent rows:   C21, C30, C39, C48, C57
_T_NAME = [18, 27, 36, 45, 54]
_T_SUITE = [19, 28, 37, 46, 55]
_T_SF    = [20, 29, 38, 47, 56]
_T_RENT  = [21, 30, 39, 48, 57]


def build_mob_exec_summary(wb):
    ws = wb.create_sheet("Executive Summary")
    ws.sheet_view.showGridLines = False

    # Columns: A=margin(2), B=label(30), C=value(16), D=value2(18), E=pct(12), F=notes(20), G=extra(12)
    for col, wd in [(1, 2), (2, 30), (3, 16), (4, 18), (5, 12), (6, 20), (7, 12)]:
        w(ws, col, wd)

    # ── Tab header ────────────────────────────────────────────────────────────
    h(ws, 1, 34)
    c = ws.cell(row=1, column=2, value="Executive Summary")
    c.font = FT_TITLE; c.alignment = AL_L
    merge(ws, 1, 2, 1, 6)

    h(ws, 2, 18)
    c = ws.cell(row=2, column=2,
                value="[Property Name  —  City, State]  ·  Multi-Tenant MOB  ·  Investment Snapshot  ·  [Analysis Date]")
    c.font = FT_NOTE; c.alignment = AL_L
    merge(ws, 2, 2, 2, 6)
    h(ws, 3, 8)

    # ── INVESTMENT SNAPSHOT ───────────────────────────────────────────────────
    sec(ws, 4, "INVESTMENT SNAPSHOT", col_start=2, ncols=5)

    # Each tuple: (label, default_or_None, fmt_or_None, is_formula)
    # is_formula=True → use frm(); False → use inp() unless default is set
    snapshot = [
        ("Property Name / Address",          None,                            None,              False),
        ("City, State",                      None,                            None,              False),
        ("Asset Type",                       "Multi-Tenant MOB / Commercial", None,              False),
        ("Year Built",                       None,                            "0",               False),
        ("Total GLA / RBA (SF)",             None,                            N0,                False),
        ("Land Area (Acres)",                None,                            "0.00",            False),
        ("Number of Tenants",                None,                            N0,                False),
        ("Occupancy Rate",                   None,                            P2,                True),
        ("WALT (Weighted Avg Lease Term)",   None,                            '0.0 "yrs"',       False),
        ("Year 1 Gross Potential Rent",      None,                            D0,                True),
        ("Year 1 NOI",                       None,                            D0,                True),
        ("Asking Cap Rate",                  None,                            P2,                False),
        ("Asking Price",                     None,                            D0,                True),
    ]

    # Row assignments: snapshot starts at row 5
    SNAP_START = 5

    for i, (label, default, fmt, is_formula) in enumerate(snapshot):
        r = SNAP_START + i
        # WALT row (i=8) needs extra height for the longer label
        h(ws, r, 28 if i == 8 else 18)
        bg = F_PALE if i % 2 == 0 else F_WHITE

        c = ws.cell(row=r, column=2, value=label)
        c.font = FT_DATA; c.fill = bg; c.alignment = AL_L
        for col in range(3, 7):
            ws.cell(row=r, column=col).fill = bg

        if default:
            # Hard-coded text default (Asset Type, etc.)
            v = ws.cell(row=r, column=3, value=default)
            v.font = FT_DATA; v.alignment = AL_L
            merge(ws, r, 3, r, 6)
        elif is_formula:
            # Formula row — determine which formula by row index
            if i == 7:
                # Occupancy Rate
                formula = f'=IFERROR(IF({_OCC}="","",{_OCC}),"")'
            elif i == 9:
                # Year 1 Gross Potential Rent
                formula = f'=IFERROR(IF({_GPR}="","",{_GPR}),"")'
            elif i == 10:
                # Year 1 NOI
                formula = f'=IFERROR(IF({_NOI}="","",{_NOI}),"")'
            elif i == 12:
                # Asking Price: cap rate is one row above (C{r-1})
                cap_ref = f"C{r - 1}"
                formula = (
                    f'=IFERROR(IF(OR({_NOI}="",{cap_ref}="",{cap_ref}=0),"",{_NOI}/{cap_ref}),"")'
                )
            else:
                formula = '=""'
            v = frm(ws, r, 3, formula, fmt=fmt, align=AL_L)
            v.fill = bg
            merge(ws, r, 3, r, 6)
        else:
            # Standard input cell
            v = inp(ws, r, 3, fmt=fmt)
            v.fill = bg
            merge(ws, r, 3, r, 6)

    SNAP_END = SNAP_START + len(snapshot) - 1  # row 17
    # Apply CF clear only to genuine input rows (not formula rows or default-text rows).
    # Formula rows (i=7 Occupancy, i=9 GPR, i=10 NOI, i=12 Asking Price) and the
    # hard-coded default row (i=2 Asset Type) always have cell content, so ISBLANK()
    # is always False → the white-fill CF would fire constantly and wipe out alternating bands.
    # Input row indices: 0,1,3,4,5,6,8,11  →  rows 5,6,8,9,10,11,13,16
    for inp_i in [0, 1, 3, 4, 5, 6, 8, 11]:
        inp_r = SNAP_START + inp_i
        add_cf_clear(ws, f"C{inp_r}:C{inp_r}")

    h(ws, SNAP_END + 1, 10)  # spacer row 18

    # ── TENANT SUMMARY TABLE ──────────────────────────────────────────────────
    TSEC_ROW = SNAP_END + 2   # row 19
    sec(ws, TSEC_ROW, "TENANT SUMMARY", col_start=2, ncols=6)

    THDR_ROW = TSEC_ROW + 1   # row 20
    h(ws, THDR_ROW, 18)
    for col, label, al in [
        (2, "TENANT",       AL_L),
        (3, "SUITE",        AL_C),
        (4, "SF",           AL_R),
        (5, "% OF GLA",     AL_R),
        (6, "ANNUAL RENT",  AL_R),
        (7, "% OF RENT",    AL_R),
    ]:
        c = ws.cell(row=THDR_ROW, column=col, value=label)
        c.font = FT_CHDR; c.fill = F_NAVY; c.alignment = al

    TDATA_START = THDR_ROW + 1   # row 21

    for i in range(5):
        r = TDATA_START + i
        h(ws, r, 18)
        bg = F_PALE if i % 2 == 0 else F_WHITE
        nr = _T_NAME[i]; sr = _T_SUITE[i]; sfr = _T_SF[i]; rntr = _T_RENT[i]

        # Tenant name
        c = ws.cell(row=r, column=2,
                    value=f"=IFERROR(IF({_A}!$C${nr}=\"\",\"—\",{_A}!$C${nr}),\"—\")")
        c.font = FT_FORM; c.fill = bg; c.alignment = AL_L

        # Suite
        c = ws.cell(row=r, column=3,
                    value=f"=IFERROR(IF({_A}!$C${sr}=\"\",\"\",{_A}!$C${sr}),\"\")")
        c.font = FT_FORM; c.fill = bg; c.alignment = AL_C

        # SF
        c = ws.cell(row=r, column=4,
                    value=f"=IFERROR(IF({_A}!$C${sfr}=\"\",\"\",{_A}!$C${sfr}),\"\")")
        c.font = FT_FORM; c.fill = bg; c.alignment = AL_R
        c.number_format = N0

        # % of GLA
        c = ws.cell(row=r, column=5,
                    value=f"=IFERROR(IF(OR({_A}!$C${sfr}=\"\",{_SF}=\"\",{_SF}=0),\"\",{_A}!$C${sfr}/{_SF}),\"\")")
        c.font = FT_FORM; c.fill = bg; c.alignment = AL_R
        c.number_format = P1

        # Annual Rent
        c = ws.cell(row=r, column=6,
                    value=f"=IFERROR(IF({_A}!$C${rntr}=\"\",\"\",{_A}!$C${rntr}),\"\")")
        c.font = FT_FORM; c.fill = bg; c.alignment = AL_R
        c.number_format = D0

        # % of Rent
        c = ws.cell(row=r, column=7,
                    value=f"=IFERROR(IF(OR({_A}!$C${rntr}=\"\",{_GPR}=\"\",{_GPR}=0),\"\",{_A}!$C${rntr}/{_GPR}),\"\")")
        c.font = FT_FORM; c.fill = bg; c.alignment = AL_R
        c.number_format = P1

    TDATA_END = TDATA_START + 4   # row 25

    # Totals row
    TTOT_ROW = TDATA_END + 1      # row 26
    h(ws, TTOT_ROW, 18)
    c = ws.cell(row=TTOT_ROW, column=2, value="TOTAL / OCCUPIED")
    c.font = FT_TOTAL; c.fill = F_TOT; c.alignment = AL_L

    # Suite: blank
    ws.cell(row=TTOT_ROW, column=3).fill = F_TOT

    # Total SF
    sf_cells = "+".join(
        f"IFERROR({_A}!$C${r},0)" for r in _T_SF
    )
    c = ws.cell(row=TTOT_ROW, column=4,
                value=f"=IFERROR({sf_cells},\"\")")
    c.font = FT_TOTAL; c.fill = F_TOT; c.alignment = AL_R; c.number_format = N0

    # Occupancy % (total SF / building SF)
    c = ws.cell(row=TTOT_ROW, column=5,
                value=f"=IFERROR(IF(OR({_OCC}=\"\",{_OCC}=0),\"\",{_OCC}),\"\")")
    c.font = FT_TOTAL; c.fill = F_TOT; c.alignment = AL_R; c.number_format = P1

    # Total Annual Rent
    rent_cells = "+".join(
        f"IFERROR({_A}!$C${r},0)" for r in _T_RENT
    )
    c = ws.cell(row=TTOT_ROW, column=6,
                value=f"=IFERROR({rent_cells},\"\")")
    c.font = FT_TOTAL; c.fill = F_TOT; c.alignment = AL_R; c.number_format = D0

    # % of Rent total (always 100%)
    c = ws.cell(row=TTOT_ROW, column=7, value=1)
    c.font = FT_TOTAL; c.fill = F_TOT; c.alignment = AL_R; c.number_format = P1

    h(ws, TTOT_ROW + 1, 10)   # spacer row 27

    # ── PRICING STRATEGY ──────────────────────────────────────────────────────
    PSEC_ROW = TTOT_ROW + 2   # row 28
    sec(ws, PSEC_ROW, "PRICING STRATEGY", col_start=2, ncols=5)

    # Recommended Asking Price sub-header
    PASK_HDR = PSEC_ROW + 1   # row 29
    h(ws, PASK_HDR, 16)
    c = ws.cell(row=PASK_HDR, column=2, value="RECOMMENDED ASKING PRICE")
    c.font = ft("Calibri", 10, b=True, c=NAVY); c.fill = F_PALE; c.alignment = AL_L
    for col in range(3, 7):
        ws.cell(row=PASK_HDR, column=col).fill = F_PALE

    # Column headers for pricing
    PCHDR_ROW = PASK_HDR + 1  # row 30
    h(ws, PCHDR_ROW, 18)
    for col, label, al in [
        (2, "",           AL_L),
        (3, "CAP RATE",   AL_C),
        (4, "PRICE",      AL_R),
        (5, "PRICE / SF", AL_R),
        (6, "NOI",        AL_R),
    ]:
        c = ws.cell(row=PCHDR_ROW, column=col, value=label)
        c.font = FT_CHDR; c.fill = F_NAVY; c.alignment = al

    # Asking price row
    PASK_ROW = PCHDR_ROW + 1  # row 31
    h(ws, PASK_ROW, 22)
    ws.cell(row=PASK_ROW, column=2, value="Asking Price").font = FT_LABEL
    ws.cell(row=PASK_ROW, column=2).fill = F_TOT
    ws.cell(row=PASK_ROW, column=2).alignment = AL_L
    inp(ws, PASK_ROW, 3, fmt=P2).fill = F_TOT
    frm(ws, PASK_ROW, 4,
        f'=IFERROR(IF(OR({_NOI}="",C{PASK_ROW}="",C{PASK_ROW}=0),"",{_NOI}/C{PASK_ROW}),"")',
        fmt=D0).fill = F_TOT
    frm(ws, PASK_ROW, 5,
        f'=IFERROR(IF(OR(D{PASK_ROW}="",{_SF}="",{_SF}=0),"",D{PASK_ROW}/{_SF}),"")',
        fmt=D2).fill = F_TOT
    frm(ws, PASK_ROW, 6,
        f'=IFERROR(IF({_NOI}="","",{_NOI}),"")',
        fmt=D0).fill = F_TOT
    add_cf_clear(ws, f"C{PASK_ROW}")

    h(ws, PASK_ROW + 1, 10)   # spacer

    # Expected Trade Range sub-header
    ETRNG_HDR = PASK_ROW + 2  # row 33
    h(ws, ETRNG_HDR, 16)
    c = ws.cell(row=ETRNG_HDR, column=2, value="EXPECTED TRADE RANGE")
    c.font = ft("Calibri", 10, b=True, c=NAVY); c.fill = F_PALE; c.alignment = AL_L
    for col in range(3, 7):
        ws.cell(row=ETRNG_HDR, column=col).fill = F_PALE

    # Column headers for trade range
    ETCHDR_ROW = ETRNG_HDR + 1  # row 34
    h(ws, ETCHDR_ROW, 18)
    for col, label, al in [
        (2, "",           AL_L),
        (3, "CAP RATE",   AL_C),
        (4, "PRICE",      AL_R),
        (5, "PRICE / SF", AL_R),
        (6, "NOTES",      AL_L),
    ]:
        c = ws.cell(row=ETCHDR_ROW, column=col, value=label)
        c.font = FT_CHDR; c.fill = F_NAVY; c.alignment = al

    # Low / High indication rows
    for i, (scenario, bg) in enumerate([
        ("Low Indication",  F_PALE),
        ("High Indication", F_WHITE),
    ]):
        r = ETCHDR_ROW + 1 + i  # rows 35, 36
        h(ws, r, 20)
        ws.cell(row=r, column=2, value=scenario).font = FT_DATA
        ws.cell(row=r, column=2).fill = bg; ws.cell(row=r, column=2).alignment = AL_L
        inp(ws, r, 3, fmt=P2).fill = bg
        frm(ws, r, 4,
            f'=IFERROR(IF(OR({_NOI}="",C{r}="",C{r}=0),"",{_NOI}/C{r}),"")',
            fmt=D0).fill = bg
        frm(ws, r, 5,
            f'=IFERROR(IF(OR(D{r}="",{_SF}="",{_SF}=0),"",D{r}/{_SF}),"")',
            fmt=D2).fill = bg
        inp(ws, r, 6).fill = bg

    ETR_LOW = ETCHDR_ROW + 1
    ETR_HIGH = ETCHDR_ROW + 2
    add_cf_clear(ws, f"C{ETR_LOW}:C{ETR_HIGH}")
    add_cf_clear(ws, f"F{ETR_LOW}:F{ETR_HIGH}")

    h(ws, ETR_HIGH + 1, 10)   # spacer

    # ── BROKER RECOMMENDATION ─────────────────────────────────────────────────
    BREC_SEC = ETR_HIGH + 2
    sec(ws, BREC_SEC, "BROKER RECOMMENDATION", col_start=2, ncols=5)
    BREC_ROW = BREC_SEC + 1
    h(ws, BREC_ROW, 80)
    c = ws.cell(row=BREC_ROW, column=2,
                value="[Enter pricing recommendation and rationale. Include comparable cap rates, "
                      "market conditions, tenant mix, lease terms remaining, occupancy trend, and the competitive "
                      "bid landscape expected to drive pricing into or beyond the indicated trade range.]")
    c.font = FT_INPUT; c.fill = F_YELL; c.alignment = AL_TL
    merge(ws, BREC_ROW, 2, BREC_ROW, 6)
    add_cf_clear(ws, f"B{BREC_ROW}")
    h(ws, BREC_ROW + 1, 10)   # spacer

    # ── ENGAGEMENT PROPOSAL ───────────────────────────────────────────────────
    ENG_SEC = BREC_ROW + 2
    sec(ws, ENG_SEC, "ENGAGEMENT PROPOSAL", col_start=2, ncols=5)

    ENG_CHDR = ENG_SEC + 1
    h(ws, ENG_CHDR, 18)
    for col, label, al in [
        (2, "TERM",   AL_L),
        (3, "DETAIL", AL_L),
    ]:
        c = ws.cell(row=ENG_CHDR, column=col, value=label)
        c.font = FT_CHDR; c.fill = F_NAVY; c.alignment = al
    for col in range(4, 7):
        ws.cell(row=ENG_CHDR, column=col).fill = F_NAVY
    merge(ws, ENG_CHDR, 3, ENG_CHDR, 6)

    engagement = [
        ("Listing Type",
         "Exclusive Right to Sell",
         20, False),
        ("Brokerage Commission",
         "[X]% of Gross Sales Price, Payable by Seller at Closing — "
         "Split 50/50 Between Procuring Broker and Listing Broker",
         52, True),
        ("Transactional Brokerage Discount",
         "[X]% of Gross Sales Price, Payable by Seller at Closing — "
         "Applicable When Team Briggs Represents Both Buyer and Seller in the Transaction",
         52, True),
        ("Listing Agreement Term",
         "Six (6) Months",
         20, False),
        ("Additional Terms",
         "All Marketing and Pursuit Costs Paid by NorthMarq Commercial in Advance "
         "Regardless of Closing",
         44, True),
    ]

    ENG_DATA_START = ENG_CHDR + 1
    for i, (label, val, ht, is_tall) in enumerate(engagement):
        r = ENG_DATA_START + i
        h(ws, r, ht)
        bg = F_PALE if i % 2 == 0 else F_WHITE
        c = ws.cell(row=r, column=2, value=label)
        c.font = FT_LABEL; c.fill = bg; c.alignment = AL_TL if is_tall else AL_L
        v = ws.cell(row=r, column=3, value=val)
        v.font = FT_INPUT; v.fill = bg
        v.alignment = AL_TL if is_tall else AL_L
        for col in range(4, 7):
            ws.cell(row=r, column=col).fill = bg
        merge(ws, r, 3, r, 6)

    ENG_DATA_END = ENG_DATA_START + len(engagement) - 1
    add_cf_clear(ws, f"C{ENG_DATA_START}:C{ENG_DATA_END}")

    # NM branded bottom
    bot_r = ENG_DATA_END + 2
    h(ws, bot_r, 10)
    from openpyxl.styles import Border, Side
    for ci in range(2, 7):
        ws.cell(row=bot_r, column=ci).border = Border(
            bottom=Side(style='medium', color=NAVY)
        )
