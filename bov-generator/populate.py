"""
populate.py — Write JSON inputs into the Assumptions & Flags cells
of a freshly built BOV workbook (before recalc).

NNN cell map (Assumptions & Flags sheet):
  Left side (col C):
    C6  = Property Address
    C7  = Tenant / Operator
    C8  = Guarantor
    C9  = Building SF
    C13 = Estimated Close Date
    C16 = Recommended Asking Cap Rate
    C24 = Year 1 Base Rent
    C25 = Annual Rent Escalation %
    C26 = Tenant Reimbursements
    C30 = Mgmt Fee %
    C31 = Capital / Replacement Reserves
    C36 = Purchase Price
  Right side (col I):
    I6  = Purchase Price
    I7  = LTV %
    I10 = Interest Rate
    I11 = Amortization (years)
    I13 = Hold Period (years)
    I20 = Exit Cap Rate

MOB cell map (Assumptions & Flags sheet):
  Left side (col C):
    C6  = Property Address / Name
    C9  = Total Building SF (GLA)
    C14 = Estimated Close Date
    Per tenant (T1–T5, 9-row stride starting at row 17):
      T1: C18=name, C19=suite, C20=SF, C21=rent, C22=esc%, C23=type, C24=reimb
      T2: C27=name, C28=suite, C29=SF, C30=rent, C31=esc%, C32=type, C33=reimb
      T3: C36..C42  |  T4: C45..C51  |  T5: C54..C60
    C63 = Vacancy / Credit Loss %
    C64 = Portfolio Avg Escalation %
    C69 = Real Estate Taxes ($/yr)
    C70 = Insurance ($/yr)
    C71 = CAM / Janitorial ($/yr)
    C72 = Management Fee %
    C76 = Capital / Replacement Reserves ($/yr)
    C94 = Purchase Price (left-side input)
  Right side (col I):
    I6=purchase_price, I7=LTV, I10=rate, I11=amort, I13=hold, I20=exit_cap

Rent Schedule tab: identification block + per-lease-period grid.
  NNN: ident E6-E12; grid rows 16-45 (C start, D end, E label, F annual, I esc, J status).
  MOB: per-tenant grid begins at row 17 + t*26 (15 rows each).
"""

from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border

SHEET = "Assumptions & Flags"


def _date(val: str):
    """Parse YYYY-MM-DD string to Python date; return None if blank/invalid."""
    if not val:
        return None
    try:
        return datetime.strptime(val, "%Y-%m-%d").date()
    except ValueError:
        return None


def _pct(val):
    """Pass through float directly — openpyxl stores percentages as fractions (0.065 = 6.5%)."""
    return val if val is not None else ""


def _num(val):
    return val if val is not None else ""


def fill_nnn_assumptions(wb, req: dict) -> None:
    """Write NNN deal inputs into Assumptions & Flags cells."""
    ws = wb[SHEET]
    prop   = req.get("property", {})
    uw     = req.get("underwriting", {})
    tenant = (req.get("tenants") or [{}])[0]

    # Property identifiers
    ws["C6"] = prop.get("address", "")
    ws["C7"] = tenant.get("name", "")
    ws["C8"] = tenant.get("guarantor", "")
    ws["C9"] = _num(prop.get("building_sf"))

    # Close date
    d = _date(prop.get("close_date", ""))
    if d:
        ws["C13"] = d

    # Pricing
    ws["C16"] = _pct(uw.get("going_in_cap"))

    # Revenue
    ws["C24"] = _num(tenant.get("year1_rent"))
    ws["C25"] = _pct(tenant.get("escalation_pct"))
    ws["C26"] = _num(tenant.get("reimbursements", 0))

    # Expenses
    ws["C30"] = _pct(tenant.get("mgmt_fee_pct", 0))
    ws["C31"] = _num(uw.get("capital_reserves", 0))

    # Acquisition
    ws["C36"] = _num(uw.get("purchase_price"))

    # Right side — debt / leverage
    ws["I6"]  = _num(uw.get("purchase_price"))
    ws["I7"]  = _pct(uw.get("ltv", 0.65))
    ws["I10"] = _pct(uw.get("interest_rate", 0.065))
    ws["I11"] = _num(uw.get("amortization_years", 25))
    ws["I13"] = _num(uw.get("hold_years", 10))
    ws["I20"] = _pct(uw.get("exit_cap"))


# Tenant row offsets within each 9-row MOB tenant block
# Block start rows (header row): 17, 26, 35, 44, 53
_MOB_TENANT_STARTS = [17, 26, 35, 44, 53]
_ROW_NAME  = 1   # name row = start + 1
_ROW_SUITE = 2
_ROW_SF    = 3
_ROW_RENT  = 4
_ROW_ESC   = 5
_ROW_TYPE  = 6
_ROW_REIMB = 7


def fill_mob_assumptions(wb, req: dict) -> None:
    """Write MOB deal inputs into Assumptions & Flags cells."""
    ws = wb[SHEET]
    prop    = req.get("property", {})
    uw      = req.get("underwriting", {})
    tenants = req.get("tenants", [])

    # Property
    ws["C6"] = prop.get("address", "")
    ws["C9"] = _num(prop.get("building_sf"))

    d = _date(prop.get("close_date", ""))
    if d:
        ws["C14"] = d

    # Per-tenant data (up to 5 tenants)
    for i, t_start in enumerate(_MOB_TENANT_STARTS):
        if i >= len(tenants):
            break
        t = tenants[i]
        ws.cell(row=t_start + _ROW_NAME,  column=3).value = t.get("name", "")
        ws.cell(row=t_start + _ROW_SUITE, column=3).value = t.get("suite", "")
        ws.cell(row=t_start + _ROW_SF,    column=3).value = _num(t.get("sf"))
        ws.cell(row=t_start + _ROW_RENT,  column=3).value = _num(t.get("year1_rent"))
        ws.cell(row=t_start + _ROW_ESC,   column=3).value = _pct(t.get("escalation_pct"))
        ws.cell(row=t_start + _ROW_TYPE,  column=3).value = t.get("lease_type", "NNN")
        ws.cell(row=t_start + _ROW_REIMB, column=3).value = _num(t.get("reimbursements", 0))

    # Vacancy / escalation portfolio averages
    ws["C63"] = _pct(uw.get("vacancy_pct", 0.05))
    if tenants:
        avg_esc = sum(t.get("escalation_pct", 0) for t in tenants) / len(tenants)
        ws["C64"] = _pct(avg_esc)

    # Operating expenses (MOB LL-responsible)
    ws["C69"] = _num(uw.get("real_estate_taxes", 0))
    ws["C70"] = _num(uw.get("insurance", 0))
    ws["C71"] = _num(uw.get("cam", 0))
    ws["C72"] = _pct(uw.get("mgmt_fee_pct", 0.04))
    ws["C76"] = _num(uw.get("capital_reserves", 0))

    # Acquisition pricing
    ws["C94"] = _num(uw.get("purchase_price"))
    ws["C88"] = _pct(uw.get("going_in_cap"))

    # Right side — debt / leverage (same as NNN)
    ws["I6"]  = _num(uw.get("purchase_price"))
    ws["I7"]  = _pct(uw.get("ltv", 0.65))
    ws["I10"] = _pct(uw.get("interest_rate", 0.065))
    ws["I11"] = _num(uw.get("amortization_years", 25))
    ws["I13"] = _num(uw.get("hold_years", 10))
    ws["I20"] = _pct(uw.get("exit_cap"))


RENT_SCHEDULE_SHEET = "Rent Schedule"


def _add_years(d, n):
    """Shift date d by n years, clamping Feb 29 to Feb 28."""
    try:
        return d.replace(year=d.year + n)
    except ValueError:
        return d.replace(year=d.year + n, day=28)


def _rent_periods_from(tenant: dict):
    """
    Build the rent-schedule rows for one tenant.
    Preferred: tenant['rent_schedule'] = [{label,start_date,end_date,annual_rent,status}, ...]
      — the exact contracted schedule (handles stepped rents and option bumps).
    Fallback: compute from year1_rent x (1+escalation)^(n-1) across the lease term.
    Returns dicts: {label, start(date|None), end(date|None), annual, esc, status}.
    """
    rows = []
    rs = tenant.get("rent_schedule")
    if rs:
        prev = None
        for i, p in enumerate(rs):
            annual = _num(p.get("annual_rent"))
            esc = None
            if i > 0 and isinstance(annual, (int, float)) and isinstance(prev, (int, float)) and prev:
                esc = round(annual / prev - 1, 4)
            rows.append({
                "label":  p.get("label") or f"Year {i + 1}",
                "start":  _date(p.get("start_date", "")),
                "end":    _date(p.get("end_date", "")),
                "annual": annual,
                "esc":    esc,
                "status": p.get("status") or "Contracted",
            })
            if isinstance(annual, (int, float)):
                prev = annual
        return rows

    # Fallback — compute from Year-1 rent and a single escalation rate
    y1 = tenant.get("year1_rent")
    if y1 is None:
        return rows
    esc = tenant.get("escalation_pct") or 0.0
    comm = _date(tenant.get("lease_commencement", ""))
    exp = _date(tenant.get("lease_expiration", ""))
    nyears = 10
    if comm and exp:
        nyears = max(1, round((exp - comm).days / 365.25))
    for n in range(nyears):
        start = _add_years(comm, n) if comm else None
        end = (_add_years(comm, n + 1) - timedelta(days=1)) if comm else None
        rows.append({
            "label":  f"Year {n + 1}",
            "start":  start,
            "end":    end,
            "annual": round(y1 * ((1 + esc) ** n)),
            "esc":    (esc if n > 0 else None),
            "status": "Contracted (computed)",
        })
    return rows


def _write_rent_grid(ws, data_start: int, periods: list, max_rows: int,
                     hide_unused: bool = True) -> int:
    """
    Write periods into grid rows: C=start D=end E=label F=annual I=esc J=status.
    First period's escalation shows "-" (no prior period to step from).
    Unused grid rows are hidden so the TOTALS bar sits directly under the last
    contracted period. Returns the number of periods written.
    """
    n = 0
    for i, p in enumerate(periods[:max_rows]):
        rr = data_start + i
        if p.get("start") is not None:
            ws.cell(row=rr, column=3).value = p["start"]
        if p.get("end") is not None:
            ws.cell(row=rr, column=4).value = p["end"]
        if p.get("label"):
            ws.cell(row=rr, column=5).value = p["label"]
        if isinstance(p.get("annual"), (int, float)):
            ws.cell(row=rr, column=6).value = p["annual"]
        if isinstance(p.get("esc"), (int, float)):
            ws.cell(row=rr, column=9).value = p["esc"]
        elif i == 0:
            ws.cell(row=rr, column=9).value = "-"   # first period: no escalation
        if p.get("status"):
            ws.cell(row=rr, column=10).value = p["status"]
        n += 1
    if hide_unused:
        for rr in range(data_start + n, data_start + max_rows):
            ws.row_dimensions[rr].hidden = True
    return n


def fill_nnn_rent_schedule(wb, req: dict) -> None:
    """Populate the single-tenant Rent Schedule tab (identification + grid)."""
    if RENT_SCHEDULE_SHEET not in wb.sheetnames:
        return
    ws = wb[RENT_SCHEDULE_SHEET]
    prop = req.get("property", {})
    tenant = (req.get("tenants") or [{}])[0]

    ws["E6"] = prop.get("address", "")
    ws["E7"] = tenant.get("name", "")
    ws["E8"] = tenant.get("guarantor", "")
    comm = _date(tenant.get("lease_commencement", ""))
    if comm:
        ws["E9"] = comm
    exp = _date(tenant.get("lease_expiration", ""))
    if exp:
        ws["E10"] = exp
        # Static remaining term on the same basis as the Lease Abstract / Exec
        # Summary (as of close date, not TODAY) so the number matches elsewhere.
        ref = _ref_date(prop)
        yrs = (exp - ref).days / 365.25
        ws["E11"] = f"{yrs:.1f} yrs" if yrs >= 0 else "Expired"
    ws["E12"] = _num(prop.get("building_sf"))

    _write_rent_grid(ws, 16, _rent_periods_from(tenant), 30)


# MOB rent-schedule grid: tenant t grid begins at row 17 + t*26 (15 rows each)
_MOB_RS_GRID_START = 17
_MOB_RS_STRIDE = 26
_MOB_RS_MAX_ROWS = 15


def fill_mob_rent_schedule(wb, req: dict) -> None:
    """Populate each per-tenant grid on the multi-tenant Rent Schedule tab."""
    if RENT_SCHEDULE_SHEET not in wb.sheetnames:
        return
    ws = wb[RENT_SCHEDULE_SHEET]
    for t, tenant in enumerate((req.get("tenants") or [])[:5]):
        data_start = _MOB_RS_GRID_START + t * _MOB_RS_STRIDE
        _write_rent_grid(ws, data_start, _rent_periods_from(tenant), _MOB_RS_MAX_ROWS)


COVER_SHEET = "Cover"
EXECSUM_SHEET = "Executive Summary"


def _analysis_date() -> str:
    return datetime.now().strftime("%B %d, %Y")


def _property_display(req: dict) -> str:
    """'{Property/Tenant Name}  —  {City, State}' for the Cover and Exec Summary subtitle."""
    prop = req.get("property", {})
    name = (prop.get("name") or "").strip()
    if not name:
        tenants = req.get("tenants") or []
        if req.get("asset_type", "").upper() != "MOB" and tenants and tenants[0].get("name"):
            name = tenants[0]["name"]
        else:
            name = prop.get("address", "")
    cs = (prop.get("city_state") or "").strip()
    return f"{name}  —  {cs}" if cs else name


def fill_cover(wb, req: dict) -> None:
    """Fill the Cover tab identity cells (B10=Property/Tenant, B12=Address, B14=Client, B18=Analysis Date)."""
    if COVER_SHEET not in wb.sheetnames:
        return
    ws = wb[COVER_SHEET]
    prop = req.get("property", {})
    client = req.get("client", {})
    ws["B10"] = _property_display(req)
    ws["B12"] = prop.get("address", "")
    ws["B14"] = client.get("last_name", "")
    ws["B18"] = _analysis_date()


def fill_exec_subtitle(wb, req: dict) -> None:
    """Fill the Executive Summary subtitle (B2), replacing the placeholder."""
    if EXECSUM_SHEET not in wb.sheetnames:
        return
    ws = wb[EXECSUM_SHEET]
    kind = "Multi-Tenant MOB" if req.get("asset_type", "").upper() == "MOB" else "Single-Tenant NNN"
    ws["B2"] = f"{_property_display(req)}  ·  {kind}  ·  Investment Snapshot  ·  {_analysis_date()}"


REAL_ESTATE_SHEET = "Real Estate"

# ── Real Estate tab cell map ──────────────────────────────────────────────────
# The NNN (bov_tabs_3_real_estate.py) and MOB (mob_tab_3_real_estate.py) Real
# Estate tabs share an identical layout, so one map serves both.
#
# Short-form summary — attribute label in col C, wide finding merged into col D.
_RE_SF = {                       # attribute        -> D-column row
    "address":       7,          # Property Address (address + city/state)
    "building_sf":   8,          # Building SF
    "year":          9,          # Year Built / Renovated
    "site_area":    10,          # Site Area (Acres)
    "zoning":       11,          # Zoning
    "flood_zone":   12,          # Flood Zone
    "parking":      13,          # Parking
    "condition":    14,          # Condition (Overall)
    "environmental":15,          # Environmental Status
    "proximity":    16,          # Proximity / Demand Generators
    "market":       17,          # Market Context
    "strengths":    18,          # Notable Strengths
    "concerns":     19,          # Notable Concerns
    "commentary":   20,          # Broker Commentary
}
# Long-form matrix — finding in col E (5), source in col F (6), notes in col G (7).
_RE_LF = {                       # attribute                      -> E-column row
    "address":       25,
    "city_state":    26,
    "county":        27,
    "msa":           28,
    "population":    29,
    "median_income": 30,
    "traffic":       31,
    "proximity":     32,
    "market_rent":   33,
    "site_area":     35,
    "parcel_apn":    36,
    "legal":         37,
    "lot_config":    38,
    "frontage":      39,
    "topography":    40,
    "flood_zone":    41,
    "utilities":     42,
    "building_sf":   44,
    "year_built":    45,
    "year_renov":    46,
    "construction":  47,
    "roof":          48,
    "hvac":          49,
    "ada":           50,
    "condition":     51,
    "deferred_maint":52,
    "zoning":        54,
    "permitted_use": 55,
    "drive_through": 56,
    "signage":       57,
    "covenants":     58,
    "easements":     59,
    "access_points": 61,
    "shared_access": 62,
    "parking_spaces":63,
    "parking_ratio": 64,
    "delivery":      65,
    "phase_i":       67,
    "phase_ii":      68,
    "known_recs":    69,
    "ust":           70,
    "comp1":         72,
    "comp2":         73,
    "comp3":         74,
    "implied_cap":   75,
    "implied_psf":   76,
}
_RE_FINDING_COL = 5   # E
_RE_SOURCE_COL  = 6   # F


def _s(v) -> str:
    """Trim to a clean string; empty for None/blank."""
    if v is None:
        return ""
    return str(v).strip()


def _fmt_int(v):
    try:
        return f"{int(round(float(v))):,}"
    except (TypeError, ValueError):
        return ""


def _fmt_pct(v):
    """Accept a fraction (0.0675) or a whole number (6.75) and render as a %."""
    try:
        f = float(v)
    except (TypeError, ValueError):
        return ""
    if f > 1:            # already expressed as a percent
        f /= 100.0
    return f"{f:.2%}"


def _year_summary(built, renov) -> str:
    b, r = _s(built), _s(renov)
    if b and r:
        return f"{b}  /  renov. {r}"
    if b:
        return b
    if r:
        return f"Renovated {r}"
    return ""


def _comp_finding(comp: dict) -> str:
    """Compose a one-line comp string from an explicit summary or its parts."""
    if _s(comp.get("summary")):
        return _s(comp["summary"])
    bits = []
    if _s(comp.get("address")):
        bits.append(_s(comp["address"]))
    tail = []
    if comp.get("sale_price") is not None:
        tail.append(f"${_fmt_int(comp['sale_price'])}")
    if comp.get("price_sf") is not None:
        tail.append(f"${_fmt_int(comp['price_sf'])}/SF")
    if comp.get("cap_rate") is not None:
        tail.append(f"{_fmt_pct(comp['cap_rate'])} cap")
    if _s(comp.get("sale_date")):
        tail.append(_s(comp["sale_date"]))
    if tail:
        bits.append(" · ".join(tail))
    return " — ".join(bits)


def fill_real_estate(wb, req: dict) -> None:
    """
    Auto-fill the Real Estate tab (shared NNN/MOB layout) from the optional
    `real_estate` block plus address/building-SF already on `property`.
    Only supplied values are written; everything else stays a broker-input cell.
    The tab's conditional formatting turns each filled yellow cell white.
    """
    if REAL_ESTATE_SHEET not in wb.sheetnames:
        return
    ws = wb[REAL_ESTATE_SHEET]

    # Fail-safe: the cell map is positional. If the tab layout has shifted,
    # skip rather than risk writing findings into the wrong rows. Anchor on the
    # short-form label (C7) and a long-form attribute (D36).
    if (_s(ws.cell(row=7, column=3).value) != "Property Address"
            or _s(ws.cell(row=36, column=4).value) != "Parcel APN / Tax ID"):
        return

    prop = req.get("property", {}) or {}
    re_d = req.get("real_estate") or {}

    address    = _s(prop.get("address"))
    city_state = _s(prop.get("city_state"))
    bsf        = prop.get("building_sf")
    building_sf_str = f"{_fmt_int(bsf)} SF" if bsf is not None else ""
    src        = _s(re_d.get("default_source"))

    # ── Derived / formatted values ──────────────────────────────────────────
    full_address = ", ".join(p for p in [address, city_state] if p)
    year_sf   = _year_summary(re_d.get("year_built"), re_d.get("year_renovated"))
    site_str  = (f"{re_d['site_area_acres']:g} acres"
                 if re_d.get("site_area_acres") is not None else "")
    frontage_str = (f"{_fmt_int(re_d['frontage_lf'])} LF"
                    if re_d.get("frontage_lf") is not None else "")

    zoning_sf = _s(re_d.get("zoning"))
    if zoning_sf and _s(re_d.get("zoning_description")):
        zoning_sf = f"{zoning_sf} — {_s(re_d['zoning_description'])}"
    elif not zoning_sf:
        zoning_sf = _s(re_d.get("zoning_description"))

    # Parking ratio: use supplied, else compute from spaces + building SF.
    spaces = re_d.get("parking_spaces")
    ratio  = re_d.get("parking_ratio")
    if ratio is None and spaces is not None and bsf:
        try:
            ratio = float(spaces) / (float(bsf) / 1000.0)
        except (TypeError, ValueError, ZeroDivisionError):
            ratio = None
    ratio_str = f"{ratio:.2f} per 1,000 SF" if isinstance(ratio, (int, float)) else ""
    if spaces is not None and ratio_str:
        parking_sf = f"{_fmt_int(spaces)} spaces  ({ratio:.1f} / 1,000 SF)"
    elif spaces is not None:
        parking_sf = f"{_fmt_int(spaces)} spaces"
    else:
        parking_sf = ""

    def _w(row: int, col: int, value: str) -> None:
        """Write a finding, and stamp the source column when a value landed."""
        value = _s(value)
        if not value:
            return
        ws.cell(row=row, column=col).value = value
        if col == _RE_FINDING_COL and src:
            ws.cell(row=row, column=_RE_SOURCE_COL).value = src

    # ── Short-form summary (col D) ──────────────────────────────────────────
    sf = _RE_SF
    _w(sf["address"],       4, full_address)
    _w(sf["building_sf"],   4, building_sf_str)
    _w(sf["year"],          4, year_sf)
    _w(sf["site_area"],     4, site_str)
    _w(sf["zoning"],        4, zoning_sf)
    _w(sf["flood_zone"],    4, _s(re_d.get("flood_zone")))
    _w(sf["parking"],       4, parking_sf)
    _w(sf["condition"],     4, _s(re_d.get("condition")))
    _w(sf["environmental"], 4, _s(re_d.get("environmental_status")) or _s(re_d.get("phase_i")))
    _w(sf["proximity"],     4, _s(re_d.get("proximity_demand_generators")))
    _w(sf["market"],        4, _s(re_d.get("market_rent_context")))
    _w(sf["strengths"],     4, _s(re_d.get("notable_strengths")))
    _w(sf["concerns"],      4, _s(re_d.get("notable_concerns")))
    _w(sf["commentary"],    4, _s(re_d.get("broker_commentary")))

    # ── Long-form diligence matrix (col E finding, col F source) ────────────
    lf = _RE_LF
    _w(lf["address"],       5, address)
    _w(lf["city_state"],    5, city_state)
    _w(lf["county"],        5, _s(re_d.get("county")))
    _w(lf["msa"],           5, _s(re_d.get("msa_submarket")))
    _w(lf["population"],    5, _s(re_d.get("population_1_3_5")))
    _w(lf["median_income"], 5, _s(re_d.get("median_hh_income")))
    _w(lf["traffic"],       5, _s(re_d.get("traffic_counts")))
    _w(lf["proximity"],     5, _s(re_d.get("proximity_demand_generators")))
    _w(lf["market_rent"],   5, _s(re_d.get("market_rent_context")))
    _w(lf["site_area"],     5, site_str)
    _w(lf["parcel_apn"],    5, _s(re_d.get("parcel_apn")))
    _w(lf["legal"],         5, _s(re_d.get("legal_description")))
    _w(lf["lot_config"],    5, _s(re_d.get("lot_configuration")))
    _w(lf["frontage"],      5, frontage_str)
    _w(lf["topography"],    5, _s(re_d.get("topography")))
    _w(lf["flood_zone"],    5, _s(re_d.get("flood_zone")))
    _w(lf["utilities"],     5, _s(re_d.get("utilities")))
    _w(lf["building_sf"],   5, building_sf_str)
    _w(lf["year_built"],    5, _s(re_d.get("year_built")))
    _w(lf["year_renov"],    5, _s(re_d.get("year_renovated")))
    _w(lf["construction"],  5, _s(re_d.get("construction_type")))
    _w(lf["roof"],          5, _s(re_d.get("roof")))
    _w(lf["hvac"],          5, _s(re_d.get("hvac")))
    _w(lf["ada"],           5, _s(re_d.get("ada_compliance")))
    _w(lf["condition"],     5, _s(re_d.get("condition")))
    _w(lf["deferred_maint"],5, _s(re_d.get("deferred_maintenance")))
    _w(lf["zoning"],        5, _s(re_d.get("zoning")))
    _w(lf["permitted_use"], 5, _s(re_d.get("permitted_use_confirmation")))
    _w(lf["drive_through"], 5, _s(re_d.get("drive_through_permitted")))
    _w(lf["signage"],       5, _s(re_d.get("signage_rights")))
    _w(lf["covenants"],     5, _s(re_d.get("restrictive_covenants")))
    _w(lf["easements"],     5, _s(re_d.get("easements")))
    _w(lf["access_points"], 5, _s(re_d.get("access_points")))
    _w(lf["shared_access"], 5, _s(re_d.get("shared_access")))
    _w(lf["parking_spaces"],5, _fmt_int(spaces) if spaces is not None else "")
    _w(lf["parking_ratio"], 5, ratio_str)
    _w(lf["delivery"],      5, _s(re_d.get("delivery_loading")))
    _w(lf["phase_i"],       5, _s(re_d.get("phase_i")))
    _w(lf["phase_ii"],      5, _s(re_d.get("phase_ii")))
    _w(lf["known_recs"],    5, _s(re_d.get("known_recs")))
    _w(lf["ust"],           5, _s(re_d.get("underground_storage_tanks")))

    # Comps — write finding, and a per-comp source overriding the default.
    comps = re_d.get("comps") or []
    for i, key in enumerate(["comp1", "comp2", "comp3"]):
        if i >= len(comps):
            break
        comp = comps[i] or {}
        finding = _comp_finding(comp)
        if finding:
            ws.cell(row=lf[key], column=5).value = finding
            comp_src = _s(comp.get("source")) or src
            if comp_src:
                ws.cell(row=lf[key], column=_RE_SOURCE_COL).value = comp_src

    _w(lf["implied_cap"], 5, _fmt_pct(re_d.get("implied_market_cap_rate"))
       if re_d.get("implied_market_cap_rate") is not None else "")
    _w(lf["implied_psf"], 5, f"${_fmt_int(re_d['implied_market_price_sf'])} / SF"
       if re_d.get("implied_market_price_sf") is not None else "")


LEASE_ABSTRACT_SHEET = "Lease Abstract"


def _fmt_date_str(v) -> str:
    """YYYY-MM-DD -> MM/DD/YYYY; pass through anything else unchanged."""
    d = _date(v) if v else None
    return d.strftime("%m/%d/%Y") if d else _s(v)


def _ref_date(prop: dict):
    """Reference point for remaining-term math: close date if given, else today."""
    return _date(prop.get("close_date", "")) or datetime.now().date()


def _remaining_term(exp_v, ref_date) -> str:
    d = _date(exp_v) if exp_v else None
    if not d:
        return ""
    yrs = (d - ref_date).days / 365.25
    return "Expired" if yrs < 0 else f"{yrs:.1f} years"


def _esc_summary(t: dict) -> str:
    rs = t.get("rent_schedule")
    if rs and len(rs) > 1:
        return "Stepped increases — see Rent Schedule"
    esc = t.get("escalation_pct")
    return f"{_fmt_pct(esc)} annually" if esc else ""


def _rent_and_psf(t: dict, prop: dict):
    y1 = t.get("year1_rent")
    if y1 is None:
        return "", ""
    sf = t.get("sf") or prop.get("building_sf")
    rent = f"${_fmt_int(y1)}"
    psf = ""
    try:
        if sf:
            psf = f"${float(y1) / float(sf):.2f}/SF"
    except (TypeError, ValueError, ZeroDivisionError):
        psf = ""
    return rent, psf


def _join(*parts) -> str:
    return " / ".join(p for p in (_s(x) for x in parts) if p)


def _lease_values(t: dict, prop: dict, ref_date):
    """Build (short-form dict, long-form dict, source) for one tenant's lease."""
    ab = t.get("abstract") or {}
    name       = _s(ab.get("tenant_of_record")) or _s(t.get("name"))
    lease_type = _s(t.get("lease_type")) or "NNN"
    structure  = _s(ab.get("lease_structure")) or lease_type
    comm       = _fmt_date_str(t.get("lease_commencement", ""))
    exp        = _fmt_date_str(t.get("lease_expiration", ""))
    remaining  = _remaining_term(t.get("lease_expiration", ""), ref_date)
    rent, psf  = _rent_and_psf(t, prop)
    rent_full  = f"{rent}  ({psf})" if rent and psf else rent
    esc        = _esc_summary(t)
    suite      = _s(t.get("suite"))
    if t.get("sf") is not None:
        leased_sf = f"{_fmt_int(t.get('sf'))} SF"
    elif prop.get("building_sf") is not None:
        leased_sf = f"{_fmt_int(prop.get('building_sf'))} SF"
    else:
        leased_sf = ""
    renew = "  ×  ".join(x for x in [_s(ab.get("num_renewal_options")),
                                     _s(ab.get("option_term_length"))] if x)

    premises = _s(prop.get("address"))
    if suite:
        premises = f"{premises}, Suite {suite}".strip(", ")

    sf_values = {
        "Tenant (Lease)":          name,
        "Guarantor":               _s(t.get("guarantor")),
        "Lease Type":              lease_type,
        "Lease Commencement":      comm,
        "Lease Expiration":        exp,
        "Remaining Lease Term":    remaining,
        "Suite / Unit":            suite,
        "Leased SF":               leased_sf,
        "Year 1 Base Rent":        rent_full,
        "Rent / SF":               psf,
        "Rent Escalations":        esc,
        "Renewal Options":         renew,
        "Renewal Rent Method":     _s(ab.get("renewal_rent_method")),
        "Expense Structure":       structure,
        "LL Responsibilities":     _s(ab.get("landlord_obligations")),
        "Early Termination":       _s(ab.get("early_termination")),
        "Assignment / Subletting": _join(ab.get("assignment_rights"), ab.get("subletting_rights")),
        "ROFR / ROFO":             _join(ab.get("rofr"), ab.get("rofo")),
        "Option to Purchase":      _s(ab.get("option_to_purchase")),
        "Key Lease Strengths":     _s(ab.get("key_lease_strengths")),
        "Key Lease Risks":         _s(ab.get("key_lease_risks")),
        "Broker Commentary":       _s(ab.get("broker_commentary")),
    }

    lf_values = {
        "Lease Type / Form":                          lease_type,
        "Execution Date":                             _fmt_date_str(ab.get("execution_date", "")),
        "Effective / Commencement Date":              comm,
        "Lease Expiration Date":                      exp,
        "Landlord of Record":                         _s(ab.get("landlord_of_record")),
        "Tenant of Record":                           name,
        "Guarantor (if any)":                         _s(t.get("guarantor")),
        "Demised Premises Address":                   premises,
        "Demised Premises / Suite":                   premises,
        "Leased SF (per lease)":                      leased_sf,
        "Leased SF (per survey)":                     _s(ab.get("leased_sf_per_survey")),
        "Exclusive Use / Permitted Use":              _s(ab.get("permitted_use")),
        "Permitted Use":                              _s(ab.get("permitted_use")),
        "Prohibited Uses":                            _s(ab.get("prohibited_uses")),
        "Base Rent — Year 1":                         rent_full,
        "Annual Rent Escalations":                    esc,
        "Rent Commencement Date":                     _fmt_date_str(ab.get("rent_commencement_date", "")),
        "Rent Abatement / Free Rent":                 _s(ab.get("rent_abatement")),
        "Percentage Rent":                            _s(ab.get("percentage_rent")),
        "Lease Structure (NNN / NN / Gross / MG)":    structure,
        "Lease Structure (NNN / Gross / MG)":         structure,
        "Real Estate Taxes — Responsibility":         _s(ab.get("taxes_responsibility")),
        "Insurance — Responsibility":                 _s(ab.get("insurance_responsibility")),
        "CAM / Maintenance — Responsibility":         _s(ab.get("cam_responsibility")),
        "Capital / Roof / Structure — Responsibility":_s(ab.get("capital_responsibility")),
        "Expense Cap (if any)":                       _s(ab.get("expense_cap")),
        "Landlord Obligations":                       _s(ab.get("landlord_obligations")),
        "Number of Renewal Options":                  _s(ab.get("num_renewal_options")),
        "Option Term Length":                         _s(ab.get("option_term_length")),
        "Renewal Rent — Method":                      _s(ab.get("renewal_rent_method")),
        "Renewal Notice Requirement":                 _s(ab.get("renewal_notice")),
        "Option to Purchase":                         _s(ab.get("option_to_purchase")),
        "Right of First Refusal (ROFR)":              _s(ab.get("rofr")),
        "Right of First Offer (ROFO)":                _s(ab.get("rofo")),
        "Assignment Rights":                          _s(ab.get("assignment_rights")),
        "Subletting Rights":                          _s(ab.get("subletting_rights")),
        "Change of Control Provisions":               _s(ab.get("change_of_control")),
        "Release of Guarantor on Assignment":         _s(ab.get("guarantor_release")),
        "Early Termination Right":                    _s(ab.get("early_termination")),
        "Termination Fee / Penalty":                  _s(ab.get("termination_fee")),
        "Default / Cure Periods":                     _s(ab.get("default_cure")),
        "Co-Tenancy Provisions":                      _s(ab.get("co_tenancy")),
        "Go-Dark Provision":                          _s(ab.get("go_dark")),
        "Co-Tenancy / Go-Dark Provision":             _join(ab.get("co_tenancy"), ab.get("go_dark")),
        "TI Allowance / Landlord Work":               _s(ab.get("ti_allowance")),
        "Signage Rights":                             _s(ab.get("signage_rights")),
        "Parking Allocation":                         _s(ab.get("parking_allocation")),
        "Subordination / SNDA / Estoppel":            _s(ab.get("snda")),
        "Condemnation Provisions":                    _s(ab.get("condemnation")),
        "Casualty / Damage Provisions":               _s(ab.get("casualty")),
        "Holdover Provisions":                        _s(ab.get("holdover")),
        "Notices":                                    _s(ab.get("notices")),
    }
    source = _s(ab.get("default_source")) or "Executed Lease"
    return sf_values, lf_values, source


def _lf_boundary(ws) -> int:
    """Row where the long-form section begins (col B header starts 'LONG-FORM')."""
    for row in range(1, ws.max_row + 1):
        if _s(ws.cell(row=row, column=2).value).upper().startswith("LONG-FORM"):
            return row
    return ws.max_row + 1


def fill_nnn_lease_abstract(wb, req: dict) -> None:
    """Single-tenant Lease Abstract: short-form (C->D) + long-form (D->H, source G)."""
    if LEASE_ABSTRACT_SHEET not in wb.sheetnames:
        return
    tenants = req.get("tenants") or []
    if not tenants:
        return
    ws = wb[LEASE_ABSTRACT_SHEET]
    sfv, lfv, source = _lease_values(tenants[0], req.get("property", {}) or {}, _ref_date(req.get("property", {}) or {}))
    lf_start = _lf_boundary(ws)

    # Short-form region: attribute label in col C (3), wide input in col D (4).
    for row in range(1, lf_start):
        label = _s(ws.cell(row=row, column=3).value)
        if label in sfv and sfv[label]:
            ws.cell(row=row, column=4).value = sfv[label]

    # Long-form region: clause label in col D (4); operative -> H (8), source -> G (7).
    for row in range(lf_start, ws.max_row + 1):
        clause = _s(ws.cell(row=row, column=4).value)
        if clause in lfv and lfv[clause]:
            ws.cell(row=row, column=8).value = lfv[clause]
            if source:
                ws.cell(row=row, column=7).value = source


def fill_mob_lease_abstract(wb, req: dict) -> None:
    """
    Multi-tenant Lease Abstract: a 5-column short-form summary table, then one
    stacked long-form section per tenant (teal header rows delimit sections).
    """
    if LEASE_ABSTRACT_SHEET not in wb.sheetnames:
        return
    tenants = (req.get("tenants") or [])[:5]
    if not tenants:
        return
    ws = wb[LEASE_ABSTRACT_SHEET]
    prop = req.get("property", {}) or {}
    ref = _ref_date(prop)
    vals = [_lease_values(t, prop, ref) for t in tenants]  # list of (sfv, lfv, source)
    lf_start = _lf_boundary(ws)

    # Short-form: provision label in col B (2); tenant columns C..G (3..7).
    for row in range(1, lf_start):
        label = _s(ws.cell(row=row, column=2).value)
        if not label:
            continue
        for i, (sfv, _lfv, _src) in enumerate(vals):
            v = sfv.get(label)
            if v:
                ws.cell(row=row, column=3 + i).value = v

    # Long-form: teal header rows carry a formula in col B and advance the tenant
    # index; within a section, clause label in col C (3) -> operative G (7), source F (6).
    sec_idx = -1
    for row in range(lf_start, ws.max_row + 1):
        b = ws.cell(row=row, column=2).value
        if isinstance(b, str) and b.startswith("="):
            sec_idx += 1
            continue
        if 0 <= sec_idx < len(vals):
            _sfv, lfv, source = vals[sec_idx]
            clause = _s(ws.cell(row=row, column=3).value)
            if clause in lfv and lfv[clause]:
                ws.cell(row=row, column=7).value = lfv[clause]
                if source:
                    ws.cell(row=row, column=6).value = source


RENT_ROLL_SHEET = "Rent Roll"
# Assumptions tenant-block header rows (7 data rows each, spacer before each).
_MOB_ASM_TENANT_STARTS = [17, 26, 35, 44, 53]
# Rent Roll tenant rows and Exec Summary tenant-summary rows.
_MOB_RR_DATA_START = 16
_MOB_ES_TENANT_START = 19   # tenant summary (asking cap/price removed from snapshot → shifted up 2)


def fill_mob_rent_roll(wb, req: dict) -> None:
    """Fill the Rent Roll's per-tenant lease commencement (G) and expiration (H) dates."""
    if RENT_ROLL_SHEET not in wb.sheetnames:
        return
    ws = wb[RENT_ROLL_SHEET]
    for t, tenant in enumerate((req.get("tenants") or [])[:5]):
        rr = _MOB_RR_DATA_START + t
        comm = _date(tenant.get("lease_commencement", ""))
        exp = _date(tenant.get("lease_expiration", ""))
        if comm:
            ws.cell(row=rr, column=7).value = comm
        if exp:
            ws.cell(row=rr, column=8).value = exp


def _hide_unused_mob_tenants(wb, n_tenants: int) -> None:
    """
    MOB templates build a fixed 5 tenant slots. Hide the rows/columns for slots
    beyond the actual tenant count so no phantom 'TENANT 5 / [Vacant]' shows.
    Uses hide (not delete) so all downstream formulas and fixed layouts are intact.
    """
    n = max(0, min(5, n_tenants))
    if n >= 5:
        return

    # 1) Assumptions & Flags — hide the spacer + 8-row block for each unused slot.
    if SHEET in wb.sheetnames:
        ws = wb[SHEET]
        for t in range(n, 5):
            s = _MOB_ASM_TENANT_STARTS[t]
            for rr in range(s - 1, s + 8):
                ws.row_dimensions[rr].hidden = True

    # 2) Rent Roll — hide the unused tenant rows (drops the "[Vacant]" phantoms).
    if RENT_ROLL_SHEET in wb.sheetnames:
        ws = wb[RENT_ROLL_SHEET]
        for t in range(n, 5):
            ws.row_dimensions[_MOB_RR_DATA_START + t].hidden = True

    # 3) Executive Summary tenant-summary — hide unused tenant rows.
    if EXECSUM_SHEET in wb.sheetnames:
        ws = wb[EXECSUM_SHEET]
        for t in range(n, 5):
            ws.row_dimensions[_MOB_ES_TENANT_START + t].hidden = True

    # 4) Lease Abstract — short-form: blank the unused tenant COLUMNS within the
    #    short-form band only (cols share the sheet with long-form, so we clear
    #    rather than hide). Long-form: hide the unused per-tenant sections.
    if LEASE_ABSTRACT_SHEET in wb.sheetnames:
        ws = wb[LEASE_ABSTRACT_SHEET]
        _blank_fill = PatternFill(fill_type=None)
        _no_border = Border()
        for col in range(3 + n, 8):            # unused tenant cols C..G
            for rr in range(6, 26):            # short-form band (header + 19 provisions)
                cell = ws.cell(row=rr, column=col)
                cell.value = None
                cell.fill = _blank_fill
                cell.border = _no_border
        lf_start = _lf_boundary(ws)
        sec_idx = -1
        for rr in range(lf_start, ws.max_row + 1):
            b = ws.cell(row=rr, column=2).value
            if isinstance(b, str) and b.startswith("="):
                sec_idx += 1
            if 0 <= sec_idx and sec_idx >= n:
                ws.row_dimensions[rr].hidden = True

    # 5) Rent Schedule — hide the unused per-tenant sections (teal header = col B formula).
    if RENT_SCHEDULE_SHEET in wb.sheetnames:
        ws = wb[RENT_SCHEDULE_SHEET]
        headers = [rr for rr in range(1, ws.max_row + 1)
                   if isinstance(ws.cell(row=rr, column=2).value, str)
                   and ws.cell(row=rr, column=2).value.startswith("=")]
        for t in range(n, len(headers)):
            start = headers[t]
            end = headers[t + 1] - 1 if t + 1 < len(headers) else start + _MOB_RS_STRIDE
            for rr in range(start, end + 1):
                ws.row_dimensions[rr].hidden = True


_ASREF = f"'{SHEET}'"


def _trade_range(ws, gic, low_row, high_row, spread=0.0025):
    """Pre-populate a cap-rate trade range around the going-in cap (broker-adjustable)."""
    if isinstance(gic, (int, float)):
        ws.cell(row=low_row,  column=3).value = round(gic + spread, 4)   # low indication = higher cap
        ws.cell(row=high_row, column=3).value = round(gic - spread, 4)   # high indication = lower cap


def fill_nnn_exec_summary(wb, req: dict) -> None:
    """Fill the NNN Executive Summary Investment Snapshot + trade-range caps."""
    if EXECSUM_SHEET not in wb.sheetnames:
        return
    ws = wb[EXECSUM_SHEET]
    prop = req.get("property", {}) or {}
    re_d = req.get("real_estate") or {}
    uw = req.get("underwriting") or {}
    t = (req.get("tenants") or [{}])[0]
    ref = _ref_date(prop)

    def put(row, val):
        if val not in (None, ""):
            ws.cell(row=row, column=3).value = val

    put(5, _s(prop.get("name")) or _s(t.get("name")))
    put(6, _s(prop.get("address")))
    put(7, _s(prop.get("city_state")))
    put(9, re_d.get("year_built"))
    put(10, prop.get("building_sf"))
    put(11, re_d.get("site_area_acres"))
    if _s(t.get("lease_type")):
        put(12, _s(t.get("lease_type")))
    comm = _date(t.get("lease_commencement", ""))
    exp = _date(t.get("lease_expiration", ""))
    if comm:
        put(13, comm)
    if exp:
        put(14, exp)
        put(15, round((exp - ref).days / 365.25, 1))
    put(16, t.get("year1_rent"))
    ws.cell(row=17, column=3).value = f'=IFERROR({_ASREF}!$C$33,"")'   # Y1 NOI (computed)
    name, guar = _s(t.get("name")), _s(t.get("guarantor"))
    put(18, f"{name}  /  {guar}" if guar else name)
    put(19, _s(t.get("credit_rating")))

    _trade_range(ws, uw.get("going_in_cap"), 28, 29)


def fill_mob_exec_summary(wb, req: dict) -> None:
    """Fill the MOB Executive Summary snapshot + trade-range caps (pricing section below)."""
    if EXECSUM_SHEET not in wb.sheetnames:
        return
    ws = wb[EXECSUM_SHEET]
    prop = req.get("property", {}) or {}
    re_d = req.get("real_estate") or {}
    uw = req.get("underwriting") or {}
    tenants = (req.get("tenants") or [])[:5]
    ref = _ref_date(prop)

    def put(row, val):
        if val not in (None, ""):
            ws.cell(row=row, column=3).value = val

    name = _s(prop.get("name"))
    addr = _s(prop.get("address"))
    put(5, "  —  ".join(x for x in [name, addr] if x) or addr)
    put(6, _s(prop.get("city_state")))
    put(8, re_d.get("year_built"))
    put(9, prop.get("building_sf"))
    put(10, re_d.get("site_area_acres"))
    put(11, len(tenants))
    # WALT — SF-weighted remaining lease term across tenants with an expiration.
    num = den = 0.0
    for t in tenants:
        exp = _date(t.get("lease_expiration", ""))
        sf = t.get("sf")
        if exp and isinstance(sf, (int, float)):
            num += float(sf) * max(0.0, (exp - ref).days / 365.25)
            den += float(sf)
    if den:
        put(13, round(num / den, 1))

    _trade_range(ws, uw.get("going_in_cap"), 33, 34)   # ETR Low / High after 2-row shift


def fill_assumptions(wb, req: dict) -> None:
    """Dispatch to the correct fill functions based on asset_type."""
    asset_type = req.get("asset_type", "NNN").upper()
    if asset_type == "MOB":
        fill_mob_assumptions(wb, req)
        fill_mob_rent_schedule(wb, req)
        fill_mob_lease_abstract(wb, req)
        fill_mob_rent_roll(wb, req)
        fill_mob_exec_summary(wb, req)
        _hide_unused_mob_tenants(wb, len(req.get("tenants") or []))
    else:
        fill_nnn_assumptions(wb, req)
        fill_nnn_rent_schedule(wb, req)
        fill_nnn_lease_abstract(wb, req)
        fill_nnn_exec_summary(wb, req)
    fill_cover(wb, req)
    fill_exec_subtitle(wb, req)
    fill_real_estate(wb, req)

