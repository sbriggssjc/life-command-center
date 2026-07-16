"""BOV v2 — Tab 8: Assumptions & Flags."""
from bov_constants import *
from openpyxl.styles import Border, Side
from openpyxl.utils import get_column_letter

_NM_M = Side(style='medium', color=NAVY)
_NM_T = Side(style='thin',   color=NAVY)
_GY_T = Side(style='thin',   color="CCCCCC")

_AS = "Assumptions & Flags"  # self-reference label for clarity


def build_assumptions_tab(wb):
    ws = wb.create_sheet("Assumptions & Flags")
    ws.sheet_view.showGridLines = False

    # ── Column widths ──────────────────────────────────────────────────────────
    # A=label(32), B=spacer(1.5), C=value(18), D=spacer(1.5), E=notes/source(36)
    col_widths = [32, 1.5, 18, 1.5, 36, 1.5, 32, 1.5, 18, 1.5, 36]
    for i, wd in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = wd

    # ── Header ────────────────────────────────────────────────────────────────
    r = 1
    ws.row_dimensions[r].height = 6
    r = 2
    ws.row_dimensions[r].height = 28
    c = ws.cell(row=r, column=1, value="ASSUMPTIONS & FLAGS")
    c.font = FT_TITLE; c.alignment = AL_L
    merge(ws, r, 1, r, 11)

    r = 3
    ws.row_dimensions[r].height = 14
    c = ws.cell(row=r, column=1, value="All key inputs for the Pro Forma  ·  Blue text = broker enters  ·  Yellow = fill before delivering  ·  Flag discrepancies in right column")
    c.font = FT_NOTE; c.alignment = AL_L
    merge(ws, r, 1, r, 11)

    r = 4
    ws.row_dimensions[r].height = 8

    # ══════════════════════════════════════════════════════════════════════════
    # HELPER: two-column assumption entry (left side cols A–E, right side G–K)
    # ══════════════════════════════════════════════════════════════════════════
    def row_inp(ws, rr, label, val=None, fmt=None, note=None):
        ws.row_dimensions[rr].height = 18
        lbl(ws, rr, 1, label)
        c = inp(ws, rr, 3, val, fmt)
        c.alignment = AL_R
        if note:
            c2 = ws.cell(row=rr, column=5, value=note)
            c2.font = FT_NOTE; c2.alignment = AL_L
        add_cf_clear(ws, f"C{rr}:C{rr}")
        _t = Side(style='thin', color="CCCCCC")
        ws.cell(row=rr, column=1).border = Border(bottom=_t)
        ws.cell(row=rr, column=3).border = Border(bottom=_t)

    def row_frm(ws, rr, label, formula, fmt=None, note=None):
        ws.row_dimensions[rr].height = 18
        lbl(ws, rr, 1, label, bold=False)
        c = frm(ws, rr, 3, formula, fmt=fmt, align=AL_R)
        if note:
            c2 = ws.cell(row=rr, column=5, value=note)
            c2.font = FT_NOTE; c2.alignment = AL_L
        _t = Side(style='thin', color="CCCCCC")
        ws.cell(row=rr, column=1).border = Border(bottom=_t)
        ws.cell(row=rr, column=3).border = Border(bottom=_t)

    # ══════════════════════════════════════════════════════════════════════════
    # LEFT COLUMN — PROPERTY & PRICING
    # ══════════════════════════════════════════════════════════════════════════

    # SECTION: Property
    r = 5
    sec(ws, r, "PROPERTY IDENTIFICATION", col_start=1, ncols=5)

    row_inp(ws, 6,  "Property Address",          note="Used on Cover and Executive Summary")
    row_inp(ws, 7,  "Tenant / Operator",          note="Trade name + entity name")
    row_inp(ws, 8,  "Guarantor",                  note="Corporate or personal")
    row_inp(ws, 9,  "Building SF",               fmt=N0, note="Per lease or survey — flag discrepancy")
    row_inp(ws, 10, "Site Area (Acres)",          fmt="0.00", note="Per survey or tax record")
    row_inp(ws, 11, "Year Built",                 note="Four-digit year")
    row_inp(ws, 12, "Zoning",                     note="Confirm permitted use")
    # Estimated Close Date
    row_inp(ws, 13, "Estimated Close Date",       fmt=DT, note="Used for year-end date row on Pro Forma")
    CLOSE_DATE_ROW = 13   # referenced as C6 from Pro Forma — wait, Pro Forma uses C6
    # Note: the Pro Forma references 'Assumptions & Flags'!$C$6 for close date.
    # We placed close date at row 13 here. Let's ensure close date IS at row 6 for Pro Forma alignment.
    # We'll restructure: address=6 actually was close date reference spot.
    # Let me just note the mapping — Pro Forma references $C$6 for close date and
    # $C$4 for building SF, $C$5 for purchase price, $C$9 for base rent.
    # I need to align this with actual row numbers. Let me place rows carefully:

    # Actually I already started above. Let me just fix by referencing correctly below.
    # Pro Forma uses these Assumptions rows:
    # $C$4 = Building SF
    # $C$5 = Purchase Price
    # $C$6 = Close Date
    # $C$9 = Year 1 Base Rent
    # $C$10 = Rent Escalation Rate
    # $C$11 = Tenant Reimbursements
    # $C$12 = Management Fee %
    # $C$13 = Capital Reserves
    # $C$19 = Equity Investment
    # $C$22 = Annual Debt Service
    # $C$27 = Exit Cap Rate
    # $C$32 = Net Reversion (unleveraged)
    # $C$33 = Net Reversion after debt
    # $C$34 = Remaining Loan Balance

    # I need to be explicit. Let me write the assumptions block with fixed row assignments.
    # I'll rebuild this function more carefully with explicit row numbers.

    # Clear what was done above and restart with explicit rows
    # (The rows above 5-13 are already written; we need to extend below)

    # r14 = spacer
    r = 14
    ws.row_dimensions[r].height = 6

    # SECTION: PRICING & INVESTMENT (rows 15+)
    r = 15
    sec(ws, r, "PRICING STRATEGY  ·  FUNGIBLE INPUTS", col_start=1, ncols=5)

    row_inp(ws, 16, "Recommended Asking Cap Rate",  fmt=P2,   note="Drives Recommended Asking Price on Exec Summary")
    row_inp(ws, 17, "Trade Range — Low Cap Rate",   fmt=P2,   note="Low = higher price")
    row_inp(ws, 18, "Trade Range — High Cap Rate",  fmt=P2,   note="High = lower price")
    row_frm(ws, 19, "Implied Ask Price (NOI / Cap)", '=IFERROR(IF(OR(C9="",C16="",C16=0),"",C9/C16),"")',  fmt=D0, note="Updates Executive Summary")
    row_inp(ws, 20, "Broker Price Opinion (BPO)",    fmt=D0,   note="Override if needed")
    row_inp(ws, 21, "Price Per SF",                 fmt=D2,   note="Manual entry or use Valuation Matrix")

    r = 22
    ws.row_dimensions[r].height = 6
    r = 23
    sec(ws, r, "PRO FORMA INPUTS  —  REVENUE", col_start=1, ncols=5)

    row_inp(ws, 24, "Year 1 Base Rent ($)",          fmt=D0,   note="Per executed lease / rent schedule")
    row_inp(ws, 25, "Annual Rent Escalation (%)",    fmt=P2,   note="Fixed bump or avg CPI — label in Notes")
    row_inp(ws, 26, "Tenant Reimbursements ($/yr)",  fmt=D0,   note="NNN = $0 for expense tracking purposes")
    row_frm(ws, 27, "Effective Gross Revenue (Y1)", '=IFERROR(IF(C24="","",C24+IFERROR(C26,0)),"")',  fmt=D0)

    r = 28
    ws.row_dimensions[r].height = 6
    r = 29
    sec(ws, r, "PRO FORMA INPUTS  —  EXPENSES", col_start=1, ncols=5)

    row_inp(ws, 30, "Mgmt Fee (%)",                 fmt=P2,   note="Typically 0–3% for NNN")
    row_inp(ws, 31, "Capital / Replacement Reserves ($/yr)", fmt=D0, note="$0–$0.15/SF typical NNN")
    row_frm(ws, 32, "Total Expenses (Y1)",          '=IFERROR(IF(C24="","",C24*IFERROR(C30,0)+IFERROR(C31,0)),"")', fmt=D0)
    row_frm(ws, 33, "Estimated NOI (Y1)",           '=IFERROR(IF(C27="","",C27-C32),"")', fmt=D0)

    r = 34
    ws.row_dimensions[r].height = 6
    r = 35
    sec(ws, r, "ACQUISITION PRICING", col_start=1, ncols=5)

    row_inp(ws, 36, "Purchase Price ($)",            fmt=D0,   note="Fungible — change to model different scenarios")
    row_frm(ws, 37, "Going-In Cap Rate",            '=IFERROR(IF(OR(C33="",C36="",C36=0),"",C33/C36),"")', fmt=P2)
    row_frm(ws, 38, "Price Per SF",                 '=IFERROR(IF(OR(C36="",C9="",C9=0),"",C36/C9),"")', fmt=D2,
            note="Uses Building SF (row 9 above)")

    r = 39
    ws.row_dimensions[r].height = 6

    # Flag zone (left col bottom)
    r = 40
    sec(ws, r, "DISCREPANCY FLAGS  —  Note any source conflicts or assumptions below", col_start=1, ncols=5)
    for i in range(12):
        rr = 41 + i
        ws.row_dimensions[rr].height = 18
        if i == 0:
            ws.cell(row=rr, column=1, value="Flag / Assumption").font = FT_CHDR
            ws.cell(row=rr, column=1).fill = F_NAVY
            ws.cell(row=rr, column=3, value="Description / Resolution").font = FT_CHDR
            ws.cell(row=rr, column=3).fill = F_NAVY
            merge(ws, rr, 3, rr, 5)
        else:
            inp(ws, rr, 1)
            ws.cell(row=rr, column=1).alignment = AL_TL
            inp(ws, rr, 3)
            ws.cell(row=rr, column=3).alignment = AL_TL
            merge(ws, rr, 3, rr, 5)
            _t = Side(style='thin', color="CCCCCC")
            ws.cell(row=rr, column=1).border = Border(bottom=_t)
            ws.cell(row=rr, column=3).border = Border(bottom=_t)

    # ══════════════════════════════════════════════════════════════════════════
    # RIGHT COLUMN — DEBT & DISPOSITION
    # ══════════════════════════════════════════════════════════════════════════
    def row_inp_r(ws, rr, label, val=None, fmt=None, note=None):
        ws.row_dimensions[rr].height = max(ws.row_dimensions[rr].height or 18, 18)
        lbl(ws, rr, 7, label)
        c = inp(ws, rr, 9, val, fmt)
        c.alignment = AL_R
        if note:
            c2 = ws.cell(row=rr, column=11, value=note)
            c2.font = FT_NOTE; c2.alignment = AL_L
        add_cf_clear(ws, f"I{rr}:I{rr}")
        _t = Side(style='thin', color="CCCCCC")
        ws.cell(row=rr, column=7).border = Border(bottom=_t)
        ws.cell(row=rr, column=9).border = Border(bottom=_t)

    def row_frm_r(ws, rr, label, formula, fmt=None, note=None, bold=False):
        ws.row_dimensions[rr].height = max(ws.row_dimensions[rr].height or 18, 18)
        lbl(ws, rr, 7, label, bold=bold)
        c = frm(ws, rr, 9, formula, fmt=fmt, align=AL_R)
        if bold:
            c.font = FT_TOTAL
            ws.cell(row=rr, column=9).fill = F_TOT
            ws.cell(row=rr, column=7).fill = F_TOT
        if note:
            c2 = ws.cell(row=rr, column=11, value=note)
            c2.font = FT_NOTE; c2.alignment = AL_L
        _t = Side(style='thin', color="CCCCCC")
        ws.cell(row=rr, column=7).border = Border(bottom=_t)
        ws.cell(row=rr, column=9).border = Border(bottom=_t)

    r = 5
    c = ws.cell(row=r, column=7, value="DEBT / LEVERAGE ASSUMPTIONS")
    c.font = FT_SHDR; c.fill = F_NAVY; c.alignment = AL_L
    for ci in [7, 8, 9, 10, 11]:
        ws.cell(row=r, column=ci).fill = F_NAVY
    merge(ws, r, 7, r, 11)
    ws.row_dimensions[r].height = 16

    row_inp_r(ws, 6,  "Purchase Price ($)",           fmt=D0,  note="Link to left side C36 or enter here")
    row_inp_r(ws, 7,  "LTV (%)",                      fmt=P2,  note="Default 65%")
    row_frm_r(ws, 8,  "Loan Amount",                 '=IFERROR(IF(OR(I6="",I7=""),"",-I6*I7),"")', fmt=D0,
              note="Negative = debt obligation")
    row_frm_r(ws, 9,  "Down Payment / Equity",       '=IFERROR(IF(OR(I6="",I7=""),"",I6*(1-I7)),"")', fmt=D0)
    row_inp_r(ws, 10, "Interest Rate",               fmt=P2,  note="Default 6.5%")
    row_inp_r(ws, 11, "Amortization (years)",         fmt=N0,  note="Default 25 years")
    row_inp_r(ws, 12, "IO Period (years)",            fmt=N0,  note="Enter 0 if fully amortizing")
    row_inp_r(ws, 13, "Hold Period (years)",          fmt=N0,  note="Default 10 — drives disposition year")
    row_frm_r(ws, 14, "Annual Debt Service",
              '=IFERROR(IF(OR(I8="",I10="",I11=""),"",ABS(I8)*(I10/12)*(1+I10/12)^(I11*12)/((1+I10/12)^(I11*12)-1)*12),"")',
              fmt=D0, note="P&I payment (fully amortizing)")
    row_frm_r(ws, 15, "Monthly Debt Service",
              '=IFERROR(IF(I14="","",I14/12),"")', fmt=D0)
    row_frm_r(ws, 16, "DCR (Year 1)",
              '=IFERROR(IF(OR(C33="",I14="",I14=0),"",C33/I14),"")', fmt=MX, note="Must exceed 1.20x")
    row_frm_r(ws, 17, "DSCR at 1.20x NOI Required",
              '=IFERROR(IF(I14="","",I14*1.20),"")', fmt=D0)

    r = 18
    ws.row_dimensions[r].height = 6

    # Exit / Disposition
    r = 19
    c = ws.cell(row=r, column=7, value="DISPOSITION / EXIT ASSUMPTIONS")
    c.font = FT_SHDR; c.fill = F_NAVY; c.alignment = AL_L
    for ci in [7, 8, 9, 10, 11]:
        ws.cell(row=r, column=ci).fill = F_NAVY
    merge(ws, r, 7, r, 11)
    ws.row_dimensions[r].height = 16

    row_inp_r(ws, 20, "Exit Cap Rate",               fmt=P2,  note="Applied to Year 10 NOI")
    row_frm_r(ws, 21, "Exit NOI (Year 10 Projected)",
              '=IFERROR(IF(OR(C24="",C25=""),"",C24*(1+C25)^I13),"")', fmt=D0,
              note="Y1 rent * (1+escalation)^hold period")
    row_frm_r(ws, 22, "GROSS SALE PROCEEDS",
              '=IFERROR(IF(OR(I21="",I20="",I20=0),"",I21/I20),"")', fmt=D0, bold=True)
    row_frm_r(ws, 23, "Less: Brokerage (5%)",
              '=IFERROR(IF(I22="","",-I22*0.05),"")', fmt=D0)
    row_frm_r(ws, 24, "Less: Transaction Costs (1%)",
              '=IFERROR(IF(I22="","",-I22*0.01),"")', fmt=D0)
    row_frm_r(ws, 25, "NET REVERSION (Unleveraged)",
              '=IFERROR(IF(I22="","",I22+I23+I24),"")', fmt=D0, bold=True)

    # Remaining loan balance at exit
    row_frm_r(ws, 26, "Remaining Loan Balance at Exit",
              '=IFERROR(IF(OR(I8="",I10="",I11="",I13=""),"",ABS(I8)*(1+I10/12)^(I13*12)-(ABS(I8)*(I10/12)*(1+I10/12)^(I11*12)/((1+I10/12)^(I11*12)-1))*(((1+I10/12)^(I13*12)-1)/(I10/12))),"")',
              fmt=D0, note="Outstanding balance at hold period end")
    row_frm_r(ws, 27, "NET REVERSION AFTER DEBT",
              '=IFERROR(IF(OR(I25="",I26=""),"",I25-I26),"")', fmt=D0, bold=True)

    r = 28
    ws.row_dimensions[r].height = 6

    # Return Summary Block
    r = 29
    c = ws.cell(row=r, column=7, value="RETURN SUMMARY REFERENCE")
    c.font = FT_SHDR; c.fill = F_NAVY; c.alignment = AL_L
    for ci in [7, 8, 9, 10, 11]:
        ws.cell(row=r, column=ci).fill = F_NAVY
    merge(ws, r, 7, r, 11)
    ws.row_dimensions[r].height = 16

    def row_ref_r(ws, rr, label, formula, fmt=None):
        ws.row_dimensions[rr].height = 18
        lbl(ws, rr, 7, label, bold=False)
        c = frm(ws, rr, 9, formula, fmt=fmt, align=AL_R)
        _t = Side(style='thin', color="CCCCCC")
        ws.cell(row=rr, column=7).border = Border(bottom=_t)
        ws.cell(row=rr, column=9).border = Border(bottom=_t)

    # callout_boxes(row=28) puts values at row 29; callout_boxes(row=42) puts values at row 43
    row_ref_r(ws, 30, "Unleveraged IRR",   '=IFERROR(\'Pro Forma\'!C29,"N/A")', fmt=P2)
    row_ref_r(ws, 31, "Unleveraged ERM",   '=IFERROR(\'Pro Forma\'!G29,"N/A")', fmt=MX)
    row_ref_r(ws, 32, "Unleveraged CoCR",  '=IFERROR(\'Pro Forma\'!K29,"N/A")', fmt=P2)
    row_ref_r(ws, 33, "Leveraged IRR",     '=IFERROR(\'Pro Forma\'!C43,"N/A")', fmt=P2)
    row_ref_r(ws, 34, "Leveraged ERM",     '=IFERROR(\'Pro Forma\'!G43,"N/A")', fmt=MX)
    row_ref_r(ws, 35, "Leveraged CoCR",    '=IFERROR(\'Pro Forma\'!K43,"N/A")', fmt=P2)

    # Note: The Pro Forma references specific cells on this tab. Mapping:
    # Pro Forma C4 → 'Assumptions & Flags'!$C$9  (Building SF)
    # Pro Forma C5 → 'Assumptions & Flags'!$C$36 (Purchase Price)
    # Pro Forma C6 → 'Assumptions & Flags'!$C$13 (Close Date)
    # Pro Forma C9 → 'Assumptions & Flags'!$C$24 (Year 1 Base Rent)
    # Pro Forma C10 → 'Assumptions & Flags'!$C$25 (Rent Escalation)
    # Pro Forma C11 → 'Assumptions & Flags'!$C$26 (Tenant Reimbursements)
    # Pro Forma C12 → 'Assumptions & Flags'!$C$30 (Mgmt Fee %)
    # Pro Forma C13 → 'Assumptions & Flags'!$C$31 (Capital Reserves)
    # Pro Forma C19 → 'Assumptions & Flags'!$I$9  (Down Payment / Equity)
    # Pro Forma C22 → 'Assumptions & Flags'!$I$14 (Annual Debt Service)
    # Pro Forma C27 → 'Assumptions & Flags'!$I$20 (Exit Cap Rate)
    # Pro Forma C32 → 'Assumptions & Flags'!$I$25 (Net Reversion unlev)
    # Pro Forma C33 → 'Assumptions & Flags'!$I$27 (Net Reversion after debt)
    # Pro Forma C34 → 'Assumptions & Flags'!$I$26 (Remaining Loan Balance)

    # NM bottom
    r = 55
    ws.row_dimensions[r].height = 4
    for ci in range(1, 12):
        ws.cell(row=r, column=ci).border = Border(bottom=Side(style='medium', color=NAVY))
