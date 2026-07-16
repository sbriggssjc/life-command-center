"""BOV v2 — Tab 5: Rent Schedule (Leg 2 continued)."""
from bov_constants import *
from openpyxl.styles import Border, Side, PatternFill, Font
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter


# Current-year highlight: pale blue (softer than input yellow)
CURR_FILL = PatternFill("solid", fgColor="D6E4F5")   # TOTBG — NM light blue
_NM_M     = Side(style='medium', color=NAVY)
_NM_T     = Side(style='thin',   color=NAVY)

MAX_ROWS = 30   # max lease period rows in the schedule


def build_rent_schedule_tab(wb):
    ws = wb.create_sheet("Rent Schedule")
    ws.sheet_view.showGridLines = False

    # ── Column widths ──────────────────────────────────────────────────────────
    # A(1)=margin  B(2)=YR#(narrow)  C(3)=START DATE  D(4)=END DATE
    # E(5)=LEASE PERIOD  F(6)=ANNUAL RENT  G(7)=MONTHLY RENT
    # H(8)=RENT/SF  I(9)=ESCALATION  J(10)=NOTES
    col_widths = [2, 6, 13, 13, 22, 14, 13, 10, 12, 34]
    for i, wd in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = wd

    # ── Header ────────────────────────────────────────────────────────────────
    r = 1
    ws.row_dimensions[r].height = 6
    r = 2
    ws.row_dimensions[r].height = 28
    c = ws.cell(row=r, column=2, value="RENT SCHEDULE")
    c.font = FT_TITLE; c.alignment = AL_L
    merge(ws, r, 2, r, 10)

    r = 3
    ws.row_dimensions[r].height = 14
    c = ws.cell(row=r, column=2,
                value="Leg 2 of 3  ·  Contractual rent by lease period  ·  Current lease year highlighted  ·  No projections — use Notes for FMV / CPI / other variable increases")
    c.font = FT_NOTE; c.alignment = AL_L
    merge(ws, r, 2, r, 10)

    r = 4
    ws.row_dimensions[r].height = 8

    # ── Property / Tenant header block ────────────────────────────────────────
    r = 5
    sec(ws, r, "PROPERTY & LEASE IDENTIFICATION", col_start=2, ncols=9)

    ident_rows = [
        ("Property Address",  None,  6),
        ("Tenant",            None,  7),
        ("Guarantor",         None,  8),
        ("Lease Commencement",None,  9),
        ("Lease Expiration",  None, 10),
        ("Remaining Term",    None, 11),
        ("Building SF",       None, 12),
    ]
    for label, _, rr in ident_rows:
        ws.row_dimensions[rr].height = 18
        lbl(ws, rr, 2, label)
        merge(ws, rr, 2, rr, 4)
        inp(ws, rr, 5)
        merge(ws, rr, 5, rr, 7)
        ws.cell(row=rr, column=5).alignment = AL_L
        add_cf_clear(ws, f"E{rr}:E{rr}")

    r = 13
    ws.row_dimensions[r].height = 8

    # ── Rent Schedule ─────────────────────────────────────────────────────────
    r = 14
    sec(ws, r, "RENT SCHEDULE BY LEASE PERIOD", col_start=2, ncols=9)

    r = 15
    ws.row_dimensions[r].height = 22
    hdrs = ["YR", "START DATE", "END DATE", "LEASE PERIOD", "ANNUAL RENT", "MONTHLY RENT", "RENT / SF", "ESCALATION", "NOTES"]
    for ci, hdr in enumerate(hdrs, 2):
        c = ws.cell(row=r, column=ci, value=hdr)
        c.font = FT_CHDR; c.fill = F_NAVY; c.alignment = AL_C
        c.border = Border(bottom=_NM_M)

    DATA_START = 16

    for i in range(MAX_ROWS):
        rr = DATA_START + i
        ws.row_dimensions[rr].height = 18

        # YR# (formula: =row()-15 relative, but let brokers set it)
        yr_c = ws.cell(row=rr, column=2, value=i + 1)
        yr_c.font = FT_DATA; yr_c.alignment = AL_C

        # Start Date
        inp(ws, rr, 3, fmt=DT)
        ws.cell(row=rr, column=3).alignment = AL_C

        # End Date
        inp(ws, rr, 4, fmt=DT)
        ws.cell(row=rr, column=4).alignment = AL_C

        # Lease Period — input text (Initial Term, Option 1, etc.)
        inp(ws, rr, 5)
        ws.cell(row=rr, column=5).alignment = AL_L

        # Annual Rent
        inp(ws, rr, 6, fmt=D0)
        ws.cell(row=rr, column=6).alignment = AL_R

        # Monthly Rent (formula = Annual / 12)
        frm(ws, rr, 7,
            f'=IFERROR(IF(F{rr}="","",F{rr}/12),"")',
            fmt=D0, align=AL_R)

        # Rent / SF (formula = Annual / Building SF)
        # Building SF is in E12 (ident block row 12, col 5)
        frm(ws, rr, 8,
            f'=IFERROR(IF(OR(F{rr}="",E12="",E12=0),"",F{rr}/E12),"")',
            fmt=D2, align=AL_R)

        # Escalation %
        inp(ws, rr, 9, fmt=P2)
        ws.cell(row=rr, column=9).alignment = AL_C

        # Notes
        inp(ws, rr, 10)
        ws.cell(row=rr, column=10).alignment = AL_TL

        # NM thin bottom border on each row
        for ci in range(2, 11):
            existing = ws.cell(row=rr, column=ci).border
            ws.cell(row=rr, column=ci).border = Border(bottom=_NM_T)

    DATA_END = DATA_START + MAX_ROWS - 1

    # ── CF: clear yellow when cells are filled ─────────────────────────────
    for col in [3, 4, 5, 6, 9, 10]:
        cl = get_column_letter(col)
        add_cf_clear(ws, f"{cl}{DATA_START}:{cl}{DATA_END}")

    # ── CF: highlight current lease year row (where Start ≤ TODAY ≤ End) ─
    # Apply NM light-blue fill to entire row when today falls in period
    _hl_fill = PatternFill(fill_type="solid", fgColor=TOTBG)
    _hl_font = Font(name="Calibri", size=10, bold=True, color=TEXT)
    for i in range(MAX_ROWS):
        rr = DATA_START + i
        # Formula checks: C{rr} (start) <= TODAY() and D{rr} (end) >= TODAY()
        # Anchored to C column (start date col)
        formula = f'AND(NOT(ISBLANK($C{rr})),NOT(ISBLANK($D{rr})),$C{rr}<=TODAY(),$D{rr}>=TODAY())'
        row_range = f"B{rr}:J{rr}"
        ws.conditional_formatting.add(
            row_range,
            FormulaRule(formula=[formula], fill=_hl_fill, font=_hl_font)
        )

    # Medium bottom border after last row
    rr = DATA_END + 1
    ws.row_dimensions[rr].height = 4
    for ci in range(2, 11):
        ws.cell(row=rr, column=ci).border = Border(top=_NM_M)

    # ── Totals row ────────────────────────────────────────────────────────────
    tot_r = DATA_END + 2
    ws.row_dimensions[tot_r].height = 18
    c = ws.cell(row=tot_r, column=2, value="TOTALS / AVERAGES")
    c.font = FT_TOTAL; c.fill = F_TOT; c.alignment = AL_L
    merge(ws, tot_r, 2, tot_r, 5)
    ws.cell(row=tot_r, column=2).border = Border(top=_NM_M, bottom=_NM_M)

    # Sum annual rent
    frm(ws, tot_r, 6,
        f'=IFERROR(SUMIF(F{DATA_START}:F{DATA_END},"<>",F{DATA_START}:F{DATA_END}),"")',
        fmt=D0, align=AL_R)
    ws.cell(row=tot_r, column=6).font  = FT_TOTAL
    ws.cell(row=tot_r, column=6).fill  = F_TOT
    ws.cell(row=tot_r, column=6).border = Border(top=_NM_M, bottom=_NM_M)

    for ci in [7, 8, 9, 10]:
        ws.cell(row=tot_r, column=ci).fill = F_TOT
        ws.cell(row=tot_r, column=ci).border = Border(top=_NM_M, bottom=_NM_M)

    # ── Notes & legend ────────────────────────────────────────────────────────
    note_r = tot_r + 2
    ws.row_dimensions[note_r].height = 14
    notes = [
        "LEASE PERIOD column:  Initial Term · Extension Term · Option 1, Option 2, etc.",
        "For rent tied to FMV, CPI, or other variable methods, leave Annual Rent blank and describe in Notes column.",
        "Current lease year row auto-highlights in blue based on today's date (Start Date ≤ Today ≤ End Date).",
        "Building SF (row 12) is used for Rent/SF calculations — update when SF is confirmed.",
    ]
    for idx, note in enumerate(notes):
        rr = note_r + idx
        ws.row_dimensions[rr].height = 14
        c = ws.cell(row=rr, column=2, value=f"▪  {note}")
        c.font = FT_NOTE; c.alignment = AL_L
        merge(ws, rr, 2, rr, 10)

    # NM bottom accent
    bot_r = note_r + len(notes) + 1
    ws.row_dimensions[bot_r].height = 4
    for ci in range(2, 11):
        ws.cell(row=bot_r, column=ci).border = Border(bottom=_NM_M)
