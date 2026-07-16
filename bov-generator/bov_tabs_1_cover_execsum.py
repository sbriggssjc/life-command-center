"""Cover and Executive Summary tabs for BOV Master Sheet v2."""
from bov_constants import *

ALL_TABS = [
    ("Cover",               "Property identification and workbook guide"),
    ("Executive Summary",   "Investment snapshot · pricing strategy · engagement proposal"),
    ("Real Estate",         "Leg 1 — Physical property, location, and market fundamentals"),
    ("Lease Abstract",      "Leg 2 — Executed lease terms, obligations, and economics"),
    ("Rent Schedule",       "Leg 2 — Year-by-year rent schedule by lease period"),
    ("Credit",              "Leg 3 — Tenant / guarantor credit, operations, and unit economics"),
    ("Pro Forma",           "10-year investment model: Revenue → NOI → Returns"),
    ("Assumptions & Flags", "All key inputs, fungible pricing assumptions, discrepancy flags"),
    ("Sensitivity Analysis","Returns across different asking cap rate / pricing scenarios"),
    ("Amortization",        "Loan amortization schedule — auto-updates with assumptions"),
]

_A = "'Assumptions & Flags'"   # shorthand for cross-sheet refs


def build_cover(wb):
    ws = wb.create_sheet("Cover")
    ws.sheet_view.showGridLines = False

    # Columns: A=margin(2), B=label/content(28), C=content/link(44), D=margin(2)
    for col, wd in [(1, 2), (2, 28), (3, 44), (4, 2)]:
        w(ws, col, wd)

    # ── Top Navy bar ──────────────────────────────────────────────────────────
    h(ws, 1, 10)
    for col in range(1, 5): ws.cell(row=1, column=col).fill = F_NAVY
    merge(ws, 1, 1, 1, 4)

    # ── Brand row ─────────────────────────────────────────────────────────────
    h(ws, 2, 24)
    c = ws.cell(row=2, column=2, value="NorthMarq  ·  Investment Sales")
    c.font = ft("Calibri", 11, b=True, c=NAVY); c.alignment = AL_L
    c2 = ws.cell(row=2, column=3, value="Team Briggs")
    c2.font = ft("Calibri", 11, b=True, c=MUTED); c2.alignment = AL_R

    # Pale accent band
    h(ws, 3, 8); h(ws, 4, 8); h(ws, 5, 8)
    for r in range(3, 6):
        for col in range(1, 5): ws.cell(row=r, column=col).fill = F_PALE
    merge(ws, 3, 1, 5, 4)

    # ── Main title ─────────────────────────────────────────────────────────────
    h(ws, 6, 52)
    c = ws.cell(row=6, column=2, value="Broker Opinion of Value")
    c.font = FT_COVER; c.alignment = Alignment(horizontal='left', vertical='center')
    merge(ws, 6, 2, 6, 3)

    h(ws, 7, 22)
    c = ws.cell(row=7, column=2,
                value="Single-Tenant Net Lease  ·  Investment Analysis Workbook")
    c.font = FT_BRAND; c.alignment = AL_L
    merge(ws, 7, 2, 7, 3)
    h(ws, 8, 14)

    # ── Property fields ────────────────────────────────────────────────────────
    fields = [
        (9,  "PROPERTY / TENANT",  "[Tenant Name  —  City, State]"),
        (11, "PROPERTY ADDRESS",   None),
        (13, "CLIENT / SELLER",    None),
        (15, "ASSET TYPE",         "Single-Tenant NNN"),
        (17, "ANALYSIS DATE",      None),
        (19, "PREPARED BY",        "Scott Briggs  ·  SVP, Commercial Investment Sales  ·  NorthMarq"),
        (21, "ANALYST",            "Sarah Martin"),
    ]
    for row, label, value in fields:
        c = ws.cell(row=row, column=2, value=label)
        c.font = ft("Calibri", 10, b=True, c=MUTED); c.alignment = AL_L
        h(ws, row, 14)
        h(ws, row + 1, 22)
        if value:
            v = ws.cell(row=row + 1, column=2, value=value)
            v.font = ft("Calibri", 11, b=True, c=TEXT); v.alignment = AL_L
        else:
            v = inp(ws, row + 1, 2, fmt=DT if "DATE" in label else None)
            v.alignment = AL_L
        merge(ws, row + 1, 2, row + 1, 3)
    h(ws, 23, 14)

    # ── Three legs of the stool + tab guide ───────────────────────────────────
    sec(ws, 24, "THREE LEGS OF THE STOOL  —  WORKBOOK STRUCTURE", col_start=2, ncols=2)

    legs = [
        ("LEG 1  —  REAL ESTATE FUNDAMENTALS",
         "Underlying value of the physical property; highest & best use; "
         "contractual vs. prevailing market rents; long-term value generator.",
         [("Real Estate", None)]),
        ("LEG 2  —  ECONOMICS OF THE LEASE",
         "Contractual terms, obligations, and protections imposed on or granted to "
         "the Landlord. Better terms → greater value.",
         [("Lease Abstract", None), ("Rent Schedule", None)]),
        ("LEG 3  —  TENANT & GUARANTOR CREDIT",
         "Corporate credit quality, unit-level performance (sales / EBITDA), "
         "and importance of this location to the enterprise.",
         [("Credit", None)]),
    ]

    r = 25
    for leg_label, leg_desc, tabs in legs:
        h(ws, r, 16)
        c = ws.cell(row=r, column=2, value=leg_label)
        c.font = ft("Calibri", 10, b=True, c=NAVY); c.fill = F_PALE; c.alignment = AL_L
        ws.cell(row=r, column=3).fill = F_PALE
        merge(ws, r, 2, r, 3)
        r += 1

        h(ws, r, 36)   # taller for wrapped description text
        c = ws.cell(row=r, column=2, value=leg_desc)
        c.font = FT_NOTE; c.fill = F_WHITE; c.alignment = AL_TL
        merge(ws, r, 2, r, 3)
        r += 1

        for sheet_name, display in tabs:
            h(ws, r, 18)
            ws.cell(row=r, column=2, value="→").font = ft("Calibri", 10, c=MUTED)
            ws.cell(row=r, column=2).alignment = AL_C
            hyperlink_cell(ws, r, 3, sheet_name)
            r += 1
        h(ws, r, 8); r += 1

    # ── Synthesis tabs ────────────────────────────────────────────────────────
    h(ws, r, 16)
    c = ws.cell(row=r, column=2, value="SYNTHESIS  —  ANALYSIS & MODEL")
    c.font = ft("Calibri", 10, b=True, c=NAVY); c.fill = F_PALE; c.alignment = AL_L
    ws.cell(row=r, column=3).fill = F_PALE
    merge(ws, r, 2, r, 3)
    r += 1

    for sheet_name, desc in [
        ("Executive Summary",    "Investment snapshot + pricing strategy + engagement proposal"),
        ("Pro Forma",            "10-year investment model"),
        ("Assumptions & Flags",  "All key inputs and discrepancy flags"),
        ("Sensitivity Analysis", "Returns across different pricing scenarios"),
        ("Amortization",         "Loan amortization schedule"),
    ]:
        h(ws, r, 18)
        ws.cell(row=r, column=2, value="→").font = ft("Calibri", 10, c=MUTED)
        ws.cell(row=r, column=2).alignment = AL_C
        hyperlink_cell(ws, r, 3, sheet_name)
        r += 1

    h(ws, r, 10); r += 1

    # ── Color legend ──────────────────────────────────────────────────────────
    sec(ws, r, "COLOR CONVENTIONS", col_start=2, ncols=2); r += 1
    legend = [
        # (swatch_fill, swatch_text_color, name, description)
        (PALE,  NAVY,   "NM Blue fill",  "Contracted / executed lease periods"),
        (GOLD,  TEXT,   "Gold fill",     "Renewal / option / projection periods"),
        (YELL,  TEXT,   "Yellow fill",   "Input cells — clears automatically when filled"),
        (WHITE, INPC,   "Blue text",     "User-entry fields"),
        (WHITE, MUTED,  "Muted italic",  "Calculated / formula cells — do not edit"),
    ]
    for bg, fg, name, desc in legend:
        h(ws, r, 18)
        s = ws.cell(row=r, column=2, value=f"  {name}")
        s.fill = PatternFill("solid", fgColor=bg)
        s.font = ft("Calibri", 10, b=True, c=fg); s.alignment = AL_L
        from openpyxl.styles import Border, Side
        s.border = Border(
            left=Side(style='thin', color="CCCCCC"),
            right=Side(style='thin', color="CCCCCC"),
            top=Side(style='thin', color="CCCCCC"),
            bottom=Side(style='thin', color="CCCCCC")
        )
        d = ws.cell(row=r, column=3, value=desc)
        d.font = FT_DATA; d.alignment = AL_L
        r += 1

    # Bottom Navy bar
    h(ws, r, 10)
    for col in range(1, 5): ws.cell(row=r, column=col).fill = F_NAVY
    merge(ws, r, 1, r, 4)

    # CF clear on input cells (property fields — every other row starting at row 10)
    for row in [10, 12, 14, 18]:
        add_cf_clear(ws, f"B{row}:B{row}")


def build_exec_summary(wb):
    ws = wb.create_sheet("Executive Summary")
    ws.sheet_view.showGridLines = False

    # Columns: A=margin(2), B=label(26), C=cap rate(16), D=price(18), E=price/sf(14), F=notes(22)
    for col, wd in [(1, 2), (2, 26), (3, 16), (4, 18), (5, 14), (6, 22)]:
        w(ws, col, wd)

    h(ws, 1, 34)
    c = ws.cell(row=1, column=2, value="Executive Summary")
    c.font = FT_TITLE; c.alignment = AL_L; merge(ws, 1, 2, 1, 6)

    h(ws, 2, 18)
    c = ws.cell(row=2, column=2,
                value="[Tenant Name  —  City, State]  ·  Investment Snapshot  ·  [Analysis Date]")
    c.font = FT_NOTE; c.alignment = AL_L; merge(ws, 2, 2, 2, 6)
    h(ws, 3, 8)

    # ── INVESTMENT SNAPSHOT ───────────────────────────────────────────────────
    sec(ws, 4, "INVESTMENT SNAPSHOT", col_start=2, ncols=5)

    snapshot = [
        ("Property Name / Tenant",     None,  None),
        ("Property Address",           None,  None),
        ("City, State",                None,  None),
        ("Asset Type",                 "Single-Tenant NNN", None),
        ("Year Built",                 None,  "0"),
        ("Total GLA / RBA (SF)",       None,  N0),
        ("Land Area (Acres)",          None,  N1),
        ("Lease Type",                 "NNN", None),
        ("Lease Commencement",         None,  DT),
        ("Lease Expiration",           None,  DT),
        ("Remaining Term",             None,  '0.0 "yrs"'),
        ("Year 1 Annual Rent",         None,  D0),
        ("Year 1 NOI",                 None,  D0),
        ("Tenant / Guarantor",         None,  None),
        ("Credit Rating",              None,  None),
    ]
    for i, (label, default, fmt) in enumerate(snapshot):
        r = 5 + i
        h(ws, r, 18)
        bg = F_PALE if i % 2 == 0 else F_WHITE
        c = ws.cell(row=r, column=2, value=label)
        c.font = FT_DATA; c.fill = bg; c.alignment = AL_L
        for col in range(3, 7): ws.cell(row=r, column=col).fill = bg
        if default:
            v = ws.cell(row=r, column=3, value=default)
            v.font = FT_DATA; v.alignment = AL_L
        else:
            v = inp(ws, r, 3, fmt=fmt)
        merge(ws, r, 3, r, 6)
    add_cf_clear(ws, "C5:C19")

    h(ws, 20, 10)

    # ── PRICING STRATEGY ──────────────────────────────────────────────────────
    sec(ws, 21, "PRICING STRATEGY", col_start=2, ncols=5)

    # Recommended Asking Price sub-header
    h(ws, 22, 16)
    c = ws.cell(row=22, column=2, value="RECOMMENDED ASKING PRICE")
    c.font = ft("Calibri", 10, b=True, c=NAVY); c.fill = F_PALE; c.alignment = AL_L
    for col in range(3, 7): ws.cell(row=22, column=col).fill = F_PALE

    # Column headers
    h(ws, 23, 18)
    for col, label, al in [
        (2, "",           AL_L),
        (3, "CAP RATE",   AL_C),
        (4, "PRICE",      AL_R),
        (5, "PRICE / SF", AL_R),
        (6, "NOI",        AL_R),
    ]:
        c = ws.cell(row=23, column=col, value=label)
        c.font = FT_CHDR; c.fill = F_NAVY; c.alignment = al

    # Asking price row — cap rate input; price derives from NOI / cap
    h(ws, 24, 22)
    ws.cell(row=24, column=2, value="Asking Price").font = FT_LABEL
    ws.cell(row=24, column=2).fill = F_TOT; ws.cell(row=24, column=2).alignment = AL_L
    inp(ws, 24, 3, fmt=P2).fill = F_TOT
    # NOI = Assumptions C33 (Estimated NOI Y1); Building SF = Assumptions C9
    frm(ws, 24, 4,
        f'=IFERROR(IF(OR({_A}!$C$33="",C24="",C24=0),"",{_A}!$C$33/C24),"")',
        fmt=D0).fill = F_TOT
    frm(ws, 24, 5,
        f'=IFERROR(IF(OR(D24="",{_A}!$C$9="",{_A}!$C$9=0),"",D24/{_A}!$C$9),"")',
        fmt=D2).fill = F_TOT
    frm(ws, 24, 6,
        f'=IFERROR(IF({_A}!$C$33="","",{_A}!$C$33),"")',
        fmt=D0).fill = F_TOT
    add_cf_clear(ws, "C24")

    h(ws, 25, 10)

    # Expected Trade Range
    h(ws, 26, 16)
    c = ws.cell(row=26, column=2, value="EXPECTED TRADE RANGE")
    c.font = ft("Calibri", 10, b=True, c=NAVY); c.fill = F_PALE; c.alignment = AL_L
    for col in range(3, 7): ws.cell(row=26, column=col).fill = F_PALE

    h(ws, 27, 18)
    for col, label, al in [
        (2, "",           AL_L),
        (3, "CAP RATE",   AL_C),
        (4, "PRICE",      AL_R),
        (5, "PRICE / SF", AL_R),
        (6, "NOTES",      AL_L),
    ]:
        c = ws.cell(row=27, column=col, value=label)
        c.font = FT_CHDR; c.fill = F_NAVY; c.alignment = al

    # Removed "Floor / Ceiling" — now "Low Indication" / "High Indication"
    for i, (scenario, bg) in enumerate([
        ("Low Indication",  F_PALE),
        ("High Indication", F_WHITE),
    ]):
        r = 28 + i
        h(ws, r, 20)
        ws.cell(row=r, column=2, value=scenario).font = FT_DATA
        ws.cell(row=r, column=2).fill = bg; ws.cell(row=r, column=2).alignment = AL_L
        inp(ws, r, 3, fmt=P2).fill = bg
        frm(ws, r, 4,
            f'=IFERROR(IF(OR({_A}!$C$33="",C{r}="",C{r}=0),"",{_A}!$C$33/C{r}),"")',
            fmt=D0).fill = bg
        frm(ws, r, 5,
            f'=IFERROR(IF(OR(D{r}="",{_A}!$C$9="",{_A}!$C$9=0),"",D{r}/{_A}!$C$9),"")',
            fmt=D2).fill = bg
        inp(ws, r, 6).fill = bg
    add_cf_clear(ws, "C28:C29")
    add_cf_clear(ws, "F28:F29")

    h(ws, 30, 10)

    # Broker recommendation
    sec(ws, 31, "BROKER RECOMMENDATION", col_start=2, ncols=5)
    h(ws, 32, 80)
    c = ws.cell(row=32, column=2,
                value="[Enter pricing recommendation and rationale. Include comparable cap rates, "
                      "market conditions, tenant credit, lease term remaining, and the competitive "
                      "bid landscape expected to drive pricing into or beyond the indicated trade range.]")
    c.font = FT_INPUT; c.fill = F_YELL; c.alignment = AL_TL
    merge(ws, 32, 2, 32, 6)
    add_cf_clear(ws, "B32")
    h(ws, 33, 10)

    # ── ENGAGEMENT PROPOSAL ───────────────────────────────────────────────────
    sec(ws, 34, "ENGAGEMENT PROPOSAL", col_start=2, ncols=5)

    h(ws, 35, 18)
    for col, label, al in [
        (2, "TERM",   AL_L),
        (3, "DETAIL", AL_L),
    ]:
        c = ws.cell(row=35, column=col, value=label)
        c.font = FT_CHDR; c.fill = F_NAVY; c.alignment = al
    for col in range(4, 7): ws.cell(row=35, column=col).fill = F_NAVY
    merge(ws, 35, 3, 35, 6)

    # Engagement rows — new structure per client notes
    engagement = [
        # (label, default_text, row_height, is_tall)
        ("Listing Type",
         "Exclusive Right to Sell",
         20, False),

        ("Brokerage Commission",
         "[X]% of Gross Sales Price, Payable by Seller at Closing — "
         "Split 50/50 Between Procuring Broker and Listing Broker",
         52, True),

        ("Transactional Brokerage Discount",
         "[X]% of Gross Sales Price, Payable by Seller at Closing — "
         "Applicable When Team Briggs Represents Both Buyer and Seller in the Transaction",
         52, True),

        ("Listing Agreement Term",
         "Six (6) Months",
         20, False),

        ("Additional Terms",
         "All Marketing and Pursuit Costs Paid by NorthMarq Commercial in Advance "
         "Regardless of Closing",
         44, True),
    ]

    for i, (label, val, ht, is_tall) in enumerate(engagement):
        r = 36 + i
        h(ws, r, ht)
        bg = F_PALE if i % 2 == 0 else F_WHITE
        c = ws.cell(row=r, column=2, value=label)
        c.font = FT_LABEL; c.fill = bg; c.alignment = AL_TL if is_tall else AL_L
        v = ws.cell(row=r, column=3, value=val)
        v.font = FT_INPUT; v.fill = bg
        v.alignment = AL_TL if is_tall else AL_L
        for col in range(4, 7): ws.cell(row=r, column=col).fill = bg
        merge(ws, r, 3, r, 6)

    add_cf_clear(ws, "C36:C40")

    # NM branded bottom
    r = 42
    h(ws, r, 10)
    from openpyxl.styles import Border, Side
    for ci in range(2, 7):
        ws.cell(row=r, column=ci).border = Border(
            bottom=Side(style='medium', color=NAVY)
        )
