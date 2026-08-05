"""
test_validate_comps_output.py — the comps conformance validator PASSES on a
freshly generated workbook and FAILS on a hand-rolled / corrupted one.

Structural checks (sheets, headers, formula-protected columns, trim + AVG ranges)
run without LibreOffice; the recalc-error check reads cached values, so it is
exercised with a synthesized error literal rather than a live recalc.
"""

import os
import sys
from pathlib import Path

import pytest
from openpyxl import load_workbook

_HERE = Path(__file__).parent
if str(_HERE) not in sys.path:
    sys.path.insert(0, str(_HERE))

from comps_generator import populate_comps
from validate_comps_output import (
    validate_comps_file, validate_comps_workbook, assert_conforms,
    CompsConformanceError,
)

TEMPLATE_DIR = _HERE / "templates"


def _dialysis_payload():
    return {
        "comp_type": "sales", "vertical": "dialysis",
        "sold": [
            {"tenant": "DaVita", "address": "100 Main St", "city": "Austin", "state": "TX",
             "rba": 8000, "chairs": 20, "patients": 120, "rent": 250000,
             "sold_price": 3500000, "date": "2026-03-15"},
            {"tenant": "Fresenius", "address": "200 Oak Ave", "city": "Dallas", "state": "TX",
             "rba": 9000, "chairs": 24, "patients": 140, "rent": 300000,
             "sold_price": 4200000, "date": "2026-01-10"},
        ],
        "on_market": [
            {"tenant": "US Renal", "address": "300 Elm Rd", "city": "Houston", "state": "TX",
             "rba": 7000, "chairs": 18, "patients": 100, "rent": 210000,
             "initial_price": 3000000, "last_price": 3000000, "on_market": "2026-05-01"},
        ],
    }


def _gen(tmp_path, payload):
    out = str(tmp_path / "comps.xlsx")
    populate_comps(payload, out, template_dir=TEMPLATE_DIR)
    return out


def test_fresh_dialysis_workbook_passes(tmp_path):
    out = _gen(tmp_path, _dialysis_payload())
    res = validate_comps_file(out, check_recalc_errors=False)
    assert res.ok, res.violations
    assert res.vertical == "dialysis"
    # header + formula + trim + avg checks all recorded
    assert "On Market:headers" in res.checks
    assert "Sold:headers" in res.checks
    assert "Sold:trimmed" in res.checks


def test_fresh_generic_and_government_pass(tmp_path):
    generic = {"comp_type": "sales", "sold": [
        {"tenant": "Walgreens", "address": "1 A St", "city": "Reno", "state": "NV",
         "rba": 14000, "rent": 400000, "sold_price": 6000000, "date": "2026-02-02"}]}
    out = _gen(tmp_path, generic)
    assert validate_comps_file(out, check_recalc_errors=False).ok

    gov = {"comp_type": "sales", "vertical": "government", "sold": [
        {"agency": "SSA", "address": "5 Fed Blvd", "city": "Tulsa", "state": "OK",
         "rba": 30000, "gross_rent": 600000, "noi": 500000, "sold_price": 8000000,
         "sold_date": "2026-04-01"}]}
    out2 = _gen(tmp_path, gov)
    res = validate_comps_file(out2, check_recalc_errors=False)
    assert res.ok, res.violations
    assert res.vertical == "government"


def test_populate_comps_embeds_conformance(tmp_path):
    out = str(tmp_path / "c.xlsx")
    summary = populate_comps(_dialysis_payload(), out, template_dir=TEMPLATE_DIR)
    assert summary["conformance"]["ok"] is True


def test_options_header_no_renewal_prefix(tmp_path):
    """Prompt 43: the column header is OPTIONS (not RENEWAL OPTIONS) on both tabs."""
    out = _gen(tmp_path, _dialysis_payload())
    wb = load_workbook(out)
    for sh in ("On Market", "Sold"):
        ws = wb[sh]
        hdrs = {str(ws.cell(5, c).value or "").strip().upper()
                for c in range(1, ws.max_column + 1)}
        assert "OPTIONS" in hdrs
        assert "RENEWAL OPTIONS" not in hdrs
    wb.close()


def test_autofit_nowrap_and_shared_widths(tmp_path):
    """Prompt 43: no wrapped cells, columns fit contents, shared widths line up."""
    out = _gen(tmp_path, _dialysis_payload())
    res = validate_comps_file(out, check_recalc_errors=False)
    assert res.ok, res.violations
    assert "On Market:no_wrap" in res.checks
    assert "Sold:no_wrap" in res.checks
    assert "On Market:widths_fit" in res.checks
    assert "Sold:widths_fit" in res.checks
    assert "shared_widths_match" in res.checks


def test_wrapped_cell_fails(tmp_path):
    from openpyxl.styles import Alignment
    out = _gen(tmp_path, _dialysis_payload())
    wb = load_workbook(out)
    wb["Sold"].cell(6, 3).alignment = Alignment(wrap_text=True)
    wb.save(out); wb.close()
    res = validate_comps_file(out, check_recalc_errors=False)
    assert not res.ok
    assert any("wrapped cell" in v for v in res.violations)


# ── FAIL cases (hand-rolled / corrupted) ──────────────────────────────────────

def test_overflow_on_market_is_capped_and_conforms(tmp_path):
    """Prompt 46: 174 on-market rows into a 100-row template must NOT overflow the
    AVG bar — the renderer caps to capacity, trims, and the workbook conforms."""
    om = [{"tenant": f"Tenant {i}", "address": f"{100+i} Market St", "city": "Houston",
           "state": "TX", "rba": 7000 + i, "rent": 210000, "initial_price": 3000000,
           "last_price": 2900000, "on_market": "2026-05-01"} for i in range(174)]
    sold = [{"tenant": f"Seller {i}", "address": f"{i} Oak Rd", "city": "Dallas",
             "state": "TX", "rba": 8000, "rent": 250000, "sold_price": 3500000,
             "date": "2026-03-15"} for i in range(120)]
    out = _gen(tmp_path, {"comp_type": "sales", "vertical": "dialysis",
                          "on_market": om, "sold": sold})
    res = validate_comps_file(out, check_recalc_errors=False)
    assert res.ok, res.violations
    # both grids capped to the 100-row template capacity, trimmed to the bar
    wb = load_workbook(out)
    for sh in ("On Market", "Sold"):
        ws = wb[sh]
        avg = next(r for r in range(6, ws.max_row + 1)
                   if str(ws.cell(r, 1).value).strip().upper() == "AVG")
        assert avg == 106, f"{sh} not capped to capacity (bar at {avg})"
    wb.close()


def test_shared_width_contract_matches_validator(tmp_path):
    """Prompt 46: the renderer sizes columns via the SAME contract the validator
    checks, so a fresh build passes both the fit and shared-width checks even with
    long, asymmetric content across On Market and Sold."""
    payload = {
        "comp_type": "sales", "vertical": "dialysis",
        "sold": [{"tenant": "DaVita Dialysis Center of Somewhereville",
                  "address": "12345 Very Long Boulevard Suite 2200", "city": "Austin",
                  "state": "TX", "rba": 8000, "chairs": 20, "patients": 1200,
                  "rent": 250000, "sold_price": 3500000, "date": "2026-03-15"}],
        "on_market": [{"tenant": "US Renal", "address": "3 Elm", "city": "Houston",
                       "state": "TX", "rba": 7000, "chairs": 18, "patients": 100,
                       "rent": 210000, "initial_price": 3000000, "last_price": 2900000,
                       "on_market": "2026-05-01"}],
    }
    out = _gen(tmp_path, payload)
    res = validate_comps_file(out, check_recalc_errors=False)
    assert res.ok, res.violations
    assert "On Market:widths_fit" in res.checks
    assert "Sold:widths_fit" in res.checks
    assert "shared_widths_match" in res.checks


def test_wrong_sheets_fail(tmp_path):
    out = _gen(tmp_path, _dialysis_payload())
    wb = load_workbook(out)
    wb.create_sheet("Methodology")     # invented extra tab
    del wb["Index"]                    # missing a required sheet
    wb.save(out); wb.close()
    res = validate_comps_file(out, check_recalc_errors=False)
    assert not res.ok
    assert any("sheet set" in v for v in res.violations)


def test_blank_chairs_untrimmed_grid_fails(tmp_path):
    """Hand-rolled: extra blank data rows pushed below the data (grid not trimmed)."""
    out = _gen(tmp_path, _dialysis_payload())
    wb = load_workbook(out)
    ws = wb["Sold"]
    # find AVG row, insert a blank row just above it → untrimmed grid
    avg = next(r for r in range(6, ws.max_row + 1)
               if str(ws.cell(r, 1).value).strip().upper() == "AVG")
    ws.insert_rows(avg, 1)  # a fully-blank data row now sits above the bar
    wb.save(out); wb.close()
    res = validate_comps_file(out, check_recalc_errors=False)
    assert not res.ok
    assert any("trimmed" in v or "range" in v for v in res.violations)


def test_literal_in_formula_column_fails(tmp_path):
    out = _gen(tmp_path, _dialysis_payload())
    wb = load_workbook(out)
    ws = wb["Sold"]
    # SOLD CAP is a formula-protected column — overwrite row-6 formula with a literal
    cap_col = next(c for c in range(1, ws.max_column + 1)
                   if str(ws.cell(5, c).value).strip().upper() == "SOLD CAP")
    ws.cell(6, cap_col).value = 0.075   # a literal, not a formula
    wb.save(out); wb.close()
    res = validate_comps_file(out, check_recalc_errors=False)
    assert not res.ok
    assert any("SOLD CAP" in v and "literal" in v for v in res.violations)


def test_wrong_headers_fail(tmp_path):
    out = _gen(tmp_path, _dialysis_payload())
    wb = load_workbook(out)
    ws = wb["On Market"]
    # blank CHAIRS header → header row no longer matches the dialysis canon
    chairs = next(c for c in range(1, ws.max_column + 1)
                  if str(ws.cell(5, c).value).strip().upper() == "CHAIRS")
    ws.cell(5, chairs).value = "SEATS"
    wb.save(out); wb.close()
    res = validate_comps_file(out, check_recalc_errors=False)
    assert not res.ok
    assert any("header row 5" in v for v in res.violations)


def test_recalc_error_literal_fails(tmp_path):
    """A cached Excel error literal (what a real recalc would surface) fails check 5."""
    out = _gen(tmp_path, _dialysis_payload())
    # inject a cached error literal into a cell value
    wb = load_workbook(out)
    wb["Sold"].cell(6, 3).value = "#REF!"
    wb.save(out); wb.close()
    res = validate_comps_file(out, check_recalc_errors=True)
    assert not res.ok
    assert any("#REF!" in v or "recalc reports" in v for v in res.violations)


def test_assert_conforms_raises(tmp_path):
    out = _gen(tmp_path, _dialysis_payload())
    wb = load_workbook(out); wb.create_sheet("Bogus"); wb.save(out); wb.close()
    with pytest.raises(CompsConformanceError):
        assert_conforms(out, check_recalc_errors=False)
