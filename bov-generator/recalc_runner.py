"""
recalc_runner.py — Thin wrapper that invokes the LibreOffice recalc logic
and raises on any error or formula failures.
"""

import sys
import os

# Make the local 'office' package importable (soffice.py shim)
_HERE = os.path.dirname(os.path.abspath(__file__))
if _HERE not in sys.path:
    sys.path.insert(0, _HERE)

# Import the recalc function directly from the skill script.
# We replicate the minimal required logic here rather than depending on
# the skill path, so the Railway container is fully self-contained.

import contextlib
import json
import re
import shutil
import subprocess
import tempfile
import time
import zipfile
from pathlib import Path

from openpyxl import load_workbook
from office.soffice import get_soffice_env, run_soffice

MACRO_FILENAME = "Module1.xba"
MAX_LOCATIONS  = 100
EXCEL_ERRORS   = ["#VALUE!", "#DIV/0!", "#REF!", "#NAME?", "#NULL!", "#NUM!", "#N/A"]

RECALCULATE_MACRO = """<?xml version="1.0" encoding="UTF-8"?>
<!DOCTYPE script:module PUBLIC "-//OpenOffice.org//DTD OfficeDocument 1.0//EN" "module.dtd">
<script:module xmlns:script="http://openoffice.org/2000/script" script:name="Module1" script:language="StarBasic">
    Sub RecalculateAndSave()
      ThisComponent.calculateAll()
      ThisComponent.store()
      ThisComponent.close(True)
    End Sub
</script:module>"""


class RecalcError(Exception):
    pass


def _stamp(path):
    st = os.stat(path)
    return st.st_mtime_ns, st.st_size


def _setup_profile(profile_dir: Path, timeout: int):
    url = profile_dir.as_uri()
    try:
        run_soffice(
            ["--headless", "--terminate_after_init", f"-env:UserInstallation={url}"],
            capture_output=True, timeout=timeout,
        )
    except FileNotFoundError:
        raise RecalcError("soffice not found — LibreOffice is required")
    except subprocess.TimeoutExpired:
        raise RecalcError("LibreOffice timed out creating its profile")

    macro_dir = profile_dir / "user" / "basic" / "Standard"
    if not macro_dir.exists():
        raise RecalcError("LibreOffice did not create a usable profile")
    (macro_dir / MACRO_FILENAME).write_text(RECALCULATE_MACRO)
    return url


def recalc_and_validate(filepath: str, timeout: int = 90) -> dict:
    """
    Recalculate all formulas in filepath using LibreOffice.
    Returns the standard result dict.
    Raises RecalcError on any failure.
    """
    abs_path = str(Path(filepath).absolute())
    if not Path(abs_path).exists():
        raise RecalcError(f"File not found: {abs_path}")
    if not os.access(abs_path, os.W_OK):
        raise RecalcError(f"File not writable: {abs_path}")

    with tempfile.TemporaryDirectory(prefix="recalc-profile-", ignore_cleanup_errors=True) as td:
        profile_url = _setup_profile(Path(td), timeout // 3)
        before = _stamp(abs_path)

        cmd = [
            "soffice", "--headless", "--norestore",
            f"-env:UserInstallation={profile_url}",
            "vnd.sun.star.script:Standard.Module1.RecalculateAndSave?language=Basic&location=application",
            abs_path,
        ]
        if shutil.which("timeout"):
            cmd = ["timeout", str(timeout)] + cmd

        try:
            result = subprocess.run(
                cmd, capture_output=True, text=True,
                env=get_soffice_env(), timeout=timeout + 15,
            )
        except subprocess.TimeoutExpired:
            raise RecalcError(f"LibreOffice timed out after {timeout}s")
        except FileNotFoundError:
            raise RecalcError("soffice not found — LibreOffice is required")

        if result.returncode == 124:
            raise RecalcError(f"LibreOffice timed out after {timeout}s")
        if result.returncode != 0:
            detail = (result.stderr or "").strip() or f"soffice exited {result.returncode}"
            raise RecalcError(f"LibreOffice recalc failed: {detail}")
        if _stamp(abs_path) == before:
            raise RecalcError("LibreOffice exited cleanly but did not rewrite the file")

    # Inspect results
    wb_vals = load_workbook(abs_path, data_only=True)
    error_details = {e: [] for e in EXCEL_ERRORS}
    total_errors  = 0

    for sheet_name in wb_vals.sheetnames:
        ws = wb_vals[sheet_name]
        if not hasattr(ws, "iter_rows"):
            continue
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    for err in EXCEL_ERRORS:
                        if err in cell.value:
                            error_details[err].append(f"{sheet_name}!{cell.coordinate}")
                            total_errors += 1
                            break
    wb_vals.close()

    wb_frm = load_workbook(abs_path, data_only=False)
    formula_count = sum(
        1
        for sn in wb_frm.sheetnames
        for row in wb_frm[sn].iter_rows()
        for cell in row
        if hasattr(wb_frm[sn], "iter_rows") and isinstance(cell.value, str) and cell.value.startswith("=")
    )
    wb_frm.close()

    summary = {}
    for err_type, locs in error_details.items():
        if locs:
            entry = {"count": len(locs), "locations": locs[:MAX_LOCATIONS]}
            if len(locs) > MAX_LOCATIONS:
                entry["locations_truncated"] = len(locs) - MAX_LOCATIONS
            summary[err_type] = entry

    result = {
        "status":         "success" if total_errors == 0 else "errors_found",
        "total_formulas": formula_count,
        "total_errors":   total_errors,
        "error_summary":  summary,
    }

    if total_errors > 0:
        raise RecalcError(
            f"Recalc completed but {total_errors} formula error(s) found: "
            + ", ".join(f"{k}({v['count']})" for k, v in summary.items())
        )

    return result
