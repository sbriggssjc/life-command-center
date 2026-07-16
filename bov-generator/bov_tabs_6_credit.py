"""BOV v2 — Tab 6: Tenant / Guarantor Credit (Leg 3).
Layout: Short-form summary on TOP, long-form credit matrix BELOW.
Both sections share the same 7-column grid (A–G).
"""
from bov_constants import *
from openpyxl.styles import Border, Side
from openpyxl.utils import get_column_letter

_NM_M = Side(style='medium', color=NAVY)
_NM_T = Side(style='thin',   color="CCCCCC")


def build_credit_tab(wb):
    ws = wb.create_sheet("Credit")
    ws.sheet_view.showGridLines = False

    # ── Unified column widths (shared by short-form and long-form) ─────────────
    # A(1)=margin  B(2)=#  C(3)=category/item-label
    # D(4)=attribute/SF-summary-anchor  E(5)=finding  F(6)=source  G(7)=notes
    col_widths = [2, 5, 26, 26, 30, 18, 26]
    for i, wd in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = wd

    # ── Tab header ────────────────────────────────────────────────────────────
    r = 1;  ws.row_dimensions[r].height = 6
    r = 2;  ws.row_dimensions[r].height = 28
    c = ws.cell(row=r, column=2, value="TENANT / GUARANTOR CREDIT")
    c.font = FT_TITLE; c.alignment = AL_L
    merge(ws, r, 2, r, 7)

    r = 3;  ws.row_dimensions[r].height = 14
    c = ws.cell(row=r, column=2,
                value="Leg 3 of 3  ·  Corporate overview · financial strength · unit economics · guaranty analysis  ·  Short-form at top · full matrix below")
    c.font = FT_NOTE; c.alignment = AL_L
    merge(ws, r, 2, r, 7)

    r = 4;  ws.row_dimensions[r].height = 8

    # ══════════════════════════════════════════════════════════════════════════
    # SHORT-FORM SUMMARY (TOP)
    # ══════════════════════════════════════════════════════════════════════════
    SF_HDR = 5
    ws.row_dimensions[SF_HDR].height = 16
    sec(ws, SF_HDR, "SHORT-FORM SUMMARY  —  For BOV, OM, or Valuation Memo", col_start=2, ncols=6)

    SF_COL_HDR = 6
    ws.row_dimensions[SF_COL_HDR].height = 20
    for ci in [2, 3]:
        ws.cell(row=SF_COL_HDR, column=ci).fill = F_NAVY
        ws.cell(row=SF_COL_HDR, column=ci).border = Border(bottom=_NM_M)
    c = ws.cell(row=SF_COL_HDR, column=3, value="ITEM")
    c.font = FT_CHDR; c.alignment = AL_C
    c2 = ws.cell(row=SF_COL_HDR, column=4, value="ABBREVIATED SUMMARY")
    c2.font = FT_CHDR; c2.fill = F_NAVY; c2.alignment = AL_C
    c2.border = Border(bottom=_NM_M)
    for ci in [5, 6, 7]:
        ws.cell(row=SF_COL_HDR, column=ci).fill = F_NAVY
        ws.cell(row=SF_COL_HDR, column=ci).border = Border(bottom=_NM_M)
    merge(ws, SF_COL_HDR, 4, SF_COL_HDR, 7)

    sf_rows = [
        "Tenant / Operator",
        "Parent Company",
        "Public / Private",
        "Credit Rating",
        "Investment Grade",
        "Total Locations",
        "Annual Revenue",
        "EBITDA / Margin",
        "Rent-to-Sales Ratio",
        "Guarantor",
        "Guaranty Type",
        "Guaranty Strength",
        "Essential / Recession-Resistant",
        "Key Credit Strengths",
        "Key Credit Risks",
        "Broker Commentary",
    ]

    SF_DATA_START = SF_COL_HDR + 1  # row 7
    for idx, item in enumerate(sf_rows):
        rr = SF_DATA_START + idx
        ws.row_dimensions[rr].height = 36
        ws.cell(row=rr, column=2).fill = F_WHITE
        c = ws.cell(row=rr, column=3, value=item)
        c.font = FT_LABEL; c.alignment = AL_TL
        # D:G merged: wide summary input
        inp(ws, rr, 4)
        ws.cell(row=rr, column=4).alignment = AL_TL
        merge(ws, rr, 4, rr, 7)
        for ci in range(2, 8):
            ws.cell(row=rr, column=ci).border = Border(bottom=_NM_T)

    SF_DATA_END = SF_DATA_START + len(sf_rows) - 1
    add_cf_clear(ws, f"D{SF_DATA_START}:D{SF_DATA_END}")

    # SF bottom accent
    sf_bot = SF_DATA_END + 1
    ws.row_dimensions[sf_bot].height = 4
    for ci in range(2, 8):
        ws.cell(row=sf_bot, column=ci).border = Border(top=_NM_M)

    sep_r = sf_bot + 1
    ws.row_dimensions[sep_r].height = 14

    # ══════════════════════════════════════════════════════════════════════════
    # LONG-FORM CREDIT MATRIX (BELOW)
    # ══════════════════════════════════════════════════════════════════════════
    LF_HDR = sep_r + 1
    ws.row_dimensions[LF_HDR].height = 16
    sec(ws, LF_HDR, "LONG-FORM CREDIT OVERVIEW  —  Corporate, financial, and unit-level diligence with source documentation", col_start=2, ncols=6)

    LF_COL_HDR = LF_HDR + 1
    ws.row_dimensions[LF_COL_HDR].height = 20
    lf_hdrs = ["#", "CATEGORY", "ATTRIBUTE", "FINDING / DETAIL", "SOURCE", "NOTES / FLAGS"]
    for ci, hdr in enumerate(lf_hdrs, 2):
        c = ws.cell(row=LF_COL_HDR, column=ci, value=hdr)
        c.font = FT_CHDR; c.fill = F_NAVY; c.alignment = AL_C
        c.border = Border(bottom=_NM_M)

    long_rows = [
        ("Corporate Overview",      "Entity Name (Lease)"),
        ("Corporate Overview",      "Entity Name (Operating / Trade)"),
        ("Corporate Overview",      "Ownership Structure"),
        ("Corporate Overview",      "Parent Company"),
        ("Corporate Overview",      "Headquarters"),
        ("Corporate Overview",      "Founded"),
        ("Corporate Overview",      "Number of Locations (Total)"),
        ("Corporate Overview",      "Number of Locations (State)"),
        ("Corporate Overview",      "Business Description"),
        ("Corporate Overview",      "Years in Operation"),
        (None, None),
        ("Credit & Ratings",        "S&P Credit Rating"),
        ("Credit & Ratings",        "Moody's Credit Rating"),
        ("Credit & Ratings",        "Investment Grade (Y/N)"),
        ("Credit & Ratings",        "Public / Private"),
        ("Credit & Ratings",        "Stock Ticker (if public)"),
        ("Credit & Ratings",        "Market Capitalization"),
        (None, None),
        ("Financial Summary",       "Revenue — Most Recent FY"),
        ("Financial Summary",       "Revenue — Prior FY"),
        ("Financial Summary",       "Revenue Growth YoY"),
        ("Financial Summary",       "EBITDA — Most Recent FY"),
        ("Financial Summary",       "EBITDA Margin"),
        ("Financial Summary",       "Net Income — Most Recent FY"),
        ("Financial Summary",       "Total Debt"),
        ("Financial Summary",       "Total Assets"),
        ("Financial Summary",       "Net Worth / Book Value"),
        ("Financial Summary",       "Cash & Equivalents"),
        ("Financial Summary",       "Debt / EBITDA"),
        ("Financial Summary",       "Source / Reporting Period"),
        (None, None),
        ("Unit Economics",          "Average Unit Volume (AUV)"),
        ("Unit Economics",          "Average Unit SF"),
        ("Unit Economics",          "Rent-to-Sales Ratio (this location)"),
        ("Unit Economics",          "Typical Store Occupancy Cost"),
        ("Unit Economics",          "Break-Even Occupancy"),
        ("Unit Economics",          "Franchise vs. Corporate"),
        ("Unit Economics",          "Local Market Performance"),
        (None, None),
        ("Guaranty",                "Guarantor Name"),
        ("Guaranty",                "Guarantor Type (Corporate / Personal)"),
        ("Guaranty",                "Guaranty Type (Full / Partial / Springing / Burn-off)"),
        ("Guaranty",                "Guaranty Cap ($)"),
        ("Guaranty",                "Burn-off / Release Conditions"),
        ("Guaranty",                "Guarantor Net Worth"),
        ("Guaranty",                "Guaranty Expiration"),
        ("Guaranty",                "Guaranty Document Source"),
        (None, None),
        ("Qualitative",             "Recession Resistance / Essential Services"),
        ("Qualitative",             "Industry / Sector Trends"),
        ("Qualitative",             "Key Risks to Tenancy"),
        ("Qualitative",             "Key Credit Strengths"),
        ("Qualitative",             "Online / Omnichannel Exposure"),
        ("Qualitative",             "Relocation / Closure History"),
        ("Qualitative",             "Broker Commentary"),
    ]

    LF_DATA_START = LF_COL_HDR + 1
    row_num = 0
    current_cat = None
    data_row_num = 0

    for entry in long_rows:
        rr = LF_DATA_START + row_num
        cat, attr = entry
        if cat is None:
            ws.row_dimensions[rr].height = 6
            row_num += 1
            continue
        ws.row_dimensions[rr].height = 36
        data_row_num += 1

        # B: row number
        c = ws.cell(row=rr, column=2, value=data_row_num)
        c.font = FT_NOTE; c.alignment = AL_C

        # C: category (shown on first row of each group)
        if cat != current_cat:
            c2 = ws.cell(row=rr, column=3, value=cat)
            c2.font = FT_LABEL; c2.alignment = AL_TL
            current_cat = cat
        else:
            ws.cell(row=rr, column=3).alignment = AL_TL

        # D: attribute
        c3 = ws.cell(row=rr, column=4, value=attr)
        c3.font = FT_DATA; c3.alignment = AL_TL

        # E: finding / detail (input)
        inp(ws, rr, 5); ws.cell(row=rr, column=5).alignment = AL_TL

        # F: source (input)
        inp(ws, rr, 6); ws.cell(row=rr, column=6).alignment = AL_TL

        # G: notes / flags (input)
        inp(ws, rr, 7); ws.cell(row=rr, column=7).alignment = AL_TL

        for ci in range(2, 8):
            ws.cell(row=rr, column=ci).border = Border(bottom=_NM_T)

        row_num += 1

    LF_DATA_END = LF_DATA_START + row_num - 1

    # CF clear on finding, source, notes columns
    for col in [5, 6, 7]:
        cl = get_column_letter(col)
        add_cf_clear(ws, f"{cl}{LF_DATA_START}:{cl}{LF_DATA_END}")

    # Medium bottom border
    lf_bot = LF_DATA_END + 1
    ws.row_dimensions[lf_bot].height = 4
    for ci in range(2, 8):
        ws.cell(row=lf_bot, column=ci).border = Border(top=_NM_M)

    # Note row
    note_r = lf_bot + 1
    ws.row_dimensions[note_r].height = 14
    c = ws.cell(row=note_r, column=2,
                value="Source: SEC filing, S&P Global, Moody's, company press release, FDD, or broker research  ·  Flag any estimated or unverified figures in Notes column")
    c.font = FT_NOTE; c.alignment = AL_L
    merge(ws, note_r, 2, note_r, 7)

    # NM bottom accent
    bot_r = note_r + 2
    ws.row_dimensions[bot_r].height = 4
    for ci in range(2, 8):
        ws.cell(row=bot_r, column=ci).border = Border(bottom=_NM_M)
