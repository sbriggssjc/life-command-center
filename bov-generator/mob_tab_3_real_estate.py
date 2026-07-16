"""BOV MOB — Tab 3: Real Estate Fundamentals (Leg 1).
Layout: Short-form summary on TOP, long-form diligence matrix BELOW.
Both sections share the same 7-column grid (A–G).
"""
from bov_constants import *
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

_NM_M = Side(style='medium', color=NAVY)
_NM_T = Side(style='thin',   color="CCCCCC")


def build_mob_real_estate_tab(wb):
    ws = wb.create_sheet("Real Estate")
    ws.sheet_view.showGridLines = False

    # ── Unified column widths (shared by short-form and long-form) ─────────────
    # A(1)=margin  B(2)=#  C(3)=category/label  D(4)=attribute/SF-summary-anchor
    # E(5)=finding  F(6)=source  G(7)=notes/flags
    col_widths = [2, 5, 24, 26, 30, 18, 28]
    for i, wd in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = wd

    # ── Tab header ────────────────────────────────────────────────────────────
    r = 1;  ws.row_dimensions[r].height = 6
    r = 2;  ws.row_dimensions[r].height = 28
    c = ws.cell(row=r, column=2, value="REAL ESTATE FUNDAMENTALS")
    c.font = FT_TITLE; c.alignment = AL_L
    merge(ws, r, 2, r, 7)

    r = 3;  ws.row_dimensions[r].height = 14
    c = ws.cell(row=r, column=2,
                value="Leg 1 of 3  ·  Physical property, location, and market fundamentals  ·  Multi-Tenant MOB / Commercial  ·  Short-form at top · full matrix below")
    c.font = FT_NOTE; c.alignment = AL_L
    merge(ws, r, 2, r, 7)

    r = 4;  ws.row_dimensions[r].height = 8

    # ══════════════════════════════════════════════════════════════════════════
    # SHORT-FORM SUMMARY (TOP)
    # ══════════════════════════════════════════════════════════════════════════
    SF_HDR = 5
    ws.row_dimensions[SF_HDR].height = 16
    sec(ws, SF_HDR, "SHORT-FORM SUMMARY  —  For BOV, OM, or Valuation Memo", col_start=2, ncols=6)

    SF_COL_HDR = 6
    ws.row_dimensions[SF_COL_HDR].height = 20
    # B: blank navy filler, C: ATTRIBUTE label col, D:G merged: FINDING / SUMMARY
    for ci in [2, 3]:
        ws.cell(row=SF_COL_HDR, column=ci).fill = F_NAVY
        ws.cell(row=SF_COL_HDR, column=ci).border = Border(bottom=_NM_M)
    c = ws.cell(row=SF_COL_HDR, column=3, value="ATTRIBUTE")
    c.font = FT_CHDR; c.alignment = AL_C
    c2 = ws.cell(row=SF_COL_HDR, column=4, value="FINDING / SUMMARY")
    c2.font = FT_CHDR; c2.fill = F_NAVY; c2.alignment = AL_C
    c2.border = Border(bottom=_NM_M)
    for ci in [5, 6, 7]:
        ws.cell(row=SF_COL_HDR, column=ci).fill = F_NAVY
        ws.cell(row=SF_COL_HDR, column=ci).border = Border(bottom=_NM_M)
    merge(ws, SF_COL_HDR, 4, SF_COL_HDR, 7)

    sf_rows = [
        "Property Address",
        "Building SF",
        "Year Built / Renovated",
        "Site Area (Acres)",
        "Zoning",
        "Flood Zone",
        "Parking",
        "Condition (Overall)",
        "Environmental Status",
        "Proximity / Demand Generators",
        "Market Context",
        "Notable Strengths",
        "Notable Concerns",
        "Broker Commentary",
    ]

    SF_DATA_START = SF_COL_HDR + 1  # row 7
    for idx, attr in enumerate(sf_rows):
        rr = SF_DATA_START + idx
        ws.row_dimensions[rr].height = 36
        ws.cell(row=rr, column=2).fill = F_WHITE
        # C: attribute label
        c = ws.cell(row=rr, column=3, value=attr)
        c.font = FT_LABEL; c.alignment = AL_TL
        # D:G merged: wide summary input
        inp(ws, rr, 4)
        ws.cell(row=rr, column=4).alignment = AL_TL
        merge(ws, rr, 4, rr, 7)
        for ci in range(2, 8):
            ws.cell(row=rr, column=ci).border = Border(bottom=_NM_T)

    SF_DATA_END = SF_DATA_START + len(sf_rows) - 1
    add_cf_clear(ws, f"D{SF_DATA_START}:D{SF_DATA_END}")

    # SF bottom accent
    sf_bot = SF_DATA_END + 1
    ws.row_dimensions[sf_bot].height = 4
    for ci in range(2, 8):
        ws.cell(row=sf_bot, column=ci).border = Border(top=_NM_M)

    # Spacer between sections
    sep_r = sf_bot + 1
    ws.row_dimensions[sep_r].height = 14

    # ══════════════════════════════════════════════════════════════════════════
    # LONG-FORM DILIGENCE MATRIX (BELOW)
    # ══════════════════════════════════════════════════════════════════════════
    LF_HDR = sep_r + 1
    ws.row_dimensions[LF_HDR].height = 16
    sec(ws, LF_HDR, "LONG-FORM DETAIL  —  Internal diligence and source documentation", col_start=2, ncols=6)

    LF_COL_HDR = LF_HDR + 1
    ws.row_dimensions[LF_COL_HDR].height = 20
    lf_hdrs = ["#", "CATEGORY", "ATTRIBUTE", "FINDING / DETAIL", "SOURCE", "NOTES / FLAGS"]
    for ci, hdr in enumerate(lf_hdrs, 2):
        c = ws.cell(row=LF_COL_HDR, column=ci, value=hdr)
        c.font = FT_CHDR; c.fill = F_NAVY; c.alignment = AL_C
        c.border = Border(bottom=_NM_M)

    long_rows = [
        ("Location & Market",       "Property Address"),
        ("Location & Market",       "City / State / Zip"),
        ("Location & Market",       "County"),
        ("Location & Market",       "MSA / Submarket"),
        ("Location & Market",       "Population (1 / 3 / 5 mi)"),
        ("Location & Market",       "Median HH Income"),
        ("Location & Market",       "Traffic Counts"),
        ("Location & Market",       "Proximity to Demand Generators"),
        ("Location & Market",       "Market Rent Context"),
        (None, None),
        ("Site Characteristics",    "Site Area (Acres)"),
        ("Site Characteristics",    "Parcel APN / Tax ID"),
        ("Site Characteristics",    "Legal Description Summary"),
        ("Site Characteristics",    "Lot Configuration / Shape"),
        ("Site Characteristics",    "Frontage (LF)"),
        ("Site Characteristics",    "Topography"),
        ("Site Characteristics",    "Flood Zone Designation"),
        ("Site Characteristics",    "Utilities Available"),
        (None, None),
        ("Building & Improvements", "Building SF"),
        ("Building & Improvements", "Year Built"),
        ("Building & Improvements", "Year Renovated"),
        ("Building & Improvements", "Construction Type"),
        ("Building & Improvements", "Roof Type / Age"),
        ("Building & Improvements", "HVAC Type / Age"),
        ("Building & Improvements", "ADA Compliance"),
        ("Building & Improvements", "Condition (Overall)"),
        ("Building & Improvements", "Deferred Maintenance"),
        (None, None),
        ("Zoning & Land Use",       "Zoning Classification"),
        ("Zoning & Land Use",       "Permitted Use Confirmation"),
        ("Zoning & Land Use",       "Drive-Through Permitted"),
        ("Zoning & Land Use",       "Signage Rights"),
        ("Zoning & Land Use",       "Restrictive Covenants / CC&Rs"),
        ("Zoning & Land Use",       "Easements (Recorded)"),
        (None, None),
        ("Ingress / Egress",        "Number of Access Points"),
        ("Ingress / Egress",        "Shared / Reciprocal Access"),
        ("Ingress / Egress",        "Parking Spaces"),
        ("Ingress / Egress",        "Parking Ratio (per 1,000 SF)"),
        ("Ingress / Egress",        "Delivery / Loading Access"),
        (None, None),
        ("Environmental",           "Phase I Status"),
        ("Environmental",           "Phase II Status"),
        ("Environmental",           "Known RECs"),
        ("Environmental",           "Underground Storage Tanks"),
        (None, None),
        ("Market Comps",            "Comparable Sale 1"),
        ("Market Comps",            "Comparable Sale 2"),
        ("Market Comps",            "Comparable Sale 3"),
        ("Market Comps",            "Implied Market Cap Rate"),
        ("Market Comps",            "Implied Market Price / SF"),
    ]

    LF_DATA_START = LF_COL_HDR + 1
    row_num = 0
    current_cat = None
    data_row_num = 0

    for entry in long_rows:
        rr = LF_DATA_START + row_num
        cat, attr = entry
        if cat is None:
            ws.row_dimensions[rr].height = 6
            row_num += 1
            continue
        ws.row_dimensions[rr].height = 36
        data_row_num += 1

        # B: sequential row number
        c = ws.cell(row=rr, column=2, value=data_row_num)
        c.font = FT_NOTE; c.alignment = AL_C

        # C: category (shown on first row of each group)
        if cat != current_cat:
            c2 = ws.cell(row=rr, column=3, value=cat)
            c2.font = FT_LABEL; c2.alignment = AL_TL
            current_cat = cat
        else:
            ws.cell(row=rr, column=3).alignment = AL_TL

        # D: attribute name
        c3 = ws.cell(row=rr, column=4, value=attr)
        c3.font = FT_DATA; c3.alignment = AL_TL

        # E: finding / detail (input)
        inp(ws, rr, 5); ws.cell(row=rr, column=5).alignment = AL_TL

        # F: source (input)
        inp(ws, rr, 6); ws.cell(row=rr, column=6).alignment = AL_TL

        # G: notes / flags (input)
        inp(ws, rr, 7); ws.cell(row=rr, column=7).alignment = AL_TL

        for ci in range(2, 8):
            ws.cell(row=rr, column=ci).border = Border(bottom=_NM_T)

        row_num += 1

    LF_DATA_END = LF_DATA_START + row_num - 1

    # CF clear on finding, source, notes columns
    for col in [5, 6, 7]:
        cl = get_column_letter(col)
        add_cf_clear(ws, f"{cl}{LF_DATA_START}:{cl}{LF_DATA_END}")

    # Medium bottom border closing the LF table
    lf_bot = LF_DATA_END + 1
    ws.row_dimensions[lf_bot].height = 4
    for ci in range(2, 8):
        ws.cell(row=lf_bot, column=ci).border = Border(top=_NM_M)

    # Note row
    note_r = lf_bot + 1
    ws.row_dimensions[note_r].height = 14
    c = ws.cell(row=note_r, column=2,
                value="Yellow cells = broker input  ·  Source: cite lease, appraisal, survey, rent roll, PM report, or LOI  ·  Flag any estimated or unverified figures")
    c.font = FT_NOTE; c.alignment = AL_L
    merge(ws, note_r, 2, note_r, 7)

    # NM bottom accent
    bot_r = note_r + 2
    ws.row_dimensions[bot_r].height = 4
    for ci in range(2, 8):
        ws.cell(row=bot_r, column=ci).border = Border(bottom=_NM_M)
