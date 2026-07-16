"""BOV MOB — Tab 10: Sensitivity Analysis (v2).

Single matrix: 6 return metrics (rows) × 11 going-in cap rates (columns).
Cap rates: asking cap (C95) ±50 bps in 10 bp steps. NOI (C85) fixed.
Loan terms (rate I10, amort I11, LTV 65%) fixed. Only price varies.

IRR metrics use hidden helper rows (LibreOffice-safe cell ranges, not array literals).

MOB-specific Assumptions cell references:
  NOI:           'Assumptions & Flags'!$C$85
  Purchase Price:'Assumptions & Flags'!$C$94
  Going-In Cap:  'Assumptions & Flags'!$C$95
  Avg Escalation:'Assumptions & Flags'!$C$64
  Exit Cap Rate: 'Assumptions & Flags'!$I$20
  Hold Period:   'Assumptions & Flags'!$I$13
  Down Payment:  'Assumptions & Flags'!$I$9
  Debt Service:  'Assumptions & Flags'!$I$14
  Loan Amount:   'Assumptions & Flags'!$I$8
  Interest Rate: 'Assumptions & Flags'!$I$10
  Amortization:  'Assumptions & Flags'!$I$11
"""
from bov_constants import *
from openpyxl.styles import Border, Side, Font, PatternFill
from openpyxl.utils import get_column_letter

_NM_M = Side(style='medium', color=NAVY)
_GY_T = Side(style='thin',   color="CCCCCC")

_AS = "Assumptions & Flags"

# Key MOB cell references
_NOI   = f"'{_AS}'!$C$85"     # Year 1 NOI
_PRICE = f"'{_AS}'!$C$94"     # Purchase Price
_GCAP  = f"'{_AS}'!$C$95"     # Going-In Cap (asking)
_ESC   = f"'{_AS}'!$C$64"     # Portfolio Avg Escalation
_XCAP  = f"'{_AS}'!$I$20"     # Exit Cap Rate
_HOLD  = f"'{_AS}'!$I$13"     # Hold Period
_RATE  = f"'{_AS}'!$I$10"     # Interest Rate
_AMORT = f"'{_AS}'!$I$11"     # Amortization (years)


def build_mob_sensitivity_tab(wb):
    ws = wb.create_sheet("Sensitivity Analysis")
    ws.sheet_view.showGridLines = False

    # ── Column widths ──────────────────────────────────────────────────────────
    # A=margin(2), B=row label(28), C-M=11 data cols(11 each), N=notes(20)
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 28
    for col in range(3, 14):
        ws.column_dimensions[get_column_letter(col)].width = 11
    ws.column_dimensions["N"].width = 20

    # ── Header ────────────────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 6
    ws.row_dimensions[2].height = 28
    c = ws.cell(row=2, column=2, value="SENSITIVITY ANALYSIS  —  RETURN IMPACT BY GOING-IN CAP RATE")
    c.font = FT_TITLE; c.alignment = AL_L
    merge(ws, 2, 2, 2, 13)

    ws.row_dimensions[3].height = 14
    c = ws.cell(row=3, column=2,
                value="Multi-Tenant MOB  ·  6 return metrics at 11 going-in cap rates (ask ±50 bps / 10 bp steps)  ·  NOI and loan terms fixed  ·  Only price varies")
    c.font = FT_NOTE; c.alignment = AL_L
    merge(ws, 3, 2, 3, 13)

    ws.row_dimensions[4].height = 8

    # ── Cap rate configuration ─────────────────────────────────────────────────
    CAP_OFFSETS   = [-0.005, -0.004, -0.003, -0.002, -0.001, 0.000, 0.001, 0.002, 0.003, 0.004, 0.005]
    CAP_DATA_COLS = list(range(3, 14))   # cols 3–13 (C through M)
    ASK_COL       = 8                    # asking cap = center column (offset index 5)
    NCOLS_DISP    = 13                   # last display column

    _GOLD_HDR = PatternFill("solid", fgColor="FFC000")   # gold header for asking cap col
    _GOLD_VAL = PatternFill("solid", fgColor="FFF2CC")   # light gold values for asking cap col
    _TEAL     = PatternFill("solid", fgColor="1F7A8C")   # unleveraged group header
    _NAVY_DRK = PatternFill("solid", fgColor="1F3F5C")   # leveraged group header

    # ── Section header ────────────────────────────────────────────────────────
    ws.row_dimensions[5].height = 6
    sec(ws, 6, "RETURN METRICS BY GOING-IN CAP RATE  ·  Center = Asking Cap (highlighted gold)  ·  NOI and Loan Terms Fixed", col_start=2, ncols=12)
    ws.row_dimensions[6].height = 16

    # ── Column header row (dynamic cap rates from Assumptions C95) ─────────────
    CAP_HDR_ROW = 7
    ws.row_dimensions[CAP_HDR_ROW].height = 22
    c = ws.cell(row=CAP_HDR_ROW, column=2, value="METRIC  /  GOING-IN CAP →")
    c.font = FT_CHDR; c.fill = F_NAVY; c.alignment = AL_C

    for i, offset in enumerate(CAP_OFFSETS):
        col = 3 + i
        is_ask = (col == ASK_COL)
        if offset == 0.0:
            formula = f'=IFERROR({_GCAP},"")'
        elif offset < 0:
            formula = f'=IFERROR({_GCAP}{offset:.3f},"")'
        else:
            formula = f'=IFERROR({_GCAP}+{offset:.3f},"")'
        c = frm(ws, CAP_HDR_ROW, col, formula, fmt=P2, align=AL_C)
        c.font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
        c.fill = _GOLD_HDR if is_ask else F_NAVY

    # ── Unleveraged group header ───────────────────────────────────────────────
    UL_GRP_ROW = 8
    ws.row_dimensions[UL_GRP_ROW].height = 18
    c = ws.cell(row=UL_GRP_ROW, column=2, value="UNLEVERAGED RETURNS")
    c.font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
    c.fill = _TEAL; c.alignment = AL_L
    merge(ws, UL_GRP_ROW, 2, UL_GRP_ROW, NCOLS_DISP)

    # Row definitions: (row_num, label, fill, font, format)
    M1_ROW = 9;  ws.row_dimensions[M1_ROW].height = 20
    M2_ROW = 10; ws.row_dimensions[M2_ROW].height = 20
    M3_ROW = 11; ws.row_dimensions[M3_ROW].height = 20

    for r, label in [(M1_ROW, "Avg Cash Cap Rate (over hold)"),
                     (M2_ROW, "Unleveraged ERM  (Cash ×)"),
                     (M3_ROW, "Cash IRR  (Unleveraged)")]:
        bg = F_PALE if r == M1_ROW else (F_TOT if r == M3_ROW else F_WHITE)
        ft = FT_TOTAL if r == M3_ROW else FT_DATA
        c = ws.cell(row=r, column=2, value=label)
        c.font = ft; c.fill = bg; c.alignment = AL_L

    # ── Leveraged group header ─────────────────────────────────────────────────
    LEV_GRP_ROW = 12
    ws.row_dimensions[LEV_GRP_ROW].height = 18
    c = ws.cell(row=LEV_GRP_ROW, column=2, value="LEVERAGED RETURNS  (65% LTV · Rate & Amortization from Assumptions)")
    c.font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
    c.fill = _NAVY_DRK; c.alignment = AL_L
    merge(ws, LEV_GRP_ROW, 2, LEV_GRP_ROW, NCOLS_DISP)

    M4_ROW = 13; ws.row_dimensions[M4_ROW].height = 20
    M5_ROW = 14; ws.row_dimensions[M5_ROW].height = 20
    M6_ROW = 15; ws.row_dimensions[M6_ROW].height = 20

    for r, label in [(M4_ROW, "Lev Avg Cash-on-Cash (over hold)"),
                     (M5_ROW, "Leveraged ERM  (Cash ×)"),
                     (M6_ROW, "Leveraged IRR")]:
        bg = F_PALE if r == M4_ROW else (F_TOT if r == M6_ROW else F_WHITE)
        ft = FT_TOTAL if r == M6_ROW else FT_DATA
        c = ws.cell(row=r, column=2, value=label)
        c.font = ft; c.fill = bg; c.alignment = AL_L

    # ── Display matrix bottom accent & notes ──────────────────────────────────
    ws.row_dimensions[16].height = 4
    for ci in range(2, NCOLS_DISP + 1):
        ws.cell(row=16, column=ci).border = Border(top=_NM_M)

    notes = [
        "NOI from C85  ·  Escalation from C64  ·  Exit Cap from I20  ·  Hold from I13  ·  LTV 65%  ·  Rate I10  ·  Amortization I11",
        "Avg Cap Rate = avg (NOI×(1+esc)^t / Price) over hold period  ·  ERM = total cash returned / price (or equity)",
        "IRR computed via year-by-year helper rows below (LibreOffice-safe — no array literals).",
    ]
    for i, note_txt in enumerate(notes):
        rr = 17 + i
        ws.row_dimensions[rr].height = 14
        c = ws.cell(row=rr, column=2, value=note_txt)
        c.font = FT_NOTE; c.alignment = AL_L
        merge(ws, rr, 2, rr, NCOLS_DISP)

    # ══════════════════════════════════════════════════════════════════════════
    # HELPER BLOCK A: UNLEVERAGED CASH FLOWS  (rows 22–32, years 0–10)
    # One row per year, one column per cap rate scenario.
    # Year 10 includes Net Reversion (0.94 × exit NOI / exit cap).
    # ══════════════════════════════════════════════════════════════════════════
    ws.row_dimensions[21].height = 6
    sec(ws, 22, "HELPER DATA  —  Year-by-Year Cash Flows for IRR (do not edit)", col_start=2, ncols=12)
    ws.row_dimensions[22].height = 12

    UL_HELP_START = 23   # Year 0 (purchase price outflow)
    UL_HELP_END   = 33   # Year 10 (NOI + net reversion)

    for yr in range(11):
        rr = UL_HELP_START + yr
        ws.row_dimensions[rr].height = 13
        c = ws.cell(row=rr, column=2, value=f"UL  Yr {yr}")
        c.font = FT_NOTE; c.alignment = AL_L

        for i in range(11):
            col = 3 + i
            cl  = get_column_letter(col)
            cap_ref = f"{cl}{CAP_HDR_ROW}"

            if yr == 0:
                # CF = −Price = −NOI / cap_rate
                formula = (f'=IFERROR(IF(OR({_NOI}="",{cap_ref}=0),"",'
                           f'-{_NOI}/{cap_ref}),"")')
            elif yr < 10:
                # CF = NOI × (1+esc)^(yr−1)
                formula = (f'=IFERROR(IF({_NOI}="","",{_NOI}*(1+{_ESC})^{yr - 1}),"")')
            else:
                # CF = NOI × (1+esc)^9 + Net Reversion (94% of gross proceeds)
                formula = (f'=IFERROR(IF(OR({_NOI}="",{cap_ref}=0),"",{_NOI}*(1+{_ESC})^9'
                           f'+({_NOI}*(1+{_ESC})^9/{_XCAP})*0.94),"")')
            c = ws.cell(row=rr, column=col, value=formula)
            c.font = FT_NOTE; c.number_format = D0; c.alignment = AL_R

    # ══════════════════════════════════════════════════════════════════════════
    # HELPER BLOCK B: LEVERAGED CASH FLOWS  (rows 35–45, years 0–10)
    # Year 0 = −Equity; years 1–9 = CFADS; year 10 = CFADS + net rev after debt.
    # ══════════════════════════════════════════════════════════════════════════
    ws.row_dimensions[34].height = 6

    LEV_HELP_START = 35
    LEV_HELP_END   = 45

    for yr in range(11):
        rr = LEV_HELP_START + yr
        ws.row_dimensions[rr].height = 13
        c = ws.cell(row=rr, column=2, value=f"Lev Yr {yr}")
        c.font = FT_NOTE; c.alignment = AL_L

        for i in range(11):
            col = 3 + i
            cl  = get_column_letter(col)
            cap_ref = f"{cl}{CAP_HDR_ROW}"

            price_f  = f"({_NOI}/{cap_ref})"
            loan_f   = f"(0.65*{_NOI}/{cap_ref})"
            equity_f = f"(0.35*{_NOI}/{cap_ref})"
            # Annual DS (positive cash outflow): −PMT(rate/12, amort×12, loan) × 12
            ds_f = f"(-PMT({_RATE}/12,{_AMORT}*12,{loan_f})*12)"

            if yr == 0:
                # CF = −Equity
                formula = (f'=IFERROR(IF(OR({_NOI}="",{cap_ref}=0),"",-{equity_f}),"")')
            elif yr < 10:
                # CF = Annual NOI_yr − Annual DS
                noi_t = f"({_NOI}*(1+{_ESC})^{yr - 1})"
                formula = (f'=IFERROR(IF(OR({_NOI}="",{cap_ref}=0),"",{noi_t}-{ds_f}),"")')
            else:
                # CF = NOI_10 − DS + Net Reversion After Debt
                noi_10  = f"({_NOI}*(1+{_ESC})^9)"
                net_rev = f"({noi_10}/{_XCAP}*0.94)"
                # Remaining loan = PV of (amort−hold) remaining payments
                rem_loan = (f"(-PV({_RATE}/12,({_AMORT}-{_HOLD})*12,"
                            f"PMT({_RATE}/12,{_AMORT}*12,{loan_f})))")
                net_rev_after = f"({net_rev}-{rem_loan})"
                formula = (f'=IFERROR(IF(OR({_NOI}="",{cap_ref}=0),"",{noi_10}-{ds_f}+{net_rev_after}),"")')
            c = ws.cell(row=rr, column=col, value=formula)
            c.font = FT_NOTE; c.number_format = D0; c.alignment = AL_R

    # ══════════════════════════════════════════════════════════════════════════
    # FILL DISPLAY MATRIX CELLS — now that helper rows are defined
    # ══════════════════════════════════════════════════════════════════════════
    PALE_ROWS = {M1_ROW, M4_ROW}
    TOT_ROWS  = {M3_ROW, M6_ROW}

    for i in range(11):
        col     = 3 + i
        cl      = get_column_letter(col)
        cap_ref = f"{cl}{CAP_HDR_ROW}"
        is_ask  = (col == ASK_COL)

        price_f  = f"({_NOI}/{cap_ref})"
        loan_f   = f"(0.65*{_NOI}/{cap_ref})"
        equity_f = f"(0.35*{_NOI}/{cap_ref})"
        ds_f     = f"(-PMT({_RATE}/12,{_AMORT}*12,{loan_f})*12)"

        def _cell(rr, fmt_, formula_):
            c = frm(ws, rr, col, formula_, fmt=fmt_, align=AL_R)
            if is_ask:
                c.fill = _GOLD_VAL
            elif rr in PALE_ROWS:
                c.fill = F_PALE
            elif rr in TOT_ROWS:
                c.fill = F_TOT
            else:
                c.fill = F_WHITE
            if rr in TOT_ROWS:
                c.font = FT_TOTAL
            c.border = Border(bottom=_GY_T)
            return c

        # ── M1: Average Cash Cap Rate over hold ──────────────────────────────
        # = cap_rate × geometric escalation factor (handles esc=0 case)
        # = average of (NOI_t / Price) over t=1..hold
        m1 = (f'=IFERROR(IF(OR({_NOI}="",{cap_ref}=0),"",IF({_ESC}=0,{cap_ref},'
              f'{cap_ref}*(POWER(1+{_ESC},{_HOLD})-1)/({_ESC}*{_HOLD}))),"")')
        _cell(M1_ROW, P2, m1)

        # ── M2: Unleveraged ERM = SUM(UL CFs yr1-10) / Price ────────────────
        ul_pos = f"SUM({cl}{UL_HELP_START + 1}:{cl}{UL_HELP_END})"
        m2 = (f'=IFERROR(IF(OR({_NOI}="",{cap_ref}=0),"",({ul_pos})/{price_f}),"")')
        _cell(M2_ROW, MX, m2)

        # ── M3: Cash IRR (Unleveraged) = IRR over full UL helper range ───────
        ul_range = f"{cl}{UL_HELP_START}:{cl}{UL_HELP_END}"
        m3 = f'=IFERROR(IF(OR({_NOI}="",{cap_ref}=0),"",IRR({ul_range})),"")'
        _cell(M3_ROW, P2, m3)

        # ── M4: Leveraged Avg Cash-on-Cash = (Avg NOI − DS) / Equity ─────────
        noi_avg = (f'IF({_ESC}=0,{_NOI},{_NOI}*(POWER(1+{_ESC},{_HOLD})-1)/({_ESC}*{_HOLD}))')
        m4 = (f'=IFERROR(IF(OR({_NOI}="",{cap_ref}=0),"",({noi_avg}-{ds_f})/{equity_f}),"")')
        _cell(M4_ROW, P2, m4)

        # ── M5: Leveraged ERM = SUM(Lev CFs yr1-10) / Equity ─────────────────
        lev_pos = f"SUM({cl}{LEV_HELP_START + 1}:{cl}{LEV_HELP_END})"
        m5 = (f'=IFERROR(IF(OR({_NOI}="",{cap_ref}=0),"",({lev_pos})/{equity_f}),"")')
        _cell(M5_ROW, MX, m5)

        # ── M6: Leveraged IRR = IRR over full Lev helper range ───────────────
        lev_range = f"{cl}{LEV_HELP_START}:{cl}{LEV_HELP_END}"
        m6 = f'=IFERROR(IF(OR({_NOI}="",{cap_ref}=0),"",IRR({lev_range})),"")'
        _cell(M6_ROW, P2, m6)

    # ── Final bottom accent ────────────────────────────────────────────────────
    ws.row_dimensions[46].height = 4
    for ci in range(2, NCOLS_DISP + 1):
        ws.cell(row=46, column=ci).border = Border(bottom=Side(style='medium', color=NAVY))
