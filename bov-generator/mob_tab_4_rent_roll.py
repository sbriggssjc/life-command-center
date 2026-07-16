"""BOV MOB — Tab 4: Rent Roll (consolidated single-page tenant view).

This is a new tab specific to the MOB template. It provides a clean,
single-page snapshot of all tenants, their suites, SF, rent, lease terms,
and lease type — suitable for BOV, OM, or quick diligence review.

All data references pull from Assumptions & Flags (for computed fields)
or are input directly here (for lease dates and notes).

The tab is organized as:
  - Property header block (address, total GLA, occupancy rate)
  - Rent Roll table: one row per tenant (5 rows + 1 vacant placeholder)
    Columns: Suite | Tenant Name | Lease Type | SF | % GLA | Comm Date | Exp Date | Rem Term | Annual Rent | Mo Rent | Rent/SF | Escalation | Notes
  - Totals / weighted average row
  - Summary stats block: Gross Potential Rent, Vacancy, EGI, Avg Rent/SF
"""
from bov_constants import *
from openpyxl.styles import Border, Side, PatternFill, Font
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter

_NM_M = Side(style='medium', color=NAVY)
_NM_T = Side(style='thin',   color=NAVY)
_GY_T = Side(style='thin',   color="CCCCCC")

_AS = "Assumptions & Flags"

# Assumptions cell references for each tenant
# (name, suite, sf, rent, esc, leasetype, reimb)
TENANT_ASM = [
    (18, 19, 20, 21, 22, 23, 24),  # T1
    (27, 28, 29, 30, 31, 32, 33),  # T2
    (36, 37, 38, 39, 40, 41, 42),  # T3
    (45, 46, 47, 48, 49, 50, 51),  # T4
    (54, 55, 56, 57, 58, 59, 60),  # T5
]

# Columns for the rent roll table
# A=margin(2), B=Suite(10), C=Tenant Name(22), D=Lease Type(10), E=SF(10),
# F=% GLA(7), G=Comm Date(11), H=Exp Date(11), I=Rem Term(8),
# J=Annual Rent(13), K=Mo Rent(12), L=Rent/SF(8), M=Escalation(10), N=Notes(24)
COL_WIDTHS = [2, 10, 22, 10, 10, 7, 11, 11, 8, 13, 12, 8, 10, 24]


def build_rent_roll_tab(wb):
    ws = wb.create_sheet("Rent Roll")
    ws.sheet_view.showGridLines = False

    for i, wd in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = wd

    # ── Header ────────────────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 6
    ws.row_dimensions[2].height = 28
    c = ws.cell(row=2, column=2, value="RENT ROLL")
    c.font = FT_TITLE; c.alignment = AL_L
    merge(ws, 2, 2, 2, 14)

    ws.row_dimensions[3].height = 14
    c = ws.cell(row=3, column=2,
                value="Consolidated single-page tenant summary  ·  Lease dates and notes entered here  ·  Rent, SF, and escalation auto-fill from Assumptions & Flags")
    c.font = FT_NOTE; c.alignment = AL_L
    merge(ws, 3, 2, 3, 14)

    ws.row_dimensions[4].height = 8

    # ── Property header block ─────────────────────────────────────────────────
    sec(ws, 5, "PROPERTY IDENTIFICATION", col_start=2, ncols=13)

    prop_rows = [
        (6,  "Property Address",    f'=IFERROR(\'{_AS}\'!$C$6,"")'),
        (7,  "Total Building SF",   f'=IFERROR(\'{_AS}\'!$C$9,"")'),
        (8,  "Number of Tenants",   f'=IFERROR(COUNTA(\'{_AS}\'!$C$18,\'{_AS}\'!$C$27,\'{_AS}\'!$C$36,\'{_AS}\'!$C$45,\'{_AS}\'!$C$54),"")'),
        (9,  "Occupied SF",         f'=IFERROR(\'{_AS}\'!$C$65,"")'),
        (10, "Occupancy Rate",      f'=IFERROR(\'{_AS}\'!$C$66,"")'),
        (11, "Gross Potential Rent",f'=IFERROR(\'{_AS}\'!$C$79,"")'),
        (12, "Year 1 NOI",          f'=IFERROR(\'{_AS}\'!$C$85,"")'),
    ]
    fmts_prop = [None, N0, N0, N0, P2, D0, D0]
    for (rr, label, formula), fmt in zip(prop_rows, fmts_prop):
        ws.row_dimensions[rr].height = 18
        lbl(ws, rr, 2, label)
        merge(ws, rr, 2, rr, 5)
        if formula:
            c = frm(ws, rr, 6, formula, fmt=fmt, align=AL_L)
        else:
            c = inp(ws, rr, 6, fmt=fmt)
            c.alignment = AL_L
        merge(ws, rr, 6, rr, 10)
        add_cf_clear(ws, f"F{rr}:F{rr}")
        ws.cell(row=rr, column=2).border = Border(bottom=_GY_T)
        ws.cell(row=rr, column=6).border = Border(bottom=_GY_T)

    ws.row_dimensions[13].height = 8

    # ── Rent Roll table ───────────────────────────────────────────────────────
    sec(ws, 14, "RENT ROLL  —  Current as of Analysis Date", col_start=2, ncols=13)

    # Column headers (row 15)
    ws.row_dimensions[15].height = 22
    hdrs = [
        "SUITE", "TENANT NAME", "LEASE\nTYPE", "LEASED\nSF", "% GLA",
        "COMM\nDATE", "EXP\nDATE", "REM\nTERM", "ANNUAL\nRENT",
        "MO RENT", "RENT\n/SF", "ESC %", "NOTES"
    ]
    for ci, hdr in enumerate(hdrs, 2):
        c = ws.cell(row=15, column=ci, value=hdr)
        c.font = FT_CHDR; c.fill = F_NAVY; c.alignment = AL_C
        c.border = Border(bottom=_NM_M)

    DATA_START = 16

    # 5 tenant rows
    ROW_COLORS = [F_WHITE, F_PALE, F_WHITE, F_PALE, F_WHITE]
    for t_idx, (name_r, suite_r, sf_r, rent_r, esc_r, type_r, reimb_r) in enumerate(TENANT_ASM):
        rr = DATA_START + t_idx
        ws.row_dimensions[rr].height = 22
        bg = ROW_COLORS[t_idx]

        # B: Suite — links from Assumptions
        c = frm(ws, rr, 2,
                f'=IFERROR(\'{_AS}\'!$C${suite_r},"")',
                align=AL_C)
        c.fill = bg

        # C: Tenant Name — links from Assumptions
        c = frm(ws, rr, 3,
                f'=IFERROR(IF(\'{_AS}\'!$C${name_r}="","[Vacant]",\'{_AS}\'!$C${name_r}),"")',
                align=AL_L)
        c.fill = bg; c.font = FT_LABEL

        # D: Lease Type — links from Assumptions
        c = frm(ws, rr, 4,
                f'=IFERROR(\'{_AS}\'!$C${type_r},"")',
                align=AL_C)
        c.fill = bg

        # E: Leased SF — links from Assumptions
        c = frm(ws, rr, 5,
                f'=IFERROR(\'{_AS}\'!$C${sf_r},"")',
                fmt=N0, align=AL_R)
        c.fill = bg

        # F: % of GLA
        c = frm(ws, rr, 6,
                f'=IFERROR(IF(OR(\'{_AS}\'!$C${sf_r}="",\'{_AS}\'!$C$9="",\'{_AS}\'!$C$9=0),"",\'{_AS}\'!$C${sf_r}/\'{_AS}\'!$C$9),"")',
                fmt=P1, align=AL_C)
        c.fill = bg

        # G: Lease Commencement Date — input here (not in Assumptions)
        inp(ws, rr, 7, fmt=DT)
        ws.cell(row=rr, column=7).fill = bg
        ws.cell(row=rr, column=7).alignment = AL_C

        # H: Lease Expiration Date — input here
        inp(ws, rr, 8, fmt=DT)
        ws.cell(row=rr, column=8).fill = bg
        ws.cell(row=rr, column=8).alignment = AL_C

        # I: Remaining Term (formula: (Exp - TODAY) / 365.25)
        c = frm(ws, rr, 9,
                f'=IFERROR(IF(H{rr}="","",(H{rr}-TODAY())/365.25),"")',
                fmt='0.0', align=AL_C)
        c.fill = bg

        # J: Annual Rent — links from Assumptions
        c = frm(ws, rr, 10,
                f'=IFERROR(\'{_AS}\'!$C${rent_r},"")',
                fmt=D0, align=AL_R)
        c.fill = bg; c.font = FT_TOTAL

        # K: Monthly Rent (Annual / 12)
        c = frm(ws, rr, 11,
                f'=IFERROR(IF(J{rr}="","",J{rr}/12),"")',
                fmt=D0, align=AL_R)
        c.fill = bg

        # L: Rent / SF
        c = frm(ws, rr, 12,
                f'=IFERROR(IF(OR(J{rr}="",E{rr}="",E{rr}=0),"",J{rr}/E{rr}),"")',
                fmt=D0, align=AL_R)
        c.fill = bg

        # M: Escalation %
        c = frm(ws, rr, 13,
                f'=IFERROR(\'{_AS}\'!$C${esc_r},"")',
                fmt=P2, align=AL_C)
        c.fill = bg

        # N: Notes — input
        inp(ws, rr, 14)
        ws.cell(row=rr, column=14).fill = bg
        ws.cell(row=rr, column=14).alignment = AL_TL

        for ci in range(2, 15):
            ws.cell(row=rr, column=ci).border = Border(bottom=_GY_T)

    # CF: clear yellow when comm/exp dates are filled
    for t_idx in range(5):
        rr = DATA_START + t_idx
        add_cf_clear(ws, f"G{rr}:G{rr}")
        add_cf_clear(ws, f"H{rr}:H{rr}")
        add_cf_clear(ws, f"N{rr}:N{rr}")

    DATA_END = DATA_START + 4  # 5 tenants = rows 16-20

    # ── Separator ─────────────────────────────────────────────────────────────
    ws.row_dimensions[DATA_END + 1].height = 4
    for ci in range(2, 15):
        ws.cell(row=DATA_END + 1, column=ci).border = Border(top=_NM_M)

    # ── Totals row ────────────────────────────────────────────────────────────
    tot_r = DATA_END + 2
    ws.row_dimensions[tot_r].height = 20
    c = ws.cell(row=tot_r, column=2, value="TOTALS / AVERAGES")
    c.font = FT_TOTAL; c.fill = F_TOT; c.alignment = AL_L
    merge(ws, tot_r, 2, tot_r, 4)

    # SF total
    c = frm(ws, tot_r, 5,
            f'=IFERROR(SUM(E{DATA_START}:E{DATA_END}),"")',
            fmt=N0, align=AL_R)
    c.font = FT_TOTAL; c.fill = F_TOT

    # % GLA (should = occupancy rate)
    c = frm(ws, tot_r, 6,
            f'=IFERROR(\'{_AS}\'!$C$66,"")',
            fmt=P1, align=AL_C)
    c.font = FT_TOTAL; c.fill = F_TOT

    # Skip date cols
    for ci in [7, 8, 9]:
        ws.cell(row=tot_r, column=ci).fill = F_TOT

    # Annual Rent total = GPR
    c = frm(ws, tot_r, 10,
            f'=IFERROR(\'{_AS}\'!$C$79,"")',
            fmt=D0, align=AL_R)
    c.font = FT_TOTAL; c.fill = F_TOT

    # Monthly total
    c = frm(ws, tot_r, 11,
            f'=IFERROR(IF(J{tot_r}="","",J{tot_r}/12),"")',
            fmt=D0, align=AL_R)
    c.font = FT_TOTAL; c.fill = F_TOT

    # Avg Rent/SF
    c = frm(ws, tot_r, 12,
            f'=IFERROR(IF(OR(\'{_AS}\'!$C$79="",\'{_AS}\'!$C$65="",\'{_AS}\'!$C$65=0),"",\'{_AS}\'!$C$79/\'{_AS}\'!$C$65),"")',
            fmt=D0, align=AL_R)
    c.font = FT_TOTAL; c.fill = F_TOT

    for ci in [13, 14]:
        ws.cell(row=tot_r, column=ci).fill = F_TOT

    for ci in range(2, 15):
        ws.cell(row=tot_r, column=ci).border = Border(top=_NM_M, bottom=_NM_M)

    # ── Summary stats block ───────────────────────────────────────────────────
    stat_r = tot_r + 2
    ws.row_dimensions[stat_r].height = 16
    sec(ws, stat_r, "NOI SUMMARY  —  From Assumptions & Flags", col_start=2, ncols=13)

    stats = [
        ("Gross Potential Rent",    f'=IFERROR(\'{_AS}\'!$C$79,"")', D0),
        ("Vacancy / Credit Loss",   f'=IFERROR(\'{_AS}\'!$C$80,"")', D0),
        ("Effective Base Rent",     f'=IFERROR(\'{_AS}\'!$C$81,"")', D0),
        ("Total Reimbursements",    f'=IFERROR(\'{_AS}\'!$C$82,"")', D0),
        ("Effective Gross Income",  f'=IFERROR(\'{_AS}\'!$C$83,"")', D0),
        ("Total Expenses",          f'=IFERROR(\'{_AS}\'!$C$84,"")', D0),
        ("NET OPERATING INCOME",    f'=IFERROR(\'{_AS}\'!$C$85,"")', D0),
    ]
    for i, (label, formula, fmt) in enumerate(stats):
        rr = stat_r + 1 + i
        ws.row_dimensions[rr].height = 18
        bg = F_PALE if i % 2 == 0 else F_WHITE
        is_total = (label == "NET OPERATING INCOME")
        lbl(ws, rr, 2, label, bold=is_total)
        merge(ws, rr, 2, rr, 5)
        for ci in [2, 3, 4, 5]:
            ws.cell(row=rr, column=ci).fill = F_TOT if is_total else bg
        c = frm(ws, rr, 6, formula, fmt=fmt, align=AL_R)
        c.fill = F_TOT if is_total else bg
        if is_total:
            c.font = FT_TOTAL
            c.border = Border(top=_NM_M, bottom=_NM_M)
        merge(ws, rr, 6, rr, 9)
        for ci in [6, 7, 8, 9]:
            ws.cell(row=rr, column=ci).fill = F_TOT if is_total else bg
        ws.cell(row=rr, column=2).border = Border(bottom=_GY_T)

    # NM bottom accent
    bot_r = stat_r + len(stats) + 2
    ws.row_dimensions[bot_r].height = 4
    for ci in range(2, 15):
        ws.cell(row=bot_r, column=ci).border = Border(bottom=Side(style='medium', color=NAVY))

    # Notes
    note_r = bot_r + 1
    notes = [
        "Lease Commencement and Expiration dates are entered directly on this tab — they do not pull from Assumptions.",
        "Suite, Tenant, SF, Rent, Escalation, and Lease Type auto-fill from Assumptions & Flags tab.",
        "Remaining Term formula: (Expiration Date − TODAY) ÷ 365.25 — updates daily.",
    ]
    for i, note in enumerate(notes):
        rr = note_r + i
        ws.row_dimensions[rr].height = 14
        c = ws.cell(row=rr, column=2, value=f"▪  {note}")
        c.font = FT_NOTE; c.alignment = AL_L
        merge(ws, rr, 2, rr, 14)
