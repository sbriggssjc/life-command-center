"""BOV v2 — Tab 4: Lease Abstract (Leg 2).
Layout: Short-form summary on TOP, long-form abstract BELOW.
Both sections share the same 7-column grid (A–G).
"""
from bov_constants import *
from openpyxl.styles import Border, Side
from openpyxl.utils import get_column_letter

_NM_M = Side(style='medium', color=NAVY)
_NM_T = Side(style='thin',   color="CCCCCC")


def build_lease_abstract_tab(wb):
    ws = wb.create_sheet("Lease Abstract")
    ws.sheet_view.showGridLines = False

    # ── Unified column widths (shared by short-form and long-form) ─────────────
    # A(1)=margin  B(2)=#  C(3)=article-section/provision-label
    # D(4)=clause-provision/SF-summary-anchor  E(5)=page(narrow)
    # F(6)=lease section ref  G(7)=document source  H(8)=operative language
    col_widths = [2, 5, 26, 26, 8, 16, 18, 40]
    for i, wd in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = wd

    # ── Tab header ────────────────────────────────────────────────────────────
    r = 1;  ws.row_dimensions[r].height = 6
    r = 2;  ws.row_dimensions[r].height = 28
    c = ws.cell(row=r, column=2, value="LEASE ABSTRACT")
    c.font = FT_TITLE; c.alignment = AL_L
    merge(ws, r, 2, r, 8)

    r = 3;  ws.row_dimensions[r].height = 14
    c = ws.cell(row=r, column=2,
                value="Leg 2 of 3  ·  Executed lease terms, obligations, and economics  ·  Short-form summary at top  ·  Full article-by-article abstract below")
    c.font = FT_NOTE; c.alignment = AL_L
    merge(ws, r, 2, r, 8)

    r = 4;  ws.row_dimensions[r].height = 8

    # ══════════════════════════════════════════════════════════════════════════
    # SHORT-FORM SUMMARY (TOP)
    # ══════════════════════════════════════════════════════════════════════════
    SF_HDR = 5
    ws.row_dimensions[SF_HDR].height = 16
    sec(ws, SF_HDR, "SHORT-FORM SUMMARY  —  For BOV, OM, or Valuation Memo", col_start=2, ncols=7)

    SF_COL_HDR = 6
    ws.row_dimensions[SF_COL_HDR].height = 20
    for ci in [2, 3]:
        ws.cell(row=SF_COL_HDR, column=ci).fill = F_NAVY
        ws.cell(row=SF_COL_HDR, column=ci).border = Border(bottom=_NM_M)
    c = ws.cell(row=SF_COL_HDR, column=3, value="PROVISION")
    c.font = FT_CHDR; c.alignment = AL_C
    c2 = ws.cell(row=SF_COL_HDR, column=4, value="ABBREVIATED SUMMARY")
    c2.font = FT_CHDR; c2.fill = F_NAVY; c2.alignment = AL_C
    c2.border = Border(bottom=_NM_M)
    for ci in [5, 6, 7, 8]:
        ws.cell(row=SF_COL_HDR, column=ci).fill = F_NAVY
        ws.cell(row=SF_COL_HDR, column=ci).border = Border(bottom=_NM_M)
    merge(ws, SF_COL_HDR, 4, SF_COL_HDR, 8)

    sf_rows = [
        "Tenant (Lease)",
        "Guarantor",
        "Lease Type",
        "Lease Commencement",
        "Lease Expiration",
        "Remaining Lease Term",
        "Year 1 Base Rent",
        "Rent Escalations",
        "Renewal Options",
        "Renewal Rent Method",
        "Expense Structure",
        "LL Responsibilities",
        "Early Termination",
        "Assignment / Subletting",
        "ROFR / ROFO",
        "Option to Purchase",
        "Key Lease Strengths",
        "Key Lease Risks",
        "Broker Commentary",
    ]

    SF_DATA_START = SF_COL_HDR + 1  # row 7
    for idx, prov in enumerate(sf_rows):
        rr = SF_DATA_START + idx
        ws.row_dimensions[rr].height = 36
        ws.cell(row=rr, column=2).fill = F_WHITE
        c = ws.cell(row=rr, column=3, value=prov)
        c.font = FT_LABEL; c.alignment = AL_TL
        # D:H merged: wide summary input
        inp(ws, rr, 4)
        ws.cell(row=rr, column=4).alignment = AL_TL
        merge(ws, rr, 4, rr, 8)
        for ci in range(2, 9):
            ws.cell(row=rr, column=ci).border = Border(bottom=_NM_T)

    SF_DATA_END = SF_DATA_START + len(sf_rows) - 1
    add_cf_clear(ws, f"D{SF_DATA_START}:D{SF_DATA_END}")

    # SF bottom accent
    sf_bot = SF_DATA_END + 1
    ws.row_dimensions[sf_bot].height = 4
    for ci in range(2, 9):
        ws.cell(row=sf_bot, column=ci).border = Border(top=_NM_M)

    sep_r = sf_bot + 1
    ws.row_dimensions[sep_r].height = 14

    # ══════════════════════════════════════════════════════════════════════════
    # LONG-FORM ABSTRACT (BELOW)
    # ══════════════════════════════════════════════════════════════════════════
    LF_HDR = sep_r + 1
    ws.row_dimensions[LF_HDR].height = 16
    sec(ws, LF_HDR, "LONG-FORM ABSTRACT  —  Article / clause reference with source document and operative language", col_start=2, ncols=7)

    LF_COL_HDR = LF_HDR + 1
    ws.row_dimensions[LF_COL_HDR].height = 20
    lf_hdrs = ["#", "ARTICLE / SECTION", "CLAUSE / PROVISION", "PAGE", "LEASE SECTION", "DOCUMENT SOURCE", "OPERATIVE LANGUAGE"]
    for ci, hdr in enumerate(lf_hdrs, 2):
        c = ws.cell(row=LF_COL_HDR, column=ci, value=hdr)
        c.font = FT_CHDR; c.fill = F_NAVY; c.alignment = AL_C
        c.border = Border(bottom=_NM_M)

    # (article_section, clause_provision)
    long_rows = [
        ("Lease Identification",    "Lease Type / Form"),
        ("Lease Identification",    "Execution Date"),
        ("Lease Identification",    "Effective / Commencement Date"),
        ("Lease Identification",    "Lease Expiration Date"),
        ("Lease Identification",    "Landlord of Record"),
        ("Lease Identification",    "Tenant of Record"),
        ("Lease Identification",    "Guarantor (if any)"),
        (None, None),
        ("Premises",                "Demised Premises Address"),
        ("Premises",                "Leased SF (per lease)"),
        ("Premises",                "Leased SF (per survey)"),
        ("Premises",                "Exclusive Use / Permitted Use"),
        ("Premises",                "Prohibited Uses"),
        (None, None),
        ("Rent",                    "Base Rent — Year 1"),
        ("Rent",                    "Annual Rent Escalations"),
        ("Rent",                    "Rent Commencement Date"),
        ("Rent",                    "Rent Abatement / Free Rent"),
        ("Rent",                    "Percentage Rent"),
        (None, None),
        ("Expense Structure",       "Lease Structure (NNN / NN / Gross / MG)"),
        ("Expense Structure",       "Real Estate Taxes — Responsibility"),
        ("Expense Structure",       "Insurance — Responsibility"),
        ("Expense Structure",       "CAM / Maintenance — Responsibility"),
        ("Expense Structure",       "Capital / Roof / Structure — Responsibility"),
        ("Expense Structure",       "Expense Cap (if any)"),
        ("Expense Structure",       "Landlord Obligations"),
        (None, None),
        ("Options & Renewals",      "Number of Renewal Options"),
        ("Options & Renewals",      "Option Term Length"),
        ("Options & Renewals",      "Renewal Rent — Method"),
        ("Options & Renewals",      "Renewal Notice Requirement"),
        ("Options & Renewals",      "Option to Purchase"),
        ("Options & Renewals",      "Right of First Refusal (ROFR)"),
        ("Options & Renewals",      "Right of First Offer (ROFO)"),
        (None, None),
        ("Assignment & Subletting", "Assignment Rights"),
        ("Assignment & Subletting", "Subletting Rights"),
        ("Assignment & Subletting", "Change of Control Provisions"),
        ("Assignment & Subletting", "Release of Guarantor on Assignment"),
        (None, None),
        ("Termination & Default",   "Early Termination Right"),
        ("Termination & Default",   "Termination Fee / Penalty"),
        ("Termination & Default",   "Default / Cure Periods"),
        ("Termination & Default",   "Co-Tenancy Provisions"),
        ("Termination & Default",   "Go-Dark Provision"),
        (None, None),
        ("Other Provisions",        "TI Allowance / Landlord Work"),
        ("Other Provisions",        "Signage Rights"),
        ("Other Provisions",        "Parking Allocation"),
        ("Other Provisions",        "Subordination / SNDA / Estoppel"),
        ("Other Provisions",        "Condemnation Provisions"),
        ("Other Provisions",        "Casualty / Damage Provisions"),
        ("Other Provisions",        "Holdover Provisions"),
        ("Other Provisions",        "Notices"),
    ]

    LF_DATA_START = LF_COL_HDR + 1
    row_num = 0
    current_art = None
    data_row_num = 0

    for entry in long_rows:
        rr = LF_DATA_START + row_num
        art, clause = entry
        if art is None:
            ws.row_dimensions[rr].height = 6
            row_num += 1
            continue
        ws.row_dimensions[rr].height = 52   # taller for operative language
        data_row_num += 1

        # B: row #
        c = ws.cell(row=rr, column=2, value=data_row_num)
        c.font = FT_NOTE; c.alignment = AL_C

        # C: article / section (shown on first row of each group)
        if art != current_art:
            c2 = ws.cell(row=rr, column=3, value=art)
            c2.font = FT_LABEL; c2.alignment = AL_TL
            current_art = art
        else:
            ws.cell(row=rr, column=3).alignment = AL_TL

        # D: clause / provision
        c3 = ws.cell(row=rr, column=4, value=clause)
        c3.font = FT_DATA; c3.alignment = AL_TL

        # E: page ref (input, centered — short entry)
        inp(ws, rr, 5); ws.cell(row=rr, column=5).alignment = AL_C

        # F: lease section ref (input, e.g. "Section 1.1(i)(a)-(c)")
        inp(ws, rr, 6); ws.cell(row=rr, column=6).alignment = AL_C

        # G: document source (input)
        inp(ws, rr, 7); ws.cell(row=rr, column=7).alignment = AL_TL

        # H: operative language / finding (input — widest)
        inp(ws, rr, 8); ws.cell(row=rr, column=8).alignment = AL_TL

        for ci in range(2, 9):
            ws.cell(row=rr, column=ci).border = Border(bottom=_NM_T)

        row_num += 1

    LF_DATA_END = LF_DATA_START + row_num - 1

    # CF clear on page, lease section, source, operative language columns
    for col in [5, 6, 7, 8]:
        cl = get_column_letter(col)
        add_cf_clear(ws, f"{cl}{LF_DATA_START}:{cl}{LF_DATA_END}")

    # Medium bottom border
    lf_bot = LF_DATA_END + 1
    ws.row_dimensions[lf_bot].height = 4
    for ci in range(2, 9):
        ws.cell(row=lf_bot, column=ci).border = Border(top=_NM_M)

    # Note row
    note_r = lf_bot + 1
    ws.row_dimensions[note_r].height = 14
    c = ws.cell(row=note_r, column=2,
                value="Page: lease page #  ·  Lease Section: cite the specific section ref (e.g. 'Section 1.1(i)(a)')  ·  Source: 'Executed Lease', 'Amendment No. 1', etc.  ·  Operative Language: quote or paraphrase the actual lease provision")
    c.font = FT_NOTE; c.alignment = AL_L
    merge(ws, note_r, 2, note_r, 8)

    # NM bottom accent
    bot_r = note_r + 2
    ws.row_dimensions[bot_r].height = 4
    for ci in range(2, 9):
        ws.cell(row=bot_r, column=ci).border = Border(bottom=_NM_M)
