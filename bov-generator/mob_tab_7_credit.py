"""BOV MOB — Tab 7: Tenant / Guarantor Credit (Leg 3).
Layout: Short-form summary table at TOP (5 tenant columns), then 5 per-tenant
long-form credit sections stacked below.
"""
from bov_constants import *
from openpyxl.styles import Border, Side, Font, PatternFill
from openpyxl.utils import get_column_letter

_NM_M = Side(style='medium', color=NAVY)
_NM_T = Side(style='thin',   color="CCCCCC")

AS = "'Assumptions & Flags'"

TENANT_MAP = [
    {"label": "TENANT 1", "name": "C18", "suite": "C19", "sf": "C20"},
    {"label": "TENANT 2", "name": "C27", "suite": "C28", "sf": "C29"},
    {"label": "TENANT 3", "name": "C36", "suite": "C37", "sf": "C38"},
    {"label": "TENANT 4", "name": "C45", "suite": "C46", "sf": "C47"},
    {"label": "TENANT 5", "name": "C54", "suite": "C55", "sf": "C56"},
]

# Short-form items (one value per tenant column)
SF_ITEMS = [
    "Tenant / Operator",
    "Parent Company",
    "Public / Private",
    "Credit Rating",
    "Investment Grade (Y/N)",
    "Total Locations",
    "Annual Revenue",
    "EBITDA / Margin",
    "Rent-to-Sales Ratio",
    "Guarantor",
    "Guaranty Type",
    "Guaranty Strength",
    "Essential / Recession-Resistant",
    "Key Credit Strengths",
    "Key Credit Risks",
]

# Long-form rows per tenant: (category, attribute)
LONG_ROWS = [
    ("Corporate Overview",      "Entity Name (Lease)"),
    ("Corporate Overview",      "Entity Name (Operating / Trade)"),
    ("Corporate Overview",      "Ownership Structure"),
    ("Corporate Overview",      "Parent Company"),
    ("Corporate Overview",      "Headquarters"),
    ("Corporate Overview",      "Founded"),
    ("Corporate Overview",      "Number of Locations (Total)"),
    ("Corporate Overview",      "Number of Locations (State)"),
    ("Corporate Overview",      "Business Description"),
    ("Corporate Overview",      "Years in Operation"),
    (None, None),
    ("Credit & Ratings",        "S&P Credit Rating"),
    ("Credit & Ratings",        "Moody's Credit Rating"),
    ("Credit & Ratings",        "Investment Grade (Y/N)"),
    ("Credit & Ratings",        "Public / Private"),
    ("Credit & Ratings",        "Stock Ticker (if public)"),
    ("Credit & Ratings",        "Market Capitalization"),
    (None, None),
    ("Financial Summary",       "Revenue — Most Recent FY"),
    ("Financial Summary",       "Revenue — Prior FY"),
    ("Financial Summary",       "Revenue Growth YoY"),
    ("Financial Summary",       "EBITDA — Most Recent FY"),
    ("Financial Summary",       "EBITDA Margin"),
    ("Financial Summary",       "Net Income — Most Recent FY"),
    ("Financial Summary",       "Total Debt"),
    ("Financial Summary",       "Net Worth / Book Value"),
    ("Financial Summary",       "Debt / EBITDA"),
    ("Financial Summary",       "Source / Reporting Period"),
    (None, None),
    ("Guaranty",                "Guarantor Name"),
    ("Guaranty",                "Guarantor Type (Corporate / Personal)"),
    ("Guaranty",                "Guaranty Type (Full / Partial / Springing / Burn-off)"),
    ("Guaranty",                "Guaranty Cap ($)"),
    ("Guaranty",                "Burn-off / Release Conditions"),
    ("Guaranty",                "Guarantor Net Worth"),
    ("Guaranty",                "Guaranty Expiration"),
    (None, None),
    ("Qualitative",             "Recession Resistance / Essential Services"),
    ("Qualitative",             "Industry / Sector Trends"),
    ("Qualitative",             "Key Risks to Tenancy"),
    ("Qualitative",             "Key Credit Strengths"),
    ("Qualitative",             "Online / Omnichannel Exposure"),
    ("Qualitative",             "Broker Commentary"),
]


def _credit_section(ws, current_row, tenant):
    """Write one per-tenant long-form credit section. Returns next available row."""

    # ── Teal tenant section header ────────────────────────────────────────────
    ws.row_dimensions[current_row].height = 20
    hdr_formula = (
        f'=IFERROR(IF({AS}!{tenant["name"]}="",'
        f'"{tenant["label"]} — Credit Overview",'
        f'"CREDIT — "&{AS}!{tenant["name"]}),"{tenant["label"]} — Credit Overview")'
    )
    frm(ws, current_row, 2, hdr_formula)
    ws.cell(row=current_row, column=2).font      = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    ws.cell(row=current_row, column=2).fill      = PatternFill("solid", fgColor="1F7A8C")
    ws.cell(row=current_row, column=2).alignment = AL_L
    merge(ws, current_row, 2, current_row, 7)
    current_row += 1

    # ── Long-form rows ────────────────────────────────────────────────────────
    sec_data_start = current_row
    row_num     = 0
    data_rnum   = 0
    current_cat = None

    for entry in LONG_ROWS:
        rr = sec_data_start + row_num
        cat, attr = entry
        if cat is None:
            ws.row_dimensions[rr].height = 6
            row_num += 1
            continue
        ws.row_dimensions[rr].height = 36
        data_rnum += 1

        # B: row number
        c = ws.cell(row=rr, column=2, value=data_rnum)
        c.font = FT_NOTE; c.alignment = AL_C

        # C: category (first row of each group)
        if cat != current_cat:
            c2 = ws.cell(row=rr, column=3, value=cat)
            c2.font = FT_LABEL; c2.alignment = AL_TL
            current_cat = cat
        else:
            ws.cell(row=rr, column=3).alignment = AL_TL

        # D: attribute
        c3 = ws.cell(row=rr, column=4, value=attr)
        c3.font = FT_DATA; c3.alignment = AL_TL

        # E: finding / detail (input)
        inp(ws, rr, 5); ws.cell(row=rr, column=5).alignment = AL_TL

        # F: source (input)
        inp(ws, rr, 6); ws.cell(row=rr, column=6).alignment = AL_TL

        # G: notes / flags (input)
        inp(ws, rr, 7); ws.cell(row=rr, column=7).alignment = AL_TL

        for ci in range(2, 8):
            ws.cell(row=rr, column=ci).border = Border(bottom=_NM_T)

        row_num += 1

    sec_data_end = sec_data_start + row_num - 1

    # CF clear on input columns
    for col in [5, 6, 7]:
        cl = get_column_letter(col)
        add_cf_clear(ws, f"{cl}{sec_data_start}:{cl}{sec_data_end}")

    current_row = sec_data_end + 1

    # Bottom accent
    ws.row_dimensions[current_row].height = 4
    for ci in range(2, 8):
        ws.cell(row=current_row, column=ci).border = Border(top=_NM_M)
    current_row += 1

    return current_row


def build_mob_credit_tab(wb):
    ws = wb.create_sheet("Credit")
    ws.sheet_view.showGridLines = False

    # ── Column widths ─────────────────────────────────────────────────────────
    # A=margin  B=item/cat  C-G=5 tenant cols (short-form) / detail cols (long-form)
    col_widths = [2, 26, 18, 18, 18, 18, 18]
    for i, wd in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = wd

    # ── Tab header ────────────────────────────────────────────────────────────
    r = 1;  ws.row_dimensions[r].height = 6
    r = 2;  ws.row_dimensions[r].height = 28
    c = ws.cell(row=r, column=2, value="TENANT / GUARANTOR CREDIT  —  Multi-Tenant")
    c.font = FT_TITLE; c.alignment = AL_L
    merge(ws, r, 2, r, 7)

    r = 3;  ws.row_dimensions[r].height = 14
    c = ws.cell(row=r, column=2,
                value="Leg 3 of 3  ·  Corporate overview · financial strength · guaranty analysis  ·  5-tenant summary at top  ·  Full per-tenant credit matrix below")
    c.font = FT_NOTE; c.alignment = AL_L
    merge(ws, r, 2, r, 7)

    r = 4;  ws.row_dimensions[r].height = 8

    # ══════════════════════════════════════════════════════════════════════════
    # SHORT-FORM SUMMARY TABLE (TOP)
    # ══════════════════════════════════════════════════════════════════════════
    SF_HDR = 5
    ws.row_dimensions[SF_HDR].height = 16
    sec(ws, SF_HDR, "SHORT-FORM SUMMARY  —  Key Credit Metrics Across All Tenants", col_start=2, ncols=6)

    SF_COL_HDR = 6
    ws.row_dimensions[SF_COL_HDR].height = 24

    # ITEM label
    c = ws.cell(row=SF_COL_HDR, column=2, value="ITEM")
    c.font = FT_CHDR; c.fill = F_NAVY; c.alignment = AL_C
    c.border = Border(bottom=_NM_M)

    # Tenant name columns
    for ti, t in enumerate(TENANT_MAP):
        col = 3 + ti
        frm(ws, SF_COL_HDR, col,
            f'=IFERROR(IF({AS}!{t["name"]}="","{t["label"]}",{AS}!{t["name"]}),"{t["label"]}")')
        ws.cell(row=SF_COL_HDR, column=col).font      = FT_CHDR
        ws.cell(row=SF_COL_HDR, column=col).fill      = F_NAVY
        ws.cell(row=SF_COL_HDR, column=col).alignment = AL_C
        ws.cell(row=SF_COL_HDR, column=col).border    = Border(bottom=_NM_M)

    SF_DATA_START = SF_COL_HDR + 1
    for idx, item in enumerate(SF_ITEMS):
        rr = SF_DATA_START + idx
        ws.row_dimensions[rr].height = 32
        c = ws.cell(row=rr, column=2, value=item)
        c.font = FT_LABEL; c.alignment = AL_TL
        ws.cell(row=rr, column=2).border = Border(bottom=_NM_T)
        for col in range(3, 8):
            inp(ws, rr, col)
            ws.cell(row=rr, column=col).alignment = AL_TL
            ws.cell(row=rr, column=col).border    = Border(bottom=_NM_T)

    SF_DATA_END = SF_DATA_START + len(SF_ITEMS) - 1
    for col in range(3, 8):
        cl = get_column_letter(col)
        add_cf_clear(ws, f"{cl}{SF_DATA_START}:{cl}{SF_DATA_END}")

    sf_bot = SF_DATA_END + 1
    ws.row_dimensions[sf_bot].height = 4
    for ci in range(2, 8):
        ws.cell(row=sf_bot, column=ci).border = Border(top=_NM_M)

    sep_r = sf_bot + 1
    ws.row_dimensions[sep_r].height = 14

    # ══════════════════════════════════════════════════════════════════════════
    # PER-TENANT LONG-FORM CREDIT SECTIONS (STACKED)
    # ══════════════════════════════════════════════════════════════════════════
    LF_HDR = sep_r + 1
    ws.row_dimensions[LF_HDR].height = 16
    sec(ws, LF_HDR, "LONG-FORM CREDIT OVERVIEW  —  Corporate, financial, and guaranty diligence per tenant", col_start=2, ncols=6)

    # Column header for long-form sections
    LF_COL_HDR = LF_HDR + 1
    ws.row_dimensions[LF_COL_HDR].height = 20
    lf_hdrs = ["#", "CATEGORY", "ATTRIBUTE", "FINDING / DETAIL", "SOURCE", "NOTES / FLAGS"]
    for ci, hdr in enumerate(lf_hdrs, 2):
        c = ws.cell(row=LF_COL_HDR, column=ci, value=hdr)
        c.font = FT_CHDR; c.fill = F_NAVY; c.alignment = AL_C
        c.border = Border(bottom=_NM_M)

    current_row = LF_COL_HDR + 1

    for ti, tenant in enumerate(TENANT_MAP):
        current_row = _credit_section(ws, current_row, tenant)
        if ti < len(TENANT_MAP) - 1:
            ws.row_dimensions[current_row].height = 10
            current_row += 1

    # ── Note row ──────────────────────────────────────────────────────────────
    note_r = current_row + 1
    ws.row_dimensions[note_r].height = 14
    c = ws.cell(row=note_r, column=2,
                value="Source: SEC filing, S&P Global, Moody's, company press release, FDD, or broker research  ·  Flag any estimated or unverified figures in Notes column")
    c.font = FT_NOTE; c.alignment = AL_L
    merge(ws, note_r, 2, note_r, 7)

    bot_r = note_r + 2
    ws.row_dimensions[bot_r].height = 4
    for ci in range(2, 8):
        ws.cell(row=bot_r, column=ci).border = Border(bottom=_NM_M)
