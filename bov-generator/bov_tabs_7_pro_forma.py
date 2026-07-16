"""BOV v2 — Tab 7: Pro Forma (10-year investment model).

Cell reference map — Assumptions & Flags tab rows used here:
  Left-side (col C):
    C9  = Building SF
    C13 = Close Date
    C24 = Year 1 Base Rent
    C25 = Annual Rent Escalation %
    C26 = Tenant Reimbursements
    C30 = Mgmt Fee %
    C31 = Capital / Replacement Reserves
    C33 = Estimated NOI Year 1
    C36 = Purchase Price
  Right-side (col I):
    I9  = Down Payment / Equity (formula)
    I14 = Annual Debt Service (formula)
    I20 = Exit Cap Rate
    I25 = Net Reversion (Unleveraged)
    I26 = Remaining Loan Balance
    I27 = Net Reversion After Debt
"""
from bov_constants import *
from openpyxl.styles import Border, Side, Font, PatternFill
from openpyxl.utils import get_column_letter

_NM_M = Side(style='medium', color=NAVY)
_NM_T = Side(style='thin',   color=NAVY)
_GY_T = Side(style='thin',   color="CCCCCC")
_AS   = "Assumptions & Flags"

# Grid: col 1=labels(30), col 2=spacer(1), cols 3-13=Year0-Year10(11 cols), col 14=notes(18)
LABEL_COL = 1
NCOLS     = 14
# Col 3 = Year 0 / Purchase; cols 4-13 = Years 1-10
YR0_COL   = 3
YR_COLS   = list(range(4, 14))   # cols 4-13 = Year 1-10


def build_pro_forma_tab(wb):
    ws = wb.create_sheet("Pro Forma")
    ws.sheet_view.showGridLines = False

    # ── Column widths ──────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 1.5
    ws.column_dimensions["C"].width = 12   # Year 0 / Purchase
    for col in range(4, 14):
        ws.column_dimensions[get_column_letter(col)].width = 11
    ws.column_dimensions["N"].width = 20

    # ══════════════════════════════════════════════════════════════════════════
    # HEADER
    # ══════════════════════════════════════════════════════════════════════════
    r = 1
    ws.row_dimensions[r].height = 6
    r = 2
    ws.row_dimensions[r].height = 28
    c = ws.cell(row=r, column=1, value="PRO FORMA  —  10-YEAR INVESTMENT MODEL")
    c.font = FT_TITLE; c.alignment = AL_L
    merge(ws, r, 1, r, NCOLS)

    r = 3
    ws.row_dimensions[r].height = 14
    c = ws.cell(row=r, column=1,
                value="Single-Tenant NNN  ·  All inputs on Assumptions & Flags tab  ·  Col C = Year 0 (purchase); Cols D–M = Operating Years 1–10")
    c.font = FT_NOTE; c.alignment = AL_L
    merge(ws, r, 1, r, NCOLS)

    r = 4
    ws.row_dimensions[r].height = 8

    # ── Column headers row ────────────────────────────────────────────────────
    r = 5
    ws.row_dimensions[r].height = 18
    # fill all cols with NAVY
    for ci in range(1, NCOLS + 1):
        ws.cell(row=r, column=ci).fill = F_NAVY
    ws.cell(row=r, column=1, value="").font = FT_CHDR
    c = ws.cell(row=r, column=YR0_COL, value="Year 0 / Purchase")
    c.font = FT_CHDR; c.fill = F_NAVY; c.alignment = AL_C
    for i, col in enumerate(YR_COLS):
        c = ws.cell(row=r, column=col, value=f"Year {i+1}")
        c.font = FT_CHDR; c.fill = F_NAVY; c.alignment = AL_C

    # ── Year-end date row ─────────────────────────────────────────────────────
    r = 6
    ws.row_dimensions[r].height = 14
    c = ws.cell(row=r, column=1, value="Year-End Date")
    c.font = FT_NOTE; c.fill = F_PALE; c.alignment = AL_L
    ws.cell(row=r, column=YR0_COL).fill = F_PALE   # blank for Year 0
    for i, col in enumerate(YR_COLS):
        n = i + 1
        c = frm(ws, r, col,
                f'=IFERROR(DATE(YEAR(\'{_AS}\'!$C$13)+{n},MONTH(\'{_AS}\'!$C$13),DAY(\'{_AS}\'!$C$13)-1),"")',
                fmt=DT, align=AL_C)
        c.fill = F_PALE; c.font = FT_NOTE
    ws.cell(row=r, column=NCOLS).fill = F_PALE

    # ══════════════════════════════════════════════════════════════════════════
    # SECTION 1: REVENUE
    # ══════════════════════════════════════════════════════════════════════════
    r = 7
    sec(ws, r, "REVENUE", col_start=1, ncols=NCOLS)

    # Base Rent: Y1 from Assumptions; Y2-Y10 escalate
    r = 8
    ws.row_dimensions[r].height = 18
    lbl(ws, r, 1, "Base Rent")
    BASE_RENT_ROW = r
    frm(ws, r, 4,   # Year 1 = col 4
        f'=IFERROR(IF(\'{_AS}\'!$C$24="","",\'{_AS}\'!$C$24),"")',
        fmt=D0, align=AL_R)
    for i, col in enumerate(YR_COLS[1:], 2):    # Years 2-10
        prev = get_column_letter(col - 1)
        frm(ws, r, col,
            f'=IFERROR(IF({prev}{r}="","",{prev}{r}*(1+\'{_AS}\'!$C$25)),"")',
            fmt=D0, align=AL_R)
    ws.cell(row=r, column=NCOLS, value="NNN base rent").font = FT_NOTE

    r = 9
    ws.row_dimensions[r].height = 18
    lbl(ws, r, 1, "Tenant Reimbursements (NNN)", bold=False)
    for col in YR_COLS:
        frm(ws, r, col,
            f'=IFERROR(IF(\'{_AS}\'!$C$26="","",\'{_AS}\'!$C$26),"")',
            fmt=D0, align=AL_R)
    ws.cell(row=r, column=NCOLS, value="From Assumptions").font = FT_NOTE

    r = 10
    ws.row_dimensions[r].height = 18
    lbl(ws, r, 1, "Other Income", bold=False)
    for col in YR_COLS:
        inp(ws, r, col, fmt=D0)
        ws.cell(row=r, column=col).alignment = AL_R
    add_cf_clear(ws, f"D10:M10")

    r = 11
    ws.row_dimensions[r].height = 18
    lbl(ws, r, 1, "EFFECTIVE GROSS REVENUE")
    ws.cell(row=r, column=1).fill = F_TOT
    EGR_ROW = r
    for col in YR_COLS:
        cl = get_column_letter(col)
        c = frm(ws, r, col,
                f'=IFERROR(IFERROR({cl}8,0)+IFERROR({cl}9,0)+IFERROR({cl}10,0),"")',
                fmt=D0, align=AL_R)
        c.font = FT_TOTAL; c.fill = F_TOT
        c.border = Border(top=_GY_T, bottom=_NM_T)

    # ══════════════════════════════════════════════════════════════════════════
    # SECTION 2: EXPENSES
    # ══════════════════════════════════════════════════════════════════════════
    r = 12
    ws.row_dimensions[r].height = 6
    r = 13
    sec(ws, r, "EXPENSES", col_start=1, ncols=NCOLS)

    r = 14
    ws.row_dimensions[r].height = 18
    lbl(ws, r, 1, "Management Fee", bold=False)
    for col in YR_COLS:
        cl = get_column_letter(col)
        frm(ws, r, col,
            f'=IFERROR(IF({cl}{EGR_ROW}="","",{cl}{EGR_ROW}*\'{_AS}\'!$C$30),"")',
            fmt=D0, align=AL_R)

    r = 15
    ws.row_dimensions[r].height = 18
    lbl(ws, r, 1, "Real Estate Taxes (LL Responsibility)", bold=False)
    for col in YR_COLS:
        inp(ws, r, col, fmt=D0)
        ws.cell(row=r, column=col).alignment = AL_R
    ws.cell(row=r, column=NCOLS, value="NNN = $0 typically").font = FT_NOTE
    add_cf_clear(ws, "D15:M15")

    r = 16
    ws.row_dimensions[r].height = 18
    lbl(ws, r, 1, "Insurance (LL Responsibility)", bold=False)
    for col in YR_COLS:
        inp(ws, r, col, fmt=D0)
        ws.cell(row=r, column=col).alignment = AL_R
    add_cf_clear(ws, "D16:M16")

    r = 17
    ws.row_dimensions[r].height = 18
    lbl(ws, r, 1, "Capital / Replacement Reserves", bold=False)
    for col in YR_COLS:
        frm(ws, r, col,
            f'=IFERROR(IF(\'{_AS}\'!$C$31="","",\'{_AS}\'!$C$31),"")',
            fmt=D0, align=AL_R)

    r = 18
    ws.row_dimensions[r].height = 18
    lbl(ws, r, 1, "TOTAL OPERATING EXPENSES")
    ws.cell(row=r, column=1).fill = F_TOT
    EXP_ROW = r
    for col in YR_COLS:
        cl = get_column_letter(col)
        c = frm(ws, r, col,
                f'=IFERROR(SUM({cl}14:{cl}17),"")',
                fmt=D0, align=AL_R)
        c.font = FT_TOTAL; c.fill = F_TOT
        c.border = Border(top=_GY_T, bottom=_NM_T)

    # ══════════════════════════════════════════════════════════════════════════
    # SECTION 3: NOI
    # ══════════════════════════════════════════════════════════════════════════
    r = 19
    ws.row_dimensions[r].height = 6
    r = 20
    ws.row_dimensions[r].height = 22
    NOI_ROW = r
    c = ws.cell(row=r, column=1, value="NET OPERATING INCOME (NOI)")
    c.font = Font(name="Calibri", size=11, bold=True, color=NAVY)
    c.fill = F_PALE
    # Do NOT merge — data cells need to be individually writable
    for col in range(1, NCOLS + 1):
        ws.cell(row=r, column=col).fill = F_PALE
    for col in YR_COLS:
        cl = get_column_letter(col)
        c = ws.cell(row=r, column=col,
                    value=f'=IFERROR({cl}{EGR_ROW}-{cl}{EXP_ROW},"")')
        c.number_format = D0
        c.font = Font(name="Calibri", size=11, bold=True, color=NAVY)
        c.alignment = AL_R; c.fill = F_PALE
        c.border = Border(top=_NM_M, bottom=_NM_M)
    ws.cell(row=r, column=NCOLS).fill = F_PALE

    # ══════════════════════════════════════════════════════════════════════════
    # SECTION 4: UNLEVERAGED ANALYSIS
    # ══════════════════════════════════════════════════════════════════════════
    r = 21
    ws.row_dimensions[r].height = 6
    r = 22
    sec(ws, r, "UNLEVERAGED INVESTMENT ANALYSIS", col_start=1, ncols=NCOLS)

    r = 23
    ws.row_dimensions[r].height = 18
    lbl(ws, r, 1, "Purchase Price (Year 0, negative outflow)")
    UL_CF_ROW = r
    c = frm(ws, r, YR0_COL,
            f'=IFERROR(IF(\'{_AS}\'!$C$36="","",-\'{_AS}\'!$C$36),"")',
            fmt=D0, align=AL_R)
    ws.cell(row=r, column=NCOLS, value="Col C = Year 0 outflow").font = FT_NOTE

    r = 24
    ws.row_dimensions[r].height = 18
    lbl(ws, r, 1, "Annual NOI Cash Flows (Years 1-10)")
    UL_NOI_ROW = r
    for col in YR_COLS:
        cl = get_column_letter(col)
        frm(ws, r, col, f'=IFERROR({cl}{NOI_ROW},"")', fmt=D0, align=AL_R)

    r = 25
    ws.row_dimensions[r].height = 18
    lbl(ws, r, 1, "Net Reversion (Year 10)")
    NET_REV_UL_ROW = r
    # Net reversion lives in Assumptions right side I25
    frm(ws, r, 13,   # col M = Year 10
        f'=IFERROR(\'{_AS}\'!$I$25,"")',
        fmt=D0, align=AL_R)
    ws.cell(row=r, column=NCOLS, value="Gross − 5% bkg − 1% costs").font = FT_NOTE

    r = 26
    ws.row_dimensions[r].height = 18
    lbl(ws, r, 1, "UNLEVERAGED CASH FLOW (for IRR)")
    ws.cell(row=r, column=1).fill = F_TOT
    UL_IRR_ROW = r
    # Col C = Year 0 (purchase, negative)
    c = frm(ws, r, YR0_COL,
            f'=IFERROR(C{UL_CF_ROW},"")', fmt=D0, align=AL_R)
    c.fill = F_TOT; c.font = FT_TOTAL
    # Cols D-L = Years 1-9 NOI only
    for col in YR_COLS[:-1]:   # D through L = Y1-Y9
        cl = get_column_letter(col)
        c = frm(ws, r, col, f'=IFERROR({cl}{UL_NOI_ROW},"")', fmt=D0, align=AL_R)
        c.fill = F_TOT; c.font = FT_TOTAL
    # Col M = Year 10 NOI + Net Reversion
    c = frm(ws, r, 13,
            f'=IFERROR(IF(M{UL_NOI_ROW}="","",M{UL_NOI_ROW}+M{NET_REV_UL_ROW}),"")',
            fmt=D0, align=AL_R)
    c.fill = F_TOT; c.font = FT_TOTAL
    for col in [YR0_COL] + YR_COLS:
        ws.cell(row=r, column=col).border = Border(top=_GY_T, bottom=_NM_T)

    # ── Unleveraged Callout Boxes ─────────────────────────────────────────────
    r = 27
    ws.row_dimensions[r].height = 6
    r = 28
    callout_boxes(ws, row=r, cf_row=UL_IRR_ROW, label_prefix="UNLEVERAGED ")
    # Labels at row 28, values at row 29

    # ══════════════════════════════════════════════════════════════════════════
    # SECTION 5: DEBT SERVICE & LEVERAGED ANALYSIS
    # ══════════════════════════════════════════════════════════════════════════
    r = 31
    ws.row_dimensions[r].height = 6
    r = 32
    sec(ws, r, "CASH FLOW AFTER DEBT SERVICE  —  LEVERAGED ANALYSIS", col_start=1, ncols=NCOLS)

    r = 33
    ws.row_dimensions[r].height = 18
    lbl(ws, r, 1, "Net Operating Income")
    DS_NOI_ROW = r
    for col in YR_COLS:
        cl = get_column_letter(col)
        frm(ws, r, col, f'=IFERROR({cl}{NOI_ROW},"")', fmt=D0, align=AL_R)

    r = 34
    ws.row_dimensions[r].height = 18
    lbl(ws, r, 1, "Annual Debt Service", bold=False)
    DS_ROW = r
    for col in YR_COLS:
        frm(ws, r, col,
            f'=IFERROR(IF(\'{_AS}\'!$I$14="","",\'{_AS}\'!$I$14),"")',
            fmt=D0, align=AL_R)
    ws.cell(row=r, column=NCOLS, value="From Assumptions — I14").font = FT_NOTE

    r = 35
    ws.row_dimensions[r].height = 18
    lbl(ws, r, 1, "CASH FLOW AFTER DEBT SERVICE")
    ws.cell(row=r, column=1).fill = F_TOT
    CFADS_ROW = r
    for col in YR_COLS:
        cl = get_column_letter(col)
        c = frm(ws, r, col,
                f'=IFERROR(IF({cl}{DS_NOI_ROW}="","",{cl}{DS_NOI_ROW}-{cl}{DS_ROW}),"")',
                fmt=D0, align=AL_R)
        c.font = FT_TOTAL; c.fill = F_TOT
        c.border = Border(top=_GY_T, bottom=_NM_T)

    r = 36
    ws.row_dimensions[r].height = 18
    lbl(ws, r, 1, "Debt Coverage Ratio (DCR)")
    DCR_ROW = r
    for col in YR_COLS:
        cl = get_column_letter(col)
        frm(ws, r, col,
            f'=IFERROR(IF(OR({cl}{DS_ROW}="",{cl}{DS_ROW}=0),"",{cl}{DS_NOI_ROW}/{cl}{DS_ROW}),"")',
            fmt=MX, align=AL_R)

    r = 37
    ws.row_dimensions[r].height = 6

    # Leveraged IRR inputs
    r = 38
    ws.row_dimensions[r].height = 18
    lbl(ws, r, 1, "Equity Investment (Year 0, negative)")
    EQ_CF_ROW = r
    frm(ws, r, YR0_COL,
        f'=IFERROR(IF(\'{_AS}\'!$I$9="","",-\'{_AS}\'!$I$9),"")',
        fmt=D0, align=AL_R)
    ws.cell(row=r, column=NCOLS, value="Down payment from I9").font = FT_NOTE

    r = 39
    ws.row_dimensions[r].height = 18
    lbl(ws, r, 1, "Net Reversion After Debt Payoff")
    NET_REV_LEV_ROW = r
    frm(ws, r, 13,
        f'=IFERROR(\'{_AS}\'!$I$27,"")',
        fmt=D0, align=AL_R)
    ws.cell(row=r, column=NCOLS, value="Net Rev − loan balance").font = FT_NOTE

    r = 40
    ws.row_dimensions[r].height = 18
    lbl(ws, r, 1, "LEVERAGED CASH FLOW (for IRR)")
    ws.cell(row=r, column=1).fill = F_TOT
    LEV_IRR_ROW = r
    c = frm(ws, r, YR0_COL,
            f'=IFERROR(C{EQ_CF_ROW},"")', fmt=D0, align=AL_R)
    c.fill = F_TOT; c.font = FT_TOTAL
    for col in YR_COLS[:-1]:    # Y1-Y9 CFADS
        cl = get_column_letter(col)
        c = frm(ws, r, col, f'=IFERROR({cl}{CFADS_ROW},"")', fmt=D0, align=AL_R)
        c.fill = F_TOT; c.font = FT_TOTAL
    c = frm(ws, r, 13,   # Y10 CFADS + Net Reversion after debt
            f'=IFERROR(IF(M{CFADS_ROW}="","",M{CFADS_ROW}+M{NET_REV_LEV_ROW}),"")',
            fmt=D0, align=AL_R)
    c.fill = F_TOT; c.font = FT_TOTAL
    for col in [YR0_COL] + YR_COLS:
        ws.cell(row=r, column=col).border = Border(top=_GY_T, bottom=_NM_T)

    # ── Leveraged Callout Boxes ────────────────────────────────────────────────
    r = 41
    ws.row_dimensions[r].height = 6
    r = 42
    callout_boxes(ws, row=r, cf_row=LEV_IRR_ROW, label_prefix="LEVERAGED ")
    # Labels at row 42, values at row 43

    # ══════════════════════════════════════════════════════════════════════════
    # SECTION 6: DISPOSITION ANALYSIS
    # ══════════════════════════════════════════════════════════════════════════
    r = 45
    ws.row_dimensions[r].height = 6
    r = 46
    sec(ws, r, "DISPOSITION ANALYSIS  —  Year 10 Exit (detail reference)", col_start=1, ncols=NCOLS)

    disp_items = [
        (47, "Exit NOI (Year 10)",             f'=IFERROR(M{NOI_ROW},"")',                                  D0,  False),
        (48, "Exit Cap Rate",                   f'=IFERROR(\'{_AS}\'!$I$20,"")',                             P2,  False),
        (49, "GROSS SALE PROCEEDS",             f'=IFERROR(IF(OR(C47="",C48="",C48=0),"",C47/C48),"")',     D0,  True),
        (50, "Less: Brokerage Fee (5%)",        f'=IFERROR(IF(C49="","",-C49*0.05),"")',                     D0,  False),
        (51, "Less: Transaction Costs (1%)",    f'=IFERROR(IF(C49="","",-C49*0.01),"")',                     D0,  False),
        (52, "NET REVERSION (Unleveraged)",     f'=IFERROR(IF(C49="","",C49+C50+C51),"")',                  D0,  True),
        (53, "Less: Remaining Loan Balance",    f'=IFERROR(\'{_AS}\'!$I$26,"")',                             D0,  False),
        (54, "NET REVERSION AFTER DEBT PAYOFF", f'=IFERROR(IF(C52="","",C52-C53),"")',                      D0,  True),
    ]
    for row_n, label, formula, fmt, is_total in disp_items:
        ws.row_dimensions[row_n].height = 18
        lbl(ws, row_n, 1, label, bold=is_total)
        if is_total:
            ws.cell(row=row_n, column=1).fill = F_TOT
        c = frm(ws, row_n, 3, formula, fmt=fmt, align=AL_R)
        if is_total:
            c.font = FT_TOTAL; c.fill = F_TOT
            c.border = Border(top=_GY_T, bottom=_NM_M)

    r = 55
    ws.row_dimensions[r].height = 14
    c = ws.cell(row=r, column=1,
                value="Brokerage: 5%  ·  Transaction Costs: 1%  ·  Gross Proceeds = Exit NOI ÷ Exit Cap Rate  ·  All inputs on Assumptions & Flags tab")
    c.font = FT_NOTE; c.alignment = AL_L
    merge(ws, r, 1, r, NCOLS)

    # ══════════════════════════════════════════════════════════════════════════
    # PRICE AT CAP RATE — dynamic 11-column valuation matrix
    # ══════════════════════════════════════════════════════════════════════════
    # 11 columns centered on going-in cap (C37 ±50bps in 10bp steps)
    _VM_OFFSETS = [-0.005, -0.004, -0.003, -0.002, -0.001, 0.000, 0.001, 0.002, 0.003, 0.004, 0.005]
    _VM_ASK_IDX = 5
    _VM_ASK_COL = 3 + _VM_ASK_IDX   # = column 8 (H)
    _VM_GCAP    = f"'{_AS}'!$C$37"
    _GOLD_HDR   = PatternFill("solid", fgColor="FFC000")
    _GOLD_VAL   = PatternFill("solid", fgColor="FFF2CC")

    r = 57
    ws.row_dimensions[r].height = 6
    r = 58
    sec(ws, r, "PRICE AT CAP RATE  —  Year 1 NOI ÷ Going-In Cap Rate", col_start=1, ncols=NCOLS)

    # Cap rate header row — dynamic formulas referencing C37 ± offset
    r = 59
    ws.row_dimensions[r].height = 18
    c = ws.cell(row=r, column=1, value="Year 1 NOI  →  Price at Going-In Cap Rate:")
    c.font = FT_LABEL; c.alignment = AL_L

    for idx, offset in enumerate(_VM_OFFSETS):
        col = 3 + idx
        if offset == 0.0:
            formula = f'=IFERROR({_VM_GCAP},"")'
        elif offset > 0:
            formula = f'=IFERROR({_VM_GCAP}+{round(offset, 4)},"")'
        else:
            formula = f'=IFERROR({_VM_GCAP}-{round(abs(offset), 4)},"")'
        frm(ws, r, col, formula, fmt=P2, align=AL_C)
        cell = ws.cell(row=r, column=col)
        cell.font = FT_CHDR
        cell.fill = _GOLD_HDR if col == _VM_ASK_COL else F_NAVY

    # Price at cap rate row (NOI ÷ cap)
    r = 60
    ws.row_dimensions[r].height = 18
    c = ws.cell(row=r, column=1, value="Price at Cap Rate (NOI ÷ Cap)")
    c.font = FT_DATA; c.alignment = AL_L
    for idx in range(11):
        col = 3 + idx
        cl  = get_column_letter(col)
        c = frm(ws, r, col,
                f'=IFERROR(IF(OR(D{NOI_ROW}="",{cl}59=0),"",D{NOI_ROW}/{cl}59),"")',
                fmt=D0, align=AL_R)
        if col == _VM_ASK_COL:
            c.fill = _GOLD_VAL

    # Price per SF row — D0 (whole dollar)
    r = 61
    ws.row_dimensions[r].height = 18
    c = ws.cell(row=r, column=1, value="Price per SF")
    c.font = FT_DATA; c.alignment = AL_L
    for idx in range(11):
        col = 3 + idx
        cl  = get_column_letter(col)
        c = frm(ws, r, col,
                f'=IFERROR(IF(OR({cl}60="",\'{_AS}\'!$C$9="",\'{_AS}\'!$C$9=0),"",{cl}60/\'{_AS}\'!$C$9),"")',
                fmt=D0, align=AL_R)
        if col == _VM_ASK_COL:
            c.fill = _GOLD_VAL

    # NM bottom accent
    r = 63
    ws.row_dimensions[r].height = 4
    for ci in range(1, NCOLS + 1):
        ws.cell(row=r, column=ci).border = Border(bottom=Side(style='medium', color=NAVY))
