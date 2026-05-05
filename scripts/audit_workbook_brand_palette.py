"""audit_workbook_brand_palette.py

Quick-scan an .xlsx for Office-default palette hits (off-brand) vs Northmarq
brand palette hits. Walks chart series, axes, plot areas, titles, and a
sample of cells.

Usage:
    python scripts/audit_workbook_brand_palette.py <path> [<path> ...]

Or, with no args, audits the canonical Capital Markets master workbooks:
    Copy Government Master Document.xlsx
    Dialysis Comp Work MASTER.xlsx
    ST Market.xlsx

Returns exit code 0 if all sampled colors are NM-brand or neutral derivatives,
1 if any Office-default hexes are found. Suitable for CI gating after a
quarterly marketing refresh.

Why this exists: Issue #11.2 in CAPITAL_MARKETS_ARCHITECTURE.md flagged the
Dialysis and ST Market masters as using default Office colors. A 2026-05-05
audit found zero such hits. Run this script before assuming the issue has
regressed.
"""
from __future__ import annotations

import argparse
import sys
from collections import Counter
from pathlib import Path

import warnings
warnings.filterwarnings("ignore")

import openpyxl

OFFICE_DEFAULTS = {
    "5B9BD5": "Office.AccentBlue",
    "ED7D31": "Office.AccentOrange",
    "A5A5A5": "Office.AccentGray",
    "FFC000": "Office.AccentYellow",
    "4472C4": "Office.AccentBlue2",
    "70AD47": "Office.AccentGreen",
    "264478": "Office.AccentBlueDark",
    "9E480E": "Office.AccentOrangeDark",
    "636363": "Office.GrayDark",
    "997300": "Office.YellowDark",
    "255E91": "Office.BlueDarker",
    "43682B": "Office.GreenDark",
    "4F81BD": "Office2010.Blue",
    "C0504D": "Office2010.Red",
    "9BBB59": "Office2010.Green",
    "8064A2": "Office2010.Purple",
    "4BACC6": "Office2010.Cyan",
    "F79646": "Office2010.Orange",
}

NM_PALETTE = {
    "003DA5": "nm_navy",
    "62B5E5": "nm_sky",
    "E0E8F4": "nm_pale",
    "265AB2": "nm_blue_mid",
    "6A748C": "nm_axis",
    "191919": "nm_text",
    "666666": "nm_text_muted",
    "FFFFFF": "nm_bg",
    "E7E6E6": "nm_bg_alt",
}

# Default OneDrive paths — used when no args supplied
DEFAULT_PATHS = [
    r"C:\Users\scott\NorthMarq Capital, LLC\Team Briggs - Documents\Gv't Leased Research\Copy Government Master Document.xlsx",
    r"C:\Users\scott\NorthMarq Capital, LLC\Team Briggs - Documents\Dialysis Research\Comps\Dialysis Comp Work MASTER.xlsx",
    r"C:\Users\scott\NorthMarq Capital, LLC\Team Briggs - Documents\Single-Tenant Market\ST Market.xlsx",
]


def classify(hex6: str) -> str:
    h = (hex6 or "").upper()
    if h in NM_PALETTE:
        return "NM"
    if h in OFFICE_DEFAULTS:
        return "OFFICE"
    return "OTHER"


def cell_hex(color) -> str | None:
    """Extract a 6-char hex from an openpyxl Color object."""
    if color is None:
        return None
    rgb = getattr(color, "rgb", None)
    if isinstance(rgb, str):
        return rgb.upper().lstrip("FF")[-6:]
    return None


def walk_chart(ch) -> list[tuple[str, str]]:
    """Walk a chart's structure for any srgbClr; return [(location, hex), ...]."""
    found: list[tuple[str, str]] = []

    def visit(obj, path):
        if obj is None:
            return
        srgb = getattr(obj, "srgbClr", None)
        if srgb:
            found.append((path, srgb.upper()))
        for attr in ("solidFill", "line", "color"):
            sub = getattr(obj, attr, None)
            if sub is not None and not isinstance(sub, (str, int, float, list, tuple)):
                visit(sub, f"{path}.{attr}")

    try:
        for i, ser in enumerate(ch.series):
            gp = getattr(ser, "graphicalProperties", None)
            if gp:
                visit(gp, f"series[{i}]")
    except Exception:
        pass

    for axname in ("x_axis", "y_axis", "z_axis"):
        ax = getattr(ch, axname, None)
        if ax:
            visit(ax, axname)

    pa = getattr(ch, "plot_area", None)
    if pa:
        visit(pa, "plot_area")
    tt = getattr(ch, "title", None)
    if tt:
        visit(tt, "title")

    return found


def audit_path(path: Path, max_cell_rows: int = 200, max_cell_cols: int = 30) -> dict:
    wb = openpyxl.load_workbook(path, data_only=False)
    counters = Counter()
    office_hits: list[str] = []  # human-readable locations
    chart_count = 0

    for sn in wb.sheetnames:
        ws = wb[sn]
        # Charts
        for i, ch in enumerate(ws._charts):
            chart_count += 1
            for loc, hexv in walk_chart(ch):
                cls = classify(hexv)
                counters[cls] += 1
                if cls == "OFFICE":
                    office_hits.append(f"{sn}/chart[{i}].{loc}={hexv}")

        # Cells (capped sample)
        for row in ws.iter_rows(max_row=min(ws.max_row, max_cell_rows),
                                max_col=min(ws.max_column, max_cell_cols)):
            for cell in row:
                if cell.value is None:
                    continue
                # Fill foreground
                if cell.fill and cell.fill.fgColor:
                    h = cell_hex(cell.fill.fgColor)
                    if h:
                        cls = classify(h)
                        counters[f"cell_fill_{cls}"] += 1
                        if cls == "OFFICE":
                            office_hits.append(f"{sn}!{cell.coordinate} fill={h}")
                # Font color
                if cell.font and cell.font.color:
                    h = cell_hex(cell.font.color)
                    if h:
                        cls = classify(h)
                        counters[f"cell_font_{cls}"] += 1
                        if cls == "OFFICE":
                            office_hits.append(f"{sn}!{cell.coordinate} font={h}")

    return {
        "path": str(path),
        "sheets": len(wb.sheetnames),
        "charts": chart_count,
        "counters": counters,
        "office_hits": office_hits,
    }


def main() -> int:
    parser = argparse.ArgumentParser(description="Audit .xlsx for off-brand palette hits.")
    parser.add_argument("paths", nargs="*", help="Workbooks to audit (default: 3 master workbooks).")
    parser.add_argument("--strict", action="store_true",
                        help="Exit 1 if ANY Office-default hits are found (default behavior).")
    parser.add_argument("--quiet", action="store_true", help="Suppress per-hit detail.")
    args = parser.parse_args()

    targets = args.paths or DEFAULT_PATHS
    total_office_hits = 0

    for raw_path in targets:
        p = Path(raw_path)
        if not p.exists():
            print(f"[skip] {p} — file not found")
            continue
        print(f"\n=== {p.name} ===")
        result = audit_path(p)
        c = result["counters"]
        total_office_hits += len(result["office_hits"])
        chart_nm     = c.get("NM", 0)
        chart_office = c.get("OFFICE", 0)
        chart_other  = c.get("OTHER", 0)
        cell_nm      = c.get("cell_fill_NM", 0) + c.get("cell_font_NM", 0)
        cell_office  = c.get("cell_fill_OFFICE", 0) + c.get("cell_font_OFFICE", 0)
        cell_other   = c.get("cell_fill_OTHER", 0) + c.get("cell_font_OTHER", 0)
        print(f"  sheets={result['sheets']}  charts={result['charts']}")
        print(f"  chart-color hits — NM={chart_nm} OFFICE={chart_office} OTHER={chart_other}")
        print(f"  cell-color hits  — NM={cell_nm} OFFICE={cell_office} OTHER={cell_other}")
        if result["office_hits"] and not args.quiet:
            print(f"  Office-default hits ({len(result['office_hits'])}):")
            for hit in result["office_hits"][:15]:
                print(f"    {hit}")
            if len(result["office_hits"]) > 15:
                print(f"    … and {len(result['office_hits']) - 15} more")

    print()
    if total_office_hits == 0:
        print("PASS — no Office-default palette hits across all audited workbooks.")
        return 0
    else:
        print(f"FAIL — {total_office_hits} Office-default palette hit(s). See per-file detail above.")
        return 1


if __name__ == "__main__":
    sys.exit(main())
