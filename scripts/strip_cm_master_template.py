#!/usr/bin/env python
"""
strip_cm_master_template.py

Takes a master Capital Markets Excel (gov or dialysis) and produces a
stripped template suitable for committing to LCC.

Why Python (openpyxl) and not Node (exceljs)?
exceljs hits a parser bug on Excel's table-autofilter date-grouping XML
(<dateGroupItem>) which is widely used in our masters. openpyxl handles
this correctly. After stripping, the output has simpler XML that exceljs
can re-load on the server for population at export time.

What's preserved:
  - All chart objects (positions, axis specs, data range references,
    formatting, brand colors, font choices)
  - Headers (rows 1-2 of every tab)
  - Formula structure on chart-source tabs
  - Brand styling, page setup, named ranges

What's removed:
  - Data rows on bulky transactional tabs (Sold, FRPP-Leased, Ownership,
    Inventory, Top Buyers, Top Sellers, Off Market Comps, Available Coverage)
  - Data rows on chart-source tabs ('All Charts', 'SSA Charts', etc.)
    beyond the structural header rows

Usage:
  python scripts/strip_cm_master_template.py <input.xlsx> <output.xlsx> [vertical]

vertical: 'gov' | 'dialysis'
"""
import warnings
warnings.filterwarnings("ignore")

import sys
import os
import openpyxl

# Tabs to clear bulky data rows from (keep first 2 rows = headers)
BULKY_DATA_TABS = {
    'gov': {'Sold', 'Ownership', 'Inventory', 'FRPP-Leased', 'Top Buyers', 'Top Sellers'},
    'dialysis': {'Sales Comps', 'Available Comps', 'Available Coverage', 'Off Market Comps',
                 'Top Buyers', 'Top Sellers', 'Sheet1'},
}

# Tabs whose data we re-populate at export time. Charts on these tabs
# reference rows below the header — we clear rows past the structural header.
# Format: { sheet_name: keep_rows_through }
CHART_SOURCE_TABS = {
    'gov': {
        'All Charts': 4,   # Keep rows 1-4 (header + summary formulas at row 1)
        'SSA Charts': 4,
    },
    'dialysis': {
        'Charts': 14,        # First 14 rows are summary formulas; data starts at row 15
        'Market Size': 4,
        'Core Cap Chart': 4,
    },
}

# Tabs to keep as-is (small static data, brand pages, etc.)
KEEP_AS_IS = {'Rent Survey', 'Competition'}


def strip_master(input_path: str, output_path: str, vertical: str = 'gov'):
    print(f"Loading {input_path} ...")
    wb = openpyxl.load_workbook(input_path)
    print(f"Sheets: {wb.sheetnames}")

    bulky = BULKY_DATA_TABS.get(vertical, set())
    chart_src = CHART_SOURCE_TABS.get(vertical, {})

    total_cleared = 0
    tables_removed = 0
    filters_cleared = 0

    for name in wb.sheetnames:
        ws = wb[name]
        before = ws.max_row

        if name in bulky:
            keep = 2
            if before > keep:
                ws.delete_rows(keep + 1, before - keep)
                cleared = before - keep
                total_cleared += cleared
                print(f"  [{name}] BULKY: cleared {cleared} data rows, kept {keep} header rows")
            else:
                print(f"  [{name}] BULKY: only {before} rows, nothing to clear")
        elif name in chart_src:
            keep = chart_src[name]
            if before > keep:
                ws.delete_rows(keep + 1, before - keep)
                cleared = before - keep
                total_cleared += cleared
                print(f"  [{name}] CHART SOURCE: cleared {cleared} data rows, kept {keep} structural rows + charts")
            else:
                print(f"  [{name}] CHART SOURCE: only {before} rows, nothing to clear")
        elif name in KEEP_AS_IS:
            print(f"  [{name}] KEEP: {before} rows preserved as-is")
        else:
            print(f"  [{name}] UNKNOWN: {before} rows preserved (review manually)")

        # Strip tables & autoFilters — they reference cleared data ranges and
        # contain dateGroupItem XML elements that exceljs (the server-side
        # library) can't parse. Charts on these sheets are NOT inside tables;
        # they live in drawings/ and reference cell ranges directly.
        if hasattr(ws, '_tables') and ws._tables:
            n_tables = len(ws._tables)
            ws._tables = {}  # type: ignore
            tables_removed += n_tables
            if n_tables > 0:
                print(f"      Removed {n_tables} table(s) (referenced cleared data + dateGroupItem filters)")

        if ws.auto_filter and ws.auto_filter.ref:
            ws.auto_filter.ref = None
            filters_cleared += 1

    print(f"\nTotal data rows cleared: {total_cleared:,}")
    print(f"Tables removed: {tables_removed}")
    print(f"AutoFilters cleared: {filters_cleared}")
    print(f"Saving to {output_path} ...")
    wb.save(output_path)

    in_size = os.path.getsize(input_path) / 1024 / 1024
    out_size = os.path.getsize(output_path) / 1024 / 1024
    reduction = (1 - out_size / in_size) * 100 if in_size else 0
    print(f"Done. Size: {in_size:.2f}MB -> {out_size:.2f}MB ({reduction:.0f}% reduction)")


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Usage: python strip_cm_master_template.py <input.xlsx> <output.xlsx> [vertical=gov]")
        sys.exit(1)
    in_path = sys.argv[1]
    out_path = sys.argv[2]
    vert = sys.argv[3] if len(sys.argv) > 3 else 'gov'
    strip_master(in_path, out_path, vert)
