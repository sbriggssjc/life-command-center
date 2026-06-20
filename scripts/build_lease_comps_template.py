"""build_lease_comps_template.py

Generates the Northmarq-branded dialysis lease-comps XLSX template at
assets/cm-templates/dialysis-lease-comps-template.xlsx.

This is the SINGLE CANONICAL lease-comps template. The property-page
"Export Lease Comps" button populates this template (detail.js
_udExportLeaseComps + the detail-lease-comps-fix.js runtime overrides), and
the Work-Product Framework restyle reads the same canonical output (see
assets/work-product-templates/comps/README.md). The two legacy
"Lease Comps Template - Briggs.xlsx" variants were retired in favor of this
one.

Canonical-merge (2026-06-20):
  - Merged the three genuinely-useful columns the export lacked
    (LEASE TYPE, OPTIONS, NOTES) into the deployed dialysis template, keeping
    every dialysis specific (DISTANCE TO SUBJECT haversine + PATIENTS count)
    and every ExcelJS-compat fix below. New 26-column layout (A..Z):
      A # | B TENANT | C OPERATOR | D ADDRESS | E CITY | F ST | G LAND |
      H BUILT | I RENO | J RBA | K SF LEASED | L OCCUPANCY | M RENT/SF |
      N CURRENT RENT | O COMM | P EXP | Q INITIAL TERM | R TERM REM |
      S LEASE TYPE | T EXPENSES | U BUMPS | V OPTIONS | W USER/OWNER |
      X DISTANCE TO SUBJECT | Y PATIENTS | Z NOTES
    LEASE TYPE (S), OPTIONS (V), NOTES (Z) are all text "@" and carry NO
    AVERAGE in the totals row. DISTANCE (was V) and PATIENTS (was W) shifted
    right to X / Y; their Comps[<name>] AVERAGE refs follow (see avg_formulas
    + _UD_TABLE_COL_MAP in detail-lease-comps-fix.js, which must stay in sync).

Round 76gn.q fix:
  - Abbreviate three header labels (STATE -> ST, RENOVATED -> RENO,
    COMMENCE -> COMM) to match the existing abbreviated style of RBA /
    EXP / TERM REM / RENT/SF / SF LEASED. The bold header
    font renders wider than the raw character count suggests, so the
    full-word versions wrapped mid-word in narrower columns (especially
    column F STATE at width 7, column I RENOVATED at width 11). The
    avg_formula reference for column I also changes from
    Comps[RENOVATED] to Comps[RENO] since Excel Table column names
    derive from the header row cell value, and the JS runtime hot-patch
    (_UD_TABLE_COL_MAP in detail-lease-comps-fix.js) is updated to
    match. Once the eventual sales-comps export reuses this scaffolding
    it inherits the same abbreviations and visual proportions.

Round 76gn.n fix:
  - Remove the "Subject" Excel table entirely. Earlier rounds wrapped
    rows 3-4 (subject section's header + data) in a 2-row Excel Table
    so the Subject cells would render with table styling, but every
    revision of the table's range/header/totals attributes triggered a
    fresh class of "Cannot read properties of undefined" or
    "workbook opens blank, repair fails" errors in ExcelJS + Excel.
    The Subject section is just two rows of formatted cells; it
    doesn't need to be a table. AVERAGE formulas only reference the
    Comps table, not Subject, so dropping the Subject Table object has
    zero impact on output data. Visual styling (navy band, header row
    fills, data row borders) all live on the cells themselves and are
    unaffected.

Round 76gn.m fix:
  - Post-process the saved XLSX to rewrite package-relationship Target
    paths from absolute ("/xl/tables/table1.xml") to relative
    ("../tables/table1.xml"). openpyxl writes the rels with absolute
    paths, but ExcelJS' loader (worksheet-xform.js line 522) keys its
    options.tables dictionary by RELATIVE paths. The mismatch makes the
    table-part lookup return undefined, leaving holes in the loaded
    worksheet's tables[] array; ExcelJS then crashes at worksheet.js:920
    with "Cannot read properties of undefined (reading 'name')" inside
    its Array.reduce. Paired with a runtime sanitize in
    detail-lease-comps-fix.js (_udSanitizeTemplateRels) so existing
    deployed binaries also load cleanly without needing a regen.

Round 76gn.i additions:
  - New PATIENTS column at W (latest_patient_count). Comps table extends
    to W. AVERAGE formula appended for column W.
  - Section band extends to column W so the heading bar covers the new
    column too.

Prior fixes (Round 76gn.f / 76gn.h):
  - Header cell number formats corrected (date format was applied to text headers,
    accounting/currency formats applied to year/occupancy/date columns).
  - Freeze pane moved from B30 to B8 so the comps header stays visible while scrolling.
  - Tab color set to NMQ blue. Header band uses white bold text on NMQ blue.
  - Section title bands ("Subject Property", "Lease Comps") get a navy band.
  - Heading/body use Calibri Light / Calibri per the brand source
    public/reports/cm_brand_tokens.json (was ad-hoc Trebuchet MS / Open Sans).
  - Comps rows alternate warm-white / white fill.

Run from repo root:
    python scripts/build_lease_comps_template.py
"""
from __future__ import annotations

import os
import shutil
import zipfile
from copy import copy
from pathlib import Path

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

NMQ = {
    "blue":      "FF003DA5",
    "navy":      "FF001159",
    "lightBlue": "FF62B5E5",
    "blueTint":  "FFE0E8F4",
    "warmWhite": "FFFAF9F5",
    "bodyText":  "FF3D4A54",
    "darkText":  "FF191919",
    "muted":     "FF6A748C",
    "border":    "FFD8DFDF",
    "white":     "FFFFFFFF",
}

# Excel deliverable fonts — Calibri Light / Calibri, per the documented brand
# source public/reports/cm_brand_tokens.json ("Excel exports must use Calibri
# Light/Calibri to match the existing master workbooks"; Northmarq brand book
# Futura PT primary, Calibri as the deliverable fallback). The prior
# Trebuchet MS / Open Sans values were ad-hoc (not Northmarq brand fonts).
# NOTE: only the Excel export is aligned here; the detail.js HTML/print report
# (NMQ_BRAND) is a separate surface where the brand source permits a different
# web heading font.
HEADING_FONT = "Calibri Light"
BODY_FONT = "Calibri"

OUT_PATH = Path("assets/cm-templates/dialysis-lease-comps-template.xlsx")

# The Work-Product Framework restyle (Dialysis src/comps_restyle.py) reads the
# lease-comps template from this path BY FILENAME. Emit the canonical template
# here too so the framework's lease-comps template IS this single generator
# output — no separate hand-maintained variant to drift. (Replaces the retired
# legacy "Lease Comps Template - Briggs.xlsx" variants.)
FRAMEWORK_OUT_PATH = Path(
    "assets/work-product-templates/comps/Lease Comps Template - Briggs.xlsx"
)

COMP_FIRST_DATA_ROW = 8
COMP_LAST_TEMPLATED_ROW = 40
COMP_TOTAL_ROW = 60
LAST_COL_LETTER = "Z"
LAST_COL_INDEX = 26  # column Z (1-based)

# Header labels intentionally use the abbreviated style established by
# existing columns (RBA, EXP, RENT/SF, TERM REM). Three labels were
# shortened in Round 76gn.q so they don't wrap mid-word under the bold
# header font: STATE -> ST, RENOVATED -> RENO, COMMENCE -> COMM.
# If any of these are renamed, also update:
#   1. avg_formulas[<letter>] below (Excel Table column refs match the
#      header cell value)
#   2. _UD_TABLE_COL_MAP in detail-lease-comps-fix.js (runtime rewrite
#      from Comps[<NAME>] to cell ranges)
COLUMNS = [
    ("counter",      "A", "",                       3.5,  "0",                         False),
    ("tenant",       "B", "TENANT",                25,    "@",                         False),
    ("operator",     "C", "OPERATOR",              22,    "@",                         False),
    ("address",      "D", "ADDRESS",               24,    "@",                         False),
    ("city",         "E", "CITY",                  14,    "@",                         False),
    ("state",        "F", "ST",                     7,    "@",                         True),
    ("land",         "G", "LAND",                  10,    '#,##0.0" ac"',              True),
    ("built",        "H", "BUILT",                  9,    "0",                         True),
    ("renovated",    "I", "RENO",                  11,    "0",                         True),
    ("rba",          "J", "RBA",                   11,    "#,##0",                     True),
    ("sfLeased",     "K", "SF LEASED",             12,    "#,##0",                     True),
    ("occupancy",    "L", "OCCUPANCY",             11,    "0%",                        True),
    ("rentPsf",      "M", "RENT/SF",               11,    '"$"#,##0.00',               True),
    ("current",      "N", "CURRENT RENT",          14,    '"$"#,##0',                  True),
    ("commence",     "O", "COMM",                  11,    "mmm-yy",                    True),
    ("exp",          "P", "EXP",                   11,    "mmm-yy",                    True),
    ("initTerm",     "Q", "INITIAL TERM",          13,    '0.0" Years"',               True),
    ("termRem",      "R", "TERM REM",              13,    '0.0" Years";"EXPIRED";"-"', True),
    # Canonical-merge additions: LEASE TYPE (S) / OPTIONS (V) / NOTES (Z) are
    # text "@" and carry NO AVERAGE. Inserting them shifts EXPENSES/BUMPS/
    # USER-OWNER/DISTANCE/PATIENTS right (see module docstring).
    ("leaseType",    "S", "LEASE TYPE",            13,    "@",                         True),
    ("expenses",     "T", "EXPENSES",              13,    "@",                         True),
    ("bumps",        "U", "BUMPS",                 14,    "@",                         True),
    ("options",      "V", "OPTIONS",               20,    "@",                         False),
    ("userOwner",    "W", "USER/OWNER",            12,    "@",                         True),
    ("distance",     "X", "DISTANCE TO SUBJECT",   16,    '#,##0.0" mi"',              True),
    ("patientCount", "Y", "PATIENTS",              11,    "#,##0",                     True),
    ("notes",        "Z", "NOTES",                 30,    "@",                         False),
]


def build():
    wb = Workbook()
    ws = wb.active
    ws.title = "Lease Comps"
    ws.sheet_properties.tabColor = NMQ["blue"][2:]

    for _, letter, _label, width, _fmt, _center in COLUMNS:
        ws.column_dimensions[letter].width = width

    ws.sheet_view.showGridLines = False

    # Top brand band: row 1 — extends across all columns B..W
    ws.row_dimensions[1].height = 30
    band = ws.cell(row=1, column=2, value="NORTHMARQ  ·  LEASE COMPS")
    band.font = Font(name=HEADING_FONT, size=14, bold=True, color=NMQ["white"])
    band.fill = PatternFill(fill_type="solid", fgColor=NMQ["blue"])
    band.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=LAST_COL_INDEX)
    for col in range(2, LAST_COL_INDEX + 1):
        c = ws.cell(row=1, column=col)
        c.fill = PatternFill(fill_type="solid", fgColor=NMQ["blue"])

    # Subject section band — extends to W so PATIENTS column header is covered.
    ws.row_dimensions[2].height = 24
    sec = ws.cell(row=2, column=2, value="Subject Property")
    sec.font = Font(name=HEADING_FONT, size=13, bold=True, color=NMQ["white"])
    sec.fill = PatternFill(fill_type="solid", fgColor=NMQ["navy"])
    sec.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=LAST_COL_INDEX)
    for col in range(2, LAST_COL_INDEX + 1):
        ws.cell(row=2, column=col).fill = PatternFill(fill_type="solid", fgColor=NMQ["navy"])

    # Subject header + data row — written as ordinary styled cells, NOT wrapped
    # in an Excel Table object (see Round 76gn.n note in module docstring).
    # The full set of columns (excluding counter) is included in the header so
    # the data row's USER/OWNER, DISTANCE, PATIENTS cells render under a
    # labeled column.
    subject_cols = [c for c in COLUMNS if c[0] != "counter"]
    _write_header_row(ws, 3, columns_subset=subject_cols)
    _write_data_row_styles(ws, 4, columns_subset=subject_cols, stripe=False)
    ws.row_dimensions[3].height = 24
    ws.row_dimensions[4].height = 22

    # Comps section
    ws.row_dimensions[5].height = 8
    ws.row_dimensions[6].height = 24
    sec2 = ws.cell(row=6, column=2, value="Lease Comps")
    sec2.font = Font(name=HEADING_FONT, size=13, bold=True, color=NMQ["white"])
    sec2.fill = PatternFill(fill_type="solid", fgColor=NMQ["navy"])
    sec2.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.merge_cells(start_row=6, start_column=2, end_row=6, end_column=LAST_COL_INDEX)
    for col in range(2, LAST_COL_INDEX + 1):
        ws.cell(row=6, column=col).fill = PatternFill(fill_type="solid", fgColor=NMQ["navy"])

    _write_header_row(ws, 7, columns_subset=COLUMNS[1:])
    ws.row_dimensions[7].height = 24

    a7 = ws.cell(row=7, column=1, value="#")
    a7.font = Font(name=HEADING_FONT, size=11, bold=True, color=NMQ["white"])
    a7.fill = PatternFill(fill_type="solid", fgColor=NMQ["blue"])
    a7.alignment = Alignment(horizontal="center", vertical="center")

    a3 = ws.cell(row=3, column=1, value="#")
    a3.font = Font(name=HEADING_FONT, size=11, bold=True, color=NMQ["white"])
    a3.fill = PatternFill(fill_type="solid", fgColor=NMQ["blue"])
    a3.alignment = Alignment(horizontal="center", vertical="center")

    a4 = ws.cell(row=4, column=1, value=1)
    a4.font = Font(name=BODY_FONT, size=10, bold=True, color=NMQ["bodyText"])
    a4.alignment = Alignment(horizontal="center", vertical="center")
    a4.number_format = "0"
    _bottom_border(a4)

    for r in range(COMP_FIRST_DATA_ROW, COMP_LAST_TEMPLATED_ROW + 1):
        ws.row_dimensions[r].height = 20
        cell_a = ws.cell(row=r, column=1)
        if r == COMP_FIRST_DATA_ROW:
            cell_a.value = 1
        else:
            cell_a.value = f"=A{r-1}+1"
        cell_a.font = Font(name=BODY_FONT, size=10, bold=True, color=NMQ["muted"])
        cell_a.alignment = Alignment(horizontal="center", vertical="center")
        cell_a.number_format = "0"

        stripe = (r - COMP_FIRST_DATA_ROW) % 2 == 1
        _write_data_row_styles(ws, r, columns_subset=COLUMNS[1:], stripe=stripe)

    # Totals / averages row
    tr = COMP_TOTAL_ROW
    ws.row_dimensions[tr].height = 22
    label = ws.cell(row=tr, column=2, value="AVERAGE")
    label.font = Font(name=HEADING_FONT, size=11, bold=True, color=NMQ["white"])
    label.fill = PatternFill(fill_type="solid", fgColor=NMQ["navy"])
    label.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.merge_cells(start_row=tr, start_column=2, end_row=tr, end_column=6)
    for col in range(2, 7):
        ws.cell(row=tr, column=col).fill = PatternFill(fill_type="solid", fgColor=NMQ["navy"])

    # Comps[X] structured references — column names must exactly match the
    # header-row cell values (Excel Tables derive column names from headers).
    # If any of these change, update the COLUMNS labels above AND the
    # _UD_TABLE_COL_MAP in detail-lease-comps-fix.js (the runtime hot-patch
    # rewrites Comps[X] to cell ranges before download).
    # NOTE: LEASE TYPE (S) / OPTIONS (V) / NOTES (Z) are text columns and get
    # NO AVERAGE. DISTANCE / PATIENTS shifted to X / Y under the canonical
    # merge — keep these letters in sync with _UD_TABLE_COL_MAP in
    # detail-lease-comps-fix.js.
    avg_formulas = {
        "G": '=IFERROR(SUBTOTAL(101,Comps[LAND]),"")',
        "H": '=IFERROR(SUBTOTAL(101,Comps[BUILT]),"")',
        "I": '=IFERROR(SUBTOTAL(101,Comps[RENO]),"")',
        "J": '=IFERROR(AVERAGE(Comps[RBA]),"")',
        "K": '=IFERROR(AVERAGE(Comps[SF LEASED]),"")',
        "L": '=IFERROR(SUBTOTAL(101,Comps[OCCUPANCY]),"")',
        "M": '=IFERROR(AVERAGE(Comps[RENT/SF]),"")',
        "N": '=IFERROR(AVERAGE(Comps[CURRENT RENT]),"")',
        "Q": '=IFERROR(SUBTOTAL(101,Comps[INITIAL TERM]),"")',
        "R": '=IFERROR(SUBTOTAL(101,Comps[TERM REM]),"")',
        "X": '=IFERROR(AVERAGE(Comps[DISTANCE TO SUBJECT]),"")',
        "Y": '=IFERROR(AVERAGE(Comps[PATIENTS]),"")',
    }
    fmt_by_letter = {letter: fmt for _, letter, _l, _w, fmt, _c in COLUMNS}
    for letter, formula in avg_formulas.items():
        c = ws[f"{letter}{tr}"]
        c.value = formula
        c.font = Font(name=BODY_FONT, size=11, bold=True, color=NMQ["darkText"])
        c.fill = PatternFill(fill_type="solid", fgColor=NMQ["blueTint"])
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.number_format = fmt_by_letter[letter]
        _bottom_border(c, top=True)

    for col in range(7, LAST_COL_INDEX + 1):
        cell = ws.cell(row=tr, column=col)
        if cell.value is None:
            cell.fill = PatternFill(fill_type="solid", fgColor=NMQ["blueTint"])
            _bottom_border(cell, top=True)
        if cell.alignment is None or cell.alignment.horizontal is None:
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # Excel tables — Comps only. The Subject section is intentionally NOT a
    # table (see Round 76gn.n note in module docstring).
    comps_tbl = Table(displayName="Comps", ref=f"B7:{LAST_COL_LETTER}{COMP_TOTAL_ROW}")
    comps_tbl.tableStyleInfo = TableStyleInfo(
        name="TableStyleLight1", showFirstColumn=False, showLastColumn=False,
        showRowStripes=False, showColumnStripes=False
    )
    comps_tbl.totalsRowCount = 1
    # Intentionally NOT setting totalsRowFunction on individual columns:
    # Excel would override the explicit IFERROR(AVERAGE(...)) formulas in row 60
    # with its own auto-computed values on open. The explicit formulas give us
    # control over what to do when a column is entirely empty (return ""
    # instead of #DIV/0!), so we keep them.
    ws.add_table(comps_tbl)

    ws.freeze_panes = "B8"

    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_options.horizontalCentered = True
    ws.page_margins.left = 0.4
    ws.page_margins.right = 0.4
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5
    ws.print_title_rows = "1:7"

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_PATH)
    _patch_rels_paths(OUT_PATH)
    print(f"wrote {OUT_PATH} ({OUT_PATH.stat().st_size} bytes)")

    # Mirror the canonical output to the Work-Product Framework comps folder so
    # the framework template never drifts from the deployed export template.
    FRAMEWORK_OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    shutil.copyfile(OUT_PATH, FRAMEWORK_OUT_PATH)
    print(f"wrote {FRAMEWORK_OUT_PATH} ({FRAMEWORK_OUT_PATH.stat().st_size} bytes)")


def _patch_rels_paths(xlsx_path: Path) -> None:
    """Rewrite package-relationship Target attributes from absolute to relative.

    openpyxl writes xl/worksheets/_rels/sheet1.xml.rels (and workbook.xml.rels)
    with Target attributes pointing at absolute paths inside the package
    ("/xl/tables/table1.xml"). ExcelJS' loader (worksheet-xform.js line 522)
    looks those up in `options.tables` which is keyed by RELATIVE path
    ("../tables/table1.xml"). The mismatch makes the lookup return undefined,
    the worksheet model gets a hole in its tables[] array, and
    Worksheet#model's reduce at worksheet.js:920 crashes with
    "Cannot read properties of undefined (reading 'name')".

    Fix: rewrite the .rels in-place so Targets are relative. This matches what
    Excel itself emits and what ExcelJS expects.
    """
    rels_paths = (
        "xl/worksheets/_rels/sheet1.xml.rels",
        "xl/worksheets/_rels/sheet2.xml.rels",
        "xl/_rels/workbook.xml.rels",
    )
    with zipfile.ZipFile(xlsx_path) as src:
        members = {name: src.read(name) for name in src.namelist()}

    mutated = False
    for path in rels_paths:
        if path not in members:
            continue
        text = members[path].decode("utf-8")
        fixed = (
            text
            .replace('Target="/xl/tables/', 'Target="../tables/')
            .replace('Target="/xl/worksheets/', 'Target="worksheets/')
            .replace('Target="/xl/', 'Target="')
        )
        if fixed != text:
            members[path] = fixed.encode("utf-8")
            mutated = True

    if not mutated:
        return

    with zipfile.ZipFile(xlsx_path, "w", zipfile.ZIP_DEFLATED) as out:
        for name, data in members.items():
            out.writestr(name, data)


def _write_header_row(ws, row, *, columns_subset):
    blue = PatternFill(fill_type="solid", fgColor=NMQ["blue"])
    for _name, letter, label, _w, _fmt, _center in columns_subset:
        c = ws[f"{letter}{row}"]
        c.value = label
        c.font = Font(name=HEADING_FONT, size=11, bold=True, color=NMQ["white"])
        c.fill = blue
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.number_format = "@"


def _write_data_row_styles(ws, row, *, columns_subset, stripe):
    fill_color = NMQ["warmWhite"] if stripe else NMQ["white"]
    bottom = Border(bottom=Side(style="thin", color=NMQ["border"]))
    for _name, letter, _label, _w, fmt, center in columns_subset:
        c = ws[f"{letter}{row}"]
        c.font = Font(name=BODY_FONT, size=10, color=NMQ["bodyText"])
        c.fill = PatternFill(fill_type="solid", fgColor=fill_color)
        c.alignment = Alignment(
            horizontal="center" if center else "left",
            vertical="center",
            indent=0 if center else 1,
        )
        c.number_format = fmt
        c.border = bottom


def _bottom_border(c, *, top=False):
    sides = {"bottom": Side(style="thin", color=NMQ["border"])}
    if top:
        sides["top"] = Side(style="thin", color=NMQ["border"])
    c.border = Border(**sides)


if __name__ == "__main__":
    os.chdir(Path(__file__).resolve().parent.parent)
    build()
